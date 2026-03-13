#!/usr/bin/env python3
"""
classify_races.py — Smart race/training classifier for activity_overrides.xlsx

Detects races from Master file using HR intensity as the primary signal,
with distance-specific HR thresholds (marathons are run at lower HR than 5Ks).

Strategy:
  1. Find all runs at standard race distances (1500m–Marathon) using
     gps_distance_km and strava_distance_km (uncontaminated by prior corrections).
     Tolerance: max(2%, 300m), parkrun: max(3%, 400m), bad GPS: max(4%, 500m).
  2. Classify using HR-for-distance model (sustained 90%+ LTHR at marathon = race)
  3. Boost/demote using activity name keywords (expanded for international events)
  4. Detect parkruns (name, or Saturday ~9am + ~5km)
  5. Detect surface from name keywords (indoor, track) and GPS bbox

Usage:
    python classify_races.py \\
        --master athletes/A005/output/Master_FULL.xlsx \\
        --overrides athletes/A005/activity_overrides.xlsx \\
        --athlete-yml athletes/A005/athlete.yml \\
        --from-master
"""

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill

# ─── Race detection patterns ──────────────────────────────────────────────

# Keywords that strongly indicate a race (expanded for international events)
RACE_KEYWORDS = re.compile(
    r'parkrun|park run|marathon|half marathon|\bhalf\b|\brace\b|championship|league|'
    r'serpentine|LFOTM|cross country|\bXC\b|assembly|relays?\b|'
    r'mob match|interclub|'
    r'vitality|bupa|great run|great north|big half|'
    # Swedish / European race names
    r'loppet|varvet|rusket|midnattsloppet|vinthund|premiärhalv|maraton|'
    # Common race series / organisers
    r"RunThrough|Brooks 5k|CityRun|STHLM 10|Winter 10k|Stockholm's Bästa|"
    r'Harry Hawkes|BMC\b|\bTT\b|mästerskap|'
    # PB/SB mentions (strong race signal)
    r'\bPB[!\s]|\bSB[!\s]',
    re.I
)

# Keywords that indicate training (not a race)
ANTI_KEYWORDS = re.compile(
    r'tempo|recovery|recover|steady|easy|\bwarm.?up\b|cool.?down|treadmill|löpband|'
    r'\bjog\b|fartlek|interval|threshold|long run|out and back|club run|'
    r'zwift|progression|hills?\b|repeat|session|drill|commute|travel|'
    r'shakeout|pre.?match|lunch run|morning run|\bWU\b|\bCD\b|'
    r'sandwich|streak|relaxed|gentle|slow|brisk|progressive|buggy|'
    r'FK Studenterna|FKS\b|\bprep\b|preamble|calibrat',
    re.I
)

# Specific race name matches (abbreviations, event names)
RACE_NAME_OVERRIDES = re.compile(
    r'\bVLM\b|London Marathon|Copenhagen Marathon|Stockholm Marathon|Stockholm maraton|'
    r'Hackney Half|Battersea Park|Djurgårdsvarvet|Höstrusket|Hässelbyloppet|'
    r'Lidingöloppet|Midnattsloppet|2 sjöar|'
    # Additional known races
    r'Royal Parks|Oxford Half|Hampton Court|Göteborgsvarvet|Premiärhalv|'
    r'Broloppet|Örebro|Big Half|Reading Half|Willow 10k|'
    r'Paddock Wood|Chichester|Maidenhead|Beckenham Assembly|'
    r'Middlesex 10k|Kent XC|Surrey league|Titsey|'
    # Mile / 1500m races
    r'Golden Stag|Bannister.*mile|Highgate.*1500|Palladino.*mile|virtual mile',
    re.I
)

# Race keywords excluding parkrun — used to check if "parkrun" was the only match
# at non-5K distances (sandwich runs, longer runs containing a parkrun)
NON_PARKRUN_RACE_KW = re.compile(
    r'marathon|half marathon|\bhalf\b|\brace\b|championship|league|'
    r'serpentine|LFOTM|cross country|\bXC\b|assembly|relays?\b|'
    r'mob match|interclub|vitality|bupa|great run|great north|big half|'
    r'loppet|varvet|rusket|midnattsloppet|vinthund|premiärhalv|maraton|'
    r"RunThrough|Brooks 5k|CityRun|STHLM 10|Winter 10k|Stockholm's Bästa|"
    r'Harry Hawkes|BMC\b|\bTT\b|\bPB[!\s]|\bSB[!\s]',
    re.I
)

# Surface detection from activity name
INDOOR_KEYWORDS = re.compile(r'\bindoor\b|\btreadmill\b|\bgym\b', re.I)
TRACK_KEYWORDS = re.compile(r'\btrack\b|\b[345]000m\b|\b1500m\b|\bmile\b.*\btrack\b|\bopen\s+1500\b|\bHighgate\b.*\b1500\b', re.I)

# Standard race distances: (official_km, label)
# Tolerances are computed dynamically by _distance_tolerance().
RACE_DISTANCES_V2 = [
    (1.5,     "1500m"),
    (1.609,   "Mile"),
    (3.0,     "3K"),
    (5.0,     "5K"),
    (10.0,    "10K"),
    (16.0934, "10M"),
    (21.097,  "HM"),
    (30.0,    "30K"),
    (42.195,  "Marathon"),
]

# Legacy format for backward compatibility (race_type lookup from distance)
RACE_DISTANCES = [(d - _tol, d + _tol, d, l)
                  for d, l in RACE_DISTANCES_V2
                  for _tol in [max(d * 0.02, 0.3)]]


def _distance_tolerance(official_km: float, is_parkrun: bool = False,
                        bad_gps: bool = False) -> float:
    """Compute distance matching tolerance.

    Base:    max(2%, 300m)
    Parkrun: max(3%, 400m)  — GPS-noisy short courses
    Bad GPS: max(4%, 500m)  — high outlier_frac or max_seg > 200m
    """
    if bad_gps:
        return max(official_km * 0.04, 0.5)
    if is_parkrun:
        return max(official_km * 0.03, 0.4)
    return max(official_km * 0.02, 0.3)


def _match_standard_distance(km: float, is_parkrun: bool = False,
                             bad_gps: bool = False):
    """Match a distance to a standard race distance.

    Returns (official_km, label) or (None, None).
    When multiple standard distances are within tolerance (e.g. 1500m vs Mile),
    picks the nearest.
    """
    if km is None or km != km or km <= 0:  # NaN/None/zero check
        return None, None
    best_dist = None
    best_label = None
    best_gap = float('inf')
    for official, label in RACE_DISTANCES_V2:
        tol = _distance_tolerance(official, is_parkrun=is_parkrun,
                                  bad_gps=bad_gps)
        gap = abs(km - official)
        if gap <= tol and gap < best_gap:
            best_dist = official
            best_label = label
            best_gap = gap
    return best_dist, best_label

# ─── HR thresholds by distance ────────────────────────────────────────────
# Default race HR thresholds as %LTHR. Loaded from athlete.yml if present.
# HR alone cannot classify races for all athletes (some train at race HR).
# Keywords do the heavy lifting; HR is the tiebreaker for unnamed runs.
# Named races (keyword/override match) get threshold - 5%.
DEFAULT_RACE_HR_THRESHOLDS = {
    "1500m": 0.98,
    "Mile":  0.98,
    "3K":    0.98,
    "5K":    0.98,
    "10K":   0.97,
    "10M":   0.95,
    "HM":    0.94,
    "30K":   0.90,
    "Marathon": 0.88,
    "Bespoke": 0.95,
}


def load_athlete_config(yml_path: str) -> dict:
    """Load LTHR, max_hr, timezone, and race HR thresholds from athlete.yml."""
    try:
        import yaml
        with open(yml_path) as f:
            cfg = yaml.safe_load(f)
        result = {
            "lthr": cfg["athlete"]["lthr"],
            "max_hr": cfg["athlete"]["max_hr"],
            "timezone": cfg["athlete"].get("timezone", "UTC"),
        }
        # Race HR thresholds: merge athlete overrides with defaults
        athlete_thresholds = cfg["athlete"].get("race_hr_thresholds_pct", {})
        merged = dict(DEFAULT_RACE_HR_THRESHOLDS)
        merged.update(athlete_thresholds)
        result["race_hr_thresholds"] = merged
        return result
    except Exception as e:
        print(f"Warning: Could not read {yml_path}: {e}")
        return {"race_hr_thresholds": dict(DEFAULT_RACE_HR_THRESHOLDS)}


def detect_surface(name: str) -> str:
    """Detect surface type from activity name. Returns None if undetected."""
    if not name:
        return None
    if INDOOR_KEYWORDS.search(name):
        return "indoor"
    if TRACK_KEYWORDS.search(name):
        return "track"
    return None


def is_parkrun(name: str, date_val, dist_km: float, start_hour: float = None) -> bool:
    """Detect parkrun from name, day-of-week, time, and distance.
    
    Parkrun sandwiches (longer run containing a parkrun) are NOT flagged
    as parkrun races — the distance check filters them out.
    """
    if name and re.search(r'parkrun|park run', name, re.I):
        if dist_km and float(dist_km) > 5.1:
            return False  # Sandwich or longer run
        return True
    
    # Saturday + ~5km + parkrun start time (9:00-9:10 or 9:30-9:40)
    if date_val and dist_km:
        try:
            if isinstance(date_val, str):
                dt = datetime.strptime(str(date_val)[:10], '%Y-%m-%d')
            elif hasattr(date_val, 'weekday'):
                dt = date_val
            else:
                return False
            if dt.weekday() == 5 and 4.8 <= float(dist_km) <= 5.1:
                if start_hour is not None:
                    # Parkrun starts: 9:00-9:10 or 9:30-9:40
                    h = float(start_hour)
                    in_window = (9.0 <= h <= 9.17) or (9.5 <= h <= 9.67)
                    if not in_window:
                        return False
                return True
        except (ValueError, TypeError):
            pass
    
    return False


def classify_run(name: str, avg_hr: float, lthr: float, max_hr: float,
                 distance_km: float, duration_min: float,
                 race_type: str,
                 avg_pace: float = None,
                 pred_5k_pace: float = None,
                 race_hr_thresholds: dict = None,
                 is_parkrun_candidate: bool = False) -> tuple:
    """Classify a single candidate as race/training/uncertain.
    
    Primary signal: HR intensity vs distance-specific threshold.
    Secondary signal: activity name keywords.
    Tertiary signal: structural parkrun detection (Saturday + 5km + start time).
    Fallback (no HR): race keyword + pace at least as fast as predicted 5K.
    
    avg_pace: average pace in min/km (used as fallback when HR is missing).
    pred_5k_pace: predicted 5K race pace in min/km (athlete-specific threshold).
    race_hr_thresholds: dict of {distance_label: pct_lthr} from athlete.yml.
    is_parkrun_candidate: True if structurally detected as parkrun (Saturday + ~5km + start time).
    """
    if race_hr_thresholds is None:
        race_hr_thresholds = DEFAULT_RACE_HR_THRESHOLDS
    
    hr_pct = avg_hr / lthr if lthr > 0 and avg_hr > 0 else 0
    
    has_race_kw = bool(RACE_KEYWORDS.search(name))
    has_race_name = bool(RACE_NAME_OVERRIDES.search(name))
    has_anti_kw = bool(ANTI_KEYWORDS.search(name))
    
    # Structural parkrun detection (Saturday + 5km + start time) is equivalent
    # to a race keyword — it's strong evidence this is a race event
    if is_parkrun_candidate and not has_race_kw:
        has_race_kw = True
    
    # "parkrun" in a non-5K run is a sandwich/longer run — don't treat as race keyword
    if has_race_kw and race_type != "5K":
        if not NON_PARKRUN_RACE_KW.search(name) and not is_parkrun_candidate:
            has_race_kw = False  # Only "parkrun" matched — not a race at this distance
    
    min_hr = race_hr_thresholds.get(race_type, 1.00)
    # Named races get a lower HR bar (name is strong evidence)
    named_hr = min_hr - 0.05
    
    # ── Decision tree ──
    
    # 0. No HR data — use pace as proxy for intensity
    #    If pace is at least as fast as predicted 5K race pace AND distance
    #    is a recognised race distance (>=3km), it's a race.
    #    Requires a race keyword to avoid false positives from tempo runs.
    no_hr = (avg_hr <= 0 or hr_pct == 0)
    if no_hr:
        has_fast_race_pace = False
        pace_str = f'{avg_pace:.2f}/km' if avg_pace and avg_pace > 0 else 'unknown'
        if (avg_pace is not None and avg_pace > 0 and
                pred_5k_pace is not None and pred_5k_pace > 0 and
                distance_km >= 2.9):  # 2.9 to allow GPS undershoot on 3000m
            has_fast_race_pace = avg_pace <= pred_5k_pace
            pace_str = f'{avg_pace:.2f}/km vs pred 5K {pred_5k_pace:.2f}/km'
        
        if (has_race_name or (has_race_kw and not has_anti_kw)) and has_fast_race_pace:
            kw_type = 'Named race' if has_race_name else 'Race keyword'
            return ('race', 'medium', f'{kw_type} + fast pace ({pace_str}), no HR')
        # No HR, not clearly fast enough or no keywords — can't classify
        if has_race_kw or has_race_name:
            return ('training', 'low', f'Race keyword but no HR, pace {pace_str} not fast enough')
        return ('training', 'low', f'No HR data, no race keywords')
    
    # 1. Specific race name override (VLM, Stockholm Marathon etc)
    #    Race names beat generic anti-keywords (e.g. "Hackney Half...steady")
    #    UNLESS a decisive training pattern is present (e.g. "VLM week 9")
    #    Pace faster than predicted 5K overrides HR threshold — handles HR ramp-up
    #    in short races (mile, 1500m) where avg HR is dragged down.
    _has_fast_pace = (avg_pace is not None and avg_pace > 0 and
                      pred_5k_pace is not None and pred_5k_pace > 0 and
                      avg_pace < pred_5k_pace)
    _pace_str = (f'pace {avg_pace:.2f} < pred 5K {pred_5k_pace:.2f}/km'
                 if _has_fast_pace else '')
    if has_race_name:
        is_training_context = bool(re.search(
            r'week\s*\d|sandwich|calibrat|shakeout|pre.?match|preamble|'
            r'\bprep\b|\bsim\b|warm.?up.*cool.?down|\bWU\b.*\bCD\b',
            name, re.I))
        if is_training_context:
            # Fall through to later steps
            pass
        elif hr_pct >= named_hr:
            return ('race', 'high', f'Named race + HR {hr_pct*100:.0f}%LTHR')
        elif _has_fast_pace:
            return ('race', 'high', f'Named race + {_pace_str}')
        else:
            return ('training', 'medium', f'Named race but HR {hr_pct*100:.0f}%LTHR below race threshold')
    
    # 2. Race keyword + no anti-keyword — use FULL HR threshold (no discount)
    #    If HR is just below threshold AND pace confirms race effort (faster than
    #    predicted 5K), classify as race. But HR must be within 5% of threshold —
    #    a parkrun at 88% LTHR is not a race even if pace is fast.
    if has_race_kw and not has_anti_kw:
        if hr_pct >= min_hr:
            return ('race', 'high', f'Race keyword + HR {hr_pct*100:.0f}%LTHR')
        elif _has_fast_pace and hr_pct >= min_hr - 0.05:
            return ('race', 'medium', f'Race keyword + {_pace_str} (HR {hr_pct*100:.0f}%LTHR just below threshold)')
        else:
            return ('training', 'medium', f'Race keyword but HR {hr_pct*100:.0f}%LTHR below race threshold')
    
    # 3. Anti-keyword only → training IF HR is below threshold
    #    If HR is above race threshold despite anti-keyword, fall through to
    #    step 5 where HR + pace/distance can classify it (e.g. "Tempo 42km" at
    #    race HR is a marathon, not a tempo run)
    if has_anti_kw and not has_race_kw and not has_race_name:
        if hr_pct < min_hr:
            return ('training', 'high', f'Training keyword: "{_first_match(ANTI_KEYWORDS, name)}", HR {hr_pct*100:.0f}%LTHR')
    
    # 4. Both race + anti keywords → check context
    if has_race_kw and has_anti_kw:
        # Decisive race keywords override any anti-keyword (e.g. FKS championship)
        decisive_race = re.search(r'championship|mästerskap', name, re.I)
        if decisive_race and hr_pct >= min_hr:
            return ('race', 'high', f'Decisive race keyword: "{decisive_race.group()}" + HR {hr_pct*100:.0f}%LTHR')
        # Decisive anti-keywords override race keywords
        decisive_anti = re.search(
            r'sandwich|calibrat|shakeout|pre.?match|preamble|'
            r'\btempo\b|\bsession\b|\binterval|FK Studenterna|FKS\b|'
            r'\bWU\b.*\bCD\b|\bCD\b.*\bWU\b|warm.?up.*cool.?down|'
            r'\bstreaks?\b|\bprep\b|buggy',
            name, re.I)
        if decisive_anti:
            return ('training', 'high', f'Decisive training keyword: "{decisive_anti.group()}"')
        if hr_pct >= min_hr:
            return ('race', 'medium', f'Mixed keywords, HR {hr_pct*100:.0f}%LTHR suggests race')
        else:
            return ('training', 'medium', f'Mixed keywords, HR {hr_pct*100:.0f}%LTHR suggests training')
    
    # 5. No keywords — HR thresholds + pace as co-signal
    #    a) Faster than predicted 5K pace at ANY distance → race (strongest pace signal,
    #       you can't run faster than 5K prediction without racing)
    #    b) HR + pace together: HR >= base threshold AND pace within 5% of distance-
    #       adjusted prediction → race (two co-signals confirming race effort)
    #    c) HR alone: require base threshold + 4% uplift (high bar, no other evidence)
    #    d) Pace rejection: even with high HR, reject if pace is >5% slower than predicted
    UNNAMED_HR_UPLIFT_BASE = 0.04  # +4% LTHR for HR-only at 5K
    UNNAMED_PACE_RATIO = 1.05  # >5% slower than predicted race pace → not race pace
    # Scale uplift by distance: longer races need less extra HR evidence
    # 5K: +4%, 10K: +3.5%, HM: +2.5%, Marathon: +1.5%, 30K+: +1%
    _dist_uplift_scale = max(0.25, 1.0 - (distance_km - 5.0) / 50.0)  # linear from 1.0 at 5K to 0.25 at 42K+
    UNNAMED_HR_UPLIFT = UNNAMED_HR_UPLIFT_BASE * _dist_uplift_scale
    
    # Compute pace evidence once
    _has_pace_data = (avg_pace and pred_5k_pace and pred_5k_pace > 0)
    _faster_than_5k_pred = False
    _pace_confirms_race = False
    _pace_rejects_race = False
    if _has_pace_data:
        # Faster than predicted 5K pace (raw, no distance scaling)
        _faster_than_5k_pred = avg_pace < pred_5k_pace
        # Distance-adjusted pace comparison
        STANDARD_KMS = {'5K': 5, '10K': 10, 'HM': 21.097, 'Marathon': 42.195,
                        '3K': 3, '1500m': 1.5, 'Mile': 1.609, '10M': 16.093, '30K': 30}
        dist_for_type = STANDARD_KMS.get(race_type, distance_km)
        dist_factor = (dist_for_type / 5.0) ** 0.06 if dist_for_type > 0 else 1.0
        expected_race_pace = pred_5k_pace * dist_factor
        pace_ratio = avg_pace / expected_race_pace if expected_race_pace > 0 else 1.0
        _pace_confirms_race = pace_ratio <= UNNAMED_PACE_RATIO
        _pace_rejects_race = pace_ratio > UNNAMED_PACE_RATIO
    
    # Path A: HR above base threshold + pace confirms → race
    #   Pace confirmation: either faster than predicted 5K (strongest) or within
    #   5% of distance-adjusted prediction. Either way, HR must also be above
    #   the base threshold — pace alone (without keywords) is not sufficient
    #   because interval sessions can average fast pace with low avg HR.
    if hr_pct >= min_hr and (_faster_than_5k_pred or _pace_confirms_race):
        pace_detail = (f'faster than pred 5K {pred_5k_pace:.2f}/km'
                       if _faster_than_5k_pred
                       else f'race pace ratio {pace_ratio:.2f}')
        return ('race', 'medium', f'HR {hr_pct*100:.0f}%LTHR ≥ {min_hr*100:.0f}% + {pace_detail} (unnamed) for {race_type} — REVIEW')
    
    # Path B: HR above uplifted threshold (HR alone is strong enough)
    if hr_pct >= min_hr + UNNAMED_HR_UPLIFT:
        if _pace_rejects_race:
            # At long distances (HM+), sustained high HR IS the race signal.
            # Nobody sustains 92% LTHR for 2+ hours in training.
            # Pace rejection only applies at shorter distances where a hard tempo
            # at race HR but slower pace is plausible.
            # Also override if HR is >5% above the base threshold — at any distance,
            # HR that far above threshold is a race regardless of pace.
            hr_margin = hr_pct - min_hr
            is_long_race = distance_km >= 20
            hr_strongly_above = hr_margin >= 0.04  # 4%+ above base threshold
            if is_long_race or hr_strongly_above:
                return ('race', 'medium',
                        f'HR {hr_pct*100:.0f}%LTHR ≥ {min_hr*100:.0f}% + '
                        f'{hr_margin*100:.1f}% margin (pace slower than predicted but '
                        f'{"sustained effort at " + f"{distance_km:.0f}km" if is_long_race else "HR strongly above threshold"}'
                        f') — REVIEW')
            return ('training', 'medium',
                    f'HR {hr_pct*100:.0f}%LTHR but pace {avg_pace:.2f} min/km is '
                    f'{(pace_ratio-1)*100:.0f}% slower than expected race pace '
                    f'{expected_race_pace:.2f} — no keywords to override')
        return ('race', 'medium', f'HR {hr_pct*100:.0f}%LTHR ≥ {(min_hr+UNNAMED_HR_UPLIFT)*100:.0f}% (unnamed) for {race_type} — REVIEW')
    
    return ('training', 'medium', f'HR {hr_pct*100:.0f}%LTHR < {(min_hr+UNNAMED_HR_UPLIFT)*100:.0f}% (unnamed) for {race_type}')


def _first_match(pattern, text):
    """Return the first regex match from text, for readable reasons."""
    m = pattern.search(text)
    return m.group(0) if m else ""


def detect_candidates_from_master(master_df: pd.DataFrame, lthr: float,
                                  max_hr: float) -> pd.DataFrame:
    """Find ALL runs at standard race distances that could be races.

    Distance matching uses gps_distance_km and strava_distance_km (both
    immune to prior official_distance_km corrections) with dynamic
    tolerances:
      - Base:    max(2%, 300m)
      - Parkrun: max(3%, 400m) for 5K when parkrun candidate
      - Bad GPS: max(4%, 500m) when gps_max_seg_m > 200 or outlier_frac > 0.5%

    If either GPS or Strava distance matches a standard distance → candidate.

    HR above 82% LTHR, OR no HR data (classify_run handles no-HR via pace).
    No pace filter. The classify_run() step handles the actual classification.
    """
    df = master_df.copy()
    has_gps_col = 'gps_distance_km' in df.columns
    has_strava_col = 'strava_distance_km' in df.columns
    has_max_seg = 'gps_max_seg_m' in df.columns
    has_outlier = 'gps_outlier_frac' in df.columns

    candidates = []
    for _, row in df.iterrows():
        hr = row.get('avg_hr', 0) or 0
        name = str(row.get('activity_name', '') or '')

        # Gather uncontaminated distance measurements
        gps_km = row['gps_distance_km'] if has_gps_col else None
        if gps_km is not None and (gps_km != gps_km or gps_km <= 0):
            gps_km = None
        strava_km = row['strava_distance_km'] if has_strava_col else None
        if strava_km is not None and (strava_km != strava_km or strava_km <= 0):
            strava_km = None

        # ── Virtual run exclusion ──
        # Strava 'Virtual Run' = Zwift/treadmill with unreliable distance. Skip entirely.
        _stype = str(row.get('strava_activity_type', '') or '').lower()
        if 'virtual' in _stype:
            continue
        
        # ── Fix B: Treadmill exclusion ──
        # Treadmill runs have no GPS trace. Keyword + no GPS → skip entirely.
        # Indoor track races (real events indoors) are NOT excluded — they still
        # have race keywords and get classified normally with distance from watch.
        _is_treadmill = False
        if gps_km is None and strava_km is None:
            _treadmill_kw = re.search(
                r'treadmill|\bTM\b|löpband|\bLB\b',
                name, re.I)
            if _treadmill_kw:
                _is_treadmill = True
        if _is_treadmill:
            continue

        # Fallback: if neither GPS nor Strava available, use distance_km
        # (may be corrected on re-runs but better than skipping entirely)
        dist_km = row.get('distance_km')
        if gps_km is None and strava_km is None:
            if dist_km is not None and dist_km == dist_km and dist_km > 0:
                gps_km = dist_km  # treat as GPS-equivalent
            else:
                continue

        # Detect GPS quality issues
        max_seg = row.get('gps_max_seg_m', 0) if has_max_seg else 0
        outlier_frac = row.get('gps_outlier_frac', 0) if has_outlier else 0
        max_seg = max_seg if (max_seg is not None and max_seg == max_seg) else 0
        outlier_frac = outlier_frac if (outlier_frac is not None and outlier_frac == outlier_frac) else 0
        bad_gps = (max_seg > 200) or (outlier_frac > 0.005)

        # Detect parkrun candidate (for wider 5K tolerance)
        date_val = row.get('date')
        _is_parkrun_candidate = is_parkrun(name, date_val,
                                           gps_km or strava_km or dist_km)

        # Try matching each distance source against standard race distances
        matched = False
        official_dist = None
        race_type = None

        # ── Fix A: Assembly League → always bespoke distance ──
        # Assembly League courses are nominally ~5K but actual distances vary
        # significantly (4.7-5.5km). Forcing them to standard 5K distorts
        # the athlete's race data. Use actual GPS distance instead.
        _is_assembly = bool(re.search(r'assembly', name, re.I))
        if _is_assembly:
            best_dist_km = gps_km or strava_km or dist_km or 0
            if best_dist_km >= 1.0:
                candidates.append({
                    'file': row['file'],
                    'official_distance_km': round(best_dist_km, 3),
                    'race_type': 'Bespoke',
                })
            continue

        for km in [gps_km, strava_km]:
            if km is None:
                continue
            odist, label = _match_standard_distance(
                km,
                is_parkrun=_is_parkrun_candidate,
                bad_gps=bad_gps)
            if label:
                matched = True
                official_dist = odist
                race_type = label
                break

        # ── Parkrun name override ──
        # If the name says "parkrun" but GPS distance fell outside tolerance
        # (e.g. 5.7km from GPS wander, or 4.5km from cutting corners),
        # force-match to 5.0km. This catches GPS-distorted parkruns that
        # is_parkrun() rejected as "sandwiches" due to the >5.1km cutoff.
        # Range 4.0–7.0km: below 4.0 is likely a different activity,
        # above 7.0 is likely a genuine sandwich/longer run.
        if not matched and name and re.search(r'parkrun|park run', name, re.I):
            best_km = gps_km or strava_km or 0
            if 4.0 <= best_km <= 7.0:
                matched = True
                official_dist = 5.0
                race_type = '5K'

        if not matched:
            # ── Bespoke distance detection ──
            # Runs at non-standard distances with race-effort HR (>= 95% LTHR)
            # become bespoke candidates at their actual GPS distance.
            # Examples: Stockholm's Bästa series, Lidingöloppet 15, relay legs.
            #
            # Bespoke candidates MUST have a race keyword or race name override.
            # HR alone at a non-standard distance produces too many false
            # positives — club sessions, tempo runs, and interval workouts
            # routinely exceed 95% LTHR for athletes whose training and racing
            # HR ranges overlap.
            # classify_run() applies keyword/anti-keyword filtering downstream.
            BESPOKE_HR_PCT = 0.95
            BESPOKE_MIN_DIST_KM = 1.0
            BESPOKE_MIN_DURATION_S = 180  # 3 minutes
            has_hr = hr > 0
            elapsed = row.get('elapsed_time_s', 0) or 0
            best_dist_km = gps_km or strava_km or 0
            has_race_signal = bool(
                RACE_KEYWORDS.search(name) or RACE_NAME_OVERRIDES.search(name)
            )
            if (has_hr and lthr > 0
                    and hr >= lthr * BESPOKE_HR_PCT
                    and best_dist_km >= BESPOKE_MIN_DIST_KM
                    and elapsed >= BESPOKE_MIN_DURATION_S
                    and has_race_signal):
                candidates.append({
                    'file': row['file'],
                    'official_distance_km': round(best_dist_km, 3),
                    'race_type': 'Bespoke',
                })
            continue

        # Minimum HR floor — below 82% LTHR nothing is a race in calibration data
        # (Easy parkruns at 83-88% are caught by parkrun name detection instead)
        # Exception: no HR data (hr == 0 or NaN) — let through for pace-based classification
        has_hr = hr > 0
        if has_hr and lthr > 0 and hr < lthr * 0.82:
            continue

        candidates.append({
            'file': row['file'],
            'official_distance_km': official_dist,
            'race_type': race_type,
        })

    return pd.DataFrame(candidates)


def enrich_and_classify(master_path: str, overrides_path: str,
                        lthr: float, max_hr: float,
                        from_master: bool = False,
                        timezone: str = "UTC",
                        race_hr_thresholds: dict = None) -> str:
    """Main function: detect, classify, enrich, and write overrides."""
    print(f"Loading Master: {master_path}")
    master = pd.read_excel(master_path)
    print(f"  {len(master)} runs, {master['date'].min().date()} to {master['date'].max().date()}")
    
    master_lookup = master.set_index('file').to_dict('index')
    
    # ── Load or generate overrides ──
    if from_master:
        print(f"\nDetecting race candidates from Master...")
        candidates = detect_candidates_from_master(master, lthr, max_hr)
        print(f"  {len(candidates)} candidates at race distances")
        
        ov = candidates[['file', 'official_distance_km', 'race_type']].copy()
        ov['race_flag'] = 1
        ov['parkrun'] = 0
        ov['surface'] = None
        ov['surface_adj'] = None
        ov['temp_override'] = None
        ov['power_override_w'] = None
        ov['official_time_s'] = None
        ov['notes'] = ''
    else:
        print(f"\nLoading overrides: {overrides_path}")
        ov = pd.read_excel(overrides_path, dtype={'file': str})
        print(f"  {len(ov)} rows")
        
        rename_map = {}
        if 'race' in ov.columns and 'race_flag' not in ov.columns:
            rename_map['race'] = 'race_flag'
        if 'distance_km' in ov.columns and 'official_distance_km' not in ov.columns:
            rename_map['distance_km'] = 'official_distance_km'
        if 'temp_c' in ov.columns and 'temp_override' not in ov.columns:
            rename_map['temp_c'] = 'temp_override'
        if rename_map:
            ov.rename(columns=rename_map, inplace=True)
        
        if 'race_flag' in ov.columns:
            ov['race_flag'] = ov['race_flag'].astype(int)
        if 'parkrun' in ov.columns:
            ov['parkrun'] = ov['parkrun'].fillna(0).astype(int)
        
        if 'race_type' not in ov.columns:
            ov['race_type'] = None
            for i, row in ov.iterrows():
                dist = row.get('official_distance_km', 0)
                if dist:
                    for odist, label in RACE_DISTANCES_V2:
                        if abs(dist - odist) < 0.5:
                            ov.at[i, 'race_type'] = label
                            break
    
    # ── Build per-run predicted 5K pace lookup ──
    # Used for: (a) no-HR fallback, (b) named race pace override (HR ramp-up in short races)
    # Prefer per-run predictions from Master_*_FULL_post (shift(1) values = prediction at race time)
    # Support both old (Master_FULL.xlsx) and new (Master_A001_FULL.xlsx) naming
    _pred_5k_by_file = {}
    _pred_source_df = master  # default
    _post_path = master_path.replace('_FULL.xlsx', '_FULL_post.xlsx')
    if Path(_post_path).exists() and _post_path != master_path:
        try:
            _post_df = pd.read_excel(_post_path, sheet_name=0)
            if any(c in _post_df.columns for c in ['pred_5k_s_gap', 'pred_5k_s_sim', 'pred_5k_s']):
                _pred_source_df = _post_df
        except Exception:
            pass
    _pred_cols = [c for c in ['pred_5k_s_gap', 'pred_5k_s_sim', 'pred_5k_s']
                  if c in _pred_source_df.columns]
    if _pred_cols:
        for _, _pr in _pred_source_df.iterrows():
            _pf = str(_pr.get('file', ''))
            if not _pf:
                continue
            # Per-row fallback: take first non-NaN prediction column
            for _pc in _pred_cols:
                _pv = _pr.get(_pc)
                if pd.notna(_pv) and _pv > 0:
                    _pred_5k_by_file[_pf] = (_pv / 60) / 5.0  # seconds → min/km
                    break
        _src_name = 'Master_FULL_post' if _pred_source_df is not master else 'Master_FULL'
        print(f"  5K pace lookup: {len(_pred_5k_by_file)} runs with predictions ({', '.join(_pred_cols)} in {_src_name})")
    else:
        print("  No 5K prediction column available — pace-based classification disabled")

    # ── Classify ──
    print(f"\nClassifying {len(ov)} candidates...")
    enriched_cols = {
        'date': [], 'activity_name': [], 'actual_dist_km': [],
        'avg_hr': [], 'hr_pct_lthr': [], 'duration_min': [],
        'verdict': [], 'confidence': [], 'reason': []
    }
    
    race_count = training_count = uncertain_count = 0
    
    for _, row in ov.iterrows():
        fname = row['file']
        m = master_lookup.get(fname, {})
        
        date = m.get('date')
        name = str(m.get('activity_name', '') or '')
        dist = m.get('distance_km', 0) or 0
        hr_raw = m.get('avg_hr', 0)
        hr = 0 if (hr_raw is None or (isinstance(hr_raw, float) and pd.isna(hr_raw))) else hr_raw
        elapsed = m.get('elapsed_time_s', 0) or 0
        duration = elapsed / 60
        pace = m.get('avg_gap_pace_min_per_km', 0) or 0
        if not pace:
            pace = m.get('avg_pace_min_per_km', 0) or 0
        race_type = row.get('race_type', '') or ''
        
        enriched_cols['date'].append(
            date.strftime('%Y-%m-%d') if hasattr(date, 'strftime') else str(date)[:10] if date else '')
        enriched_cols['activity_name'].append(name)
        enriched_cols['actual_dist_km'].append(round(dist, 2) if dist else None)
        enriched_cols['avg_hr'].append(int(hr) if hr else None)
        enriched_cols['hr_pct_lthr'].append(round(hr / lthr * 100, 0) if hr and lthr else None)
        enriched_cols['duration_min'].append(round(duration, 1) if duration else None)
        
        # Detect structural parkrun (Saturday + ~5km + start time)
        _start_hour = None
        if hasattr(date, 'hour'):
            _start_hour = date.hour + date.minute / 60
        _parkrun_cand = is_parkrun(name, date, dist, _start_hour)
        
        verdict, conf, reason = classify_run(name, hr, lthr, max_hr, dist, duration, race_type,
                                               avg_pace=pace,
                                               pred_5k_pace=_pred_5k_by_file.get(fname),
                                               race_hr_thresholds=race_hr_thresholds,
                                               is_parkrun_candidate=_parkrun_cand)
        
        enriched_cols['verdict'].append(f"{verdict} ({conf})")
        enriched_cols['confidence'].append(conf)
        enriched_cols['reason'].append(reason)
        
        if verdict == 'race':
            race_count += 1
        elif verdict == 'training':
            training_count += 1
        else:
            uncertain_count += 1
    
    for col, values in enriched_cols.items():
        ov[col] = values
    
    # ── Apply verdicts to race_flag ──
    for i, row in ov.iterrows():
        v = row['verdict']
        if 'training' in v:
            ov.at[i, 'race_flag'] = 0
        elif 'race (high)' in v:
            ov.at[i, 'race_flag'] = 1
        elif 'race (medium)' in v:
            ov.at[i, 'race_flag'] = 1
            if 'REVIEW' in row.get('reason', ''):
                ov.at[i, 'notes'] = f"REVIEW: {row['reason']}"
        elif 'uncertain' in v:
            ov.at[i, 'race_flag'] = 0
            ov.at[i, 'notes'] = f"REVIEW: {row['reason']}"
    
    # ── Deflag virtual/Zwift races ──
    # Strava 'Virtual Run' activity type = treadmill with unreliable distance.
    # These should not count as races regardless of HR/keyword signals.
    virtual_count = 0
    for i, row in ov.iterrows():
        if row.get('race_flag', 0) != 1:
            continue
        fname = row['file']
        m = master_lookup.get(fname, {})
        _stype = str(m.get('strava_activity_type', '') or '').lower()
        if 'virtual' in _stype:
            ov.at[i, 'race_flag'] = 0
            ov.at[i, 'notes'] = f"Virtual run (Strava type: {_stype}) — deflagged"
            virtual_count += 1
    if virtual_count:
        print(f"  Deflagged {virtual_count} virtual/Zwift race(s)")
    
    # ── Demote races with large elapsed/moving ratio ──
    # If elapsed_time >> moving_time, the activity has significant pauses
    # (rest intervals, relay exchanges, etc.), suggesting reps not a race.
    #
    # Two thresholds:
    #   - Named races (keywords): generous 1.50 ratio (relay exchanges, split recordings)
    #   - Unnamed runs: strict 1.20 ratio (no keywords = higher evidence bar)
    # Plus bbox guard: if bbox < 100,000m² for dist >= 3km, the run is on
    # a tiny loop (reps, track session) regardless of elapsed/moving ratio.
    ELAPSED_MOVING_MAX_RATIO = 1.50        # for named races
    ELAPSED_MOVING_UNNAMED_RATIO = 1.20    # for unnamed runs (no keywords)
    BBOX_MIN_FOR_RACE_M2 = 100_000         # minimum bbox for a real race >= 3km
    demoted_count = 0
    for i, row in ov.iterrows():
        if row.get('race_flag', 0) != 1:
            continue
        fname = row['file']
        m = master_lookup.get(fname, {})
        elapsed = m.get('elapsed_time_s', 0) or 0
        moving = m.get('moving_time_s', 0) or 0
        name_for_demote = str(m.get('activity_name', '') or '')
        _has_kw = bool(RACE_KEYWORDS.search(name_for_demote) or
                       RACE_NAME_OVERRIDES.search(name_for_demote))
        
        # Elapsed/moving ratio check
        if moving > 0 and elapsed > 0:
            ratio = elapsed / moving
            threshold = ELAPSED_MOVING_MAX_RATIO if _has_kw else ELAPSED_MOVING_UNNAMED_RATIO
            if ratio > threshold:
                ov.at[i, 'race_flag'] = 0
                ov.at[i, 'notes'] = f"REVIEW: elapsed/moving ratio {ratio:.2f}x > {threshold:.2f} ({'named' if _has_kw else 'unnamed'}) — likely reps/session"
                demoted_count += 1
                continue
        
        # Bbox guard: tiny area + medium distance = reps on a loop
        bbox = m.get('gps_bbox_m2')
        dist_km = m.get('distance_km', 0) or 0
        if (bbox is not None and pd.notna(bbox) and bbox < BBOX_MIN_FOR_RACE_M2
                and dist_km >= 3.0 and not _has_kw):
            ov.at[i, 'race_flag'] = 0
            ov.at[i, 'notes'] = f"REVIEW: bbox {bbox:.0f}m² < {BBOX_MIN_FOR_RACE_M2} for {dist_km:.1f}km — likely reps on loop"
            demoted_count += 1
            continue
    if demoted_count:
        print(f"  Demoted {demoted_count} race(s) with elapsed/moving > {ELAPSED_MOVING_MAX_RATIO}x")
    
    # ── Detect parkruns ──
    # parkrun=1 marks the event (always set if it's a parkrun).
    # race_flag is set independently by the HR-based classification above.
    # A run can be parkrun=1, race_flag=0 (jogged/buggy parkrun).
    parkrun_count = 0
    for i, row in ov.iterrows():
        fname = row['file']
        m = master_lookup.get(fname, {})
        name = str(m.get('activity_name', '') or '')
        date_val = m.get('date')
        dist = m.get('distance_km')
        start_hour = None
        if hasattr(date_val, 'hour'):
            start_hour = date_val.hour + date_val.minute / 60
        if is_parkrun(name, date_val, dist, start_hour):
            ov.at[i, 'parkrun'] = 1
            # race_flag already set by classify_run — don't override
            parkrun_count += 1
        elif name and re.search(r'parkrun|park run', name, re.I):
            # Name says parkrun but is_parkrun() rejected it (GPS distance
            # outside tolerance). If official_distance is 5.0 (set by the
            # parkrun name override in detect_candidates), still flag it.
            if row.get('official_distance_km') == 5.0:
                ov.at[i, 'parkrun'] = 1
                parkrun_count += 1
    
    # ── Detect surface from activity name ──
    surface_count = 0
    for i, row in ov.iterrows():
        if pd.notna(row.get('surface')) and row['surface']:
            continue
        fname = row['file']
        m = master_lookup.get(fname, {})
        name = str(m.get('activity_name', '') or '')
        detected = detect_surface(name)
        if detected:
            ov.at[i, 'surface'] = detected
            surface_count += 1
    
    # ── Detect track races from GPS bounding box ──
    # A race on a 400m outdoor track has a GPS footprint of ~8,000 m².
    # A 200m indoor track is ~1,800 m². Road races are 100,000+ m².
    # For flagged races at track distances (3K/5K/10K) with no surface set:
    #   - No GPS data (NaN bbox)    → INDOOR_TRACK (indoors = no GPS signal)
    #   - bbox < 5,000 m²           → INDOOR_TRACK (200m oval)
    #   - bbox < 50,000 m²          → TRACK (400m oval)
    # Skip entirely if gps_bbox_m2 is not in the Master (e.g. GAP-mode athletes).
    INDOOR_BBOX_THRESHOLD_M2 = 5_000
    TRACK_BBOX_THRESHOLD_M2 = 50_000
    TRACK_RACE_DISTANCES = {'1500m', 'Mile', '3K', '5K', '10K'}
    bbox_track_count = 0
    bbox_indoor_count = 0
    has_bbox = 'gps_bbox_m2' in master.columns if master is not None else False
    if has_bbox:
        # Only trust NaN-as-indoor if the column is actually populated for most runs
        # (i.e. the pipeline computes bbox). If >80% NaN, the column is just empty.
        bbox_fill_rate = master['gps_bbox_m2'].notna().mean() if master is not None else 0
        if bbox_fill_rate < 0.2:
            has_bbox = False
            print(f"  GPS bbox detection skipped — gps_bbox_m2 only {bbox_fill_rate:.0%} populated")
    if not has_bbox:
        print("  GPS bbox detection skipped — gps_bbox_m2 not in Master")
    else:
        for i, row in ov.iterrows():
            if pd.notna(row.get('surface')) and row['surface']:
                continue  # Surface already set (by name or manual override)
            if row.get('race_flag', 0) != 1:
                continue  # Only check flagged races
            race_type = row.get('race_type', '')
            if race_type not in TRACK_RACE_DISTANCES:
                continue  # Only track-plausible distances
            fname = row['file']
            m = master_lookup.get(fname, {})
            bbox = m.get('gps_bbox_m2')
            if bbox is None or pd.isna(bbox) or bbox == 0:
                # No GPS data — indoor venue (GPS doesn't work indoors)
                ov.at[i, 'surface'] = 'indoor'
                bbox_indoor_count += 1
            elif bbox < INDOOR_BBOX_THRESHOLD_M2:
                ov.at[i, 'surface'] = 'indoor'
                bbox_indoor_count += 1
            elif bbox < TRACK_BBOX_THRESHOLD_M2:
                ov.at[i, 'surface'] = 'track'
                bbox_track_count += 1
    if bbox_track_count or bbox_indoor_count:
        surface_count += bbox_track_count + bbox_indoor_count
        print(f"  Track detected from GPS bbox: {bbox_track_count} outdoor, {bbox_indoor_count} indoor")
    
    # ── Clear official_distance_km for non-races ──
    # Candidates at race distances get official_distance_km during detection,
    # but only actual races (race_flag=1) should keep it. Without this,
    # StepB applies distance corrections to training runs, distorting speed/RE/RF.
    non_race_dist_cleared = 0
    for i, row in ov.iterrows():
        if row.get('race_flag', 0) != 1 and pd.notna(row.get('official_distance_km')):
            ov.at[i, 'official_distance_km'] = None
            non_race_dist_cleared += 1
    if non_race_dist_cleared:
        print(f"  Cleared official_distance_km from {non_race_dist_cleared} non-race rows")

    # ── Sort by date ──
    ov.sort_values('date', inplace=True)
    
    # ── Summary ──
    final_races = (ov['race_flag'] == 1).sum()
    final_training = (ov['race_flag'] == 0).sum()
    review_count = ov['notes'].str.contains('REVIEW', na=False).sum()
    
    print(f"\n{'='*60}")
    print(f"  Classification results")
    print(f"{'='*60}")
    print(f"  Race (high confidence):     {race_count}")
    print(f"  Training:                   {training_count}")
    print(f"  Uncertain:                  {uncertain_count}")
    print(f"{'─'*60}")
    print(f"  Final race_flag=1:          {final_races}")
    print(f"  Final race_flag=0:          {final_training}")
    print(f"  Parkruns detected:          {parkrun_count}")
    print(f"  Surface detected:           {surface_count}")
    print(f"  Rows marked REVIEW:         {review_count}")
    print(f"{'='*60}")
    
    # ── Write output ──
    output_path = overrides_path
    core_cols = ['file', 'race_flag', 'parkrun', 'official_distance_km',
                 'surface', 'surface_adj', 'temp_override', 'power_override_w', 'official_time_s']
    enrich_cols = ['date', 'activity_name', 'actual_dist_km', 'avg_hr',
                   'hr_pct_lthr', 'duration_min', 'verdict', 'reason', 'notes']
    
    for c in core_cols + enrich_cols:
        if c not in ov.columns:
            ov[c] = None
    
    out_cols = core_cols + enrich_cols
    ov_out = ov[out_cols].copy()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "overrides"
    
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    review_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    race_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    training_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    
    for col_idx, header in enumerate(out_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
    
    for row_idx, (_, row) in enumerate(ov_out.iterrows(), 2):
        for col_idx, col in enumerate(out_cols, 1):
            val = row[col]
            if isinstance(val, (np.integer,)):
                val = int(val)
            elif isinstance(val, (np.floating,)):
                val = float(val) if not np.isnan(val) else None
            elif pd.isna(val):
                val = None
            ws.cell(row=row_idx, column=col_idx, value=val)
        
        verdict = str(row.get('verdict', ''))
        notes = str(row.get('notes', '') or '')
        if 'REVIEW' in notes:
            for c in range(1, len(out_cols) + 1):
                ws.cell(row=row_idx, column=c).fill = review_fill
        elif 'race' in verdict:
            for c in range(1, len(out_cols) + 1):
                ws.cell(row=row_idx, column=c).fill = race_fill
        elif 'training' in verdict:
            for c in range(1, len(out_cols) + 1):
                ws.cell(row=row_idx, column=c).fill = training_fill
    
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 60)
    
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    
    wb.save(output_path)
    print(f"\nSaved: {output_path}")
    print(f"  Green = race, Red = training, Yellow = REVIEW needed")
    
    return output_path


def _backfill_parkruns(ov, master_path, overrides_path):
    """Lightweight parkrun detection — runs even when full classification is skipped."""
    if 'parkrun' not in ov.columns:
        ov['parkrun'] = 0
    
    try:
        master = pd.read_excel(master_path)
    except Exception as e:
        print(f"  Could not load Master for parkrun backfill: {e}")
        return
    
    master_lookup = {}
    for _, row in master.iterrows():
        fname = str(row.get('file', ''))
        if fname:
            master_lookup[fname] = {
                'activity_name': str(row.get('activity_name', '') or ''),
                'date': row.get('date', ''),
                'distance_km': row.get('distance_km', None),
            }
    
    filled = 0
    for i, row in ov.iterrows():
        if row.get('parkrun', 0) == 1:
            continue
        fname = str(row.get('file', ''))
        m = master_lookup.get(fname, {})
        name = m.get('activity_name', '')
        date_val = m.get('date', '')
        dist = m.get('distance_km', None)
        if is_parkrun(name, date_val, dist):
            ov.at[i, 'parkrun'] = 1
            if 'race_flag' in ov.columns:
                ov.at[i, 'race_flag'] = 1
            filled += 1
    
    if filled > 0:
        ov.to_excel(overrides_path, index=False)
        print(f"  Parkrun backfill: marked {filled} new parkrun(s).")
    else:
        print(f"  Parkrun backfill: no new parkruns found.")


def main():
    p = argparse.ArgumentParser(description="Smart race classifier for activity overrides")
    p.add_argument("--master", required=True, help="Path to Master_FULL.xlsx")
    p.add_argument("--overrides", required=True, help="Path to activity_overrides.xlsx")
    p.add_argument("--athlete-yml", help="Path to athlete.yml (for LTHR/max HR)")
    p.add_argument("--lthr", type=int, help="Lactate threshold HR (overrides athlete.yml)")
    p.add_argument("--max-hr", type=int, help="Max HR (overrides athlete.yml)")
    p.add_argument("--from-master", action="store_true",
                   help="Generate candidates from Master instead of reading existing overrides")
    p.add_argument("--skip-if-classified", action="store_true",
                   help="Skip if overrides already has verdict column (preserves athlete edits)")
    args = p.parse_args()
    
    if args.skip_if_classified and Path(args.overrides).exists():
        try:
            existing = pd.read_excel(args.overrides)
            skip = False
            if 'verdict' in existing.columns:
                print(f"Overrides already classified (has 'verdict' column) — skipping.")
                skip = True
            elif 'race_flag' in existing.columns and existing['race_flag'].notna().any():
                n_races = (existing['race_flag'] == 1).sum()
                print(f"Overrides already has {n_races} race flags set — skipping.")
                skip = True
            if skip:
                _backfill_parkruns(existing, args.master, args.overrides)
                return
        except Exception:
            pass
    
    lthr = args.lthr
    max_hr = args.max_hr
    timezone = "UTC"
    race_hr_thresholds = None
    
    if args.athlete_yml and (lthr is None or max_hr is None):
        cfg = load_athlete_config(args.athlete_yml)
        if lthr is None:
            lthr = cfg.get("lthr", 160)
        if max_hr is None:
            max_hr = cfg.get("max_hr", 185)
        timezone = cfg.get("timezone", "UTC")
        race_hr_thresholds = cfg.get("race_hr_thresholds")
    
    if lthr is None:
        lthr = 160
        print(f"Warning: using default LTHR={lthr}")
    if max_hr is None:
        max_hr = 185
        print(f"Warning: using default max_hr={max_hr}")
    
    print(f"HR thresholds: LTHR={lthr}, max HR={max_hr}")
    if race_hr_thresholds:
        print(f"Race HR thresholds (from athlete.yml): { {k: f'{v:.0%}' for k,v in race_hr_thresholds.items()} }")
    else:
        print(f"Race HR thresholds: defaults")
    
    enrich_and_classify(args.master, args.overrides, lthr, max_hr,
                        from_master=args.from_master, timezone=timezone,
                        race_hr_thresholds=race_hr_thresholds)


if __name__ == "__main__":
    main()
