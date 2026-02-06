# File: rebuild_from_fit_zip_v49.py
# Purpose: Rebuild master log from FIT ZIP (+ optional per-second cache).
# Date: 2026-02-05
#
# Changelist v49 (2026-02-05):
# - Reduced shrinkage on power_adjuster_to_S4 regression:
#   * Measured eras (v1, repl, air): k=20 (was 120). Data gets ~60-95% weight.
#   * Simulated eras (pre_stryd, v1_late): k=60 (was 120). More cautious.
#   * S5: k=50 (was 300). Small sample but measured.
#   * Previous k=120 was crushing real calibration offsets (v1: 8.2% -> 1.8%)
#
# Changelist v45 (2026-01-31):
#
# Changelist v44 (2026-01-27):
# - Consolidated Stryd eras: removes arbitrary "early/late" splits
#   * Now uses 6 actual hardware eras: pre_stryd, v1, repl, air, s4, s5
#   * Era transition dates are configurable via stryd_era_config.json
#   * Eliminates artificial step changes in power_adjuster_to_S4
#   * Default dates based on Paul's known pod transition dates
#
# Changelist v41.2 (2026-01-23):
# - Sort by date fix: ensures output is always sorted chronologically
#   * Critical for append/UPDATE mode where new runs were appended at end unsorted
#   * Sort applied in both normal path and early-exit path (no new FITs)
# - Pending activities support: temporary activity names before Strava sync
#   * New arg: --pending-activities (CSV with file,activity_name,shoe columns)
#   * Auto-uses pending_activities.csv in output directory if present
#   * Names from pending file used when no Strava match, before falling back to generic time-based names
#   * Allows clearing file after Strava data refresh
# - Encoding fix: pending_activities.csv now handles Windows cp1252 encoding
# - Strava columns preserved: no longer drops strava_moving_time_s, strava_elapsed_time_s,
#   strava_elev_gain_m, strava_elev_loss_m after using them for canonical overrides
#
# Changelist v41.1 (2026-01-22):
# - Cache file naming: now uses sanitized FIT filename instead of epoch timestamp
#   * e.g., "524737937055_token_AAAAAGlrWM6ylmLr.npz" instead of "1737108436.npz"
#   * Special chars (&, ?, =) replaced with underscore for filesystem compatibility
# - Added fit_filename to cache metadata for easier identification
# - GPS speed spike filter: filters unrealistic GPS speeds before power simulation
#   * Caps speed at 7.5 m/s (2:13/km pace)
#   * Filters accelerations > 3 m/s² (physically impossible for running)
#   * Interpolates bad samples from surrounding good data
#   * Prevents garbage simulated power from GPS glitches
#
# Changelist v41 (2026-01-21):
# - Added strava_distance_km column: Strava's corrected distance for better matching
#   with legacy data and override file generation.
# - This allows matching on user-corrected distances rather than raw FIT/Stryd distances.
#
# Changelist v39 (2026-01-20):
# - Incremental cache-update mode: update per-second cache only for newly added runs (no full rebuild).
#   New args:
#     --cache-only                Only write/update per-second cache (.npz); skip master rebuild.
#     --cache-incremental         Skip runs that already have a cache file (default True).
#     --cache-rewrite-if-newer    Rewrite cache if the FIT file timestamp is newer than the cache file.
#     --cache-force               Always rewrite cache.

# Source: rebuild_from_fit_zip_v36_fix1_postfill_REavg_from_REn.py (provided by ChatGPT on 2026-01-15)
# Note: This file may be renamed locally; see Source line for provenance.

# - Weather stage speed-up: batch by rounded location (lat/lon) and fetch one hourly block per location.
# - Weather stage progress output: prints per-location fetch progress + per-run fill progress.
# - Fix time-weighting bug causing AttributeError("numpy.timedelta64" has no total_seconds).
# - Very recent runs: Open-Meteo archive can lag; fallback to Forecast API with past_days for the last ~7 days.

# ------------------------------------------------------------------------------
\
#!/usr/bin/env python
"""
Rebuild master log from a ZIP of FIT files (v37: loop-aware grade detrend for altitude drift) (+ optional Strava activities.csv).

v27 changes:
- Adds zero-setup per-run weather via Open-Meteo archive API (time-weighted averages):
    avg_temp_c, avg_humidity_pct, avg_wind_ms, avg_wind_dir_deg
- Weather cached under <out_dir>/_weather_cache_openmeteo for fast reruns.

v21 changes:
- RE_avg and RE_normalised are now computed using GPS-derived speed (gps_distance_km / moving_time_s) instead of device speed.

v19 changes vs v18:
- Extracts Stryd-related device identifiers (when present) from FIT:
    stryd_manufacturer
    stryd_product
    stryd_serial_number
    stryd_ant_device_number
  These are appended after the GPS quality columns in the output.
- Strategy:
    Parse all device_info messages and keep those that look like Stryd (manufacturer/product text contains "stryd"),
    OR those that match Garmin "connect iq" / "development" cases are ignored.
    If nothing obvious says Stryd, we fall back to choosing the device_info record whose serial/device number
    appears most consistently across the file and isn't the Garmin watch (best-effort).
  Note: FITs often do NOT include explicit "Stryd 4.0/5.0" model strings; these fields help identify *pod A vs pod B*.

Retained:
- GPS quality diagnostics: gps_distance_km, gps_coverage, gps_distance_ratio, gps_max_seg_m, gps_p99_speed_mps, gps_outlier_frac
- Extracts FITs into fixed folder: <out_dir>/_fit_extract and clears only *.fit each run
- HR correction: strap-step drop detection + power:HR ratio back-projection (15s rolling median)
- RF: v12 window selector with dead-time = (power < 2.5 W/kg) OR (speed < 0.3 m/s) plus timestamp-gap dead time
- RF fallback to nPower:HR when RF adjusted values are NaN
- Strava join: match by UTC start time + distance tolerance, adds Activity Name + Shoe (emoji/icons stripped)
- Output: date | activity_name | shoe | ... | file (file moved to end), session_type removed
- RF columns rounded to 3 decimals; column A widened; progress line count; output sorted by date

Usage:
  python rebuild_from_fit_zip_with_strava_v19_device_ids.py --fit-zip "history.zip" --template "Master_Rebuilt.xlsx" --strava "activities.csv" --out "Master_History_GPSQ_ID.xlsx"
"""

import argparse
import datetime as dt
import random
import time
import os
import re
import zipfile
import json
import hashlib
import sqlite3
from pathlib import Path



# -----------------------------
# Weather date-range helpers
# -----------------------------
def _clamp_end_to_today(end_date):
    """Clamp an end date (date/datetime/ISO string) to today (local date). Returns a `datetime.date`."""
    from datetime import date, datetime
    if end_date is None:
        return date.today()
    if isinstance(end_date, datetime):
        d = end_date.date()
    elif isinstance(end_date, date):
        d = end_date
    else:
        try:
            d = datetime.fromisoformat(str(end_date)[:10]).date()
        except Exception:
            # If we can't parse, safest is today (prevents future queries)
            return date.today()
    today = date.today()
    return d if d <= today else today


def _split_date_ranges(start_date: str, end_date: str, max_days: int = 366):
    """
    Split inclusive ISO date range [start_date, end_date] into chunks of at most `max_days` days.
    Returns list of (start_iso, end_iso) pairs.
    """
    from datetime import datetime, timedelta
    s = datetime.fromisoformat(str(start_date)[:10]).date()
    e = datetime.fromisoformat(str(end_date)[:10]).date()
    if e < s:
        return []
    out = []
    cur = s
    while cur <= e:
        chunk_end = min(e, cur + timedelta(days=max_days - 1))
        out.append((cur.isoformat(), chunk_end.isoformat()))
        cur = chunk_end + timedelta(days=1)
    return out

"""Optional dependency handling.

The rebuild script can call Open-Meteo over HTTP. Earlier versions used the
third-party `requests` package, but some Python installs (like yours) don't
have it.

We therefore support both:
  - `requests` if installed
  - a standard-library `urllib.request` fallback otherwise
"""

try:
    import requests  # type: ignore
except ModuleNotFoundError:
    requests = None

import urllib.parse
import urllib.request
import urllib.error
from copy import copy
from datetime import timezone
from zoneinfo import ZoneInfo

import numpy as np
import traceback

# --- v37 fix2: track per-second cache write health ---
CACHE_WRITE_OK = 0
CACHE_WRITE_FAIL = 0
FIRST_CACHE_WRITE_ERROR = None
FIRST_CACHE_WRITE_ERROR_PATH = None

# --- v39: incremental cache-update controls (set from CLI in main) ---
CACHE_INCREMENTAL = True
CACHE_REWRITE_IF_NEWER = True
CACHE_FORCE = False

def _should_write_cache(cache_path: str, fit_path: str) -> bool:
    """Return True if we should (re)write the per-second cache for this FIT."""
    try:
        if CACHE_FORCE:
            return True
        if not CACHE_INCREMENTAL:
            return True
        if not os.path.exists(cache_path):
            return True
        if CACHE_REWRITE_IF_NEWER:
            try:
                fit_mtime = os.path.getmtime(fit_path)
                cache_mtime = os.path.getmtime(cache_path)
                if fit_mtime > cache_mtime + 1.0:
                    return True
            except Exception:
                # If timestamps can't be compared, play safe: rewrite
                return True
        return False
    except Exception:
        return True

def _safe_nanmean(x):
    """Safe nanmean: returns float mean ignoring NaNs, or np.nan if empty/all-NaN."""
    # Convert to numpy array (handles lists/Series/Index etc.)
    arr = np.asarray(x, dtype=float) if x is not None else np.asarray([], dtype=float)
    if arr.size == 0:
        return np.nan
    # Fast path: if all NaN, avoid RuntimeWarning from np.nanmean
    if np.isnan(arr).all():
        return np.nan
    return float(np.nanmean(arr))

import pandas as pd

# -----------------------------
# Weather SQLite cache (hourly)
# -----------------------------
def _wx_db_init(db_path: str):
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    con = sqlite3.connect(db_path)
    try:
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS wx_hourly (
                lat_r REAL NOT NULL,
                lon_r REAL NOT NULL,
                time TEXT NOT NULL,
                temperature_2m REAL,
                relative_humidity_2m REAL,
                wind_speed_10m REAL,
                wind_direction_10m REAL,
                PRIMARY KEY (lat_r, lon_r, time)
            )
            """
        )
        con.execute("CREATE INDEX IF NOT EXISTS ix_wx_time ON wx_hourly(time)")
        con.commit()
    finally:
        con.close()


def _wx_db_get_hourly(db_path: str, lat_r: float, lon_r: float, t_start: pd.Timestamp, t_end: pd.Timestamp) -> pd.DataFrame:
    """Fetch cached hourly rows for [t_start-1h, t_end+1h] (UTC-naive)."""
    if not db_path:
        return pd.DataFrame()
    try:
        con = sqlite3.connect(db_path)
    except Exception:
        return pd.DataFrame()
    try:
        # pad 1h to support weighting and boundary inclusion
        a = (pd.Timestamp(t_start) - pd.Timedelta(hours=1)).strftime("%Y-%m-%dT%H:%M")
        b = (pd.Timestamp(t_end) + pd.Timedelta(hours=1)).strftime("%Y-%m-%dT%H:%M")
        q = """
            SELECT time, temperature_2m, relative_humidity_2m, wind_speed_10m, wind_direction_10m
            FROM wx_hourly
            WHERE lat_r=? AND lon_r=? AND time>=? AND time<=?
            ORDER BY time
        """
        rows = con.execute(q, (float(lat_r), float(lon_r), a, b)).fetchall()
        if not rows:
            return pd.DataFrame()
        dfh = pd.DataFrame(rows, columns=["time","temperature_2m","relative_humidity_2m","wind_speed_10m","wind_direction_10m"])
        dfh["time"] = pd.to_datetime(dfh["time"], errors="coerce")
        dfh = dfh.dropna(subset=["time"]).sort_values("time").drop_duplicates(subset=["time"], keep="last").reset_index(drop=True)
        return dfh
    finally:
        con.close()


def _wx_db_upsert_hourly(db_path: str, lat_r: float, lon_r: float, dfh: pd.DataFrame):
    """Upsert hourly rows into SQLite cache."""
    if not db_path or dfh is None or dfh.empty:
        return
    dfh2 = dfh.copy()
    dfh2["time"] = pd.to_datetime(dfh2["time"], errors="coerce")
    dfh2 = dfh2.dropna(subset=["time"])
    if dfh2.empty:
        return
    dfh2["time_s"] = dfh2["time"].dt.strftime("%Y-%m-%dT%H:%M")
    recs = list(
        zip(
            [float(lat_r)] * len(dfh2),
            [float(lon_r)] * len(dfh2),
            dfh2["time_s"].tolist(),
            pd.to_numeric(dfh2["temperature_2m"], errors="coerce").tolist(),
            pd.to_numeric(dfh2["relative_humidity_2m"], errors="coerce").tolist(),
            pd.to_numeric(dfh2["wind_speed_10m"], errors="coerce").tolist(),
            pd.to_numeric(dfh2["wind_direction_10m"], errors="coerce").tolist(),
        )
    )
    con = sqlite3.connect(db_path)
    try:
        con.executemany(
            """
            INSERT INTO wx_hourly(lat_r, lon_r, time, temperature_2m, relative_humidity_2m, wind_speed_10m, wind_direction_10m)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(lat_r, lon_r, time) DO UPDATE SET
                temperature_2m=excluded.temperature_2m,
                relative_humidity_2m=excluded.relative_humidity_2m,
                wind_speed_10m=excluded.wind_speed_10m,
                wind_direction_10m=excluded.wind_direction_10m
            """,
            recs,
        )
        con.commit()
    finally:
        con.close()


from fitparse import FitFile
import openpyxl
from openpyxl.utils import get_column_letter


# ---------- constants ----------


class SkipFitFile(Exception):
    """Raised to skip non-running or otherwise unsupported FIT files."""
    pass

WEIGHT_KG_DEFAULT = 76.0
PWR_DEAD_WKG = 2.5
SPD_DEAD_MS = 0.3

RF_TARGET_LEN_S = 1200
RF_MIN_WINDOW_S = 300
RF_MIN_VALID_SECONDS = 60
RF_DEAD_FRAC_PASSES = (0.20, 0.35)
BACKWARD_STEP_S = 5
FORWARD_STEP_S = 5



# ---------- weather (Open-Meteo, zero setup) ----------
# Uses Open-Meteo APIs (no key required for normal use).
# Primary: Archive/Historical endpoint (can lag for very recent dates).
# Fallback: Forecast endpoint with past_days for recent runs.
OPEN_METEO_ARCHIVE = "https://archive-api.open-meteo.com/v1/archive"
OPEN_METEO_FORECAST = "https://api.open-meteo.com/v1/forecast"

WEATHER_COLS = ["avg_temp_c", "avg_humidity_pct", "avg_wind_ms", "avg_wind_dir_deg"]

# Round GPS coordinates to 1 decimal place (~10 km) for weather grouping.
# This clusters nearby runs into the same location, reducing API calls from
# ~1900 unique 3-dp locations to ~20-30 groups.  Weather doesn't vary
# meaningfully within 10 km, so this is safe.
WX_GPS_ROUND_DP = 1




def apply_weather_overrides(df: pd.DataFrame, override_file: str) -> tuple[pd.DataFrame, int]:
    """Apply weather overrides from activity_overrides.xlsx for indoor runs.

    Reads the override file and finds rows where temp_override is set.
    For those runs, sets:
      avg_temp_c = temp_override value
      avg_humidity_pct = 50% (indoor default)
      weather_source = 'override'
      is_indoor = True
    """
    if not override_file:
        return df, 0
    p = Path(override_file)
    if not p.exists():
        return df, 0
    if "file" not in df.columns:
        return df, 0

    try:
        ov = pd.read_excel(p, dtype={'file': str})
    except Exception:
        return df, 0
    if "file" not in ov.columns or "temp_override" not in ov.columns:
        return df, 0

    # Only rows with a temp_override value
    ov = ov.dropna(subset=["file"]).copy()
    ov["file"] = ov["file"].astype(str)
    ov = ov[ov["file"].str.len() > 0]
    ov = ov[ov["temp_override"].notna()]
    ov = ov.drop_duplicates(subset=["file"], keep="last")

    if len(ov) == 0:
        return df, 0

    if "weather_source" not in df.columns:
        df["weather_source"] = ""
    if "is_indoor" not in df.columns:
        df["is_indoor"] = False

    df["file"] = df["file"].astype(str)
    override_map = dict(zip(ov["file"], ov["temp_override"].astype(float)))

    applied = 0
    for i, row in df.iterrows():
        fname = str(row.get("file", ""))
        if fname in override_map:
            df.at[i, "avg_temp_c"] = override_map[fname]
            df.at[i, "avg_humidity_pct"] = 50.0
            df.at[i, "weather_source"] = "override"
            df.at[i, "is_indoor"] = True
            applied += 1

    return df, applied
def _delta_seconds(delta) -> float:
    """Convert timedelta-like (pandas/py/np) to seconds."""
    if delta is None:
        return 0.0
    if hasattr(delta, "total_seconds"):
        return float(delta.total_seconds())
    # numpy.timedelta64
    try:
        return float(delta / np.timedelta64(1, "s"))
    except Exception:
        return 0.0


def _circular_mean_deg(deg: np.ndarray, weights: np.ndarray) -> float:
    deg = np.asarray(deg, float)
    weights = np.asarray(weights, float)
    rad = np.deg2rad(deg)
    s = float(np.sum(weights * np.sin(rad)))
    c = float(np.sum(weights * np.cos(rad)))
    if s == 0.0 and c == 0.0:
        return float("nan")
    return float((np.degrees(np.arctan2(s, c)) + 360.0) % 360.0)


def _time_weighted_mean(values: np.ndarray, times: np.ndarray, start: pd.Timestamp, end: pd.Timestamp) -> tuple[float, float]:
    """
    Piecewise-constant weighting between hourly timestamps.
    times must be pandas Timestamps (or convertible).
    Returns (mean, covered_seconds).
    """
    if len(values) == 0 or len(times) == 0:
        return float("nan"), 0.0

    # Ensure pandas timestamps
    times = pd.to_datetime(times)
    start = pd.Timestamp(start)
    end = pd.Timestamp(end)

    total = 0.0
    acc = 0.0
    for i in range(len(times)):
        t0 = times[i]
        t1 = times[i + 1] if i + 1 < len(times) else (times[i] + pd.Timedelta(hours=1))
        seg_start = max(t0, start)
        seg_end = min(t1, end)
        if seg_end <= seg_start:
            continue
        w = _delta_seconds(seg_end - seg_start)
        total += w
        acc += w * float(values[i])

    if total == 0.0:
        return float("nan"), 0.0
    return acc / total, total


def _segment_weights(times: np.ndarray, start: pd.Timestamp, end: pd.Timestamp) -> np.ndarray:
    times = pd.to_datetime(times)
    start = pd.Timestamp(start)
    end = pd.Timestamp(end)

    weights = []
    for i in range(len(times)):
        t0 = times[i]
        t1 = times[i + 1] if i + 1 < len(times) else (times[i] + pd.Timedelta(hours=1))
        seg_start = max(t0, start)
        seg_end = min(t1, end)
        weights.append(max(0.0, _delta_seconds(seg_end - seg_start)))
    return np.asarray(weights, float)


def _wx_cache_key(lat: float, lon: float, start_date: str | None, end_date: str | None, endpoint: str, extra_params: dict | None = None) -> str:
    parts = [
        str(endpoint),
        f"{float(lat):.4f}",
        f"{float(lon):.4f}",
        (str(start_date)[:10] if start_date else ""),
        (str(end_date)[:10] if end_date else ""),
        "temperature_2m,relative_humidity_2m,wind_speed_10m,wind_direction_10m",
    ]
    if extra_params:
        for k in sorted(extra_params.keys()):
            parts.append(f"{k}={extra_params[k]}")
    s = "|".join(parts)
    return hashlib.sha1(s.encode("utf-8")).hexdigest()


def _fetch_open_meteo(
    endpoint: str,
    lat: float,
    lon: float,
    start_date: str | None,
    end_date: str | None,
    cache_dir: str,
    extra_params: dict | None = None,
    max_retries: int = 8,
    base_sleep_s: float = 0.75,
    throttle_s: float = 0.05,
) -> dict:
    """Fetch Open-Meteo hourly data with an on-disk JSON cache.

    Retries on HTTP 429 (Too Many Requests) with exponential backoff (+ jitter) and honors Retry-After if present.
    IMPORTANT: The Forecast API returns HTTP 400 if you combine start_date/end_date with past_days/forecast_days.
    For forecast calls using past_days, pass start_date=None and end_date=None (and we will omit them from params).
    """
    os.makedirs(cache_dir, exist_ok=True)

    key = _wx_cache_key(lat, lon, start_date, end_date, endpoint, extra_params=extra_params)
    cache_path = os.path.join(cache_dir, f"openmeteo_{key}.json")

    if os.path.exists(cache_path):
        try:
            with open(cache_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass

    params = {
        "latitude": float(lat),
        "longitude": float(lon),
        "hourly": "temperature_2m,relative_humidity_2m,wind_speed_10m,wind_direction_10m",
        "timezone": "UTC",
        "wind_speed_unit": "ms",
        "temperature_unit": "celsius",
    }
    if start_date is not None:
        params["start_date"] = str(start_date)[:10]
    if end_date is not None:
        params["end_date"] = str(end_date)[:10]
    if extra_params:
        params.update(extra_params)

    # Gentle global throttle to reduce burstiness
    if throttle_s and throttle_s > 0:
        time.sleep(float(throttle_s))

    last_exc = None
    for attempt in range(max_retries):
        try:
            # --- HTTP GET (requests if available, else urllib) ---
            if requests is not None:
                r = requests.get(endpoint, params=params, timeout=60)
                status = int(getattr(r, "status_code", 0) or 0)
                headers = dict(getattr(r, "headers", {}) or {})
                text = r.text
                if status == 429:
                    ra = headers.get("Retry-After")
                    if ra is not None:
                        try:
                            sleep_s = float(ra)
                        except Exception:
                            sleep_s = base_sleep_s * (2 ** attempt)
                    else:
                        sleep_s = base_sleep_s * (2 ** attempt)
                    sleep_s = min(120.0, sleep_s) * (1.0 + 0.15 * random.random())
                    time.sleep(sleep_s)
                    continue
                r.raise_for_status()
                data = r.json()
            else:
                # Standard-library fallback
                qs = urllib.parse.urlencode(params)
                url = endpoint + ("&" if "?" in endpoint else "?") + qs
                req = urllib.request.Request(url, headers={"User-Agent": "python-urllib"})
                try:
                    with urllib.request.urlopen(req, timeout=60) as resp:
                        status = int(getattr(resp, "status", 200))
                        headers = dict(resp.headers.items())
                        raw = resp.read().decode("utf-8", errors="replace")
                except urllib.error.HTTPError as e:
                    status = int(getattr(e, "code", 0) or 0)
                    headers = dict(getattr(e, "headers", {}).items()) if getattr(e, "headers", None) else {}
                    raw = (e.read().decode("utf-8", errors="replace") if hasattr(e, "read") else "")

                if status == 429:
                    ra = headers.get("Retry-After")
                    if ra is not None:
                        try:
                            sleep_s = float(ra)
                        except Exception:
                            sleep_s = base_sleep_s * (2 ** attempt)
                    else:
                        sleep_s = base_sleep_s * (2 ** attempt)
                    sleep_s = min(120.0, sleep_s) * (1.0 + 0.15 * random.random())
                    time.sleep(sleep_s)
                    continue
                if status and status >= 400:
                    raise RuntimeError(f"Open-Meteo HTTP {status}: {raw[:200]}")
                data = json.loads(raw) if raw else {}

            try:
                with open(cache_path, "w", encoding="utf-8") as f:
                    json.dump(data, f)
            except Exception:
                pass

            return data

        except Exception as e:
            # Requests-specific retry behavior when available; otherwise treat generically.
            last_exc = e
            # Retry transient errors
            if attempt < max_retries - 1:
                sleep_s = min(60.0, base_sleep_s * (2 ** attempt)) * (1.0 + 0.15 * random.random())
                time.sleep(sleep_s)
                continue
            raise

    if last_exc:
        raise last_exc
    raise RuntimeError("Open-Meteo request failed without exception")


def _hourly_to_df(hourly: dict) -> pd.DataFrame:
    """Convert Open-Meteo hourly dict to DataFrame (UTC-naive timestamps)."""
    if not hourly or not hourly.get("time"):
        return pd.DataFrame(columns=["time","temperature_2m","relative_humidity_2m","wind_speed_10m","wind_direction_10m"])
    dfh = pd.DataFrame({
        "time": hourly.get("time", []),
        "temperature_2m": hourly.get("temperature_2m", []),
        "relative_humidity_2m": hourly.get("relative_humidity_2m", []),
        "wind_speed_10m": hourly.get("wind_speed_10m", []),
        "wind_direction_10m": hourly.get("wind_direction_10m", []),
    })
    dfh["time"] = pd.to_datetime(dfh["time"], utc=True).dt.tz_convert("UTC").dt.tz_localize(None)
    for c in ["temperature_2m","relative_humidity_2m","wind_speed_10m","wind_direction_10m"]:
        dfh[c] = pd.to_numeric(dfh[c], errors="coerce")
    dfh = dfh.dropna(subset=["time"]).sort_values("time")
    dfh = dfh.drop_duplicates(subset=["time"], keep="last").reset_index(drop=True)
    return dfh


def _df_to_hourly(dfh: pd.DataFrame) -> dict:
    """Convert hourly DataFrame back to dict (ISO timestamps)."""
    if dfh is None or dfh.empty:
        return {"time": [], "temperature_2m": [], "relative_humidity_2m": [], "wind_speed_10m": [], "wind_direction_10m": []}
    t = pd.to_datetime(dfh["time"]).dt.strftime("%Y-%m-%dT%H:%M").tolist()
    return {
        "time": t,
        "temperature_2m": dfh["temperature_2m"].tolist(),
        "relative_humidity_2m": dfh["relative_humidity_2m"].tolist(),
        "wind_speed_10m": dfh["wind_speed_10m"].tolist(),
        "wind_direction_10m": dfh["wind_direction_10m"].tolist(),
    }


def fetch_open_meteo_hourly_archive(lat: float, lon: float, start_date: str, end_date: str, cache_dir: str) -> dict:
    return _fetch_open_meteo(OPEN_METEO_ARCHIVE, lat, lon, start_date, end_date, cache_dir, extra_params=None)



def fetch_open_meteo_hourly_forecast(lat: float, lon: float, start_date: str, end_date: str, cache_dir: str) -> dict:
    """Forecast API wrapper using past_days (to cover the recent past when archive lags).

    IMPORTANT: Do NOT combine start_date/end_date with past_days/forecast_days (Open-Meteo returns HTTP 400).
    """
    from datetime import date, datetime

    today = date.today()
    try:
        s = datetime.fromisoformat(str(start_date)[:10]).date()
    except Exception:
        s = today

    window_start = min(s, today)
    days_back = max(1, (today - window_start).days + 1)
    days_back = min(92, days_back)

    return _fetch_open_meteo(
        OPEN_METEO_FORECAST,
        lat, lon,
        start_date=None, end_date=None,
        cache_dir=cache_dir,
        extra_params={"past_days": int(days_back), "forecast_days": 1},
    )


def compute_weather_averages_from_hourly(hourly: dict, start_utc: pd.Timestamp, end_utc: pd.Timestamp) -> dict:
    hourly = hourly or {}
    times = hourly.get("time") or []
    if not times:
        return {c: np.nan for c in WEATHER_COLS}

    tt = pd.to_datetime(times, utc=True).tz_convert("UTC").tz_localize(None)
    T = np.asarray(hourly.get("temperature_2m") or [], float)
    RH = np.asarray(hourly.get("relative_humidity_2m") or [], float)
    W = np.asarray(hourly.get("wind_speed_10m") or [], float)
    D = np.asarray(hourly.get("wind_direction_10m") or [], float)

    # overlap select (pad 1h)
    startp = pd.Timestamp(start_utc)
    endp = pd.Timestamp(end_utc)
    sel = (tt >= (startp - pd.Timedelta(hours=1))) & (tt <= (endp + pd.Timedelta(hours=1)))
    if not bool(np.any(sel)):
        return {c: np.nan for c in WEATHER_COLS}

    tt2 = tt[sel]
    T2, RH2, W2, D2 = T[sel], RH[sel], W[sel], D[sel]

    avg_t, _ = _time_weighted_mean(T2, tt2, startp, endp)
    avg_rh, _ = _time_weighted_mean(RH2, tt2, startp, endp)
    avg_w, _ = _time_weighted_mean(W2, tt2, startp, endp)

    weights = _segment_weights(tt2, startp, endp)
    avg_dir = _circular_mean_deg(D2, weights) if weights.sum() > 0 else np.nan

    return {
        "avg_temp_c": float(avg_t) if np.isfinite(avg_t) else np.nan,
        "avg_humidity_pct": float(avg_rh) if np.isfinite(avg_rh) else np.nan,
        "avg_wind_ms": float(avg_w) if np.isfinite(avg_w) else np.nan,
        "avg_wind_dir_deg": float(avg_dir) if np.isfinite(avg_dir) else np.nan,
    }


def compute_weather_averages(lat: float, lon: float, start_utc: pd.Timestamp, end_utc: pd.Timestamp, cache_dir: str) -> dict:
    """
    Fetch hourly data (archive, with forecast fallback for recent runs) and compute time-weighted averages.
    """
    if not np.isfinite(lat) or not np.isfinite(lon) or end_utc <= start_utc:
        return {c: np.nan for c in WEATHER_COLS}

    startp = pd.Timestamp(start_utc)
    endp = pd.Timestamp(end_utc)

    # Pad dates to ensure boundary hours are included
    date_min = (startp.normalize() - pd.Timedelta(days=1)).date().isoformat()
    date_max = (endp.normalize() + pd.Timedelta(days=1)).date().isoformat()

    lat_r = float(np.round(lat, WX_GPS_ROUND_DP))
    lon_r = float(np.round(lon, WX_GPS_ROUND_DP))

    # Try archive first
    data = fetch_open_meteo_hourly_archive(lat_r, lon_r, date_min, date_max, cache_dir)
    hourly = (data.get("hourly") or {})

    wx = compute_weather_averages_from_hourly(hourly, startp, endp)
    if all((not np.isfinite(wx[c])) for c in WEATHER_COLS):
        # Fallback for recent runs (archive can lag)
        data2 = fetch_open_meteo_hourly_forecast(lat_r, lon_r, date_min, date_max, cache_dir)
        wx2 = compute_weather_averages_from_hourly((data2.get("hourly") or {}), startp, endp)
        return wx2

    return wx

    # Pad dates to ensure boundary hours are included
    date_min = (start_utc.normalize() - pd.Timedelta(days=1)).date().isoformat()
    date_max = (end_utc.normalize() + pd.Timedelta(days=1)).date().isoformat()

    # Round coordinates a bit to improve cache hits without materially changing weather
    lat_r = float(np.round(lat, WX_GPS_ROUND_DP))
    lon_r = float(np.round(lon, WX_GPS_ROUND_DP))

    data = fetch_open_meteo_hourly(lat_r, lon_r, date_min, date_max, cache_dir)
    hourly = data.get("hourly") or {}
    times = hourly.get("time") or []
    if not times:
        return {c: np.nan for c in WEATHER_COLS}

    tt = pd.to_datetime(times, utc=True).tz_localize(None)  # force UTC then drop tz -> UTC-naive
    # Ensure pandas Timestamps (naive UTC). Keep start/end as Timestamps (naive ok if consistent)
    tt = tt.to_numpy()

    T = np.asarray(hourly.get("temperature_2m") or [], float)
    RH = np.asarray(hourly.get("relative_humidity_2m") or [], float)
    W = np.asarray(hourly.get("wind_speed_10m") or [], float)
    D = np.asarray(hourly.get("wind_direction_10m") or [], float)

    # Select overlap (pad 1h)
    s = start_utc.to_datetime64()
    e = end_utc.to_datetime64()
    pad1 = np.timedelta64(1, "h")
    sel = (tt >= (s - pad1)) & (tt <= (e + pad1))
    if not np.any(sel):
        return {c: np.nan for c in WEATHER_COLS}

    tt2 = tt[sel]
    T2, RH2, W2, D2 = T[sel], RH[sel], W[sel], D[sel]

    # Convert start/end into numpy datetime64 for comparisons
    start64 = start_utc.to_datetime64()
    end64 = end_utc.to_datetime64()

    # For weighting, use pandas Timestamps
    ttp = pd.to_datetime(tt2)
    startp = pd.Timestamp(start_utc)
    endp = pd.Timestamp(end_utc)

    avg_t, _ = _time_weighted_mean(T2, ttp.to_numpy(), startp, endp)
    avg_rh, _ = _time_weighted_mean(RH2, ttp.to_numpy(), startp, endp)
    avg_w, _ = _time_weighted_mean(W2, ttp.to_numpy(), startp, endp)

    weights = _segment_weights(ttp.to_numpy(), startp, endp)
    avg_dir = _circular_mean_deg(D2, weights) if weights.sum() > 0 else np.nan

    return {
        "avg_temp_c": float(avg_t) if np.isfinite(avg_t) else np.nan,
        "avg_humidity_pct": float(avg_rh) if np.isfinite(avg_rh) else np.nan,
        "avg_wind_ms": float(avg_w) if np.isfinite(avg_w) else np.nan,
        "avg_wind_dir_deg": float(avg_dir) if np.isfinite(avg_dir) else np.nan,
    }

# ---------- helpers ----------
_EMOJI_RE = re.compile(
    "["  # broad-ish emoji/icon ranges
    "\U0001F1E0-\U0001F1FF"
    "\U0001F300-\U0001F5FF"
    "\U0001F600-\U0001F64F"
    "\U0001F680-\U0001F6FF"
    "\U0001F700-\U0001F77F"
    "\U0001F780-\U0001F7FF"
    "\U0001F800-\U0001F8FF"
    "\U0001F900-\U0001F9FF"
    "\U0001FA00-\U0001FAFF"
    "\u2600-\u26FF"
    "\u2700-\u27BF"
    "]+",
    flags=re.UNICODE,
)


def sanitize_text(s: object) -> str:
    if s is None or (isinstance(s, float) and not np.isfinite(s)):
        return ""
    s = str(s)
    s = s.replace("\uFFFD", "")
    s = re.sub(r"[\r\n\t]+", " ", s)
    s = _EMOJI_RE.sub("", s)
    s = "".join(ch for ch in s if ch.isprintable())
    s = re.sub(r"\s+", " ", s).strip()
    return s


def rolling_median_arr(x: np.ndarray, k: int) -> np.ndarray:
    x = np.asarray(x, float)
    out = np.full_like(x, np.nan)
    h = k // 2
    for i in range(len(x)):
        lo = max(0, i - h)
        hi = min(len(x), i + h + 1)
        w = x[lo:hi]
        if not np.isfinite(w).any():
            out[i] = np.nan
        else:
            out[i] = float(np.nanmedian(w))
    return out


def npower_w(x: np.ndarray) -> float:
    x = np.asarray(x, float)
    x = x[np.isfinite(x)]
    if len(x) == 0:
        return float("nan")
    return float(np.power(np.mean(np.power(x, 4)), 0.25))


def elevation_gain_loss_simple(elev: np.ndarray) -> tuple[float, float]:
    """Naive gain/loss (kept for reference).

    This tends to overcount on noisy altitude streams.
    """
    elev = np.asarray(elev, float)
    if len(elev) == 0:
        return (float("nan"), float("nan"))
    elev_s = rolling_median_arr(elev, 5)
    gain = 0.0
    loss = 0.0
    last = elev_s[0]
    for a in elev_s[1:]:
        if not np.isfinite(a) or not np.isfinite(last):
            last = a
            continue
        da = a - last
        if da > 0:
            gain += da
        elif da < 0:
            loss += -da
        last = a
    return float(gain), float(loss)


def elevation_gain_loss_strava_like(
    elev_m: np.ndarray,
    speed_mps: np.ndarray | None = None,
    *,
    smooth_win: int = 21,
    min_speed_mps: float = 0.5,
    min_step_m: float = 2.0,
) -> tuple[float, float]:
    """Strava-like elevation gain/loss from a per-second altitude stream.

    Approach:
      - heavy smoothing of altitude
      - compute gain/loss only while moving (optional speed mask)
      - ignore micro undulations via an accumulator threshold (min_step_m)

    This dramatically reduces inflated gain on flat/indoor routes.
    """
    e = np.asarray(elev_m, float)
    if e.size < 2 or (not np.isfinite(e).any()):
        return (float("nan"), float("nan"))

    e_s = rolling_median_arr(e, int(max(3, smooth_win)))

    if speed_mps is not None and len(speed_mps) == len(e_s):
        v = np.asarray(speed_mps, float)
        m = np.isfinite(e_s) & np.isfinite(v) & (v >= float(min_speed_mps))
        e_s = e_s[m]
    else:
        e_s = e_s[np.isfinite(e_s)]

    if e_s.size < 2:
        return (float("nan"), float("nan"))

    gain = 0.0
    loss = 0.0
    acc = 0.0
    thr = float(min_step_m)

    # accumulator thresholding
    for dz in np.diff(e_s):
        if not np.isfinite(dz):
            continue
        acc += float(dz)
        if acc >= thr:
            gain += acc
            acc = 0.0
        elif acc <= -thr:
            loss += -acc
            acc = 0.0

    return float(gain), float(loss)


def grade_from_dist_elev(dist_m: np.ndarray, elev_m: np.ndarray) -> np.ndarray:
    dist = np.asarray(dist_m, float)
    elev = np.asarray(elev_m, float)
    g = np.full_like(dist, np.nan)
    m = np.isfinite(dist) & np.isfinite(elev)
    if m.sum() < 20:
        return g
    dd = np.diff(dist)
    de = np.diff(elev)
    dd[dd == 0] = np.nan
    gr = de / dd
    gr = np.clip(gr, -0.35, 0.35)
    g[1:] = pd.Series(gr).rolling(5, center=True, min_periods=2).mean().to_numpy()
    return g


def clean_elevation_series(ts_s: np.ndarray, elev_m: np.ndarray, max_vz_mps: float = 1.2) -> tuple[np.ndarray, str]:
    """Clean per-second elevation to avoid crazy grade from broken altitude streams.

    Returns (elev_clean, alt_quality):
      - ok: no significant correction needed
      - corrected: clipping applied
      - flat_fallback: elevation deemed unusable; returns NaNs (grade should be treated as 0)
    """
    ts = np.asarray(ts_s, float)
    e = np.asarray(elev_m, float)

    if len(e) < 20 or (not np.isfinite(e).any()):
        return e, "flat_fallback"

    # Smooth small noise first
    e_med = rolling_median_arr(e, 5)

    # Compute dt (cap to avoid huge gaps dominating)
    dt = np.diff(ts)
    dt = np.clip(dt, 0.5, 5.0)
    de = np.diff(e_med)

    # Clip vertical speed
    max_de = max_vz_mps * dt
    de_clipped = np.clip(de, -max_de, max_de)
    clipped = np.sum(np.abs(de - de_clipped) > 1e-6)

    e_clean = np.empty_like(e_med)
    e_clean[0] = e_med[0]
    e_clean[1:] = e_clean[0] + np.cumsum(de_clipped)

    frac = clipped / max(1, len(de))
    # Hard failure: repeated huge jumps -> flat fallback
    if frac > 0.20:
        return np.full_like(e, np.nan, dtype=float), "flat_fallback"

    if frac > 0.01:
        return e_clean, "corrected"
    return e_clean, "ok"


def grade_from_dist_elev_safe(dist_m: np.ndarray, elev_m: np.ndarray, alt_quality: str) -> np.ndarray:
    """Compute grade; if altitude unusable, return zeros (flat) instead of NaN spikes."""
    if alt_quality == "flat_fallback":
        return np.zeros_like(np.asarray(dist_m, float), dtype=float)
    g = grade_from_dist_elev(dist_m, elev_m)
    # If grade is all-NaN, also fall back to flat
    if not np.isfinite(g).any():
        return np.zeros_like(np.asarray(dist_m, float), dtype=float)
    return g

def heading_from_latlon(lat_deg: np.ndarray, lon_deg: np.ndarray) -> np.ndarray:
    """Per-second heading (degrees), NaN if insufficient GPS."""
    lat = np.asarray(lat_deg, float)
    lon = np.asarray(lon_deg, float)
    out = np.full_like(lat, np.nan, dtype=float)
    if lat.size < 2:
        return out

def time_bucket(dt: pd.Timestamp) -> str:
    h = dt.hour + dt.minute / 60.0
    if h < 12:
        return "Morning run"
    if h < 17:
        return "Afternoon run"
    return "Evening run"


def filter_gps_speed_spikes(t: np.ndarray, v: np.ndarray, max_speed: float = 7.5, max_accel: float = 3.0) -> np.ndarray:
    """Filter GPS speed spikes using speed cap and acceleration limit.
    
    GPS can produce unrealistic speed spikes (e.g., 24 m/s when running).
    This filter:
    1. Marks samples as "bad" if speed > max_speed OR acceleration > max_accel
    2. Expands bad regions by 1 sample each side (neighboring samples often affected)
    3. Interpolates bad samples from surrounding good data
    
    Args:
        t: timestamps (seconds)
        v: speed values (m/s)
        max_speed: maximum realistic running speed (default 7.5 m/s = 2:13/km)
        max_accel: maximum realistic acceleration (default 3.0 m/s²)
    
    Returns:
        Filtered speed array with spikes replaced by interpolated values
    """
    t = np.asarray(t, dtype=float)
    v = np.asarray(v, dtype=float)
    v_filtered = v.copy()
    n = len(v)
    
    if n < 3:
        return np.clip(v_filtered, 0, max_speed)
    
    # Mark bad samples
    bad = np.zeros(n, dtype=bool)
    
    # Bad if speed too high
    bad |= (v > max_speed) | ~np.isfinite(v)
    
    # Bad if acceleration too high
    dt = np.diff(t)
    dt = np.maximum(dt, 0.1)  # Avoid division by zero
    dv = np.diff(v)
    accel = np.zeros(n)
    accel[1:] = dv / dt
    bad |= np.abs(accel) > max_accel
    
    # Expand bad regions by 1 sample each side (neighboring samples often also affected)
    bad_expanded = bad.copy()
    for i in range(1, n - 1):
        if bad[i]:
            bad_expanded[i - 1] = True
            bad_expanded[i + 1] = True
    
    # Interpolate bad samples from good neighbors
    good_idx = np.where(~bad_expanded)[0]
    bad_idx = np.where(bad_expanded)[0]
    
    if len(good_idx) > 1 and len(bad_idx) > 0:
        v_filtered[bad_idx] = np.interp(t[bad_idx], t[good_idx], v[good_idx])
    elif len(good_idx) <= 1:
        # Not enough good samples - just cap
        v_filtered = np.clip(v, 0, max_speed)
    
    # Final safety cap
    v_filtered = np.clip(v_filtered, 0, max_speed)
    
    return v_filtered


def haversine_m(lat1, lon1, lat2, lon2) -> float:
    R = 6371000.0
    phi1 = np.radians(lat1)
    phi2 = np.radians(lat2)
    dphi = np.radians(lat2 - lat1)
    dl = np.radians(lon2 - lon1)
    a = np.sin(dphi / 2.0) ** 2 + np.cos(phi1) * np.cos(phi2) * np.sin(dl / 2.0) ** 2
    return float(2.0 * R * np.arcsin(np.sqrt(a)))

# ---------- grade detrend helper (v37) ----------
def _loop_like_from_track(df: pd.DataFrame, dist_m_est: float) -> tuple[float, bool]:
    """Return (displacement_m, loop_like).

    loop_like: start/end horizontal displacement is small relative to run distance.
    Uses a robust start/end location estimate (median of first/last valid GPS points)
    to reduce the impact of early/late GPS jitter or a delayed fix.
    Returns (nan, False) if insufficient GPS.
    """
    if df is None or df.empty:
        return float('nan'), False

    # Prefer an explicit gps_valid flag if present; otherwise accept finite lat/lon.
    valid = df.get('gps_valid')
    if valid is None:
        latv = pd.to_numeric(df.get('lat', df.get('position_lat')), errors='coerce')
        lonv = pd.to_numeric(df.get('lon', df.get('position_long')), errors='coerce')
        valid = (latv.notna() & lonv.notna())
    valid = valid.astype(bool).to_numpy()

    if valid.sum() < 2 or not (dist_m_est and dist_m_est > 0):
        return float('nan'), False

    lat = pd.to_numeric(df.loc[valid, 'lat'] if 'lat' in df.columns else df.loc[valid, 'position_lat'], errors='coerce').to_numpy(dtype=float)
    lon = pd.to_numeric(df.loc[valid, 'lon'] if 'lon' in df.columns else df.loc[valid, 'position_long'], errors='coerce').to_numpy(dtype=float)
    if lat.size < 2 or lon.size < 2:
        return float('nan'), False

    # Robust start/end: median of first/last N valid points.
    N = 5
    n = int(min(N, lat.size))
    lat0 = float(np.nanmedian(lat[:n]))
    lon0 = float(np.nanmedian(lon[:n]))
    lat1 = float(np.nanmedian(lat[-n:]))
    lon1 = float(np.nanmedian(lon[-n:]))

    disp = haversine_m(lat0, lon0, lat1, lon1)

    # Slightly more forgiving threshold to catch out-and-backs / loops with GPS jitter.
    thr = max(400.0, 0.08 * float(dist_m_est))
    return float(disp), bool(disp < thr)

# ---------- calibration eras (v44: consolidated hardware eras) ----------
# v44: Removed arbitrary "early/late" splits that were based on guesswork.
# Now uses 6 actual hardware eras with configurable transition dates.
#
# The dates below are the first run date with each new pod.
# These can be overridden via stryd_era_config.json in the output directory.

HARDWARE_ERAS = ["pre_stryd", "v1", "repl", "air", "s4", "s5"]

# Default era transition dates (first run with each pod)
# Format: era_name -> start_date (end is implicitly the next era's start)
DEFAULT_ERA_TRANSITIONS = {
    "pre_stryd": "1900-01-01",  # Everything before first Stryd
    "v1":        "2017-05-05",  # First Stryd pod (working)
    "v1_late":   "2017-08-01",  # v1 pod dying - simulate power like pre_stryd
    "repl":      "2017-09-12",  # Replacement pod when v1 broke
    "air":       "2019-09-07",  # Air power model
    "s4":        "2023-01-03",  # Stryd 4.0
    "s5":        "2025-12-17",  # Stryd 5.0
}

# v1_late is treated like pre_stryd for power simulation
# because the pod was failing and power data is unreliable
SIMULATED_POWER_ERAS = {"pre_stryd", "v1_late"}

# v1_late is treated like pre_stryd for power simulation
# because the pod was failing and power data is unreliable
SIMULATED_POWER_ERAS = {"pre_stryd", "v1_late"}

# Global era config (loaded from file or defaults)
_ERA_TRANSITIONS = None
_ERA_BOUNDARIES_CACHE = None

def load_era_config(out_dir: str) -> dict:
    """
    Load era transition dates from stryd_era_config.json if it exists.
    Falls back to DEFAULT_ERA_TRANSITIONS.
    """
    global _ERA_TRANSITIONS, _ERA_BOUNDARIES_CACHE
    config_file = Path(out_dir) / "stryd_era_config.json"
    
    if config_file.exists():
        try:
            with open(config_file, "r", encoding="utf-8") as f:
                loaded = json.load(f)
            # Merge with defaults (in case new eras added)
            _ERA_TRANSITIONS = DEFAULT_ERA_TRANSITIONS.copy()
            _ERA_TRANSITIONS.update(loaded)
            print(f"[era] Loaded era config from {config_file.name}")
        except Exception as e:
            print(f"[era] Warning: Could not load era config: {e}")
            _ERA_TRANSITIONS = DEFAULT_ERA_TRANSITIONS.copy()
    else:
        _ERA_TRANSITIONS = DEFAULT_ERA_TRANSITIONS.copy()
        # Save defaults for user to customize
        save_era_config(out_dir, _ERA_TRANSITIONS)
    
    # Clear cache so boundaries are recalculated
    _ERA_BOUNDARIES_CACHE = None
    return _ERA_TRANSITIONS

def save_era_config(out_dir: str, config: dict):
    """Save era transition dates to stryd_era_config.json."""
    config_file = Path(out_dir) / "stryd_era_config.json"
    try:
        with open(config_file, "w", encoding="utf-8") as f:
            json.dump(config, f, indent=2)
        print(f"[era] Saved era config to {config_file.name}")
        print(f"[era] Edit this file to adjust pod transition dates if needed.")
    except Exception as e:
        print(f"[era] Warning: Could not save era config: {e}")

def get_era_boundaries() -> list:
    """
    Convert era transitions dict to list of (name, start, end) tuples.
    """
    global _ERA_BOUNDARIES_CACHE
    
    if _ERA_BOUNDARIES_CACHE is not None:
        return _ERA_BOUNDARIES_CACHE
    
    if _ERA_TRANSITIONS is None:
        transitions = DEFAULT_ERA_TRANSITIONS
    else:
        transitions = _ERA_TRANSITIONS
    
    # Sort by date to build boundaries
    sorted_eras = sorted(transitions.items(), key=lambda x: x[1])
    
    boundaries = []
    for i, (era, start) in enumerate(sorted_eras):
        if i + 1 < len(sorted_eras):
            end = sorted_eras[i + 1][1]
        else:
            end = "2100-01-01"
        boundaries.append((era, start, end))
    
    _ERA_BOUNDARIES_CACHE = boundaries
    return boundaries

def print_era_config():
    """Print current era configuration for verification."""
    print("\n" + "=" * 60)
    print("STRYD ERA CONFIGURATION (v45)")
    print("=" * 60)
    
    for era, start, end in get_era_boundaries():
        print(f"  {era:12s}  {start} -> {end}")
    
    print("=" * 60)
    print("To adjust dates, edit stryd_era_config.json")
    print("=" * 60 + "\n")

def assign_calibration_era_v44(ts: pd.Timestamp) -> str:
    """
    Assign calibration era based on date.
    Uses consolidated 6-era model (no early/late splits).
    """
    if ts is None or (isinstance(ts, float) and not np.isfinite(ts)):
        return ""
    t = pd.Timestamp(ts)
    
    for era, start, end in get_era_boundaries():
        if t >= pd.Timestamp(start) and t < pd.Timestamp(end):
            return era
    return ""

# Legacy function for backwards compatibility
def assign_calibration_era(ts: pd.Timestamp) -> str:
    """Date-based era assignment - just calls v44 version."""
    return assign_calibration_era_v44(ts)



# ---------- device ids ----------
def extract_stryd_device_ids(fit: FitFile) -> dict:
    """
    Best-effort extraction of Stryd pod identifiers from device_info messages.
    Returns keys: stryd_manufacturer, stryd_product, stryd_serial_number, stryd_ant_device_number
    Values may be blank if not present.
    """
    infos = []
    for msg in fit.get_messages("device_info"):
        try:
            manu = msg.get_value("manufacturer")
            prod = msg.get_value("product")
            serial = msg.get_value("serial_number")
            devnum = msg.get_value("device_number")  # ANT device number (often)
        except Exception:
            continue
        infos.append({"manufacturer": manu, "product": prod, "serial_number": serial, "device_number": devnum})

    if not infos:
        return {"stryd_manufacturer": "", "stryd_product": "", "stryd_serial_number": "", "stryd_ant_device_number": ""}

    def as_text(v):
        return sanitize_text(v).lower()

    # Prefer entries that explicitly mention stryd in manufacturer/product text
    stryd_like = []
    for inf in infos:
        mtxt = as_text(inf["manufacturer"])
        ptxt = as_text(inf["product"])
        if "stryd" in mtxt or "stryd" in ptxt:
            stryd_like.append(inf)

    chosen = None
    if stryd_like:
        # If multiple, pick the one with a non-null serial_number, else first
        stryd_like.sort(key=lambda d: (d["serial_number"] is None, d["device_number"] is None))
        chosen = stryd_like[0]
    else:
        # Fallback: choose a non-Garmin sensor record if possible
        # Garmin watches typically show manufacturer "garmin" and product numbers; pick something else.
        non_garmin = [inf for inf in infos if "garmin" not in as_text(inf["manufacturer"])]
        if non_garmin:
            non_garmin.sort(key=lambda d: (d["serial_number"] is None, d["device_number"] is None))
            chosen = non_garmin[0]
        else:
            chosen = infos[0]

    return {
        "stryd_manufacturer": sanitize_text(chosen.get("manufacturer")),
        "stryd_product": sanitize_text(chosen.get("product")),
        "stryd_serial_number": "" if chosen.get("serial_number") is None else str(chosen.get("serial_number")),
        "stryd_ant_device_number": "" if chosen.get("device_number") is None else str(chosen.get("device_number")),
    }


# ---------- HR correction ----------
def _has_time_gap(t: np.ndarray, i0: int, i1: int, gap_s: float = 1.5) -> bool:
    if i1 <= i0:
        return False
    dt = np.diff(t[i0:i1+1])
    return bool(np.any(dt > gap_s))


def _moving_mask(sp: np.ndarray, p: np.ndarray, h: np.ndarray) -> np.ndarray:
    """Robust moving/effort mask.

    - Outdoors: use speed > 0.5 m/s
    - Indoors / speed missing: fall back to power > 0
    """
    sp = np.asarray(sp, float)
    p = np.asarray(p, float)
    h = np.asarray(h, float)
    sp_ok = np.isfinite(sp) & (sp > 0.5)
    if np.nanmax(np.where(np.isfinite(sp), sp, 0.0)) <= 0.2:
        # speed missing/zero: use power
        return np.isfinite(p) & (p > 0) & np.isfinite(h) & (h > 30)
    return sp_ok & np.isfinite(p) & (p > 0) & np.isfinite(h) & (h > 30)


def detect_hr_drop_event(
    t: np.ndarray,
    hr: np.ndarray,
    power: np.ndarray,
    speed: np.ndarray,
    max_window_s: float = 10.0,
    drop_bpm: float = 12.0,
    power_tol: float = 0.12,
) -> int | None:
    """Detect a sharp HR drop while power is stable.

    This is intended for *low-HR dropouts*, not early high artifacts.
    Includes a pause/gap firewall: ignores candidates spanning timestamp gaps.
    """
    t = np.asarray(t, float)
    h = np.asarray(hr, float)
    p = np.asarray(power, float)
    sp = np.asarray(speed, float)

    valid = _moving_mask(sp, p, h)
    if valid.sum() < 50:
        return None

    h_rm = rolling_median_arr(h, 15)

    for i in range(1, len(t)):
        if not valid[i]:
            continue
        j = i - 1
        while j >= 0 and (t[i] - t[j]) <= max_window_s:
            if not valid[j]:
                j -= 1
                continue
            # reject if any timestamp gap inside window
            if _has_time_gap(t, j, i, gap_s=1.5):
                j -= 1
                continue

            dh = h_rm[j] - h_rm[i]
            if np.isfinite(dh) and dh >= drop_bpm:
                pseg = p[j : i + 1]
                p_med = np.nanmedian(pseg)
                if p_med > 0 and (np.nanmax(np.abs(pseg - p_med)) / p_med) <= power_tol:
                    return i
            j -= 1
    return None


def _detect_early_high_segment(t: np.ndarray, h: np.ndarray, p: np.ndarray, sp: np.ndarray,
                               early_end_s: float = 720.0,
                               min_excess_bpm: float = 15.0,
                               min_len_s: float = 60.0) -> tuple[int, int, float] | None:
    """Detect an early high-HR artifact that settles to a stable baseline.

    Returns (i_start, i_end, baseline_hr).
    """
    t = np.asarray(t, float)
    h = np.asarray(h, float)
    p = np.asarray(p, float)
    sp = np.asarray(sp, float)

    moving = _moving_mask(sp, p, h)
    if t.size < 200:
        return None

    early_mask = (t <= early_end_s) & moving
    if early_mask.sum() < 60:
        return None

    # baseline from later stable portion (after early_end_s), up to +10 min if available
    later_mask = (t >= early_end_s) & (t <= early_end_s + 600) & moving
    if later_mask.sum() < 120:
        later_mask = (t >= early_end_s) & moving
    if later_mask.sum() < 120:
        return None

    baseline = float(np.nanmedian(h[later_mask]))
    if not np.isfinite(baseline) or baseline < 60:
        return None

    # Candidate points: HR well above baseline while power is not spiking wildly
    # Use rolling median power to suppress noise
    p_rm = rolling_median_arr(p, 15)
    cand = early_mask & np.isfinite(h) & (h > baseline + min_excess_bpm) & np.isfinite(p_rm) & (p_rm > 0)

    if cand.sum() < 30:
        return None

    # Find longest contiguous segment in early window (by indices) with no time gaps
    idx = np.where(cand)[0]
    best = None  # (len, s, e)
    s = idx[0]
    prev = idx[0]
    for k in idx[1:]:
        if (k == prev + 1) and (t[k] - t[prev] <= 1.5):
            prev = k
            continue
        # break
        if best is None or (prev - s) > best[0]:
            best = (prev - s, s, prev)
        s = k
        prev = k
    if best is None or (prev - s) > best[0]:
        best = (prev - s, s, prev)

    if best is None:
        return None
    _, i0, i1 = best
    dur = float(t[i1] - t[i0])
    if dur < min_len_s:
        return None

    # Extra sanity: ensure power in this segment isn't dramatically higher than later baseline segment
    p_seg = p_rm[i0:i1+1]
    p_late = p_rm[later_mask]
    if np.isfinite(p_seg).any() and np.isfinite(p_late).any():
        if float(np.nanmedian(p_seg)) > 1.25 * float(np.nanmedian(p_late)):
            # early segment is actually higher effort: don't treat as artifact
            return None

    return i0, i1, baseline


def correct_hr_backproject(
    t: np.ndarray,
    power: np.ndarray,
    hr: np.ndarray,
    speed: np.ndarray,
) -> tuple[np.ndarray, bool, float | None, str | None, float | None, float | None]:
    """HR correction v4.

    Handles two artifact types:
      A) Early-high artifact (dirty strap): early HR too high then settles by ~12min.
         -> lower only that early segment toward baseline (no change elsewhere).
      B) Low-HR dropout (classic): HR collapses while effort steady.
         -> fill only the low segment by interpolation; never invent extreme max HR.

    Includes pause/gap firewall and caps to prevent runaway max HR.
    """
    t = np.asarray(t, float)
    p = np.asarray(power, float)
    h = np.asarray(hr, float)
    sp = np.asarray(speed, float)

    h_corr = h.copy()
    applied = False
    t_event = None
    corr_type: str | None = None
    r_pre: float | None = None
    r_post: float | None = None

    moving = _moving_mask(sp, p, h)

    # ---- A) Early-high artifact ----
    early = _detect_early_high_segment(t, h_corr, p, sp, early_end_s=720.0)
    if early is not None:
        i0, i1, baseline = early
        # taper target from (baseline+5) down to baseline across the segment
        target = np.linspace(baseline + 5.0, baseline, i1 - i0 + 1)
        seg = h_corr[i0:i1+1].astype(float)
        # only reduce (never increase)
        seg2 = np.minimum(seg, target)
        h_corr[i0:i1+1] = seg2
        applied = True
        t_event = float(t[i1])
        corr_type = "early_step_down"
        corr_type = "early_step_down"

    # ---- B) Low-HR dropout ----
    # Baseline for plausibility: use median HR after 12 min if possible, else overall moving median
    baseline_mask = moving & (t >= 720.0) & np.isfinite(h_corr)
    if baseline_mask.sum() < 120:
        baseline_mask = moving & np.isfinite(h_corr)
    baseline2 = float(np.nanmedian(h_corr[baseline_mask])) if baseline_mask.any() else np.nan

    drop_i = detect_hr_drop_event(t, h_corr, p, sp)
    if drop_i is not None and np.isfinite(baseline2):
        t_drop = float(t[drop_i])
        # Require HR after drop to be implausibly low for the effort (prevents indoor false positives)
        # Use 60s medians pre/post, but also guard against standing/walking rests:
        # - If post power collapses, HR can legitimately fall and we must NOT treat it as a sensor dropout.
        post_mask = moving & (t >= t_drop) & (t <= t_drop + 60) & np.isfinite(h_corr)
        pre_mask  = moving & (t >= t_drop - 120) & (t < t_drop) & np.isfinite(h_corr)
        hr_post = float(np.nanmedian(h_corr[post_mask])) if post_mask.any() else np.nan
        p_post  = float(np.nanmedian(p[post_mask])) if post_mask.any() else np.nan
        p_pre   = float(np.nanmedian(p[pre_mask])) if pre_mask.any() else np.nan

        # Extra rest guard: require "continued effort" after the drop
        # - post median power must be reasonably high
        # - post/pre power ratio must not indicate a large stop
        continued_effort = (
            np.isfinite(p_post) and (p_post >= 180.0) and
            np.isfinite(p_pre) and (p_pre > 0) and ((p_post / p_pre) >= 0.85)
        )

        if np.isfinite(hr_post) and np.isfinite(p_post):
            low_gate = (
                continued_effort and
                (hr_post < min(130.0, baseline2 - 15.0)) and
                (p_post > 200.0)
            )
        else:
            low_gate = False

        if low_gate:
            # Identify low segment: from drop_i forward while HR remains low
            low_thr = baseline2 - 12.0
            j = drop_i
            while j < len(t) and (moving[j]) and np.isfinite(h_corr[j]) and h_corr[j] < low_thr and (t[j] - t_drop) <= 600:
                # stop if a time gap appears
                if j > drop_i and (t[j] - t[j-1] > 1.5):
                    break
                j += 1
            i2 = j - 1

            # Need endpoints for interpolation
            i_pre = drop_i - 1
            while i_pre >= 0 and (not (moving[i_pre] and np.isfinite(h_corr[i_pre]))):
                i_pre -= 1
            i_post = i2 + 1
            while i_post < len(t) and (not (moving[i_post] and np.isfinite(h_corr[i_post]))):
                i_post += 1

            if i_pre >= 0 and i_post < len(t) and i2 > drop_i:
                # Linear interpolation between endpoints
                h0 = float(h_corr[i_pre])
                h1 = float(h_corr[i_post])
                n = i_post - i_pre
                interp = np.linspace(h0, h1, n + 1)
                h_corr[i_pre:i_post+1] = interp
                applied = True
                t_event = t_drop if t_event is None else t_event
                # This branch is specifically a low-HR dropout -> correct upward
                corr_type = "late_dropout" if corr_type is None else corr_type
                corr_type = "late_dropout"

    # ---- Compute P/HR ratios for audit note (2-min windows around detected event) ----
    if applied and t_event is not None:
        try:
            # find closest index to t_event
            i_evt = int(np.nanargmin(np.abs(t - float(t_event))))
            a0 = max(0, i_evt - 120)
            a1 = min(len(t), i_evt)
            b0 = i_evt
            b1 = min(len(t), i_evt + 120)
            pre_m = moving[a0:a1] & np.isfinite(p[a0:a1]) & np.isfinite(h_corr[a0:a1]) & (h_corr[a0:a1] > 60)
            post_m = moving[b0:b1] & np.isfinite(p[b0:b1]) & np.isfinite(h_corr[b0:b1]) & (h_corr[b0:b1] > 60)
            if int(pre_m.sum()) >= 30:
                r_pre = float(np.nanmedian(p[a0:a1][pre_m] / h_corr[a0:a1][pre_m]))
            if int(post_m.sum()) >= 30:
                r_post = float(np.nanmedian(p[b0:b1][post_m] / h_corr[b0:b1][post_m]))
        except Exception:
            pass

    # ---- Caps: never invent crazy max HR ----
    if applied:
        # cap corrected HR to not exceed original max + 5 bpm
        raw_max = float(np.nanmax(h[np.isfinite(h)])) if np.isfinite(h).any() else np.inf
        cap = min(raw_max + 5.0, 195.0)
        h_corr = np.minimum(h_corr, cap)
        # smooth lightly
        h_corr = rolling_median_arr(h_corr, 5)

        # Compute P/HR sanity ratios for notes/debug: 2-minute windows pre/post the detected event
        try:
            te = float(t_event) if t_event is not None else None
            if te is not None:
                pre_w = moving & (t >= te - 120.0) & (t < te) & np.isfinite(p) & np.isfinite(h_corr) & (h_corr > 60)
                post_w = moving & (t >= te) & (t < te + 120.0) & np.isfinite(p) & np.isfinite(h_corr) & (h_corr > 60)
                if int(pre_w.sum()) >= 30:
                    r_pre = float(np.nanmedian(p[pre_w] / h_corr[pre_w]))
                if int(post_w.sum()) >= 30:
                    r_post = float(np.nanmedian(p[post_w] / h_corr[post_w]))
        except Exception:
            pass

    return h_corr, bool(applied), (float(t_event) if t_event is not None else None), corr_type, r_pre, r_post

# ---------- RF v12 ----------
def rf_score_v12(df: pd.DataFrame, hr_corr: pd.Series) -> dict:
    def _end_for(start_s: int, end_s: int) -> int:
        return int(min(end_s, start_s + RF_TARGET_LEN_S))

    def _dead_time_components(win: pd.DataFrame) -> tuple[int, int]:
        dead_sig = int(((win["pwkg"] < PWR_DEAD_WKG) | (win["speed"] < SPD_DEAD_MS)).sum())
        ts = win["ts"].values
        dts = np.diff(ts)
        dead_gap = 0
        for dt in dts:
            if dt > 1.5:
                dead_gap += int(round(dt - 1))
        return dead_sig, dead_gap

    def _compute_rf_in_window(start_s: int, end_s: int) -> dict | None:
        win = df[(df["t"] >= start_s) & (df["t"] <= end_s)].copy()
        if len(win) < 30:
            return None
        total_time = int(round(end_s - start_s + 1))
        if total_time < RF_MIN_WINDOW_S:
            return None

        dead_sig, dead_gap = _dead_time_components(win)
        dead_total = dead_sig + dead_gap
        dead_frac = (dead_total / total_time) if total_time > 0 else np.nan

        valid_mask = ~((win["pwkg"] < PWR_DEAD_WKG) | (win["speed"] < SPD_DEAD_MS))
        win_valid = win[valid_mask].copy()
        if len(win_valid) < RF_MIN_VALID_SECONDS:
            return None

        p30 = win_valid["power_w"].rolling(30, min_periods=15).mean()
        hr_v = hr_corr.loc[win_valid.index].astype(float)
        ratio = (p30 / hr_v).replace([np.inf, -np.inf], np.nan).dropna()
        if len(ratio) < 30:
            return None

        score_mean = float(ratio.mean())
        score_med = float(ratio.median())

        dead_capped = min(dead_total, total_time // 2)
        k = 0.20
        adj_factor = 1.0 - k * (dead_capped / total_time)
        adj_mean = score_mean * adj_factor
        adj_med = score_med * adj_factor

        t_rel = (win_valid.loc[ratio.index, "t"].values - start_s).astype(float)
        y = ratio.values.astype(float)

        if len(y) >= 60:
            b, a = np.polyfit(t_rel, y, 1)
            yhat = a + b * t_rel
            ss_res = float(np.sum((y - yhat) ** 2))
            ss_tot = float(np.sum((y - float(np.mean(y))) ** 2))
            r2 = 1.0 - (ss_res / ss_tot) if ss_tot > 0 else np.nan
            drift_pct_per_min = (b / float(np.mean(y))) * 60.0 * 100.0
        else:
            drift_pct_per_min = np.nan
            r2 = np.nan

        return {
            "RF_window_start_s": int(start_s),
            "RF_window_end_s": int(end_s),
            "RF_window_shifted": bool(start_s != 600),
            "RF_dead_frac": float(dead_frac) if np.isfinite(dead_frac) else np.nan,
            "RF_adjusted_mean_W_per_bpm": float(adj_mean),
            "RF_adjusted_median_W_per_bpm": float(adj_med),
            "RF_drift_pct_per_min": float(drift_pct_per_min) if np.isfinite(drift_pct_per_min) else np.nan,
            "RF_drift_r2": float(r2) if np.isfinite(r2) else np.nan,
        }

    def _is_valid(rec: dict | None, max_dead_frac: float) -> bool:
        if rec is None:
            return False
        dfc = rec.get("RF_dead_frac", np.nan)
        return np.isfinite(dfc) and dfc <= max_dead_frac

    t_last = df["t"].dropna()
    if len(t_last) == 0:
        return {}
    end_run = int(round(float(t_last.iloc[-1])))

    def _with_code(rec: dict, code: str) -> dict:
        out = dict(rec)
        out["RF_select_code"] = code
        return out

    if end_run >= 300 + RF_MIN_WINDOW_S:
        for pass_i, max_dead_frac in enumerate(RF_DEAD_FRAC_PASSES, start=1):
            if end_run >= 600 + RF_MIN_WINDOW_S:
                e = _end_for(600, end_run)
                rec = _compute_rf_in_window(600, e)
                if _is_valid(rec, max_dead_frac):
                    return _with_code(rec, f"{pass_i}.1")
                for s in range(600 - BACKWARD_STEP_S, 299, -BACKWARD_STEP_S):
                    if end_run < s + RF_MIN_WINDOW_S:
                        continue
                    e = _end_for(s, end_run)
                    cand = _compute_rf_in_window(s, e)
                    if _is_valid(cand, max_dead_frac):
                        return _with_code(cand, f"{pass_i}.2")
                max_start = max(600 + FORWARD_STEP_S, end_run - RF_MIN_WINDOW_S)
                for s in range(600 + FORWARD_STEP_S, max_start + 1, FORWARD_STEP_S):
                    if s < 300:
                        continue
                    if end_run < s + RF_MIN_WINDOW_S:
                        continue
                    e = _end_for(s, end_run)
                    cand = _compute_rf_in_window(s, e)
                    if _is_valid(cand, max_dead_frac):
                        return _with_code(cand, f"{pass_i}.3")
            else:
                max_start = max(300, end_run - RF_MIN_WINDOW_S)
                for s in range(300, max_start + 1, FORWARD_STEP_S):
                    e = _end_for(s, end_run)
                    cand = _compute_rf_in_window(s, e)
                    if _is_valid(cand, max_dead_frac):
                        return _with_code(cand, f"{pass_i}.3")

    return {
        "RF_window_start_s": np.nan,
        "RF_window_end_s": np.nan,
        "RF_window_shifted": False,
        "RF_select_code": "0.0",
        "RF_dead_frac": np.nan,
        "RF_adjusted_mean_W_per_bpm": np.nan,
        "RF_adjusted_median_W_per_bpm": np.nan,
        "RF_drift_pct_per_min": np.nan,
        "RF_drift_r2": np.nan,
    }


# ---------- FIT parsing ----------
def parse_fit_to_df_and_ids(fit_path: str, weight_kg: float) -> tuple[pd.DataFrame, pd.Timestamp, dict]:
    fit = FitFile(fit_path)
    ids = extract_stryd_device_ids(fit)

    sess = next(iter(fit.get_messages("session")), None)
    # Skip non-running activities (e.g., cycling, hiking)
    sport = None
    try:
        sport = sess.get_value("sport") if sess else None
    except Exception:
        sport = None
    if sport is not None:
        s = str(sport).lower()
        if ("run" not in s) and ("running" not in s):
            raise SkipFitFile(f"Non-running sport: {sport}")

    start = sess.get_value("start_time") if sess else None
    if start is None:
        rec0 = next(iter(fit.get_messages("record")), None)
        start = rec0.get_value("timestamp") if rec0 else None
    if start is None:
        raise ValueError(f"No start_time in {fit_path}")
    start_utc = start.replace(tzinfo=timezone.utc) if start.tzinfo is None else start.astimezone(timezone.utc)
    start_utc = pd.Timestamp(start_utc)

    rows = []
    for rec in fit.get_messages("record"):
        ts = rec.get_value("timestamp")
        if ts is None:
            continue
        ts = ts.replace(tzinfo=timezone.utc) if ts.tzinfo is None else ts.astimezone(timezone.utc)
        ts_s = ts.timestamp()

        hr = rec.get_value("heart_rate")
        cad = rec.get_value("cadence")
        # Power: prefer Stryd developer power over native (Garmin) power
        pw_native = rec.get_value("power")
        pw_stryd = rec.get_value("Power")  # Stryd developer field is often exactly "Power"
        pw = None

        # If Stryd power is present and not all-zero, prefer it.
        if pw_stryd is not None:
            pw = pw_stryd
        else:
            # Some files use other developer field names; search record fields for a best "power" candidate.
            candidates = []
            for f in rec.fields:
                try:
                    name = (f.name or "")
                    val = f.value
                    units = getattr(f, "units", None)
                except Exception:
                    continue
                if val is None:
                    continue
                # select numeric
                try:
                    v = float(val)
                except Exception:
                    continue
                nlow = name.lower().strip()
                if "power" not in nlow:
                    continue
                # exclude other power-like fields
                if "air" in nlow or "form" in nlow:
                    continue
                if "watts" in nlow and ("air" in nlow or "form" in nlow):
                    continue
                # units heuristic
                if units is not None and str(units).lower() not in ("w", "watts"):
                    # allow missing/unknown units
                    pass
                candidates.append((nlow, v))
            # Prefer exact 'power' then 'stryd power' then largest median-ish value
            if candidates:
                # sort by name preference then by value
                candidates.sort(key=lambda x: (0 if x[0] == "power" else (1 if "stryd" in x[0] else 2), -abs(x[1])))
                pw = candidates[0][1]
            else:
                pw = pw_native

        # If we ended up with native power that is exactly 0 but Stryd power exists under a different label,
        # we'll still carry 0 here; downstream averaging will treat 0 as real. Better to set 0 -> None.
        if pw is not None:
            try:
                if float(pw) == 0.0:
                    pw = None
            except Exception:
                pass
        sp = rec.get_value("enhanced_speed") or rec.get_value("speed")
        dist = rec.get_value("distance")
        elev = rec.get_value("enhanced_altitude") or rec.get_value("altitude")
        air = rec.get_value("Air Power")

        la = rec.get_value("position_lat")
        lo = rec.get_value("position_long")
        if la is not None and lo is not None:
            la_f = float(la)
            lo_f = float(lo)
            # FIT semicircles are signed int32; some parsers surface unsigned values.
            if la_f > 2**31:
                la_f -= 2**32
            if lo_f > 2**31:
                lo_f -= 2**32
            lat = (la_f * (180.0 / 2**31))
            lon = (lo_f * (180.0 / 2**31))
        else:
            lat = np.nan
            lon = np.nan

        rows.append((ts_s, hr, cad, pw, sp, dist, elev, air, lat, lon))

    d = pd.DataFrame(rows, columns=["ts", "hr", "cadence", "power_w", "speed", "dist_m", "elev_m", "air_w", "lat", "lon"])
    for c in ["ts", "hr", "cadence", "power_w", "speed", "dist_m", "elev_m", "air_w", "lat", "lon"]:
        d[c] = pd.to_numeric(d[c], errors="coerce")
    d = d.dropna(subset=["ts"]).sort_values("ts").reset_index(drop=True)

    d["t"] = d["ts"] - d["ts"].iloc[0]

    # --- Stryd mass correction (v47) -------------------------------------------
    # Stryd calculates power using configured athlete mass. During some periods the
    # configured mass was wrong.  Scale power_w so it reflects ATHLETE_MASS_KG.
    from config import STRYD_MASS_HISTORY, ATHLETE_MASS_KG as _ATH_MASS
    run_date = pd.Timestamp(start_utc).tz_localize(None) if start_utc.tzinfo else start_utc
    for _ms_start, _ms_end, _cfg_kg in STRYD_MASS_HISTORY:
        if pd.Timestamp(_ms_start) <= run_date <= pd.Timestamp(_ms_end):
            _scale = _ATH_MASS / _cfg_kg
            d["power_w"] = d["power_w"] * _scale
            d["air_w"]   = d["air_w"]   * _scale
            break
    # ---------------------------------------------------------------------------
    
    # v41.1: Filter GPS speed spikes before any downstream calculations
    # GPS can produce unrealistic speeds (e.g., 24 m/s) which corrupt simulated power
    d["speed"] = filter_gps_speed_spikes(d["t"].values, d["speed"].values, max_speed=7.5, max_accel=3.0)
    
    d["pwkg"] = d["power_w"] / weight_kg
    d["air_wkg"] = d["air_w"] / weight_kg
    d["moving"] = d["speed"] > 0.5
    # Elevation cleaning for robust grade on broken altitude streams
    elev_clean, alt_quality = clean_elevation_series(d["ts"].values, d["elev_m"].values)
    d["elev_m_clean"] = elev_clean
    d["alt_quality"] = alt_quality
    d["grade"] = grade_from_dist_elev_safe(d["dist_m"].values, d["elev_m_clean"].values, alt_quality)
    d["gps_valid"] = np.isfinite(d["lat"]) & np.isfinite(d["lon"])
    d["heading_deg"] = heading_from_latlon(d["lat"].values, d["lon"].values)

    return d, start_utc, ids


def compute_gps_metrics(df: pd.DataFrame, outlier_speed_mps: float) -> tuple[float, float, float, float, float]:
    if df.empty:
        return (float("nan"), 0.0, float("nan"), float("nan"), float("nan"))

    valid = df["gps_valid"].values.astype(bool)
    coverage = float(np.mean(valid)) if len(valid) else 0.0
    if valid.sum() < 2:
        return (float("nan"), coverage, float("nan"), float("nan"), float("nan"))

    lat = df.loc[valid, "lat"].to_numpy(dtype=float)
    lon = df.loc[valid, "lon"].to_numpy(dtype=float)
    ts = df.loc[valid, "ts"].to_numpy(dtype=float)

    seg_m = []
    seg_speed = []
    for i in range(1, len(lat)):
        d_m = haversine_m(lat[i - 1], lon[i - 1], lat[i], lon[i])
        dt = ts[i] - ts[i - 1]
        if dt <= 0:
            continue
        seg_m.append(d_m)
        seg_speed.append(d_m / dt)

    if not seg_m:
        return (float("nan"), coverage, float("nan"), float("nan"), float("nan"))

    seg_m = np.asarray(seg_m, float)
    seg_speed = np.asarray(seg_speed, float)

    gps_distance_km = float(np.nansum(seg_m) / 1000.0)
    gps_max_seg_m = float(np.nanmax(seg_m)) if np.isfinite(seg_m).any() else float("nan")
    gps_p99_speed_mps = float(np.nanpercentile(seg_speed, 99)) if np.isfinite(seg_speed).any() else float("nan")
    gps_outlier_frac = float(np.mean(seg_speed > outlier_speed_mps)) if seg_speed.size else float("nan")

    return (gps_distance_km, coverage, gps_max_seg_m, gps_p99_speed_mps, gps_outlier_frac)


def summarize_fit(fit_path: str, tz_out: str, weight_kg: float, gps_outlier_speed_mps: float, persec_cache_dir: str = "", cache_only: bool = False) -> dict:
    tz = ZoneInfo(tz_out)
    df, start_utc, ids = parse_fit_to_df_and_ids(fit_path, weight_kg)
    start_local = start_utc.to_pydatetime().astimezone(tz).replace(tzinfo=None)

    # v44: Determine era for speed source selection
    # For Stryd-era runs, prefer GPS distance over per-second speed from FIT
    # because Stryd speed can have dropout artifacts (especially v1 pod)
    v1_start = pd.Timestamp("2017-05-05", tz="UTC")
    is_stryd_era = (start_utc >= v1_start)

    # --- v37: loop-aware grade detrending (baseline drift removal) ---
    # Rationale: on loop-like routes, net gravitational work should be ~0; a large non-zero
    # mean grade is almost always altitude drift / bad baro streams, especially pre-2016.
    grade_detrended = False
    grade_mean_before = float('nan')
    grade_mean_after = float('nan')
    loop_displacement_m = float('nan')
    loop_like = False

    try:
        # Distance estimate (meters): prefer device distance; fall back to integrating speed.
        dist_m_est = float('nan')
        dist_ser = pd.to_numeric(df.get('dist_m'), errors='coerce')
        if dist_ser.notna().any():
            dist_m_est = float(dist_ser.dropna().iloc[-1] - dist_ser.dropna().iloc[0])
            if not (dist_m_est > 0):
                dist_m_est = float(dist_ser.dropna().max())
        if not (dist_m_est > 0):
            ts = pd.to_numeric(df.get('ts'), errors='coerce').to_numpy(dtype=float)
            sp = pd.to_numeric(df.get('speed'), errors='coerce').to_numpy(dtype=float)
            if ts.size >= 2 and sp.size == ts.size:
                dt = np.diff(ts)
                dt = np.clip(dt, 0.5, 5.0)
                v = np.where(np.isfinite(sp), sp, 0.0)
                dist_m_est = float(np.sum(v[1:] * dt))

        loop_displacement_m, loop_like = _loop_like_from_track(df, dist_m_est)

        # Mean grade computed on moving seconds (same mask as downstream RF/RE generally use).
        moving = (pd.to_numeric(df.get('speed'), errors='coerce') > 0.5).fillna(False).to_numpy(dtype=bool)
        g = pd.to_numeric(df.get('grade'), errors='coerce').to_numpy(dtype=float)
        if g.size and moving.size == g.size:
            mg = float(np.nanmean(g[moving])) if np.isfinite(g[moving]).any() else float('nan')
        else:
            mg = float('nan')

        grade_mean_before = mg

        # Apply detrend only when the bias is clearly implausible on loop-like geometry.
        # Threshold: 0.45% mean grade (|mg| > 0.0045).
        alt_q = str(df.get('alt_quality', pd.Series([''])).iloc[0]) if len(df) else ''
        # Secondary safeguard: if the implied net elevation change is absurdly large (e.g. -300 m over 6 km),
        # treat it as baseline drift even if geometry isn't confidently loop-like (older GPS can make start/end noisy).
        gross_bias = (np.isfinite(mg) and abs(mg) > 0.02 and (dist_m_est == dist_m_est) and (dist_m_est > 0) and (dist_m_est < 30000))
        if (((loop_like and (np.isfinite(mg) and abs(mg) > 0.0045)) or gross_bias) and (alt_q != 'flat_fallback')):
            g2 = g - mg
            # Keep a conservative cap (matches sim-power step assumptions)
            g2 = np.clip(g2, -0.12, 0.12)
            df['grade'] = g2
            grade_detrended = True
            grade_mean_after = float(np.nanmean(g2[moving])) if np.isfinite(g2[moving]).any() else float('nan')
        else:
            grade_mean_after = mg

    except Exception:
        pass

    # Optional per-second cache for downstream RE modelling / sim-power.
    # Written AFTER grade detrend (v37) so sim-power sees the corrected grade series.
    if persec_cache_dir:
        global CACHE_WRITE_OK, CACHE_WRITE_FAIL, FIRST_CACHE_WRITE_ERROR, FIRST_CACHE_WRITE_ERROR_PATH
        try:
            os.makedirs(persec_cache_dir, exist_ok=True)
            # v41.1: Use FIT filename (sanitized) for cache filename
            fit_basename = os.path.basename(fit_path)
            fit_filename_no_ext = os.path.splitext(fit_basename)[0]
            # Sanitize filename for filesystem (replace problematic chars)
            safe_name = fit_filename_no_ext.replace('&', '_').replace('?', '_').replace('=', '_')
            cache_path = os.path.join(persec_cache_dir, safe_name + '.npz')
            if not _should_write_cache(cache_path, fit_path):
                # Cache exists and is considered up-to-date
                if cache_only:
                    return {"file": os.path.basename(fit_path), "date": start_local, "cache_written": False}
            else:
                np.savez_compressed(
                    cache_path,
                    run_id=fit_filename_no_ext,  # Original name (may contain &)
                    fit_filename=fit_basename,    # v41.1: Full FIT filename with extension
                    ts=df['ts'].values.astype(float),
                    t=df['t'].values.astype(float),
                    speed_mps=df['speed'].values.astype(float),
                    dist_m=df['dist_m'].values.astype(float),
                    elev_m_clean=df['elev_m_clean'].values.astype(float) if 'elev_m_clean' in df.columns else df['elev_m'].values.astype(float),
                    grade=df['grade'].values.astype(float) if 'grade' in df.columns else np.full(len(df), np.nan),
                    heading_deg=df['heading_deg'].values.astype(float) if 'heading_deg' in df.columns else np.full(len(df), np.nan),
                    hr_bpm=df['hr'].values.astype(float),
                    power_w=df['power_w'].values.astype(float) if 'power_w' in df.columns else np.full(len(df), np.nan),
                    # NOTE: use .iloc[0] (not [0]) because the DataFrame index is not guaranteed
                    # to be 0..N-1 after upstream processing.
                    alt_quality=str(df['alt_quality'].iloc[0]) if ('alt_quality' in df.columns and len(df)) else '',
                    gps_valid=df['gps_valid'].values.astype(bool) if 'gps_valid' in df.columns else np.zeros(len(df), dtype=bool),
                    grade_detrended=bool(grade_detrended),
                    grade_mean_before=float(grade_mean_before) if grade_mean_before==grade_mean_before else float('nan'),
                    grade_mean_after=float(grade_mean_after) if grade_mean_after==grade_mean_after else float('nan'),
                    loop_like=bool(loop_like),
                    loop_displacement_m=float(loop_displacement_m) if loop_displacement_m==loop_displacement_m else float('nan'),
                    weight_kg=float(weight_kg),
                )
            CACHE_WRITE_OK += 1
        except Exception:
            CACHE_WRITE_FAIL += 1
            if FIRST_CACHE_WRITE_ERROR is None:
                FIRST_CACHE_WRITE_ERROR = traceback.format_exc()
                FIRST_CACHE_WRITE_ERROR_PATH = str(fit_path)

    if cache_only:
        return {"file": os.path.basename(fit_path), "date": start_local, "cache_written": True}
    # --- v38: HR correction disabled in rebuild (raw HR passthrough) ---
    hr_corr = pd.to_numeric(df["hr"], errors="coerce").copy()
    hr_applied = False
    t_drop = None
    hr_corr_type = None
    hr_rpre = None
    hr_rpost = None

    moving = df["moving"].values.astype(bool)
    moving_recalc = moving  # default
    moving_mask_final = moving
    moving_time_recalc_applied = False
    moving_time_s_recalc = np.nan

    # --- v32: robust moving time ---
    # Compute timestamp deltas early so all moving-time calculations are dt-weighted
    # (handles non-1Hz sampling and GPS dropout gaps where dt > 1s)
    t_dt = pd.to_datetime(df["ts"], unit="s", utc=True, errors="coerce")  # FIT epoch seconds
    if t_dt.notna().any():
        dt = t_dt.diff().dt.total_seconds().fillna(0.0).clip(lower=0.0, upper=10.0).to_numpy(dtype=float)
    else:
        # Fallback: assume 1 Hz if timestamps absent
        dt = np.ones(len(df), dtype=float)

    moving_time_s_original = int(round(float(np.sum(dt[moving]))))
    elapsed_s = int(round(float(df["t"].iloc[-1]))) if len(df) else 0

    # Distance estimate for plausibility checks (device distance if present, else GPS)
    dist_km_tmp = float(df["dist_m"].diff().fillna(0).clip(lower=0).sum() / 1000.0) if df["dist_m"].notna().any() else np.nan
    gps_km_tmp, gps_cov_tmp, gps_max_seg_m_tmp, gps_p99_speed_mps_tmp, gps_outlier_frac_tmp = compute_gps_metrics(df, gps_outlier_speed_mps)
    if (not np.isfinite(dist_km_tmp)) and np.isfinite(gps_km_tmp):
        dist_km_tmp = float(gps_km_tmp)

    implied_speed = (dist_km_tmp * 1000.0 / moving_time_s_original) if (np.isfinite(dist_km_tmp) and moving_time_s_original > 0) else np.nan
    suspicious = (
        (moving_time_s_original == 0)
        or (elapsed_s > 0 and moving_time_s_original < 0.85 * elapsed_s)
        or (np.isfinite(implied_speed) and implied_speed > 7.0)
    )

    moving_time_s_recalc = np.nan
    moving_time_recalc_applied = False
    moving_mask_final = moving

    if suspicious and len(df) > 10:
        sp = pd.to_numeric(df.get("speed"), errors="coerce").fillna(0.0).to_numpy(dtype=float)
        cad = pd.to_numeric(df.get("cadence"), errors="coerce").fillna(0.0).to_numpy(dtype=float)
        if df["dist_m"].notna().any():
            dd = pd.to_numeric(df["dist_m"], errors="coerce").diff().fillna(0.0).clip(lower=0.0).to_numpy(dtype=float)
        else:
            dd = np.zeros(len(df), dtype=float)

        moving_recalc = (sp >= SPD_DEAD_MS) | (dd >= 0.25) | (cad >= 30.0)

    mt_s = float(np.sum(dt[moving_recalc]))
    mt = int(round(mt_s))

    # Only accept if it looks sane (older files can be sparse, so be less strict than 0.85)
    if suspicious and elapsed_s > 0 and (0.70 * elapsed_s) <= mt_s <= (1.05 * elapsed_s):
        moving_time_s_recalc = mt
        moving_time_recalc_applied = True
        moving_mask_final = moving_recalc
    else:
        moving_time_s_recalc = mt  # keep for diagnostics even if not applied


    # Use final mask & time for all downstream per-run calculations
    moving = moving_mask_final
    # Use dt-weighted sum so GPS dropout gaps (dt=2s) are counted correctly
    # (np.sum(moving) only counts records, not elapsed seconds)
    moving_s = int(round(float(np.sum(dt[moving]))))

    # Compute GPS-derived distance/quality metrics (works even when FIT lacks dist_m) (works even when FIT lacks dist_m)
    gps_km, gps_cov, gps_max_seg_m, gps_p99_speed_mps, gps_outlier_frac = compute_gps_metrics(df, gps_outlier_speed_mps)

    # GPS metrics (needed for distance fallback + quality)
    dist_km = float(df["dist_m"].diff().fillna(0).clip(lower=0).sum() / 1000.0) if df["dist_m"].notna().any() else np.nan
    # Pre-Stryd / older FITs sometimes lack a usable distance field; fall back to GPS distance
    if (not np.isfinite(dist_km)) and np.isfinite(gps_km):
        dist_km = float(gps_km)
    gps_ratio = (gps_km / dist_km) if (np.isfinite(gps_km) and np.isfinite(dist_km) and dist_km > 0) else np.nan

    # --- v37: loop-aware grade detrend (fix altitude drift on loop-like runs) ---
    # If a run starts and ends at (nearly) the same location, net elevation should be ~0.
    # When the altitude stream drifts, the derived grade series can have a non-zero mean,
    # which distorts simulated power and spreads RE. We remove only the *bias* (mean grade),
    # preserving relative hills.
    try:
        dist_m_est = float(dist_km_tmp * 1000.0) if np.isfinite(dist_km_tmp) else (float(gps_km_tmp * 1000.0) if np.isfinite(gps_km_tmp) else float('nan'))
        disp_m, loop_like = _loop_like_from_track(df, dist_m_est)
        alt_q = str(df.get('alt_quality', pd.Series([''])).iloc[0]) if 'alt_quality' in df.columns and len(df) else ''
        moving0 = df['moving'].values.astype(bool) if 'moving' in df.columns else np.ones(len(df), dtype=bool)
        g = pd.to_numeric(df.get('grade'), errors='coerce').to_numpy(dtype=float)
        mean_g = float(np.nanmean(g[moving0])) if (g.size and np.isfinite(g[moving0]).any()) else float('nan')
        grade_detrended = False
        if loop_like and alt_q != 'flat_fallback' and np.isfinite(mean_g) and abs(mean_g) > 0.0045:
            g2 = g - mean_g
            g2 = np.clip(g2, -0.12, 0.12)
            df['grade'] = g2
            grade_detrended = True
            if 'notes' in df.columns:
                pass
            # store in df attrs for later note/cache
            df.attrs['grade_detrended'] = True
            df.attrs['grade_mean_before'] = mean_g
            df.attrs['grade_mean_after'] = float(np.nanmean(g2[moving0])) if np.isfinite(g2[moving0]).any() else 0.0
            df.attrs['loop_displacement_m'] = disp_m
    except Exception:
        pass
    # --- end v37 grade detrend ---
    pace = (moving_s / 60.0) / dist_km if dist_km and dist_km > 0 else np.nan

    gps_km, gps_cov, gps_max_seg_m, gps_p99_speed_mps, gps_outlier_frac = compute_gps_metrics(df, gps_outlier_speed_mps)

    # Representative location for weather: median of valid GPS trackpoints
    lat_med = float(np.nanmedian(df.loc[df['gps_valid'], 'lat'])) if df['gps_valid'].any() else np.nan
    lon_med = float(np.nanmedian(df.loc[df['gps_valid'], 'lon'])) if df['gps_valid'].any() else np.nan
    # gps_ratio computed above; do not overwrite (needed for GPS sanity + adjusters)

    pw_mov = df.loc[moving, "power_w"] if np.any(moving) else None
    avg_power = float(pw_mov.mean()) if (pw_mov is not None and pw_mov.notna().any()) else np.nan
    nP = npower_w(df.loc[moving, "power_w"].values) if np.any(moving) else np.nan

    hr_mov = hr_corr.loc[moving] if np.any(moving) else None
    avg_hr = float(hr_mov.mean()) if (hr_mov is not None and hr_mov.notna().any()) else np.nan
    max_hr = float(np.nanmax(hr_corr.loc[moving])) if np.any(moving) else np.nan
    nPower_HR = (nP / avg_hr) if np.isfinite(nP) and np.isfinite(avg_hr) and avg_hr > 0 else np.nan

    pwkg_mean = _safe_nanmean(df.loc[moving, "pwkg"]) if np.any(moving) else np.nan


    # Speed used for RE and any downstream sim-calibration should be robust to GPS anomalies
    # (e.g. train after warmup) and robust to moving-time quirks in older FITs.
    dev_speed_mps = (dist_km * 1000.0 / moving_s) if (np.isfinite(dist_km) and moving_s and moving_s > 0) else np.nan

    gps_ratio = (gps_km / dist_km) if (np.isfinite(gps_km) and np.isfinite(dist_km) and dist_km > 0) else np.nan
    
    # v44: Smart speed source selection
    # Goal: Use the most reliable distance source for RE calculation, considering:
    # - GPS can be bad: tree cover (under-reads), tracks, indoor, train jumps
    # - Stryd can be bad: dropouts, recalibration, v1 pod issues
    
    # Detect problematic GPS scenarios
    gps_has_teleport = np.isfinite(gps_max_seg_m) and gps_max_seg_m > 500  # >500m single segment = train/teleport
    gps_poor_coverage = not (np.isfinite(gps_cov) and gps_cov >= 0.80)
    gps_has_outliers = np.isfinite(gps_outlier_frac) and gps_outlier_frac > 0.05  # raised from 0.02
    gps_speed_crazy = np.isfinite(gps_p99_speed_mps) and gps_p99_speed_mps > 8.0  # raised from 6.5 (fast races have high p99)
    
    # GPS is suspect if any of these are true
    gps_suspect = gps_has_teleport or gps_poor_coverage or gps_has_outliers or gps_speed_crazy
    
    # Detect problematic Stryd scenarios (for Stryd era)
    # Stryd under-reads: gps_ratio > 1.015 means GPS measured MORE than Stryd
    # Stryd over-reads: gps_ratio < 0.985 means GPS measured LESS than Stryd
    # (Thresholds tightened from 3% to 1.5% based on data: 75th percentile is 1.011)
    stryd_suspect = False
    if is_stryd_era and np.isfinite(gps_ratio):
        # If Stryd and GPS differ by more than 1.5%, one of them is wrong
        # In tree cover, GPS under-reads (ratio < 0.985) - trust Stryd
        # With Stryd dropouts, Stryd under-reads (ratio > 1.015) - trust GPS
        stryd_under_reading = gps_ratio > 1.015  # GPS > Stryd by >1.5%
        stryd_over_reading = gps_ratio < 0.985   # GPS < Stryd by >1.5%
        stryd_suspect = stryd_under_reading  # Stryd dropouts cause under-reading
    
    # Determine best speed source
    # Priority logic:
    # 1. If GPS is clearly broken (teleport, poor coverage), use Stryd/device
    # 2. If Stryd appears to have dropouts (under-reading vs GPS), use GPS
    # 3. If GPS and Stryd agree within 3%, prefer GPS for Stryd era (protects against recalibration)
    # 4. If GPS under-reads vs Stryd (tree cover), use Stryd
    
    gps_speed_mps = (gps_km * 1000.0 / moving_s) if (np.isfinite(gps_km) and gps_km > 0 and moving_s > 0) else np.nan
    
    # Per-second speed median (from FIT file - source depends on watch/pod setup)
    sp_sec = pd.to_numeric(df.get("speed"), errors="coerce").to_numpy(dtype=float)
    sp_mask = moving & np.isfinite(sp_sec) & (sp_sec >= 0.3) & (sp_sec <= 6.5)
    speed_med_mps = float(np.nanmedian(sp_sec[sp_mask])) if np.sum(sp_mask) >= 120 else np.nan

    # Make the decision
    speed_source = "unknown"
    
    if gps_suspect:
        # GPS is unreliable - use Stryd/device
        if np.isfinite(speed_med_mps):
            speed_for_re = speed_med_mps
            speed_source = "stryd_persec"
        else:
            speed_for_re = dev_speed_mps
            speed_source = "device"
    elif is_stryd_era:
        # Stryd era with usable GPS
        if stryd_suspect:
            # Stryd appears to have dropouts - use GPS
            speed_for_re = gps_speed_mps
            speed_source = "gps_stryd_suspect"
        elif np.isfinite(gps_ratio) and gps_ratio < 0.985:
            # GPS under-reading (tree cover?) - use Stryd
            if np.isfinite(speed_med_mps):
                speed_for_re = speed_med_mps
                speed_source = "stryd_gps_under"
            else:
                speed_for_re = dev_speed_mps
                speed_source = "device_gps_under"
        else:
            # GPS and Stryd roughly agree, or GPS reads more - use GPS
            # (protects against Stryd recalibration)
            if np.isfinite(gps_speed_mps):
                speed_for_re = gps_speed_mps
                speed_source = "gps"
            elif np.isfinite(speed_med_mps):
                speed_for_re = speed_med_mps
                speed_source = "stryd_persec_fallback"
            else:
                speed_for_re = dev_speed_mps
                speed_source = "device_fallback"
    else:
        # Pre-Stryd era - original logic
        if np.isfinite(speed_med_mps):
            speed_for_re = speed_med_mps
            speed_source = "persec"
        elif np.isfinite(gps_speed_mps):
            speed_for_re = gps_speed_mps
            speed_source = "gps"
        else:
            speed_for_re = dev_speed_mps
            speed_source = "device"

    # Guard against RE blow-ups when power coverage is poor (dropouts/0s removed/etc.).
    pw_sec = pd.to_numeric(df.get("power_w"), errors="coerce").to_numpy(dtype=float)
    pw_mask = moving & np.isfinite(pw_sec) & (pw_sec > 0)
    power_cov = float(np.mean(pw_mask)) if len(df) else 0.0
    # Compute mean W/kg from active seconds (more robust for interval sessions with standing rests)
    pwkg_mean_active = np.nan
    if np.isfinite(weight_kg) and weight_kg > 0 and pw_mask.any():
        pwkg_mean_active = float(np.nanmean(pw_sec[pw_mask]) / weight_kg)

    # Relax power coverage gating for measured-power runs (non pre_stryd)
    is_measured_power = False
    try:
        era = str(calibration_era_id).lower()
        is_measured_power = (era != "pre_stryd")
    except Exception:
        is_measured_power = False


    re_note_msg = None

    pwkg_for_re = pwkg_mean_active if np.isfinite(pwkg_mean_active) and pwkg_mean_active > 0 else pwkg_mean


    # RE_avg: speed / mean W/kg (only if coverage is good)
    if np.isfinite(speed_for_re) and np.isfinite(pwkg_for_re) and pwkg_for_re > 0 and (power_cov >= 0.80 or is_measured_power):
        RE_avg = (speed_for_re / pwkg_for_re)
    else:
        RE_avg = np.nan

    # RE_normalised: use normalised power (W) and speed_for_re
    pwkg_np = (nP / weight_kg) if (np.isfinite(nP) and weight_kg and weight_kg > 0) else np.nan
    if np.isfinite(speed_for_re) and np.isfinite(pwkg_np) and pwkg_np > 0 and (power_cov >= 0.80 or is_measured_power):
        RE_norm = (speed_for_re / pwkg_np)
    else:
        RE_norm = np.nan

    # Plausibility cap (avoid "off-the-scale" RE). Keep only values in a realistic band.
    if np.isfinite(RE_avg) and (RE_avg < 0.50 or RE_avg > 1.05):
        RE_avg = np.nan
        re_note_msg = "RE skipped (implausible)"
    if np.isfinite(RE_norm) and (RE_norm < 0.50 or RE_norm > 1.05):
        RE_norm = np.nan
        re_note_msg = (re_note_msg + "; " if re_note_msg else "") + "REn skipped (implausible)"
    air_wkg_arr = df.loc[moving, "air_wkg"].to_numpy(dtype=float) if np.any(moving) else np.array([])
    air_w_arr = df.loc[moving, "air_w"].to_numpy(dtype=float) if np.any(moving) else np.array([])
    avg_air_wkg = _safe_nanmean(air_wkg_arr) if air_wkg_arr.size and np.isfinite(air_wkg_arr).any() else np.nan
    avg_air_w = _safe_nanmean(air_w_arr) if air_w_arr.size and np.isfinite(air_w_arr).any() else np.nan

    gain = loss = np.nan
    # v37_fix7: Strava-like gain/loss from FIT altitude (heavier smoothing + thresholding + moving-only)
    elev_for_gain = df["elev_m_clean"].values if "elev_m_clean" in df.columns else df["elev_m"].values
    if np.isfinite(elev_for_gain).any():
        v_for_gain = df["speed_mps"].to_numpy(dtype=float) if "speed_mps" in df.columns else None
        gain, loss = elevation_gain_loss_strava_like(
            elev_for_gain,
            v_for_gain,
            smooth_win=21,
            min_speed_mps=0.5,
            min_step_m=2.0,
        )
    grade_mean_pct = _safe_nanmean(df.loc[moving, "grade"]) * 100.0 if (df["grade"].notna().any() and np.any(moving)) else np.nan

    rf = rf_score_v12(df, hr_corr)
    if not np.isfinite(rf.get("RF_adjusted_mean_W_per_bpm", np.nan)) and np.isfinite(nPower_HR):
        rf["RF_adjusted_mean_W_per_bpm"] = float(nPower_HR)
        rf["RF_adjusted_median_W_per_bpm"] = float(nPower_HR)
        rf["RF_select_code"] = "0.0"
    note = None  # v38: do not emit any HR correction notes (rebuild uses raw HR only)

    # v37: add grade detrend note if applied
    if grade_detrended:
        msg = "grade_detrended"
        if np.isfinite(grade_mean_before) and np.isfinite(grade_mean_after) and np.isfinite(loop_displacement_m):
            msg += f" (mean {grade_mean_before*100:.2f}%->{grade_mean_after*100:.2f}%, disp {loop_displacement_m:.0f}m)"
        note = (note + "; " if note else "") + msg


    # Append any RE-related diagnostics gathered earlier
    try:
        if 're_note_msg' in locals() and re_note_msg:
            note = (note + "; " if note else "") + str(re_note_msg)
    except Exception:
        pass

    out = {
        "date": start_local,
        "file": os.path.basename(fit_path),
        "notes": note,
        "distance_km": round(dist_km, 3) if np.isfinite(dist_km) else np.nan,
        "gps_distance_km": round(gps_km, 3) if np.isfinite(gps_km) else np.nan,
        "gps_coverage": round(gps_cov, 3) if np.isfinite(gps_cov) else np.nan,
        "gps_distance_ratio": round(gps_ratio, 3) if np.isfinite(gps_ratio) else np.nan,
        "gps_max_seg_m": round(gps_max_seg_m, 1) if np.isfinite(gps_max_seg_m) else np.nan,
        "gps_p99_speed_mps": round(gps_p99_speed_mps, 2) if np.isfinite(gps_p99_speed_mps) else np.nan,
        "gps_outlier_frac": round(gps_outlier_frac, 4) if np.isfinite(gps_outlier_frac) else np.nan,
        "speed_source": speed_source,  # v44: track which speed source was used for RE
        "stryd_manufacturer": ids.get("stryd_manufacturer",""),
        "stryd_product": ids.get("stryd_product",""),
        "stryd_serial_number": ids.get("stryd_serial_number",""),
        "stryd_ant_device_number": ids.get("stryd_ant_device_number",""),
        "elapsed_time_s": elapsed_s,
        "moving_time_s": moving_s,
        "moving_time_s_original": moving_time_s_original,
        "moving_time_s_recalc": moving_time_s_recalc,
        "moving_time_recalc_applied": bool(moving_time_recalc_applied),
        "avg_pace_min_per_km": round(pace, 2) if np.isfinite(pace) else np.nan,
        "avg_power_w": round(avg_power, 1) if np.isfinite(avg_power) else np.nan,
        "npower_w": round(nP, 1) if np.isfinite(nP) else np.nan,
        "avg_air_power_wkg": round(avg_air_wkg, 3) if np.isfinite(avg_air_wkg) else np.nan,
        "avg_air_power_w": round(avg_air_w, 1) if np.isfinite(avg_air_w) else np.nan,
        "avg_hr": round(avg_hr, 1) if np.isfinite(avg_hr) else np.nan,
        "max_hr": int(round(max_hr)) if np.isfinite(max_hr) else np.nan,
        "hr_corrected": False,  # v38: rebuild never corrects HR
        "RE_avg": round(RE_avg, 3) if np.isfinite(RE_avg) else np.nan,
        "RE_normalised": round(RE_norm, 3) if np.isfinite(RE_norm) else np.nan,
        "elev_gain_m": round(gain, 1) if np.isfinite(gain) else np.nan,
        "elev_loss_m": round(loss, 1) if np.isfinite(loss) else np.nan,
        "grade_mean_pct": round(grade_mean_pct, 2) if np.isfinite(grade_mean_pct) else np.nan,
        "alt_quality": df.get("alt_quality", pd.Series([""])).iloc[0] if "alt_quality" in df.columns and len(df) else "",
        "nPower_HR": round(nPower_HR, 3) if np.isfinite(nPower_HR) else np.nan,
        **rf,
        "start_time_utc": start_utc,
    }
    # Ensure GPS medians are carried into the per-run row (used for weather stage)
    out["gps_lat_med"] = lat_med
    out["gps_lon_med"] = lon_med
    # Ensure weather columns exist (filled in Weather stage)
    for _c in WEATHER_COLS:
        if _c not in out:
            out[_c] = np.nan

    return out




# ---------- Pending activities (temporary names before Strava sync) ----------
def load_pending_activities(pending_csv: str) -> pd.DataFrame:
    """
    Load pending_activities.csv for temporary activity names.
    
    Returns DataFrame indexed by 'file' column with columns:
        - activity_name (str)
        - shoe (str, optional)
    
    Used when no Strava match exists - provides user-friendly names
    instead of generic "Morning run" / "Afternoon run" / "Evening run".
    """
    if not pending_csv or not Path(pending_csv).exists():
        return pd.DataFrame(columns=['file', 'activity_name', 'shoe']).set_index('file')
    
    try:
        # Try UTF-8 first, fall back to cp1252 (Windows) if that fails
        try:
            df = pd.read_csv(pending_csv, encoding='utf-8')
        except UnicodeDecodeError:
            df = pd.read_csv(pending_csv, encoding='cp1252')
    except Exception as e:
        print(f"Warning: Could not read pending activities file {pending_csv}: {e}")
        return pd.DataFrame(columns=['file', 'activity_name', 'shoe']).set_index('file')
    
    if 'file' not in df.columns:
        print(f"Warning: Pending activities file missing 'file' column")
        return pd.DataFrame(columns=['file', 'activity_name', 'shoe']).set_index('file')
    
    # Ensure columns exist
    if 'activity_name' not in df.columns:
        df['activity_name'] = ''
    if 'shoe' not in df.columns:
        df['shoe'] = ''
    
    # Remove empty rows
    df = df[df['file'].notna() & (df['file'] != '')]
    
    if len(df) > 0:
        print(f"Loaded {len(df)} pending activity names from {pending_csv}")
    
    return df.set_index('file')


# ---------- Strava join ----------
def load_strava(activities_csv: str, tz_local: str) -> pd.DataFrame:
    """
    Load Strava activities.csv and prepare two alternative UTC interpretations of 'Activity Date':
      - ActivityDate_utc_assume_utc: treat the timestamp as UTC
      - ActivityDate_utc_assume_local: treat the timestamp as tz_local (e.g. Europe/Stockholm) then convert to UTC

    This makes matching robust to differing Strava export behaviors across years.
    """
    act = pd.read_csv(activities_csv)

    dt_naive = pd.to_datetime(act.get("Activity Date"), errors="coerce")

    # Interpretation 1: assume UTC
    act["ActivityDate_utc_assume_utc"] = dt_naive.dt.tz_localize("UTC", ambiguous="NaT", nonexistent="NaT")

    # Interpretation 2: assume local timezone then convert to UTC
    try:
        tz = ZoneInfo(str(tz_local))
    except Exception:
        tz = ZoneInfo("Europe/Stockholm")
    act["ActivityDate_utc_assume_local"] = dt_naive.dt.tz_localize(tz, ambiguous="NaT", nonexistent="NaT").dt.tz_convert("UTC")

    act["Distance_km"] = pd.to_numeric(act.get("Distance"), errors="coerce")
    if act["Distance_km"].median(skipna=True) > 1000:
        act["Distance_km"] = act["Distance_km"] / 1000.0

    def _to_seconds(x):
        return pd.to_numeric(x, errors="coerce")

    if "Moving Time" in act.columns:
        act["MovingTime_s"] = _to_seconds(act["Moving Time"])
    elif "Moving Time (seconds)" in act.columns:
        act["MovingTime_s"] = _to_seconds(act["Moving Time (seconds)"])
    else:
        act["MovingTime_s"] = np.nan

    if "Elapsed Time" in act.columns:
        act["ElapsedTime_s"] = _to_seconds(act["Elapsed Time"])
    elif "Elapsed Time (seconds)" in act.columns:
        act["ElapsedTime_s"] = _to_seconds(act["Elapsed Time (seconds)"])
    else:
        act["ElapsedTime_s"] = np.nan

    if "Elevation Gain" in act.columns:
        act["ElevGain_m"] = pd.to_numeric(act["Elevation Gain"], errors="coerce")
    elif "Total Elevation Gain" in act.columns:
        act["ElevGain_m"] = pd.to_numeric(act["Total Elevation Gain"], errors="coerce")
    else:
        act["ElevGain_m"] = np.nan

    if "Elevation Loss" in act.columns:
        act["ElevLoss_m"] = pd.to_numeric(act["Elevation Loss"], errors="coerce")
    else:
        act["ElevLoss_m"] = np.nan

    act["Activity Name Clean"] = act.get("Activity Name", "").map(sanitize_text)
    act["Activity Gear Clean"] = act.get("Activity Gear", pd.Series([None] * len(act))).map(sanitize_text)
    act["Activity Type Lower"] = act.get("Activity Type", "").astype(str).str.lower()
    act = act[act["Activity Type Lower"].str.contains("run")].copy()

    # Prefer rows with any usable timestamp
    act = act[(act["ActivityDate_utc_assume_utc"].notna()) | (act["ActivityDate_utc_assume_local"].notna())].copy()
    print(f"Strava: loaded {len(act)} run activities from {activities_csv}")
    return act



def match_strava(master: pd.DataFrame, act: pd.DataFrame, tz_local: str, pending: pd.DataFrame = None) -> pd.DataFrame:
    """
    Robust Strava matching.

    Strategy:
      1) Compute dt_s = min(|assume_utc - start|, |assume_local - start|)
      2) Prefer close time matches; use distance only as a tie-break (NOT a hard reject) for marathon-like runs.
      3) Progressive widening windows:
           - pass1: dt <= 5 min
           - pass2: dt <= 30 min
           - pass3: dt <= 3 h (time-only fallback)
      4) Tag match quality:
           - 'matched' (time+distance used)
           - 'time_only' (matched on time window only)
           - 'unmatched'
      5) After matching, tag rows whose strava_activity_id is shared by multiple master rows as 'merged_component'
         (these often arise when you merged multiple FITs into one Strava activity).
      6) For unmatched rows, check pending_activities for temporary names before falling back to generic time-based names.
    
    Args:
        master: DataFrame with FIT data
        act: DataFrame with Strava activities
        tz_local: Timezone string
        pending: Optional DataFrame (indexed by 'file') with temporary activity names
    """
    names, shoes, ids = [], [], []
    mt_s, et_s, eg_m, el_m = [], [], [], []
    dist_km = []  # v41: Strava distance
    mtype = []

    # Precompute numpy arrays for dt to speed up repeated operations
    a_utc = act["ActivityDate_utc_assume_utc"]
    a_loc = act["ActivityDate_utc_assume_local"]
    dist_a = pd.to_numeric(act.get("Distance_km"), errors="coerce")

    # Helper: get UTC start timestamp for a master row
    tz = ZoneInfo(str(tz_local)) if tz_local else ZoneInfo("Europe/Stockholm")

    for _, r in master.iterrows():
        t_utc = r.get("start_time_utc", None)

        if t_utc is None or (isinstance(t_utc, float) and not np.isfinite(t_utc)):
            # Fallback to local naive 'date' column if present
            t_local = r.get("date", None)
            if t_local is None or (isinstance(t_local, float) and not np.isfinite(t_local)):
                names.append(r.get("activity_name", None)); shoes.append(r.get("shoe", "")); ids.append(None)
                mt_s.append(np.nan); et_s.append(np.nan); eg_m.append(np.nan); el_m.append(np.nan)
                dist_km.append(np.nan)  # v42
                mtype.append("unmatched")
                continue
            t_local = pd.Timestamp(t_local)
            if getattr(t_local, "tzinfo", None) is None:
                t_local = t_local.tz_localize(tz, ambiguous="NaT", nonexistent="shift_forward")
            t_utc = t_local.tz_convert("UTC")
        else:
            t_utc = pd.Timestamp(t_utc)
            if getattr(t_utc, "tzinfo", None) is None:
                t_utc = t_utc.tz_localize("UTC")
            else:
                t_utc = t_utc.tz_convert("UTC")

        dist = pd.to_numeric(r.get("distance_km", np.nan), errors="coerce")
        elapsed_s = pd.to_numeric(r.get("elapsed_time_s", np.nan), errors="coerce")
        marathon_like = (np.isfinite(dist) and dist >= 38.0) or (np.isfinite(elapsed_s) and elapsed_s >= 9000)

        # distance tolerance for tie-break / filters (non-marathon only)
        tol_km = 0.35
        if marathon_like and np.isfinite(dist):
            tol_km = max(0.35, 0.05 * float(dist))

        # Compute minimal dt across both timestamp interpretations
        dt1 = (a_utc - t_utc).abs().dt.total_seconds()
        dt2 = (a_loc - t_utc).abs().dt.total_seconds()
        dt = np.nanmin(np.vstack([dt1.to_numpy(dtype=float), dt2.to_numpy(dtype=float)]), axis=0)

        # Candidate indices for each pass
        cand_idx = np.where(np.isfinite(dt) & (dt <= 300))[0]
        pass_used = 1
        if cand_idx.size == 0:
            cand_idx = np.where(np.isfinite(dt) & (dt <= 1800))[0]
            pass_used = 2
        if cand_idx.size == 0:
            cand_idx = np.where(np.isfinite(dt) & (dt <= 3 * 3600))[0]
            pass_used = 3

        if cand_idx.size == 0:
            names.append(r.get("activity_name", None)); shoes.append(r.get("shoe", "")); ids.append(None)
            mt_s.append(np.nan); et_s.append(np.nan); eg_m.append(np.nan); el_m.append(np.nan)
            dist_km.append(np.nan)  # v42
            mtype.append("unmatched")
            continue

        cand = act.iloc[cand_idx].copy()
        cand["dt_s"] = dt[cand_idx]

        # Distance handling:
        # - For marathon_like: never hard-filter by distance (London GPS distortion)
        # - For others: filter by distance in pass1/pass2 only; pass3 is time-only.
        if np.isfinite(dist):
            cand["dd_km"] = (pd.to_numeric(cand["Distance_km"], errors="coerce") - float(dist)).abs()
            if (not marathon_like) and pass_used in (1, 2):
                cand = cand[cand["dd_km"] <= tol_km]

        if len(cand) == 0:
            # fall back to time-only within the pass window
            cand = act.iloc[cand_idx].copy()
            cand["dt_s"] = dt[cand_idx]
            cand["dd_km"] = np.nan
            match_type = "time_only"
        else:
            match_type = "matched" if pass_used in (1, 2) else "time_only"

        # Ensure dd_km exists even when master distance is missing (so sort_values cannot KeyError)
        if "dd_km" not in cand.columns:
            cand["dd_km"] = np.nan
        cand = cand.sort_values(["dt_s", "dd_km"], na_position="last")
        best = cand.iloc[0]

        names.append(best.get("Activity Name Clean"))
        shoes.append(best.get("Activity Gear Clean", ""))
        try:
            ids.append(int(best.get("Activity ID")))
        except Exception:
            ids.append(None)
        mt_s.append(float(best.get("MovingTime_s", np.nan)))
        et_s.append(float(best.get("ElapsedTime_s", np.nan)))
        eg_m.append(float(best.get("ElevGain_m", np.nan)))
        el_m.append(float(best.get("ElevLoss_m", np.nan)))
        dist_km.append(float(best.get("Distance_km", np.nan)))  # v41: Strava distance
        mtype.append(match_type)

    out = master.copy()
    out["activity_name"] = names
    out["shoe"] = shoes
    out["strava_activity_id"] = ids
    out["strava_moving_time_s"] = mt_s
    out["strava_elapsed_time_s"] = et_s
    out["strava_elev_gain_m"] = eg_m
    out["strava_elev_loss_m"] = el_m
    out["strava_distance_km"] = dist_km  # v41: Strava corrected distance
    out["strava_match_type"] = mtype

    # Tag merged components (multiple master rows map to same Strava activity id)
    if "strava_activity_id" in out.columns:
        dup = out["strava_activity_id"].duplicated(keep=False) & out["strava_activity_id"].notna()
        out.loc[dup, "strava_match_type"] = "merged_component"

    # v41: For unmatched rows, check pending_activities before falling back to generic names
    if pending is not None and len(pending) > 0:
        pending_used = 0
        for idx, row in out.iterrows():
            if pd.isna(row.get("activity_name")) or row.get("activity_name") == "":
                fit_file = row.get("file", "")
                if fit_file in pending.index:
                    pend_row = pending.loc[fit_file]
                    # v49: Handle duplicate index entries (loc returns DataFrame instead of Series)
                    if isinstance(pend_row, pd.DataFrame):
                        pend_row = pend_row.iloc[0]
                    pend_name = pend_row.get("activity_name")
                    if pd.notna(pend_name) and pend_name != "":
                        out.at[idx, "activity_name"] = pend_name
                        pending_used += 1
                    pend_shoe = pend_row.get("shoe")
                    if pd.notna(pend_shoe) and pend_shoe != "" and (pd.isna(row.get("shoe")) or row.get("shoe") == ""):
                        out.at[idx, "shoe"] = pend_shoe
        if pending_used > 0:
            print(f"  Applied {pending_used} pending activity names")

    out["activity_name"] = out["activity_name"].fillna(out["date"].apply(lambda x: time_bucket(pd.Timestamp(x))))
    return out


# ---------- Excel output ----------
def write_master(template_xlsx: str, out_xlsx: str, df: pd.DataFrame, df_fail: pd.DataFrame | None = None):
    """
    Write the rebuilt master to Excel using the template for styling.
    - Preserves template column order where possible
    - Inserts GPS quality columns after distance_km
    - Inserts calibration columns after RE_normalised
    - Appends v32 extra columns (HF + sim power + moving_time fix flags) if present
    - De-duplicates headers to avoid duplicate columns
    """
    wb_t = openpyxl.load_workbook(template_xlsx)
    ws_t = wb_t["Master_Rebuilt"]
    orig_headers = [c.value for c in ws_t[1]]

    # Drop legacy columns we no longer want, but keep 'file' for traceability (move to end)
    base_headers = [h for h in orig_headers if h not in ("session_type", "file")]

    out_cols: list[str] = []
    for h in base_headers:
        out_cols.append(h)
        if h == "distance_km":
            out_cols.extend(["gps_distance_km","gps_coverage","gps_distance_ratio","gps_max_seg_m","gps_p99_speed_mps","gps_outlier_frac","speed_source"])
        if h == "RE_normalised":
            out_cols.extend(["calibration_era_id","power_adjuster_to_S4"])

    # Ensure weather columns exist in output
    for wcol in WEATHER_COLS:
        if wcol not in out_cols:
            out_cols.append(wcol)

    # Extra v32 columns (only added if present in df)

    # Minimal extra columns (keep master readable)
    v33_cols = [
        "time_source",     # 'strava' | 'fit' | 'repaired'
        "elev_source",     # 'strava' | 'fit'
        "alt_quality",     # 'ok' | 'corrected' | 'flat_fallback'
        "power_source",    # 'stryd' | 'sim_v1' (future) | ''
        "sim_model_id",    # model/version id (future)
        "sim_ok",          # TRUE/FALSE (future)
        "hf_parkrun",       # Highbury Fields parkrun flag

        # Strava join outputs (match_strava)
        "strava_match_type",     # matched|time_only|merged_component|unmatched
        "strava_activity_id",
        "strava_distance_km",    # v41: Strava corrected distance
        "strava_moving_time_s",
        "strava_elapsed_time_s",
        "strava_elev_gain_m",
        "strava_elev_loss_m",
    ]
    for c in v33_cols:
        if c in df.columns and c not in out_cols:
            out_cols.append(c)

    # Move Stryd ID columns to the end for readability (but keep before 'file')
    STRYD_COLS = ['stryd_manufacturer','stryd_product','stryd_serial_number','stryd_ant_device_number']
    out_cols = [c for c in out_cols if c not in STRYD_COLS] + [c for c in STRYD_COLS if c in df.columns or c in out_cols]

    # Required lead columns + file at end
    new_headers = ["date", "activity_name", "shoe"] + [h for h in out_cols if h not in ("date","activity_name","shoe")] + ["file"]

    # Do not auto-append extra dataframe columns; keep master schema stable/readable.
    # De-duplicate while preserving order
    seen = set()
    new_headers = [h for h in new_headers if not (h in seen or seen.add(h))]

    # Round RF columns as agreed
    rf_cols = [
        "RF_adjusted_mean_W_per_bpm",
        "RF_adjusted_median_W_per_bpm",
        "RF_dead_frac",
        "RF_drift_pct_per_min",
        "RF_drift_r2",
    ]
    for c in rf_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").round(3)

    # Ensure Strava join columns exist
    if "activity_name" not in df.columns:
        df["activity_name"] = df["date"].apply(lambda x: time_bucket(pd.Timestamp(x)))
    if "shoe" not in df.columns:
        df["shoe"] = ""

    wb_new = openpyxl.Workbook()
    ws = wb_new.active
    ws.title = "Master_Rebuilt"

    col_idx_template = {h: i + 1 for i, h in enumerate(orig_headers)}

    def _template_style_col(h: str) -> int:
        if h in col_idx_template:
            return col_idx_template[h]
        # fall back to distance_km styling (generic numeric column)
        return col_idx_template.get("distance_km", 1)

    # Header row with template styles
    for j, h in enumerate(new_headers, start=1):
        src_col = _template_style_col(h)
        src = ws_t.cell(row=1, column=src_col)
        src_w = ws_t.column_dimensions[get_column_letter(src_col)].width

        if h == "activity_name":
            src_w = 28
        elif h == "shoe":
            src_w = 24
        elif h.startswith("gps_"):
            src_w = max(src_w or 12, 14)
        elif h.startswith("stryd_"):
            src_w = max(src_w or 12, 18)
        elif h in ("time_source","elev_source","alt_quality","power_source","sim_model_id","sim_ok"):
            src_w = max(src_w or 12, 14)

        c = ws.cell(row=1, column=j, value=h)
        c.font = copy(src.font)
        c.fill = copy(src.fill)
        c.border = copy(src.border)
        c.alignment = copy(src.alignment)
        c.number_format = src.number_format
        c.protection = copy(src.protection)
        ws.column_dimensions[get_column_letter(j)].width = src_w

    ws.row_dimensions[1].height = ws_t.row_dimensions[1].height

    # Data rows
    records = df.to_dict("records")
    for i, r in enumerate(records, start=2):
        for j, h in enumerate(new_headers, start=1):
            src_col = _template_style_col(h)
            src = ws_t.cell(row=2, column=src_col)

            cell = ws.cell(row=i, column=j)
            cell.font = copy(src.font)
            cell.fill = copy(src.fill)
            cell.border = copy(src.border)
            cell.alignment = copy(src.alignment)
            cell.number_format = src.number_format
            cell.protection = copy(src.protection)

            val = r.get(h)

            # Excel-friendly TRUE/FALSE for booleans
            if h in ("hr_corrected", "RF_window_shifted", "sim_ok") and val is not None:
                val = "TRUE" if bool(val) else "FALSE"

            if isinstance(val, float) and not np.isfinite(val):
                val = None
            cell.value = val

    # Widen date column
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 22)

    # Failures sheet: prefer current-run failures; otherwise copy from template if present.
    if df_fail is not None and isinstance(df_fail, pd.DataFrame) and (len(df_fail) > 0):
        ws_fail = wb_new.create_sheet("Failures")
        ws_fail.append(["file", "reason"])
        for _, rr in df_fail.iterrows():
            ws_fail.append([rr.get("file"), rr.get("reason")])
        ws_fail.column_dimensions["A"].width = 48
        ws_fail.column_dimensions["B"].width = 90
    elif "Failures" in wb_t.sheetnames:
        ws_ft = wb_t["Failures"]
        ws_fail = wb_new.create_sheet("Failures")
        for row in ws_ft.iter_rows():
            for cell0 in row:
                dst = ws_fail.cell(row=cell0.row, column=cell0.col_idx, value=cell0.value)
                dst.font = copy(cell0.font)
                dst.fill = copy(cell0.fill)
                dst.border = copy(cell0.border)
                dst.alignment = copy(cell0.alignment)
                dst.number_format = cell0.number_format
                dst.protection = copy(cell0.protection)
        for col, dim in ws_ft.column_dimensions.items():
            ws_fail.column_dimensions[col].width = dim.width

    wb_new.save(out_xlsx)


# ---------- main ----------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--fit-zip", required=True, help="Path to ZIP containing .fit files")
    ap.add_argument("--template", default="", help="Path to Master_Rebuilt.xlsx template (required unless --cache-only)")
    ap.add_argument("--out", default="", help="Output master .xlsx path (required unless --cache-only)")
    ap.add_argument("--strava", default="", help="Optional Strava activities.csv for name/shoe join")
    ap.add_argument("--activities", default="", help="Alias for --strava (activities.csv)")
    ap.add_argument("--tz", default="Europe/Stockholm", help="Timezone for local 'date' column (default Europe/Stockholm)")
    ap.add_argument("--persec-cache-dir", default="", help="Directory to write per-second caches (.npz) for RE/sim-power pipeline")

    # v39: incremental cache-update mode
    ap.add_argument("--cache-only", action="store_true", help="Only write/update per-second cache (.npz); skip master rebuild and weather")
    ap.add_argument("--cache-incremental", dest="cache_incremental", action="store_true", default=True,
                    help="Skip runs that already have a cache file (default)")
    ap.add_argument("--no-cache-incremental", dest="cache_incremental", action="store_false",
                    help="Disable incremental cache skipping (rewrite all)")
    ap.add_argument("--cache-rewrite-if-newer", dest="cache_rewrite_if_newer", action="store_true", default=True,
                    help="Rewrite cache if FIT appears newer than cache (default)")
    ap.add_argument("--no-cache-rewrite-if-newer", dest="cache_rewrite_if_newer", action="store_false",
                    help="Disable rewrite-if-newer check")
    ap.add_argument("--cache-force", action="store_true", help="Always rewrite cache")
    ap.add_argument("--append-master-in", default="", help="Existing master .xlsx to append only new runs (skip full rebuild).")
    ap.add_argument("--weather-cache-db", default="", help="Optional SQLite cache for Open-Meteo hourly data (recommended). If blank, uses <out_dir>/_weather_cache_openmeteo/openmeteo_cache.sqlite")
    ap.add_argument("--override-file", default="activity_overrides.xlsx", help="Activity overrides Excel file (indoor runs identified by temp_override column)")
    ap.add_argument("--pending-activities", default="", help="Optional CSV for temporary activity names (file,activity_name,shoe). Used when no Strava match. If blank, auto-uses pending_activities.csv if present.")
    ap.add_argument("--weight", type=float, default=WEIGHT_KG_DEFAULT, help="Default weight kg used for W/kg calculations")
    ap.add_argument("--gps-outlier-speed", type=float, default=7.0, help="Speed threshold (m/s) for gps_outlier_frac (default 7.0)")
    args = ap.parse_args()
    if getattr(args, "activities", "") and not getattr(args, "strava", ""):
        args.strava = args.activities

    # Set cache controls (globals used by summarize_fit)
    global CACHE_INCREMENTAL, CACHE_REWRITE_IF_NEWER, CACHE_FORCE
    CACHE_INCREMENTAL = bool(getattr(args, "cache_incremental", True))
    CACHE_REWRITE_IF_NEWER = bool(getattr(args, "cache_rewrite_if_newer", True))
    CACHE_FORCE = bool(getattr(args, "cache_force", False))

    if args.cache_only:
        if not args.persec_cache_dir:
            raise SystemExit("--cache-only requires --persec-cache-dir")
    else:
        if not args.template or not args.out:
            raise SystemExit("--template and --out are required unless --cache-only")

    print(f"RUNNING SCRIPT: {__file__}")

    if args.cache_only:
        # In cache-only mode we don't necessarily have an --out; anchor temp extract folder next to cache dir.
        out_dir = os.path.dirname(os.path.abspath(args.persec_cache_dir)) if args.persec_cache_dir else os.getcwd()
    else:
        out_dir = os.path.dirname(os.path.abspath(args.out))
    extract_dir = os.path.join(out_dir, "_fit_extract")
    os.makedirs(extract_dir, exist_ok=True)

    # v41: Load pending activities (temporary names before Strava sync)
    pending_csv = args.pending_activities
    if not pending_csv:
        # Auto-use pending_activities.csv if present in out_dir
        auto_pending = os.path.join(out_dir, "pending_activities.csv")
        if os.path.exists(auto_pending):
            pending_csv = auto_pending
    pending_activities = load_pending_activities(pending_csv) if pending_csv else None

    # Clear extracted FITs from previous runs
    for root, _, files in os.walk(extract_dir):
        for f in files:
            if f.lower().endswith(".fit"):
                try:
                    os.remove(os.path.join(root, f))
                except OSError:
                    pass

    # v39.9: append mode (update existing master with only new run_ids)
    append_mode = (not args.cache_only) and bool(getattr(args, "append_master_in", ""))
    base_master_df = None
    known_run_ids = set()
    if append_mode:
        if not os.path.exists(args.append_master_in):
            print(f"WARNING: --append-master-in not found: {args.append_master_in}. Disabling append mode.")
            append_mode = False
        else:
            try:
                base_master_df = pd.read_excel(args.append_master_in)
                if "run_id" in base_master_df.columns:
                    known_run_ids = set(base_master_df["run_id"].astype(str))
                elif "file" in base_master_df.columns:
                    known_run_ids = set(base_master_df["file"].astype(str))
                else:
                    known_run_ids = set()
                print(f"Append mode: base master rows={len(base_master_df)}; known run_ids={len(known_run_ids)}")
            except Exception as e:
                print("WARNING: Could not read --append-master-in master; disabling append mode:", repr(e))
                append_mode = False
                base_master_df = None
                known_run_ids = set()


    # --- Extract FITs ---
    # In cache-only + incremental mode, extract only those that need cache updates.
    with zipfile.ZipFile(args.fit_zip, "r") as z:
        if args.cache_only and args.persec_cache_dir and CACHE_INCREMENTAL:
            todo = 0
            for info in z.infolist():
                n = info.filename
                if not n.lower().endswith(".fit"):
                    continue
                # v41.1: Use sanitized FIT filename for cache path
                fit_filename_no_ext = os.path.splitext(os.path.basename(n))[0]
                safe_name = fit_filename_no_ext.replace('&', '_').replace('?', '_').replace('=', '_')
                cache_path = os.path.join(args.persec_cache_dir, safe_name + ".npz")
                # Use zip internal modified time for rewrite-if-newer checks.
                need = True
                if not CACHE_FORCE and os.path.exists(cache_path):
                    need = False
                    if CACHE_REWRITE_IF_NEWER:
                        try:
                            zdt = dt.datetime(*info.date_time)
                            z_epoch = zdt.replace(tzinfo=dt.timezone.utc).timestamp()
                            c_epoch = os.path.getmtime(cache_path)
                            if z_epoch > c_epoch + 1.0:
                                need = True
                        except Exception:
                            need = True
                if need:
                    z.extract(n, extract_dir)
                    todo += 1
            print(f"Cache-only incremental: extracted {todo} new/updated FIT(s) into {extract_dir}")
        else:
            extracted = 0
            skipped_known = 0
            for info in z.infolist():
                n = info.filename
                if not n.lower().endswith(".fit"):
                    continue
                if append_mode and known_run_ids:
                    run_id = os.path.splitext(os.path.basename(n))[0]
                    if (run_id in known_run_ids) or (os.path.basename(n) in known_run_ids):
                        skipped_known += 1
                        continue
                z.extract(n, extract_dir)
                extracted += 1
            if append_mode:
                print(f"Append mode: extracted {extracted} new FIT(s) into {extract_dir} (skipped known={skipped_known})")


    fit_paths = []
    for root, _, files in os.walk(extract_dir):
        for f in files:
            if f.lower().endswith(".fit"):
                fit_paths.append(os.path.join(root, f))
    fit_paths.sort()

    # Append mode: if nothing new was extracted, optionally refresh Strava join and write out (no heavy recompute).
    if append_mode and base_master_df is not None and len(fit_paths) == 0:
        print("Append mode: no new FITs to append.")
        df_out = base_master_df.copy()

        # Optionally (re)run Strava match even when there are no new FITs.
        # This is important when older masters were created without Strava columns.
        if args.strava:
            try:
                act = load_strava(args.strava, args.tz)
                df_out = match_strava(df_out, act, args.tz, pending=pending_activities)
                try:
                    vc = df_out.get("strava_match_type").value_counts(dropna=False)
                    print("Strava match summary:", dict(vc))
                except Exception:
                    pass
            except Exception as e:
                print("WARNING: Strava join failed in append mode:", repr(e))

        # Write output (do NOT just copy the base master, as that would drop newly-created Strava columns)
        # v41: Sort by date before writing (fixes chronological order after manual additions)
        df_out = df_out.sort_values("date").reset_index(drop=True)
        out_path = args.out if args.out else args.append_master_in
        try:
            write_master(args.template, out_path, df_out, df_fail=pd.DataFrame())
            print(f"Append mode complete. Wrote: {out_path}")
        except Exception as e:
            print("ERROR: Append mode write failed:", repr(e))
            raise
        return 0

    if args.cache_only:
        # Only write caches; no master build.
        total = len(fit_paths)
        if total == 0:
            print("Cache-only: no FITs to process (nothing extracted).")
            return 0
        wrote = 0
        skipped = 0
        failures = 0
        for i, fp in enumerate(fit_paths, start=1):
            try:
                # Force a cache write for each extracted FIT in this mode
                old_force = CACHE_FORCE
                CACHE_FORCE = True
                rec = summarize_fit(fp, args.tz, args.weight, args.gps_outlier_speed, args.persec_cache_dir, cache_only=True)
                CACHE_FORCE = old_force
                if rec.get("cache_written", False):
                    wrote += 1
                else:
                    skipped += 1
                if (i % 50) == 0 or i == total:
                    print(f"Cache-only progress: {i}/{total} (wrote {wrote}, skipped {skipped}, failures {failures})")
            except Exception as e:
                failures += 1
                print(f"Cache-only failed ({i}/{total}): {os.path.basename(fp)}  |  {repr(e)}")
        print(f"Cache-only done: processed {total} FIT(s); wrote {wrote}, skipped {skipped}, failures {failures}.")
        return 0

    total = len(fit_paths)
    rows = []
    failures = []
    for i, fp in enumerate(fit_paths, start=1):
        try:
            rec = summarize_fit(fp, args.tz, args.weight, args.gps_outlier_speed, args.persec_cache_dir)
            try:
                dt_str = pd.Timestamp(rec.get("date")).strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                dt_str = str(rec.get("date"))
            print(f"Processed ({i}/{total}): {dt_str}  |  {rec.get('file')}")
            rows.append(rec)
        except SkipFitFile as e:
            failures.append({"file": os.path.basename(fp), "reason": str(e)})
            print(f"Skipped ({i}/{total}): {os.path.basename(fp)}  |  {e}")
        except Exception as e:
            failures.append({"file": os.path.basename(fp), "reason": f"ERROR: {repr(e)}"})
            print(f"Failed  ({i}/{total}): {os.path.basename(fp)}  |  {repr(e)}")

    df = pd.DataFrame(rows).sort_values("date").reset_index(drop=True)
    if append_mode and base_master_df is not None:
        df = pd.concat([base_master_df, df], ignore_index=True)

    df_fail = pd.DataFrame(failures) if failures else pd.DataFrame(columns=['file','reason'])

    # ---------- Weather (Open-Meteo, zero setup) ----------
    # Speed: group runs by rounded location and fetch one hourly block per location (cached on disk).
    wx_cache_dir = os.path.join(out_dir, "_weather_cache_openmeteo")

    wx_db_path = args.weather_cache_db.strip() if hasattr(args, "weather_cache_db") and args.weather_cache_db else os.path.join(wx_cache_dir, "openmeteo_cache.sqlite")
    try:
        _wx_db_init(wx_db_path)
        print(f"Weather SQLite cache: {wx_db_path}")
    except Exception as e:
        wx_db_path = ""
        print(f"Weather SQLite cache disabled (init failed): {e}")


    for c in WEATHER_COLS:
        if c not in df.columns:
            df[c] = np.nan

    gps_lat_nonnull = pd.to_numeric(df.get("gps_lat_med"), errors="coerce")
    gps_lon_nonnull = pd.to_numeric(df.get("gps_lon_med"), errors="coerce")
    have_gps_loc = int(np.isfinite(gps_lat_nonnull).sum()) if gps_lat_nonnull is not None else 0
    print(f"GPS location medians present: {have_gps_loc}/{len(df)}")
    if have_gps_loc > 0:
        sample_lat = pd.to_numeric(df.get("gps_lat_med"), errors="coerce").dropna().head(3).to_list()
        sample_lon = pd.to_numeric(df.get("gps_lon_med"), errors="coerce").dropna().head(3).to_list()
        print(f"First 3 gps_lat_med/gps_lon_med: {list(zip(sample_lat, sample_lon))}")

    wx_filled = 0
    wx_skipped_no_gps = 0
    wx_failed = 0
    first_wx_error = None

    # Prepare per-run UTC-naive windows
    start_naive = []
    end_naive = []
    for _, r in df.iterrows():
        st = r.get("start_time_utc", None)
        el = pd.to_numeric(r.get("elapsed_time_s"), errors="coerce")
        if st is None or (isinstance(st, float) and not np.isfinite(st)) or not (np.isfinite(el) and el > 0):
            start_naive.append(None); end_naive.append(None); continue
        st = pd.Timestamp(st)
        if getattr(st, "tzinfo", None) is not None:
            st = st.tz_convert("UTC").tz_localize(None)
        en = st + pd.Timedelta(seconds=float(el))
        start_naive.append(st); end_naive.append(en)

    df["_wx_start_utc"] = start_naive
    df["_wx_end_utc"] = end_naive

    # Group indices by rounded location AND year.
    # This avoids fetching 13 years of hourly data for a single location —
    # each sub-group only needs ~1 API call (≤366 days) instead of ~13.
    loc_year_key = []
    for i, r in df.iterrows():
        lat = pd.to_numeric(r.get("gps_lat_med"), errors="coerce")
        lon = pd.to_numeric(r.get("gps_lon_med"), errors="coerce")
        st = df.at[i, "_wx_start_utc"] if "_wx_start_utc" in df.columns else None
        if not (np.isfinite(lat) and np.isfinite(lon)) or st is None:
            loc_year_key.append(None)
        else:
            yr = pd.Timestamp(st).year
            loc_year_key.append((float(np.round(lat, WX_GPS_ROUND_DP)), float(np.round(lon, WX_GPS_ROUND_DP)), yr))
    df["_wx_loc"] = loc_year_key

    groups = {}
    for i, k in enumerate(df["_wx_loc"].tolist()):
        if k is None:
            continue
        groups.setdefault(k, []).append(i)

    total_groups = len(groups)
    done_groups = 0

    for (lat_r, lon_r, year), idxs in groups.items():
        done_groups += 1
        # date window for this location+year group (pad 1 day)
        starts = [df.at[i, "_wx_start_utc"] for i in idxs if df.at[i, "_wx_start_utc"] is not None]
        ends = [df.at[i, "_wx_end_utc"] for i in idxs if df.at[i, "_wx_end_utc"] is not None]
        if not starts or not ends:
            continue
        date_min = (min(starts).normalize() - pd.Timedelta(days=1)).date().isoformat()
        date_max = (max(ends).normalize() + pd.Timedelta(days=1)).date().isoformat()

        print(f"Weather group {done_groups}/{total_groups}: lat={lat_r:.1f}, lon={lon_r:.1f}, {date_min}..{date_max} ({len(idxs)} runs)")

        # Fetch per-location hourly data efficiently:
        # - Archive for historical (clamped to today)
        # - Forecast only for the recent tail if the group window extends beyond today (e.g. "today" runs with +1 day padding)
        try:
            today = pd.Timestamp.utcnow().date()

            # Archive window (cannot exceed today)
            date_max_archive = pd.to_datetime(_clamp_end_to_today(date_max)).date()

            dfh_all = pd.DataFrame()

            # Chunk archive requests to avoid oversized windows
            for s_iso, e_iso in _split_date_ranges(date_min, date_max_archive.isoformat(), max_days=366):
                data = fetch_open_meteo_hourly_archive(lat_r, lon_r, s_iso, e_iso, wx_cache_dir)
                dfh = _hourly_to_df((data.get("hourly") or {}))
                if wx_db_path and not dfh.empty:
                    _wx_db_upsert_hourly(wx_db_path, lat_r, lon_r, dfh)
                if not dfh.empty:
                    dfh_all = pd.concat([dfh_all, dfh], ignore_index=True)

            if not dfh_all.empty:
                dfh_all = dfh_all.sort_values("time").drop_duplicates(subset=["time"], keep="last").reset_index(drop=True)

            # Fetch FORECAST for the recent tail (archive can lag for the last day or two)
            date_max_req = pd.to_datetime(date_max).date()
            date_min_req = pd.to_datetime(date_min).date()
            # If this location-group touches the last 10 days (including today), fetch forecast for that recent span.
            if date_max_req >= (today - dt.timedelta(days=10)):
                recent_start = max(date_min_req, today - dt.timedelta(days=10))
                s2 = recent_start.isoformat()
                e2 = max(date_max_req, today).isoformat()
                data2 = fetch_open_meteo_hourly_forecast(lat_r, lon_r, s2, e2, wx_cache_dir)
                dfh2 = _hourly_to_df((data2.get("hourly") or {}))
                if wx_db_path and not dfh2.empty:
                    _wx_db_upsert_hourly(wx_db_path, lat_r, lon_r, dfh2)
                if not dfh2.empty:
                    dfh_all = pd.concat([dfh_all, dfh2], ignore_index=True)
                    dfh_all = dfh_all.sort_values("time").drop_duplicates(subset=["time"], keep="last").reset_index(drop=True)

            hourly = _df_to_hourly(dfh_all)

        except Exception as e:
            wx_failed += len(idxs)
            if first_wx_error is None:
                first_wx_error = ("GROUP", lat_r, lon_r, date_min, repr(e))
            continue

        # Fill each run from the group hourly block
        for j, i in enumerate(idxs, start=1):
            st = df.at[i, "_wx_start_utc"]
            en = df.at[i, "_wx_end_utc"]
            if st is None or en is None:
                wx_failed += 1
                continue
            try:
                wx = compute_weather_averages_from_hourly(hourly, st, en)
            except Exception as e:
                wx_failed += 1
                if first_wx_error is None:
                    first_wx_error = (i, lat_r, lon_r, st, repr(e))
                continue

            if all((not np.isfinite(wx[c])) for c in WEATHER_COLS):
                # Per-run fallback (handles archive lag + tight window)
                try:
                    wx = compute_weather_averages(lat_r, lon_r, st, en, wx_cache_dir)
                except Exception:
                    pass
            if all((not np.isfinite(wx[c])) for c in WEATHER_COLS):
                # Attempt forecast fallback for this run only
                try:
                    date_min2 = (st.normalize() - pd.Timedelta(days=1)).date().isoformat()
                    date_max2 = (en.normalize() + pd.Timedelta(days=1)).date().isoformat()
                    data2 = fetch_open_meteo_hourly_forecast(lat_r, lon_r, date_min2, date_max2, wx_cache_dir)
                    wx2 = compute_weather_averages_from_hourly((data2.get("hourly") or {}), st, en)
                    wx = wx2
                except Exception:
                    pass

            if all((not np.isfinite(wx[c])) for c in WEATHER_COLS):
                wx_failed += 1
                continue

            for k, v in wx.items():
                df.at[i, k] = v
            wx_filled += 1

            if j % 25 == 0:
                print(f"  ...filled {j}/{len(idxs)} in this group (total filled={wx_filled})")

    # Runs without GPS
    wx_skipped_no_gps = int(df["_wx_loc"].isna().sum())

    print(f"Weather(Open-Meteo): filled={wx_filled}, skipped_no_gps={wx_skipped_no_gps}, failed={wx_failed}")

    # Round weather outputs for readability (agreed: wind speed 1 dp, others 0 dp)
    if "avg_wind_ms" in df.columns:
        df["avg_wind_ms"] = pd.to_numeric(df["avg_wind_ms"], errors="coerce").round(1)
    for c in ["avg_temp_c", "avg_humidity_pct", "avg_wind_dir_deg"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").round(0)
    if first_wx_error is not None:
        print(f"First weather error: {first_wx_error}")


    # Clean helper columns
    df = df.drop(columns=["_wx_start_utc","_wx_end_utc","_wx_loc"], errors="ignore")


    if args.strava:
        act = load_strava(args.strava, args.tz)
        df = match_strava(df, act, args.tz, pending=pending_activities)
        try:
            vc = df.get("strava_match_type").value_counts(dropna=False)
            print("Strava match summary:", dict(vc))
        except Exception:
            pass


    # ---------- Strava canonical overrides (time + elevation) ----------
    # Keep ONE truth in the master:
    #   - elapsed_time_s, moving_time_s from Strava when present/sane, else FIT/repaired
    #   - elev_gain_m/elev_loss_m from Strava for pre-Stryd when present, else FIT-derived
    df["time_source"] = "fit"
    df["elev_source"] = "fit"

    # Canonical times: Strava if available and sane
    mt = pd.to_numeric(df.get("strava_moving_time_s"), errors="coerce")
    et = pd.to_numeric(df.get("strava_elapsed_time_s"), errors="coerce")
    mt_fit = pd.to_numeric(df.get("moving_time_s"), errors="coerce")
    et_fit = pd.to_numeric(df.get("elapsed_time_s"), errors="coerce")
    dist_m = pd.to_numeric(df.get("distance_km"), errors="coerce") * 1000.0

    speed_imp = dist_m / mt.replace(0, np.nan)
    sane_mt = mt.notna() & (mt > 0) & ((et.isna()) | (mt <= et * 1.02)) & (speed_imp < 7.0)
    use_strava_time = sane_mt

    df.loc[use_strava_time, "moving_time_s"] = mt[use_strava_time]
    df.loc[use_strava_time & et.notna(), "elapsed_time_s"] = et[use_strava_time & et.notna()]
    df.loc[use_strava_time, "time_source"] = "strava"

    # Canonical elevation totals: for pre-Stryd runs, prefer Strava where available
    eg = pd.to_numeric(df.get("strava_elev_gain_m"), errors="coerce")
    el = pd.to_numeric(df.get("strava_elev_loss_m"), errors="coerce")
    is_pre = (df.get("calibration_era_id", "") == "pre_stryd") if "calibration_era_id" in df.columns else (pd.to_datetime(df["date"], errors="coerce") < pd.Timestamp("2017-05-05"))

    use_strava_eg = is_pre & eg.notna()
    use_strava_el = is_pre & el.notna()

    df.loc[use_strava_eg, "elev_gain_m"] = eg[use_strava_eg]
    df.loc[use_strava_el, "elev_loss_m"] = el[use_strava_el]
    df.loc[use_strava_eg | use_strava_el, "elev_source"] = "strava"

    # Drop Strava helper columns to keep master tidy
    drop_cols = [
    "moving_time_s_fit","elapsed_time_s_fit","ascent_m_fit","descent_m_fit",
    "moving_time_source","moving_time_s_best","ascent_m_best","descent_m_best",
    "gps_speed_mps","weather_source","is_indoor"
    ]
    df = df.drop(columns=[c for c in drop_cols if c in df.columns], errors="ignore")


    # ---------- compute per-era power adjusters to baseline (Stryd 4.0) ----------
    # These adjusters align power-derived metrics across settings/pod eras.
    # IMPORTANT: we do not overwrite any raw power; we only emit a per-run multiplier.

    # v44: Load era config (consolidated 6-era model)
    out_dir_path = str(Path(args.out).parent) if args.out else "."
    load_era_config(out_dir_path)
    print_era_config()
    
    # Assign calibration era using date-based approach (v44: no early/late splits)
    df["calibration_era_id"] = pd.to_datetime(df["date"], errors="coerce").apply(assign_calibration_era_v44)

    # Assign coarse pod model (date-based) for fallback if an era has too little data
    def _pod_model(ts: pd.Timestamp) -> str:
        t = pd.Timestamp(ts)
        if t < pd.Timestamp("2017-05-05"):
            return "Pre-Stryd"
        if t < pd.Timestamp("2017-09-12"):
            return "Stryd v1"
        if t < pd.Timestamp("2019-09-07"):
            return "Replacement (v1?)"
        if t < pd.Timestamp("2023-01-03"):
            return "Air Power model"
        if t < pd.Timestamp("2025-12-17"):
            return "Stryd 4.0"
        return "Stryd 5.0"

    df["pod_model"] = pd.to_datetime(df["date"], errors="coerce").apply(_pod_model)

    # Compute GPS speed used for calibration
    df["gps_speed_mps"] = (pd.to_numeric(df.get("gps_distance_km"), errors="coerce") * 1000.0) / pd.to_numeric(df.get("moving_time_s"), errors="coerce")

    # GPSQ + steady-ish cohort for fitting (relaxed)
    cohort = df.copy()
    cohort = cohort[np.isfinite(cohort["gps_speed_mps"]) & (cohort["gps_speed_mps"] > 0)]
    cohort = cohort[np.isfinite(pd.to_numeric(cohort.get("npower_w"), errors="coerce")) & (pd.to_numeric(cohort.get("npower_w"), errors="coerce") > 0)]
    cohort = cohort[pd.to_numeric(cohort.get("gps_distance_ratio"), errors="coerce").between(0.98, 1.02)]
    cohort = cohort[pd.to_numeric(cohort.get("gps_outlier_frac"), errors="coerce") <= 0.005]
    cohort = cohort[pd.to_numeric(cohort.get("gps_p99_speed_mps"), errors="coerce") <= 6.5]
    cohort = cohort[pd.to_numeric(cohort.get("gps_max_seg_m"), errors="coerce") <= 80]
    cohort = cohort[pd.to_numeric(cohort.get("RF_dead_frac"), errors="coerce") <= 0.08]
    cohort = cohort[pd.to_numeric(cohort.get("moving_time_s"), errors="coerce") >= 1200]
    cohort = cohort[pd.to_numeric(cohort.get("avg_pace_min_per_km"), errors="coerce").between(4.3, 5.8)]
    cohort = cohort[pd.to_numeric(cohort.get("grade_mean_pct"), errors="coerce").abs() <= 2.0]

    v_std = 3.333333333  # 5:00/km
    logv = np.log(v_std)

    def _predict_at_v(sub: pd.DataFrame) -> float:
        x = np.log(sub["gps_speed_mps"].to_numpy(dtype=float))
        y = np.log(pd.to_numeric(sub["npower_w"], errors="coerce").to_numpy(dtype=float))
        msk = np.isfinite(x) & np.isfinite(y)
        if msk.sum() < 10:
            return np.nan
        X = np.vstack([np.ones(msk.sum()), x[msk]]).T
        beta, *_ = np.linalg.lstsq(X, y[msk], rcond=None)
        a, b = beta[0], beta[1]
        return float(np.exp(a + b * logv))

        # Era-level predictions (and sample sizes) for power adjusters
    preds_era = {}
    n_era = {}
    for era, sub in cohort.groupby("calibration_era_id"):
        if era == "" or len(sub) < 15:
            continue
        pred = _predict_at_v(sub)
        if np.isfinite(pred) and pred > 0:
            preds_era[era] = float(pred)
            n_era[era] = int(len(sub))

    # Pod-level predictions fallback (and sample sizes)
    preds_pod = {}
    n_pod = {}
    for pm, sub in cohort.groupby("pod_model"):
        if len(sub) < 40:
            continue
        pred = _predict_at_v(sub)
        if np.isfinite(pred) and pred > 0:
            preds_pod[pm] = float(pred)
            n_pod[pm] = int(len(sub))

    # Baseline prediction: S4 is the anchor (power_adjuster_to_S4 = 1.0)
    # v44: Changed from "s4_late" to "s4" since early/late splits are removed
    base_pred = np.nan
    if "s4" in preds_era and np.isfinite(preds_era["s4"]):
        base_pred = float(preds_era["s4"])
    elif "Stryd 4.0" in preds_pod and np.isfinite(preds_pod["Stryd 4.0"]):
        base_pred = float(preds_pod["Stryd 4.0"])
    else:
        vals = [v for v in list(preds_era.values()) + list(preds_pod.values()) if np.isfinite(v) and v > 0]
        base_pred = float(np.nanmedian(vals)) if vals else np.nan

    # --- Convert predictions to adjusters with light shrinkage (prevents small-sample noise) ---
    # raw_adjust = base_pred / pred  (multiplier to map that era onto S4 scale)
    # v49: Dramatically reduced shrinkage.  Previous k=120/300 was over-aggressive:
    #   - v1 era (n=36) had only 23% data weight, crushing a real +8% offset to +1.8%
    #   - The 95% CI for v1 was (1.067-1.096), entirely above 1.0
    #   - Shrinkage was pulling the value OUTSIDE its own confidence interval
    # New approach:
    #   - Measured eras (real pod data): k=20  (light guard against noise)
    #   - Simulated eras (pre_stryd, v1_late): k=60  (sim model noise warrants more caution)
    #   - S5: k=50 with prior=0.995  (small sample, measured)
    SIMULATED_ERAS = {"pre_stryd", "v1_late"}
    K_MEASURED = 20.0
    K_SIMULATED = 60.0
    K_S5 = 50.0

    def _shrink_adjust(raw_adj: float, n: int, prior_center: float, k: float) -> float:
        if not (np.isfinite(raw_adj) and raw_adj > 0):
            return np.nan
        if not (np.isfinite(prior_center) and prior_center > 0):
            prior_center = 1.0
        w = float(n) / float(n + k) if (n is not None and n >= 0 and k > 0) else 1.0
        w = float(np.clip(w, 0.0, 1.0))
        return float(np.exp(w * np.log(raw_adj) + (1.0 - w) * np.log(prior_center)))

    adjust_era = {}
    adjust_pod = {}
    if np.isfinite(base_pred) and base_pred > 0:
        for era, pred in preds_era.items():
            n = int(n_era.get(era, 0))
            raw = float(base_pred / pred)
            if era == "s4":  # v44: changed from s4_late
                adj = 1.0
            elif era == "s5":
                adj = _shrink_adjust(raw, n=n, prior_center=0.995, k=K_S5)
                adj = float(np.clip(adj, 0.985, 1.010))
            elif era in SIMULATED_ERAS:
                adj = _shrink_adjust(raw, n=n, prior_center=1.0, k=K_SIMULATED)
                adj = float(np.clip(adj, 0.90, 1.12))
            else:
                adj = _shrink_adjust(raw, n=n, prior_center=1.0, k=K_MEASURED)
                adj = float(np.clip(adj, 0.90, 1.12))
            adjust_era[era] = adj

        for pm, pred in preds_pod.items():
            n = int(n_pod.get(pm, 0))
            raw = float(base_pred / pred)
            if pm == "Stryd 4.0":
                adj = 1.0
            elif pm == "Stryd 5.0":
                adj = _shrink_adjust(raw, n=n, prior_center=0.995, k=K_S5)
                adj = float(np.clip(adj, 0.985, 1.010))
            else:
                adj = _shrink_adjust(raw, n=n, prior_center=1.0, k=K_MEASURED)
                adj = float(np.clip(adj, 0.90, 1.12))
            adjust_pod[pm] = adj

    # --- v49: Always overwrite power_adjuster_to_S4 with freshly computed adjusters ---
    # Previous logic only filled NaNs, so stale values from per-row processing survived.
    # Prefer era mapping (calibration_era_id); fall back to pod_model; finally 1.0.
    df["power_adjuster_to_S4"] = df["calibration_era_id"].map(adjust_era)
    df["power_adjuster_to_S4"] = df["power_adjuster_to_S4"].where(
        np.isfinite(df["power_adjuster_to_S4"]),
        df["pod_model"].map(adjust_pod),
    )

    df["power_adjuster_to_S4"] = pd.to_numeric(df["power_adjuster_to_S4"], errors="coerce")
    df["power_adjuster_to_S4"] = df["power_adjuster_to_S4"].where(np.isfinite(df["power_adjuster_to_S4"]), 1.0)
    
    # v44: v1_late era (Aug 1 - Sept 12, 2017) is treated like pre_stryd
    # Power is simulated in StepA because the pod was failing during this period.
    # v1 (early, May-Aug) uses measured power with normal era adjustment.
    # -----------------------------------------------------------------


    # ---------- Highbury Fields parkrun detection + GPS-distance correction ----------
    # Goal: correct GPS under-distance on Highbury Fields (tree cover) for pre-Stryd parkruns,
    # using Stryd-era HF parkruns as reference (distance_km ~ Stryd, gps_distance_km ~ GPS track).
    HF_LAT, HF_LON = 51.5501, -0.1028
    HF_RADIUS_M = 1500.0

    # Helper: time-of-day window in Stockholm local time (London parkrun 09:00 ~= Stockholm 10:00)
    def _tod_ok(ts: pd.Timestamp) -> bool:
        try:
            h = int(ts.hour); mi = int(ts.minute)
        except Exception:
            return False
        # generous window around 10:00
        return (h == 10 and 0 <= mi <= 10) or (h == 9 and mi >= 55)

    # Likely parkrun by name and basic sanity
    name_series = df.get("activity_name", pd.Series([""] * len(df)))
    is_parkrun = name_series.astype(str).str.contains("parkrun", case=False, na=False)

    dist_km_s = pd.to_numeric(df.get("distance_km"), errors="coerce")
    gps_km_s = pd.to_numeric(df.get("gps_distance_km"), errors="coerce")
    mov_s = pd.to_numeric(df.get("moving_time_s"), errors="coerce")
    elap_s = pd.to_numeric(df.get("elapsed_time_s"), errors="coerce")
    stop_s = (elap_s - mov_s)

    # Time-of-day (Stockholm-local in 'date' column)
    dt_local = pd.to_datetime(df.get("date"), errors="coerce")
    tod_ok = dt_local.apply(lambda x: _tod_ok(x) if pd.notna(x) else False)

    # Location near Highbury Fields
    lat_s = pd.to_numeric(df.get("gps_lat_med"), errors="coerce")
    lon_s = pd.to_numeric(df.get("gps_lon_med"), errors="coerce")
    dist_to_hf = []
    for la, lo in zip(lat_s.to_numpy(dtype=float), lon_s.to_numpy(dtype=float)):
        if np.isfinite(la) and np.isfinite(lo):
            dist_to_hf.append(haversine_m(HF_LAT, HF_LON, la, lo))
        else:
            dist_to_hf.append(np.nan)
    dist_to_hf = pd.Series(dist_to_hf, index=df.index, dtype=float)
    near_hf = dist_to_hf <= HF_RADIUS_M

    # Duration/distance sanity using ELAPSED time (moving_time can be unreliable in older FITs)
    elapsed_s = pd.to_numeric(df.get("elapsed_time_s"), errors="coerce")
    dist_ok = dist_km_s.between(4.7, 5.2)
    elapsed_ok = elapsed_s.between(18*60, 25*60)
    # Implied pace from elapsed time to avoid false negatives due to bad moving_time
    implied_pace = (elapsed_s / 60.0) / dist_km_s.replace({0: np.nan})
    pace_ok = implied_pace.between(3.4, 5.0)

    hf_parkrun = is_parkrun & tod_ok & near_hf & dist_ok & elapsed_ok & pace_ok
    df["hf_parkrun"] = hf_parkrun.fillna(False)

    # Reference ratio from Stryd-era HF: distance_km (Stryd) / gps_distance_km (GPS)
    # Note: ratio > 1 implies GPS under-distance.
    ref = df[df["hf_parkrun"]].copy()
    ref_ratio = (pd.to_numeric(ref["distance_km"], errors="coerce") /
                 pd.to_numeric(ref["gps_distance_km"], errors="coerce"))
    ref = ref[np.isfinite(ref_ratio) & (ref_ratio > 0.9) & (ref_ratio < 1.3)].copy()
    ref["hf_ratio"] = ref_ratio.loc[ref.index]
    ref = ref.sort_values("date")
    if len(ref) >= 5:
        ref["hf_ratio_smooth"] = ref["hf_ratio"].rolling(window=9, center=True, min_periods=5).median()
        ref["hf_ratio_smooth"] = ref["hf_ratio_smooth"].fillna(ref["hf_ratio"])
    else:
        ref["hf_ratio_smooth"] = ref["hf_ratio"]

    default_hf_ratio = float(np.nanmedian(ref["hf_ratio_smooth"])) if len(ref) else np.nan

    # Assign factor to every HF parkrun via nearest-in-time reference
    df_hf = df[df["hf_parkrun"]].copy().sort_values("date")
    hf_factor = pd.Series(np.nan, index=df.index, dtype=float)
    if len(ref) >= 1:
        ref2 = ref[["date", "hf_ratio_smooth"]].copy().sort_values("date")
        # merge_asof expects datetime
        ref2["date"] = pd.to_datetime(ref2["date"], errors="coerce")
        df_hf2 = df_hf[["date"]].copy()
        df_hf2["date"] = pd.to_datetime(df_hf2["date"], errors="coerce")
        merged = pd.merge_asof(df_hf2.sort_values("date"), ref2, on="date", direction="nearest", tolerance=pd.Timedelta(days=3650))
        hf_factor.loc[merged.index] = merged["hf_ratio_smooth"].to_numpy(dtype=float)
    if np.isfinite(default_hf_ratio):
        hf_factor = hf_factor.fillna(default_hf_ratio)

    df["hf_gps_factor"] = hf_factor

    # Apply correction only where the run distance appears GPS-derived (gps_distance_ratio ~ 1)
    gps_ratio_s = pd.to_numeric(df.get("gps_distance_ratio"), errors="coerce")
    dist_is_gps = gps_ratio_s.between(0.97, 1.03)

    df["hf_distance_km_corrected"] = dist_km_s
    m_apply = df["hf_parkrun"] & dist_is_gps & np.isfinite(df["hf_gps_factor"])
    df.loc[m_apply, "hf_distance_km_corrected"] = dist_km_s.loc[m_apply] * df.loc[m_apply, "hf_gps_factor"]

    # Canonical correction: Highbury Fields parkruns often come up short under trees.
    # IMPORTANT: Do NOT overwrite the core recorded distance used by physics (per-second v/grade, RE simulation, etc.).
    # Instead, keep:
    #   - distance_km: recorded distance (as parsed earlier)
    #   - hf_distance_km_corrected: Strava-like corrected distance estimate for HF parkruns (may still be <5k if GPS was very short)
    #   - hf_distance_km_display: distance used for *display/pace* purposes, clamped to >=5.000 km
    # This prevents "forcing 5k" from inflating RE summary metrics on HF parkruns.

    # Preserve recorded distance in a dedicated column for clarity/auditing.
    if "distance_km_recorded" not in df.columns:
        df["distance_km_recorded"] = pd.to_numeric(df.get("distance_km"), errors="coerce")

    # Display distance: corrected HF distance, clamped to 5.000 km for parkrun reporting.
    df["hf_distance_km_display"] = pd.to_numeric(df.get("hf_distance_km_corrected"), errors="coerce")
    df.loc[df["hf_parkrun"].fillna(False), "hf_distance_km_display"] = df.loc[df["hf_parkrun"].fillna(False), "hf_distance_km_display"].clip(lower=5.000)

    # Convenience: official distance for HF parkrun rows (so downstream can choose to use it).
    df["official_distance_km"] = np.nan
    df.loc[df["hf_parkrun"].fillna(False), "official_distance_km"] = 5.000

    # Strava-like corrected pace over the display distance.
    df["hf_pace_min_per_km_corrected"] = (mov_s / 60.0) / pd.to_numeric(df["hf_distance_km_display"], errors="coerce")

    # (Legacy run-average simulated power removed in v33; per-second sim power comes next)

    # Ensure calibration columns exist (if adjustment computation skipped for any reason)
    if "calibration_era_id" not in df.columns:
        df["calibration_era_id"] = ""
    if "power_adjuster_to_S4" not in df.columns:
        df["power_adjuster_to_S4"] = 1.0
   

    # Power source flag (for RF interpretation)
    df["power_source"] = np.where(pd.to_numeric(df.get("avg_power_w"), errors="coerce").notna(), "stryd", "")
    df["sim_model_id"] = ""
    df["sim_ok"] = False

    # Apply weather overrides from activity_overrides.xlsx (indoor runs with temp_override)
    override_file = args.override_file.strip() if hasattr(args, "override_file") and args.override_file else ""
    if not override_file:
        cand = os.path.join(out_dir, "activity_overrides.xlsx")
        if os.path.exists(cand):
            override_file = cand
    if override_file:
        df, n_applied = apply_weather_overrides(df, override_file)
        if n_applied:
            print(f"Applied weather overrides: {n_applied} indoor run(s) from {override_file}")

    # Canonical avg pace: recompute from final distance + moving time to avoid early-era quirks
    if "distance_km" in df.columns and "moving_time_s" in df.columns:
        dist = pd.to_numeric(df["distance_km"], errors="coerce")
        movs = pd.to_numeric(df["moving_time_s"], errors="coerce")
        pace = (movs / 60.0) / dist
        pace = pace.where((dist > 0.05) & np.isfinite(pace))
        df["avg_pace_min_per_km"] = pace.round(2)


    # -------------------------
    # Presentation rounding (readability)
    # -------------------------
    # Round *watts* columns to 0dp, but keep W/kg at 3dp (e.g., avg_air_power_wkg).
    power_cols_0dp = []
    for _c in df.columns:
        cl = str(_c).lower()
        if "wkg" in cl:
            continue
        # Typical watt columns end with _w (or are named explicitly)
        if cl.endswith("_w") or cl in ("npower_w","avg_power_w","power_mean_w","power_median_w","avg_air_power_w"):
            power_cols_0dp.append(_c)

    for c in power_cols_0dp:
        df[c] = pd.to_numeric(df[c], errors="coerce").round(0)

    cols_3dp = ["gps_distance_ratio","avg_pace_min_per_km","power_adjuster_to_S4","nPower_HR",
                "RF_adjusted_mean_W_per_bpm","RF_adjusted_median_W_per_bpm","RF_dead_frac",
                "RF_drift_pct_per_min","RF_drift_r2","RE_avg","RE_normalised","grade_mean_pct",
                "hf_gps_factor"]
    for c in cols_3dp:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").round(3)


        # --- Pre-Stryd elevation sanity (Policy B: trust gain) ---

        try:

            if "calibration_era_id" in df.columns and "elev_gain_m" in df.columns and "elev_loss_m" in df.columns:

                pre = df["calibration_era_id"].astype(str).eq("pre_stryd")

                gain = pd.to_numeric(df["elev_gain_m"], errors="coerce")

                loss = pd.to_numeric(df["elev_loss_m"], errors="coerce")


                fill_gain = pre & (~np.isfinite(gain) | (gain <= 0)) & (np.isfinite(loss) & (loss > 0))

                if fill_gain.any():

                    df.loc[fill_gain, "elev_gain_m"] = loss[fill_gain]

                    if "notes" in df.columns:

                        df.loc[fill_gain, "notes"] = (df.loc[fill_gain, "notes"].fillna("").astype(str) + " elev_gain_filled_from_loss").str.strip()

                    else:

                        df["notes"] = ""

                        df.loc[fill_gain, "notes"] = "elev_gain_filled_from_loss"


                gain = pd.to_numeric(df["elev_gain_m"], errors="coerce")

                loss_bad = pre & (np.isfinite(gain) & (gain > 0)) & (~np.isfinite(loss) | (loss <= 0) | (np.abs(loss - gain) > 0.25 * gain))

                if loss_bad.any():

                    df.loc[loss_bad, "elev_loss_m"] = gain[loss_bad]

                    if "notes" in df.columns:

                        df.loc[loss_bad, "notes"] = (df.loc[loss_bad, "notes"].fillna("").astype(str) + " elev_loss_set_to_gain").str.strip()

                    else:

                        df["notes"] = ""

                        df.loc[loss_bad, "notes"] = "elev_loss_set_to_gain"

        except Exception:

            pass

        # --- end elevation sanity ---


        # --- Post-fill: for measured-power eras, if RE_avg is blank but RE_normalised is plausible, fill RE_avg from RE_normalised.

        # This primarily fixes interval/workout runs where avg-power-based RE can be flagged while REn remains sensible.

        try:

            if "calibration_era_id" in df.columns and "RE_avg" in df.columns and "RE_normalised" in df.columns:

                era = df["calibration_era_id"].astype(str).str.lower()

                ren = pd.to_numeric(df["RE_normalised"], errors="coerce")

                reavg = pd.to_numeric(df["RE_avg"], errors="coerce")

                mfill = (era != "pre_stryd") & (~np.isfinite(reavg)) & (np.isfinite(ren)) & (ren >= 0.50) & (ren <= 1.05)

                if mfill.any():

                    df.loc[mfill, "RE_avg"] = ren[mfill]

                    if "notes" in df.columns:

                        df.loc[mfill, "notes"] = (df.loc[mfill, "notes"].fillna("").astype(str) + " RE_avg_filled_from_REn").str.strip()

                    else:

                        df["notes"] = ""

                        df.loc[mfill, "notes"] = "RE_avg_filled_from_REn"

        except Exception:

            pass

        # --- end post-fill ---

    # v41: Ensure output is sorted by date (critical for append mode where new runs get concatenated at end)
    df = df.sort_values("date").reset_index(drop=True)

    write_master(args.template, args.out, df, df_fail=df_fail)
    # --- v37 fix2: cache-write health summary ---
    if args.persec_cache_dir:
        print(f"Per-second cache writes: ok={CACHE_WRITE_OK}, failed={CACHE_WRITE_FAIL}")
        if CACHE_WRITE_FAIL and FIRST_CACHE_WRITE_ERROR:
            err_path = FIRST_CACHE_WRITE_ERROR_PATH or "(unknown)"
            print("First cache write failure path:", err_path)
            print("First cache write failure traceback:\n" + FIRST_CACHE_WRITE_ERROR)
    print(f"Wrote: {args.out}")

if __name__ == "__main__":
    main()
