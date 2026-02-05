#!/usr/bin/env python3
"""
tag_new_activities.py - Interactive override tagger for newly fetched activities.

Called by Daily_Update.bat between FIT download/zip and the pipeline run.
Shows new activities and lets you tag any with overrides (race, surface, etc.)
before the pipeline processes them.

Usage:
    python tag_new_activities.py                    (check FIT_downloads for new)
    python tag_new_activities.py --fit-dir DIR      (custom dir)

Exit codes:
    0 = done (user finished tagging or skipped)
    1 = error
"""

import os
import sys
import pandas as pd
from pathlib import Path
from datetime import datetime

# ============================================================================
# CONFIGURATION — matches add_run.py / add_override.py
# ============================================================================
OVERRIDE_FILE = r"activity_overrides.xlsx"
PENDING_CSV = r"pending_activities.csv"
FIT_DOWNLOAD_DIR = r"FIT_downloads"
TOTAL_HISTORY_ZIP = r"TotalHistory.zip"

VALID_SURFACES = {'TRACK', 'INDOOR_TRACK', 'TRAIL', 'SNOW', 'HEAVY_SNOW'}


# ---- Flag parsing (shared logic with add_run.py / add_override.py) --------

def parse_flags(flags_str: str) -> dict:
    """Parse comma-separated flags string into override fields."""
    result = {
        'race_flag': 0,
        'parkrun': 0,
        'surface': '',
        'surface_adj': None,
        'official_distance_km': None,
        'temp_override': None,
    }
    if not flags_str:
        return result

    for flag in [f.strip() for f in flags_str.split(',')]:
        upper = flag.upper()
        if upper == 'RACE':
            result['race_flag'] = 1
        elif upper == 'PARKRUN':
            result['race_flag'] = 1
            result['parkrun'] = 1
            if result['official_distance_km'] is None:
                result['official_distance_km'] = 5.0
        elif upper in VALID_SURFACES:
            result['surface'] = upper
        elif upper.startswith('ADJ=') or upper.startswith('SADJ='):
            try:
                result['surface_adj'] = float(flag.split('=')[1])
            except (ValueError, IndexError):
                print(f"  WARNING: Invalid surface_adj '{flag}'")
        elif upper.startswith('TEMP='):
            try:
                result['temp_override'] = float(flag.split('=')[1])
            except (ValueError, IndexError):
                print(f"  WARNING: Invalid temp_override '{flag}'")
        else:
            try:
                val = float(flag)
                if 0.5 < val < 1.5:
                    result['surface_adj'] = val
                elif val > 1.5:
                    result['official_distance_km'] = val
            except ValueError:
                print(f"  WARNING: Unknown flag '{flag}'")
    return result


def flags_summary(flags: dict) -> str:
    """Human-readable summary of parsed flags."""
    parts = []
    if flags['race_flag']:      parts.append('RACE')
    if flags['parkrun']:        parts.append('PARKRUN')
    if flags['surface']:        parts.append(f"surface={flags['surface']}")
    if flags.get('surface_adj') is not None:
        parts.append(f"adj={flags['surface_adj']}")
    if flags.get('temp_override') is not None:
        parts.append(f"temp={flags['temp_override']}C")
    if flags['official_distance_km']:
        parts.append(f"dist={flags['official_distance_km']}km")
    return ', '.join(parts) if parts else '(none)'


# ---- Override file writing (matches add_run.py pattern) -------------------

def write_override(file_key: str, flags: dict):
    """Add or update an override entry in activity_overrides.xlsx."""
    from openpyxl import load_workbook

    if not os.path.exists(OVERRIDE_FILE):
        df = pd.DataFrame(columns=['file', 'race_flag', 'parkrun',
                                    'official_distance_km', 'surface',
                                    'surface_adj', 'temp_override', 'notes'])
    else:
        df = pd.read_excel(OVERRIDE_FILE, dtype={'file': str})

    # Remove any existing entry for this key
    df = df[df['file'].astype(str) != file_key]

    new_row = pd.DataFrame([{
        'file': file_key,
        'race_flag': flags['race_flag'],
        'parkrun': flags['parkrun'],
        'official_distance_km': flags['official_distance_km'],
        'surface': flags['surface'],
        'surface_adj': flags.get('surface_adj'),
        'temp_override': flags.get('temp_override'),
        'notes': f"Tagged via Daily_Update on {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    }])

    df = new_row if len(df) == 0 else pd.concat([df, new_row], ignore_index=True)
    df.to_excel(OVERRIDE_FILE, index=False)

    # Force Text format on file column (prevents Excel date auto-format)
    wb = load_workbook(OVERRIDE_FILE)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=1).number_format = '@'
    wb.save(OVERRIDE_FILE)


# ---- Discover new activities ----------------------------------------------

def get_new_activities(fit_dir: str) -> list:
    """
    Get list of new FIT files from download dir, enriched with activity names
    from pending_activities.csv.

    Returns list of dicts: {filename, date, name, already_tagged}
    """
    fit_path = Path(fit_dir)
    if not fit_path.exists():
        return []

    fit_files_raw = sorted(
        list(fit_path.glob("*.FIT")) + list(fit_path.glob("*.fit")),
        key=lambda f: f.name
    )
    # Deduplicate (Windows is case-insensitive, so *.FIT and *.fit match same files)
    seen = set()
    fit_files = []
    for f in fit_files_raw:
        key = f.name.lower()
        if key not in seen:
            seen.add(key)
            fit_files.append(f)
    if not fit_files:
        return []

    # Load pending_activities for activity names
    pending_names = {}
    if os.path.exists(PENDING_CSV):
        try:
            pdf = pd.read_csv(PENDING_CSV, dtype=str)
            for _, row in pdf.iterrows():
                fn = str(row.get('file', '')).strip()
                name = str(row.get('activity_name', '')).strip()
                if fn and name and name != 'nan':
                    pending_names[fn.lower()] = name
        except Exception:
            pass

    # Load existing overrides to show which are already tagged
    existing_overrides = set()
    if os.path.exists(OVERRIDE_FILE):
        try:
            odf = pd.read_excel(OVERRIDE_FILE, dtype={'file': str})
            for _, row in odf.iterrows():
                existing_overrides.add(str(row.get('file', '')).strip().lower())
        except Exception:
            pass

    # Load TotalHistory.zip to identify already-processed activities
    already_in_zip = set()
    if os.path.exists(TOTAL_HISTORY_ZIP):
        import zipfile
        try:
            with zipfile.ZipFile(TOTAL_HISTORY_ZIP, 'r') as zf:
                for name in zf.namelist():
                    already_in_zip.add(name.lower())
        except Exception:
            pass

    activities = []
    for f in fit_files:
        filename = f.name
        stem = f.stem

        # Skip if already in TotalHistory.zip (already processed)
        if filename.lower() in already_in_zip:
            continue

        # Parse date from YYYY-MM-DD_HH-MM-SS filename pattern
        date_str = stem[:10] if (len(stem) >= 10
                                  and stem[4] == '-'
                                  and stem[7] == '-') else ""

        name = pending_names.get(filename.lower(), "")
        already = (filename.lower() in existing_overrides
                   or date_str.lower() in existing_overrides)

        activities.append({
            'filename': filename,
            'date': date_str,
            'name': name,
            'already_tagged': already,
        })

    return activities


# ---- Interactive prompt ---------------------------------------------------

def interactive_tag(activities: list) -> int:
    """
    Show activities and prompt for override tags.
    Returns number of activities tagged.
    """
    print()
    print("  New activities:")
    print("  " + "-" * 62)
    for i, act in enumerate(activities, 1):
        tag_marker = " [TAGGED]" if act['already_tagged'] else ""
        name_str = f"  {act['name']}" if act['name'] else ""
        print(f"  {i}. {act['filename']}{name_str}{tag_marker}")
    print("  " + "-" * 62)
    print()
    print("  Flags: RACE  PARKRUN  SNOW  HEAVY_SNOW  TRAIL  TRACK  INDOOR_TRACK")
    print("         <dist> (e.g. 5.0)  ADJ=<val> (e.g. 1.05)  TEMP=<val>")
    print()
    print("  Type: <number> <flags>     e.g.  1 RACE,10.0")
    print("        <range> <flags>      e.g.  1-3 SNOW")
    print("        Enter to skip all and continue")
    print()

    tagged = 0
    while True:
        try:
            line = input("  Tag> ").strip()
        except (EOFError, KeyboardInterrupt):
            print()
            break

        if not line:
            break

        # Split into selector and flags
        parts = line.split(None, 1)
        if len(parts) < 2:
            print("  Need: <number> <flags>   (e.g. 1 RACE,10.0)")
            continue

        selector, flags_str = parts

        # Determine which activity indices
        indices = []
        if '-' in selector:
            try:
                lo, hi = selector.split('-', 1)
                indices = list(range(int(lo) - 1, int(hi)))
            except ValueError:
                print(f"  Invalid range: {selector}")
                continue
        else:
            try:
                idx = int(selector) - 1
                if 0 <= idx < len(activities):
                    indices = [idx]
                else:
                    print(f"  Invalid number (use 1-{len(activities)})")
                    continue
            except ValueError:
                print(f"  Invalid: {selector}")
                continue

        flags = parse_flags(flags_str)

        for idx in indices:
            act = activities[idx]
            # Use date as override key (StepB resolves date -> filename)
            file_key = act['date'] if act['date'] else act['filename']

            write_override(file_key, flags)
            act['already_tagged'] = True
            name_str = f" ({act['name']})" if act['name'] else ""
            print(f"    -> {file_key}{name_str}: {flags_summary(flags)}")
            tagged += 1

    return tagged


# ---- Main ----------------------------------------------------------------

def main():
    import argparse
    parser = argparse.ArgumentParser(description="Tag new activities with overrides")
    parser.add_argument("--fit-dir", default=FIT_DOWNLOAD_DIR)
    args = parser.parse_args()

    activities = get_new_activities(args.fit_dir)

    if not activities:
        print("  No new activities to tag.")
        return 0

    tagged = interactive_tag(activities)

    if tagged:
        print(f"\n  {tagged} override(s) saved to {OVERRIDE_FILE}")
        # Signal to Daily_Update.bat that new tags were added
        try:
            with open('_new_tags_added.tmp', 'w') as f:
                f.write(str(tagged))
        except Exception:
            pass
    else:
        print("  No overrides added — continuing.")

    return 0


if __name__ == "__main__":
    sys.exit(main())
