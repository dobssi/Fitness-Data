#!/usr/bin/env python3
# File: StepB_PostProcess_v51.py
#
# Changelist v51 (2026-02-07):
#   - Easy RF EMA: exponentially weighted mean (span=15) of easy-run RF_adj
#     (avg_hr 120-150, dist >= 4km, non-race). Forward-filled to all runs.
#   - Easy_RF_z: z-score of each easy run vs trailing 30-run baseline
#   - RFL_Trend_Delta: run-to-run change in RFL_Trend (trend mover flagging)
#   - Easy_RFL_Gap: normalised Easy_RF_EMA minus RFL_Trend (divergence metric)
#   - Alert system (console output): Alert 1 (CTL up/RFL down), Alert 1b
#     (taper quality), Alert 2 (deep TSB), Alert 3b (easy outlier),
#     Alert 5 (easy RF divergence). All thresholds in config.py.
#
# Changelist v50 (2026-02-06):
#   - Fast cache index: uses _cache_index.json via build_cache_index_fast()
#     instead of opening 3000+ .npz files to read timestamps (~30s -> <1s)
#   - Merged iterrows passes: Temp_Trend + adjustment factors now calculated
#     in a single pass instead of two separate iterrows loops
#   - Pre-allocated all output columns before main loop to avoid fragmentation
#   - run() in run_pipeline.py now handles exit code 2 (no new runs) to skip
#     Steps 2-4 entirely in UPDATE mode
#
# Changelist v49 (2026-02-05):
#   - Era_Adj now uses regression-based power_adjuster_to_S4 (speed-controlled)
#     instead of RE-median ratio (which was confounded by fitness/speed changes)
#   - calc_era_adj() reads power_adjuster_to_S4 per-row, falls back to RE-median
#   - Impact: repl era correction drops from +4.9% to +3.5%, peak RF_Trend -1.3%
#   - RFL improves ~+1.4pp (historical baseline less inflated)
#
# Changelist v48 (2026-02-05):
#   - Consolidated weather_overrides.csv into activity_overrides.xlsx
#   - Removed dead --weather-overrides arg
#   - Renamed terrain_saturation_cap -> terrain_linear_cap (matches v47 algorithm)
#   - Removed legacy terrain constants from config.py
#   - Fixed PEAK_CP_WATTS fallback (375, was 370)
#   - Unicode emoji replaced with ASCII in print statements (cp1252 safe)
#
# Changelist v47 (2026-02-04):
#   - RF_WINDOW_DURATION_S raised from 1200 to 2400 (20 -> 40 mins)
#   - Terrain: replaced exponential saturation with linear-with-cap
#   - Strava elevation gate: 8 m/km minimum for terrain adj
#   - Stryd mass correction (config.py STRYD_MASS_HISTORY)
#   - Duration penalty damping (asymmetric: 0.33 for penalty side)
#
# Changelist v46 (2026-02-03):
#   - Terrain adjustment: replaced linear formula with exponential saturation curve
#     * Old: adj = 1 + 0.012 * (und - 4), capped at 1.12
#     * New: adj = 1 + 0.06 * (1 - exp(-0.18 * und))
#     * Prevents over-compensation at moderate undulation (4-8) while preserving
#       calibration for high-undulation terrain
#     * Trail surface penalty now handled separately via surface_adj overrides
#   - Strava elevation sanity check for terrain adjustment
#     * FIT barometric elevation can be inflated by noise (especially older devices)
#     * If Strava corrected elevation is < 8 m/km, terrain adj is disabled
#     * Catches ~25 runs with phantom elevation from barometric drift
#   - Strava elevation matching: populates strava_elev_gain_m/loss_m from activities.csv
#     * Master dates (local time) matched to Strava dates (UTC) within 5 min tolerance
#     * Runs on every StepB invocation when --strava is provided
#   - Intensity_Adj disabled (set to 1.0) - was inflating RF by +30.7pp
#
# Changelist v45 (2026-01-31):
#   - Activity name refresh from Strava (--strava activities.csv)
#     * Names updated from activities.csv on every StepB run
#     * Matches by strava_activity_id for reliable updates
#     * Eliminates need for full rebuild to update names
#   - Duration-scaled heat adjustments:
#     * RF Temp_Adj scaled to RF window (~30 min): heat_mult = 0.33
#       - RF is measured early, before full heat accumulation
#       - Previously applied 100% heat adjustment to all runs
#     * Power Score heat adjustment scales with full run duration
#       - Full effect at 90 mins, capped at 150%
#       - Formula: ps_heat_adj = 1 + (Temp_Adj - 1) * min(1.5, duration/90min)
#     * Temp_Adj column still stores full value (for diagnostics/Power Score)
#   - Power Score now uses distance (not time) for Riegel factor
#     * Prevents slower finishers getting undeserved boost at same distance
#     * Reference: 5K (distance_factor = 1.0 at 5km)
#     * Uses official_distance_km when available (race corrections)
#   - Air power diminishing returns for Power Score
#     * Air power can be inflated at high paces or in headwinds
#     * Above 4% threshold: excess air power cut in half
#     * e.g., 9.3% air -> 4% + 2.65% = 6.65% effective
#   - RF_Trend now uses 42-day window with quadratic decay
#     * Replaces flat 21-day window that caused cliff-edge spikes
#     * Decay formula: weight = 1 - (days_ago/42)^2
#     * Day 7: 97%, Day 14: 89%, Day 21: 75%, Day 28: 56%, Day 42: 0%
#     * Matches CTL 42-day time constant for consistency
#   - Days-off penalty for RF_Trend decay
#     * Extended breaks (>7 days) accelerate decay for prior activities
#     * Penalty: (days_off - 7) * 2 added to effective days ago
#     * Multiple breaks in window accumulate penalties
#     * Reflects genuine fitness loss during extended inactivity
#   - Hill power metrics (diagnostic columns for terrain_adj refinement)
#     * uphill_power_mean, downhill_power_mean, uphill_downhill_ratio
#     * uphill_pct, flat_power_mean
#     * Used to distinguish hill intervals (ratio >1.8) from rolling terrain (~1.3)
#   - Terrain_Adj improvements:
#     * Now calculated from RF window only (not whole run)
#     * Gated by uphill/downhill ratio: if ratio > 1.5, no terrain boost
#     * Hill intervals (hard up, easy down) no longer get undeserved terrain boosts
#     * Rolling terrain like LL30 (consistent effort) still gets full terrain boost
#     * REMOVED RE gating - was too aggressive for races with good conditions
#   - Intensity_Adj terrain influence:
#     * Undulating terrain now lowers the intensity threshold
#     * Running at 80% CP on LL30 terrain treated like 88% CP on flat
#     * Formula: threshold reduced by 0.5% per undulation point above 4
#     * e.g., RF_und=15 -> threshold drops from 88% to 82.5%
#
# Changelist v44.5 (2026-01-31):
#   - Simplified Power Score integration:
#     * Power Score now sets a FLOOR on RF_adj: RF_adj = max(RF_adj, Power_Score/180)
#     * Power Score boosts Factor by 50% for runs above threshold
#     * Removed RFL_Combined, RF_Combined - Power Score affects RF directly
#     * This eliminates the confusing disconnect between RFL and RFL_Combined
#   - Central config.py for shared constants (PEAK_CP_WATTS, POWER_SCORE_* params)
#   - Dashboard now shows RFL_Trend (Power Score influence baked into RF_adj)
#
# Changelist v44.4 (2026-01-30):
#   - Power Score tuning for natural RFL variation:
#     * Rolling 2-year peak for Power Score (was all-time)
#     * 180-day half-life (was 60)
#     * Averaging formula for RFL_Combined
#     * Capped at 100% to prevent exceeding historical peaks
#   - Peak CP changed to 370W (was 375W)
#   - Power Score reference time 20min, k=0.08
#
# Changelist v44.3 (2026-01-29):
#   - Added Power Score and RFL_Combined metrics:
#     * Power_Score = (avg_power * era_adj) * (duration / 20min)^0.08
#
# Changelist v44.2 (2026-01-28):
#   - Fixed RE calculation to use overall distance/time (not per-second speed integration)
#   - Fixed dt clipping from 0.5-1.5s to 0.5-10s (handles sparse data)
#   - Added RE gating for Terrain_Adj (only applies if RE below era median)
#   - Dashboard independence: summary sheets generated from Master
#
# Changelist v44 (2026-01-27):
#   - Updated for data-driven Stryd era detection (serial-based)
#   - Consolidated eras: pre_stryd, v1, repl, air, s4, s5
#   - Changed reference_era from 's4_late' to 's4'
#   - Updated DEFAULT_ERA_ADJUSTERS for consolidated eras
#   - Added terrain undulation metrics (v44.1):
#     * grade_std_pct: Standard deviation of per-second grade
#     * elev_reversals_per_km: Elevation direction changes per km
#     * undulation_score: Combined metric for Lidingöloppet-style terrain
#
# Changelist v43 (2026-01-26):
#   - MAJOR: Added RF adjustment calculations from BFW
#     * Temp_Adj: temperature + humidity adjustment
#     * Terrain_Adj: undulation-based adjustment (replaces RE_Adj)  
#     * Era_Adj: Stryd era normalization to S4 baseline
#     * Total_Adj: combined adjustment factor
#   - MAJOR: Added rolling metrics (previously in BFW)
#     * RF_adj: adjusted RF with max jump cap
#     * Factor: weighting for RF trend
#     * RF_Trend: 21-day weighted average
#     * RFL, RFL_Trend: % of all-time peak
#     * TSS: Training Stress Score per activity
#     * CTL, ATL, TSB: Chronic/Acute Training Load
#   - MAJOR: Added incremental processing mode (--incremental)
#     * Only loads cache for new/changed rows
#     * Recalculates rolling metrics from earliest change
#     * ~100x faster for adding single run
#   - Added parkrun flag support (override file)
#   - Added surface column (TRACK, INDOOR_TRACK, TRAIL, SNOW, HEAVY_SNOW)
#   - Added non_running_tss.csv support for CTL/ATL
#   - Improved BFW refresh verification
#
# Changelist v42 (2026-01-25):
#   - Fixed HR reliability sparse data check inconsistency:
#     * _calc_hr_cv applies 90% truncation to RF window (excludes cooldown)
#     * Caller's sparse check was using all samples from rf_start_s
#     * This caused old Garmin runs (0.34 Hz) with 193-199 samples in truncated
#       RF window to be incorrectly marked "unreliable" when they had 200+ total
#     * Fix: Caller now applies same 90% truncation as _calc_hr_cv
#     * Affected runs: ~15-25 min runs from 2013 era with sparse recording
#   - Fixed sample rate assumption in RF and HR CV calculations:
#     * Previously assumed 1Hz data (lag_samples = HR_LAG_S)
#     * Now calculates actual sample rate from time array
#     * Correctly handles old Garmin ~0.33Hz and modern 1Hz data
#     * Smoothing and lag windows now scale properly with sample rate
#
# Changelist v41 (2026-01-22):
#   - Added canonical column ordering to ensure consistent output regardless of add order
#   - Columns now in logical groups: identification, distance, time, power, HR, efficiency, etc.
#   - Humidity default: when temp_override is set (indoor run), humidity defaults to 50%
#     to prevent outdoor humidity affecting indoor run adjustments
#   - Duplicate override handling: warns and uses first entry if duplicate filenames in override file
#   - Strava distance respects overrides: Strava distance correction skips rows where
#     official_distance_km is already set from override file
#
# Changelist v40.1 (2026-01-22):
#   - HR reliability check: blank RF for unreliable HR data
#     * Calculates CV (coefficient of variation) of power:HR ratio
#     * Threshold adjusted based on RAW power CV: 12% + max(0, power_cv - 8%) × 1.2, capped at 35%
#       - Steady runs (power CV ~5%): threshold = 12%
#       - Yasso 800s (power CV ~13%): threshold = 18%
#       - Hard intervals (power CV ~35%+): threshold = 35% (capped)
#     * Uses RF window: starts at 600s for runs >= 15min (skips HR ramp-up)
#     * Uses RAW power CV (not smoothed) to detect interval structure
#     * Skips runs < 8 minutes (HR ramp-up dominates)
#     * Excludes 60s around pauses (HR lag around stops/starts)
#     * If over threshold, marks hr_corr_type as "unreliable" and blanks RF values
#     * New column: hr_cv_pct (for diagnostics)
#   - HR lag compensation: HR responds to power with ~15s delay
#     * New constants: HR_LAG_S=15, POWER_SMOOTH_S=30
#     * RF calculation now uses 30s smoothed power shifted by 15s
#     * Late_dropout HR correction uses same 30s/15s smoothing/lag
#     * Results in smoother corrected HR that better matches pre-dropout variability
#   - Per-second speed scaling for distance-corrected runs:
#     * When official_distance differs from gps_distance, scale per-second speeds
#       by (official_dist / gps_dist)
#     * For pre-Stryd runs: corrects simulated power, RE, and RF
#     * For post-Stryd runs: corrects RE (power is measured, RF unaffected)
#     * Only applies when difference > 0.5%
#   - HR correction: improved late_dropout vs early_artifact discrimination:
#     * Late dropout: post-drop ratio > 2.5 W/bpm (HR unrealistically LOW)
#     * Early artifact: pre-drop ratio < 1.5 W/bpm (HR was too HIGH), post-drop normal
#     * Ambiguous cases: default to early_artifact if <8min, else leave uncorrected
#     * Fixes false late_dropout detections that were lifting HR to impossible values (>200)
#
# Changelist v40 (2026-01-21):
#   - Override file integration (activity_overrides_v41.xlsx):
#     * Reads race_flag, official_distance_km, surface_adj, temp_override
#     * Applies distance corrections for pre-Stryd activities BEFORE RE recalculation
#     * Outputs new columns: race_flag, official_distance_km, surface_adj, 
#       distance_corrected, gps_distance_error_m
#   - Excel output formatting: columns A-D width 25, rest width 12, 
#     top row frozen, zoom 70%
#
# Changelist v39:
#   - Progress output: --progress-every N and --progress-per-run.
#   - Dtype-safe assignments: cast numeric output columns to float to avoid pandas FutureWarning.
#
# Changelist v38.7:
#   - RE active-only: recompute RE_avg and RE_normalised using ACTIVE running only.
#
# Changelist v38.6:
#   - Early artifact robustness: allow post-drop ratio windows to be truncated before later watch pauses.
#
# Changelist v38.5:
#   - Early artifact detection improved: post-drop consistency window starts at +60s.
#
# Changelist v38.4:
#   - False late-dropout guard: reject late-dropout corrections around transient slowdowns.
#
# Changelist v38.3:
#   - Pause-gap exclusion strengthened: ignore drop candidates near watch pauses.
#
# Changelist v38.2:
#   - HR correction running/rest mask fixed: do NOT use RF dead-time 2.5 w/kg as HR correction rest threshold.
#
# Changelist v38.1:
#   - Drop legacy column hr_corr_applied (BI); Step B uses only hr_corrected.

from __future__ import annotations

import argparse
import bisect
import hashlib
import os
from pathlib import Path
from typing import List, Tuple, Optional, Dict

import numpy as np
import pandas as pd

from sim_power_pipeline import REModel, air_density_kg_m3, simulate_power_series, wind_along_track

# Import age grade calculations
try:
    from age_grade import calc_age_grade, calc_race_prediction, format_time, PEAK_CP_WATTS
    AGE_GRADE_AVAILABLE = True
except ImportError:
    AGE_GRADE_AVAILABLE = False
    print("Warning: age_grade module not found - age grade calculations disabled")


# ============================================================================
# PIPELINE CONSTANTS
# ============================================================================
HR_LAG_S = 15  # HR lags power by ~15 seconds; used in RF calculation and late_dropout correction
POWER_SMOOTH_S = 30  # Power smoothing window for HR-related calculations
HR_CV_THRESHOLD = 12.0  # If power:HR ratio CV > this %, mark HR as unreliable and blank RF

# ============================================================================
# v43: RF ADJUSTMENT CONSTANTS (from BFW Constants sheet)
# ============================================================================

# Import central config for shared constants
try:
    from config import (PEAK_CP_WATTS, POWER_SCORE_RIEGEL_K, POWER_SCORE_REFERENCE_DIST_KM,
                        POWER_SCORE_RF_DIVISOR, POWER_SCORE_FACTOR_BOOST,
                        POWER_SCORE_AIR_THRESHOLD, POWER_SCORE_AIR_EXCESS_FACTOR,
                        CTL_TIME_CONSTANT, ATL_TIME_CONSTANT, RF_TREND_WINDOW,
                        RF_WINDOW_DURATION_S,
                        TERRAIN_LINEAR_SLOPE, TERRAIN_LINEAR_CAP, TERRAIN_STRAVA_ELEV_MIN,
                        DURATION_PENALTY_DAMPING,
                        # v51: Easy RF & Alert parameters
                        EASY_RF_HR_MIN, EASY_RF_NP_CP_MAX, EASY_RF_VI_MAX, EASY_RF_DIST_MIN_KM,
                        EASY_RF_EMA_SPAN, EASY_RF_Z_WINDOW,
                        ALERT1_RFL_DROP, ALERT1_CTL_RISE, ALERT1_WINDOW_DAYS,
                        ALERT1B_RFL_GAP, ALERT1B_PEAK_WINDOW_DAYS, ALERT1B_RACE_WINDOW_DAYS,
                        ALERT2_TSB_THRESHOLD, ALERT2_COUNT, ALERT2_WINDOW,
                        ALERT3B_Z_THRESHOLD, ALERT5_GAP_THRESHOLD)
except ImportError:
    PEAK_CP_WATTS = 372  # Fallback — must match config.py
    POWER_SCORE_RIEGEL_K = 0.08
    POWER_SCORE_REFERENCE_DIST_KM = 5.0
    POWER_SCORE_RF_DIVISOR = 180
    POWER_SCORE_FACTOR_BOOST = 0.5
    POWER_SCORE_AIR_THRESHOLD = 0.04
    POWER_SCORE_AIR_EXCESS_FACTOR = 0.5
    CTL_TIME_CONSTANT = 42
    ATL_TIME_CONSTANT = 7
    RF_TREND_WINDOW = 42
    RF_WINDOW_DURATION_S = 2400.0
    TERRAIN_LINEAR_SLOPE = 0.002
    TERRAIN_LINEAR_CAP = 0.05
    TERRAIN_STRAVA_ELEV_MIN = 8.0
    DURATION_PENALTY_DAMPING = 0.33
    # v51 fallbacks
    EASY_RF_HR_MIN = 120
    EASY_RF_NP_CP_MAX = 0.85
    EASY_RF_VI_MAX = 1.10
    EASY_RF_DIST_MIN_KM = 4.0
    EASY_RF_EMA_SPAN = 15
    EASY_RF_Z_WINDOW = 30
    ALERT1_RFL_DROP = 0.02
    ALERT1_CTL_RISE = 3.0
    ALERT1_WINDOW_DAYS = 28
    ALERT1B_RFL_GAP = 0.02
    ALERT1B_PEAK_WINDOW_DAYS = 90
    ALERT1B_RACE_WINDOW_DAYS = 7
    ALERT2_TSB_THRESHOLD = -15
    ALERT2_COUNT = 3
    ALERT2_WINDOW = 5
    ALERT3B_Z_THRESHOLD = -2.0
    ALERT5_GAP_THRESHOLD = -0.03

RF_CONSTANTS = {
    # Rolling periods
    'days_back': RF_TREND_WINDOW,  # v45: Extended from 21 to 42 with quadratic decay (from config.py)
    
    # Temperature adjustment
    'temp_adj_enabled': True,
    'temp_baseline': 10,  # °C - no adjustment below this
    'temp_factor_1': 0.003,  # Penalty per °C (10-20°C range)
    'temp_factor_2': 0.01,  # Penalty per °C (above 20°C)
    'min_acclimatized_temp': 16,
    
    # Humidity adjustment  
    'humidity_adj_enabled': True,
    'humidity_temp_threshold': 20,  # Only apply above this temp
    'humidity_baseline': 50,  # % where penalty starts
    'humidity_factor': 0.005,  # Penalty per 10% above baseline
    
    # Terrain adjustment (boost undulating terrain runs)
    'terrain_adj_enabled': True,
    'terrain_threshold': 4.0,  # Undulation score where adjustment starts (legacy, kept for intensity_adj)
    'terrain_factor': 0.012,  # Legacy linear factor (replaced by linear-with-cap in v47)
    'terrain_adj_max': 1.12,  # Legacy max (replaced by terrain_linear_cap)
    'terrain_linear_cap': TERRAIN_LINEAR_CAP,        # v48: max boost as fraction (from config.py)
    'terrain_linear_slope': TERRAIN_LINEAR_SLOPE,     # v48: linear boost per undulation unit (from config.py)
    'duration_penalty_damping': DURATION_PENALTY_DAMPING,  # v47: fraction of R2 for penalty side
    'terrain_strava_elev_min': TERRAIN_STRAVA_ELEV_MIN,  # v46: Strava elev gate (from config.py)
    'terrain_hill_interval_ratio': 1.5,  # Uphill/downhill power ratio above which = hill intervals (no terrain boost)
    
    # Elevation adjustment (boost for big sustained climbs)
    'elevation_adj_enabled': True,
    'elevation_threshold': 25.0,  # Gain/km where adjustment starts
    'elevation_factor': 0.005,  # Boost per m/km above threshold
    'elevation_adj_max': 1.10,  # Maximum elevation adjustment (10% cap)
    
    # Intensity adjustment (boost high-intensity efforts)
    'intensity_adj_enabled': False,  # v46: disabled - was inflating RF by +30.7pp avg
    'intensity_threshold': 0.88,  # Start boosting above 88% of CP
    'intensity_factor': 0.25,  # Boost per 1% above threshold
    'intensity_adj_max': 1.10,  # Maximum intensity adjustment (10% cap)
    'intensity_default_rfl': 0.85,  # Fallback RFL when no RF_Trend available
    'peak_cp_watts': PEAK_CP_WATTS,  # From config.py
    
    # Duration adjustment (boost long efforts to compensate for cardiac drift)
    'duration_adj_enabled': True,
    'duration_threshold_min': 60,  # Reward: only apply above 60 minutes
    'duration_penalty_threshold_min': 75,  # Penalty: only apply above 75 minutes
    'duration_factor': 0.00037,  # Boost per minute above threshold
    'duration_drift_threshold': -0.3,  # Drift (%/min) at which adjustment is zeroed
    
    # Heat scaling (shared by RF temp_adj and Power Score heat_adj)
    'heat_reference_mins': 90.0,  # Duration at which heat has full effect
    'heat_max_multiplier': 1.5,   # Cap on heat multiplier (150% of full effect)
    
    # RF measurement window
    'rf_window_duration_s': RF_WINDOW_DURATION_S,  # From config.py (40 mins)
    'rf_window_pause_threshold_s': 10.0,  # Gaps longer than this are pauses
    
    # HR correction
    'hr_correction_floor_bpm': 120.0,  # HR floor for late dropout detection
    'hr_correction_hard_floor_s': 300.0,  # Minimum seconds before HR correction applies
    
    # Factor (weighting) calculation
    'factor_hr_baseline': 100,
    'factor_hr_boost_threshold': 160,
    'factor_hr_boost_power': 5,
    'factor_divisor': 1000,
    'factor_reduction_max_pct': 90,  # Max % reduction to Factor for below-trend runs
    
    # RF max jump
    'rf_max_jump_pct': 10,  # Max % RF can increase run-to-run
    
    # TSS calculation
    'tss_hr_baseline': 90,
    'tss_divisor': 150000,
    
    # CTL/ATL
    'ctl_time_constant': CTL_TIME_CONSTANT,  # days (from config.py)
    'atl_time_constant': ATL_TIME_CONSTANT,   # days (from config.py)
    
    # Days-off penalty (RF_Trend with gaps)
    'days_off_gap_threshold': 7,  # Days without activity before penalty applies
    'days_off_penalty_rate': 2,   # Extra effective days added per day beyond threshold
    
    # Power Score threshold (v44.5: PS sets floor on RF_adj and boosts Factor)
    # Other PS params in config.py: POWER_SCORE_RIEGEL_K, POWER_SCORE_REFERENCE_DIST_KM, etc.
    'power_score_threshold': 310,  # Only runs above this get Factor boost
    
    # Factor reduction for RF below trend
    'factor_reduction_gap_days': 7,  # Don't reduce factor if gap since last run > this
}

# Default era adjusters (used as fallback only)
# v44: Now 7 eras - v1_late added for dying pod period (simulated power)
# These are calculated dynamically from actual data; fallback values are rough estimates
DEFAULT_ERA_ADJUSTERS = {
    'pre_stryd': {'re_median': 0.913, 'adj_factor': 1.019},
    'v1': {'re_median': 0.976, 'adj_factor': 1.089},       # Early v1 only (May-Aug 2017)
    'v1_late': {'re_median': 0.913, 'adj_factor': 1.019},  # Same as pre_stryd (simulated)
    'repl': {'re_median': 0.912, 'adj_factor': 1.018},     # Combined from repl_early/late
    'air': {'re_median': 0.920, 'adj_factor': 1.027},      # Combined from air_early/late
    's4': {'re_median': 0.896, 'adj_factor': 1.000},       # Reference era
    's5': {'re_median': 0.888, 'adj_factor': 0.991},
}

# Valid surface values
VALID_SURFACES = {'TRACK', 'INDOOR_TRACK', 'TRAIL', 'SNOW', 'HEAVY_SNOW'}


# ============================================================================
# v43/v44: RF ADJUSTMENT FUNCTIONS
# ============================================================================

def calculate_era_adjusters_from_data(df: pd.DataFrame, era_col: str = 'calibration_era_id', 
                                       re_col: str = 'RE_avg', reference_era: str = 's4') -> dict:
    """
    Calculate era adjusters dynamically from actual data.
    
    Logic (same as BFW):
    1. For each era, calculate median RE_avg
    2. adj_factor = era_re_median / reference_era_re_median
    
    Returns dict: era_id -> {'re_median': float, 'adj_factor': float}
    """
    # Find the era column (might be 'Era' or 'calibration_era_id')
    actual_era_col = None
    for col in [era_col, 'Era', 'era', 'calibration_era_id']:
        if col in df.columns:
            actual_era_col = col
            break
    
    if actual_era_col is None:
        print("  WARNING: No era column found, using default adjusters")
        return DEFAULT_ERA_ADJUSTERS.copy()
    
    # Find RE column
    actual_re_col = None
    for col in [re_col, 'RE_avg', 're_avg']:
        if col in df.columns:
            actual_re_col = col
            break
    
    if actual_re_col is None:
        print("  WARNING: No RE_avg column found, using default adjusters")
        return DEFAULT_ERA_ADJUSTERS.copy()
    
    # Calculate median RE by era
    # Match BFW: exclude RE_avg > 1.1, but include NaN rows (median ignores them)
    # Create a copy with >1.1 values set to NaN
    df_for_median = df.copy()
    df_for_median.loc[df_for_median[actual_re_col] > 1.1, actual_re_col] = np.nan
    
    medians = df_for_median.groupby(actual_era_col)[actual_re_col].median()
    
    # Debug info
    na_count = df[actual_re_col].isna().sum()
    over_1_1 = (df[actual_re_col] > 1.1).sum()
    valid_count = len(df) - na_count - over_1_1
    print(f"  Debug: {len(df)} total rows, {valid_count} with valid RE_avg (NA:{na_count}, >1.1:{over_1_1})")
    
    medians = df_for_median.groupby(actual_era_col)[actual_re_col].median()
    
    # Get reference era median
    reference_era_lower = reference_era.lower()
    ref_median = None
    for era_id, median in medians.items():
        if str(era_id).lower() == reference_era_lower:
            ref_median = median
            break
    
    if ref_median is None or ref_median <= 0:
        print(f"  WARNING: Reference era '{reference_era}' not found or invalid, using default adjusters")
        return DEFAULT_ERA_ADJUSTERS.copy()
    
    # Build adjusters dict
    adjusters = {}
    for era_id, re_median in medians.items():
        era_key = str(era_id).lower().strip()
        adj_factor = re_median / ref_median
        adjusters[era_key] = {
            're_median': float(re_median),
            'adj_factor': float(adj_factor),
        }
    
    print(f"  Calculated era adjusters from {valid_count} activities:")
    for era_id in sorted(adjusters.keys()):
        data = adjusters[era_id]
        print(f"    {era_id}: RE_median={data['re_median']:.4f}, adj_factor={data['adj_factor']:.4f}")
    
    return adjusters


def export_era_adjusters_csv(adjusters: dict, output_path: str) -> None:
    """Export era adjusters to CSV for transparency."""
    rows = []
    for era_id in sorted(adjusters.keys()):
        data = adjusters[era_id]
        rows.append({
            'era_id': era_id,
            're_median': data['re_median'],
            'adj_factor': data['adj_factor'],
        })
    
    df = pd.DataFrame(rows)
    df.to_csv(output_path, index=False)
    print(f"  Exported era adjusters to: {output_path}")


def load_era_adjusters(csv_path: str) -> dict:
    """Load era adjusters from CSV, or use defaults if not found."""
    if os.path.exists(csv_path):
        try:
            df = pd.read_csv(csv_path)
            adjusters = {}
            for _, row in df.iterrows():
                era_id = str(row.get('era_id', '')).lower().strip()
                if era_id:
                    adjusters[era_id] = {
                        're_median': float(row.get('re_median', 0.95)),
                        'adj_factor': float(row.get('adj_factor', 1.0)),
                    }
            print(f"  Loaded {len(adjusters)} era adjusters from {csv_path}")
            return adjusters
        except Exception as e:
            print(f"  WARNING: Could not load {csv_path}: {e}, using defaults")
    return DEFAULT_ERA_ADJUSTERS.copy()


def calc_temp_adj(temp_c: float, humidity_pct: float, acclimatized_temp: float = None) -> float:
    """
    Calculate temperature/humidity adjustment factor.
    Returns multiplier > 1.0 for hot/humid conditions.
    """
    if not RF_CONSTANTS['temp_adj_enabled']:
        return 1.0
    
    if temp_c is None or not np.isfinite(temp_c):
        return 1.0
    
    if acclimatized_temp is None:
        acclimatized_temp = RF_CONSTANTS['min_acclimatized_temp']
    
    # BFW uses: base_temp = MAX(Temp_Trend, min_acclimatized_temp)
    # min_acclimatized_temp is 16 in BFW
    base_temp = max(acclimatized_temp, RF_CONSTANTS['min_acclimatized_temp'])
    
    # Temperature penalty
    if temp_c <= base_temp:
        temp_penalty = 0.0
    elif base_temp >= 20:
        temp_penalty = (temp_c - base_temp) * RF_CONSTANTS['temp_factor_2']
    elif temp_c <= 20:
        temp_penalty = (temp_c - base_temp) * RF_CONSTANTS['temp_factor_1']
    else:
        # Split: base_temp to 20 at factor_1, 20+ at factor_2
        temp_penalty = ((20 - base_temp) * RF_CONSTANTS['temp_factor_1'] + 
                       (temp_c - 20) * RF_CONSTANTS['temp_factor_2'])
    
    # Humidity penalty (only above temp threshold)
    humid_penalty = 0.0
    if RF_CONSTANTS['humidity_adj_enabled'] and temp_c > RF_CONSTANTS['humidity_temp_threshold']:
        if humidity_pct is not None and np.isfinite(humidity_pct):
            humid_penalty = max(0, (humidity_pct - RF_CONSTANTS['humidity_baseline']) / 10 * 
                              RF_CONSTANTS['humidity_factor'])
    
    return 1.0 + temp_penalty + humid_penalty


def calc_terrain_adj(undulation_score: float) -> float:
    """
    Calculate terrain adjustment for undulating runs.
    
    v47: Replaced exponential saturation with linear-with-cap formula.
    v46: adj = 1 + cap * (1 - exp(-rate * und))  [over-compensated mid-range]
    v47: adj = 1 + min(cap, slope * und)
    
    Targets: ~2% for typical hilly RF windows (und ~10), cap at +5%.
    Deliberately conservative — use surface_adj overrides for known tough courses.
    
    The exponential curve gave Sweden countryside (und ~5) nearly the same
    boost as Lidingöloppet (und ~10), despite LL30 having 34% more elevation
    per km and genuinely harder terrain. Linear scaling properly differentiates
    mild from severe undulation.
    
    High undulation (constant direction changes) suppresses RF because:
    - Muscles can't find a rhythm, increasing energy cost
    - Power fluctuates constantly, but median RF doesn't capture the extra effort
    - Similar to running on soft sand - higher cost, same apparent speed
    """
    if not RF_CONSTANTS['terrain_adj_enabled']:
        return 1.0
    
    if undulation_score is None or not np.isfinite(undulation_score):
        return 1.0
    
    threshold = RF_CONSTANTS['terrain_threshold']
    
    if undulation_score <= threshold:
        return 1.0
    
    # v47: Linear with cap — boost = min(cap, slope × undulation)
    # Conservative: ~2% for und=10 (LL30-type), use surface_adj overrides for tough courses
    cap = RF_CONSTANTS['terrain_linear_cap']
    slope = RF_CONSTANTS['terrain_linear_slope']
    adj = 1.0 + min(cap, slope * undulation_score)
    
    return adj


def calc_elevation_adj(elev_gain_per_km: float) -> float:
    """
    Calculate elevation adjustment for big sustained climbs.
    
    Mountain runs with low reversals don't get much undulation boost,
    but the sustained climbing still suppresses RF. This provides
    a separate adjustment based purely on elevation gain per km.
    
    Thresholds (from analysis):
    - LL30: ~12-14 m/km (no adjustment needed)
    - Mountain runs: 30-40+ m/km (need boost)
    - Threshold set at 25 m/km to separate these
    """
    if not RF_CONSTANTS['elevation_adj_enabled']:
        return 1.0
    
    if elev_gain_per_km is None or not np.isfinite(elev_gain_per_km):
        return 1.0
    
    threshold = RF_CONSTANTS['elevation_threshold']
    
    if elev_gain_per_km <= threshold:
        return 1.0
    
    # Linear boost above threshold
    excess = elev_gain_per_km - threshold
    adj = 1.0 + excess * RF_CONSTANTS['elevation_factor']
    
    return min(adj, RF_CONSTANTS['elevation_adj_max'])


def calc_era_adj(era_id: str, era_adjusters: dict, power_adj_to_s4: float = None) -> float:
    """Get era adjustment factor to normalize to S4 baseline.
    
    v49: Prefer power_adjuster_to_S4 from rebuild (regression-based, speed-controlled).
    Falls back to RE-median adjusters only if power_adjuster_to_S4 is unavailable.
    """
    # v49: Use regression-based adjuster if available
    if power_adj_to_s4 is not None and np.isfinite(power_adj_to_s4) and power_adj_to_s4 > 0:
        return float(power_adj_to_s4)
    # Fallback to RE-median adjusters
    era_id = str(era_id).lower().strip() if era_id else ''
    era_data = era_adjusters.get(era_id, {})
    return era_data.get('adj_factor', 1.0)


def calc_intensity_adj(avg_power: float, rfl_trend: float, undulation_score: float = None) -> float:
    """
    Calculate intensity adjustment for high-intensity efforts.
    
    At high intensity (power close to CP), HR drifts up relative to power,
    which artificially suppresses RF (power/HR). This adjustment compensates.
    
    v45: Undulating terrain lowers the threshold at which intensity adjustment kicks in.
    Running at 80% CP on LL30 terrain is as demanding as 88% CP on flat.
    
    Args:
        avg_power: Average power for the run (W)
        rfl_trend: Current RFL_Trend (0-1 scale, e.g., 0.85 = 85% of peak)
        undulation_score: RF window undulation score (optional)
    
    Returns:
        Adjustment factor >= 1.0 (1.0 = no adjustment)
    """
    if not RF_CONSTANTS.get('intensity_adj_enabled', True):
        return 1.0
    
    if avg_power is None or not np.isfinite(avg_power) or avg_power <= 0:
        return 1.0
    
    if rfl_trend is None or not np.isfinite(rfl_trend) or rfl_trend <= 0:
        return 1.0
    
    # Calculate current CP from peak CP and RFL_Trend
    peak_cp = RF_CONSTANTS.get('peak_cp_watts', 372)
    current_cp = peak_cp * rfl_trend
    
    if current_cp <= 0:
        return 1.0
    
    # Calculate intensity as fraction of CP
    intensity = avg_power / current_cp
    
    # Base threshold
    base_threshold = RF_CONSTANTS.get('intensity_threshold', 0.88)
    
    # v45: Lower threshold for undulating terrain
    # Reduce by 0.5% per undulation point above 4
    # e.g., und=15 -> reduce threshold by 5.5% (88% -> 82.5%)
    terrain_threshold_reduction = 0.0
    if undulation_score is not None and np.isfinite(undulation_score) and undulation_score > 4:
        terrain_threshold_reduction = (undulation_score - 4) * 0.005
    
    adjusted_threshold = base_threshold - terrain_threshold_reduction
    # Don't let threshold go below 70%
    adjusted_threshold = max(adjusted_threshold, 0.70)
    
    if intensity <= adjusted_threshold:
        return 1.0
    
    # Calculate boost: (intensity - threshold) * factor
    # e.g., at 95% CP with threshold 88%: (0.95 - 0.88) * 0.25 = 0.0175 -> 1.75% boost
    factor = RF_CONSTANTS.get('intensity_factor', 0.25)
    max_adj = RF_CONSTANTS.get('intensity_adj_max', 1.10)
    
    adj = 1.0 + (intensity - adjusted_threshold) * factor
    
    return min(adj, max_adj)


def calc_duration_adj(moving_time_s: float, rf_drift_pct_per_min: float,
                      full_run_drift: float = None, full_run_drift_r2: float = None) -> float:
    """
    Calculate duration adjustment for long efforts using full-run P/HR drift.
    
    v47: Uses full-run drift (measured post-warmup to pre-cooldown) instead of
    the 20-minute RF window drift. The formula rewards runs where P/HR was
    well-held over the full duration, and mildly penalises clear cardiac drift.
    
    Formula: base = 1.2 + drift * 2
      - Neutral point at drift = -0.10 (typical long-run cardiac drift)
      - If base >= 1: adj = base ^ R2           (reward well-held runs)  
      - If base < 1:  adj = base ^ (R2 * 0.33)  (dampened penalty)
    
    R2 acts as a confidence dial — low R2 (noisy, no real trend) pushes adj
    towards 1.0 regardless of drift value. High R2 gives the full effect.
    
    Falls back to 1.0 if no full-run drift data is available.
    
    Args:
        moving_time_s: Moving time in seconds
        rf_drift_pct_per_min: RF window drift (%/min) — unused in v47, kept for signature
        full_run_drift: Full-run drift rate (%/min), or None
        full_run_drift_r2: R-squared of the full-run drift regression, or None
    
    Returns:
        Adjustment factor (can be > 1 or < 1)
    """
    if not RF_CONSTANTS.get('duration_adj_enabled', True):
        return 1.0
    
    if moving_time_s is None or not np.isfinite(moving_time_s) or moving_time_s <= 0:
        return 1.0
    
    threshold_min = RF_CONSTANTS.get('duration_threshold_min', 60)
    penalty_threshold_min = RF_CONSTANTS.get('duration_penalty_threshold_min', 75)
    duration_min = moving_time_s / 60.0
    
    if duration_min <= threshold_min:
        return 1.0
    
    # Need full-run drift data
    if full_run_drift is None or not np.isfinite(full_run_drift):
        return 1.0
    if full_run_drift_r2 is None or not np.isfinite(full_run_drift_r2):
        return 1.0
    
    # Base: neutral at drift = -0.10 (typical long-run drift)
    # Better than -0.10 -> reward, worse -> mild penalty
    base = 1.2 + full_run_drift * 2.0
    
    # Clamp base to prevent extreme values
    base = max(0.5, min(1.5, base))
    
    r2 = max(0.0, full_run_drift_r2)
    
    if base >= 1.0:
        # Reward: use full R2
        adj = base ** r2
    else:
        # Penalty: only for runs > penalty_threshold_min (75 min)
        # Avoids penalising 60-70 min structured workouts
        if duration_min <= penalty_threshold_min:
            return 1.0
        penalty_damping = RF_CONSTANTS.get('duration_penalty_damping', 0.33)
        adj = base ** (r2 * penalty_damping)
    
    return adj


def calc_total_adj(temp_adj: float, terrain_adj: float, era_adj: float, own_adj: float = 1.0,
                   intensity_adj: float = 1.0, duration_adj: float = 1.0, 
                   elevation_adj: float = 1.0) -> float:
    """Combine all adjustment factors."""
    base = temp_adj * terrain_adj * era_adj * (own_adj if own_adj and np.isfinite(own_adj) else 1.0)
    base = base * (intensity_adj if intensity_adj and np.isfinite(intensity_adj) else 1.0)
    base = base * (elevation_adj if elevation_adj and np.isfinite(elevation_adj) else 1.0)
    return base * (duration_adj if duration_adj and np.isfinite(duration_adj) else 1.0)


def calc_rf_adj(rf_raw: float, total_adj: float, prev_rf_trend: float = None) -> float:
    """
    Calculate adjusted RF with max jump limit.
    """
    if rf_raw is None or not np.isfinite(rf_raw):
        return np.nan
    
    rf_adj = rf_raw * total_adj
    
    # Cap at max jump from previous trend
    if prev_rf_trend is not None and np.isfinite(prev_rf_trend) and prev_rf_trend > 0:
        max_rf = prev_rf_trend * (1 + RF_CONSTANTS['rf_max_jump_pct'] / 100)
        rf_adj = min(rf_adj, max_rf)
    
    return rf_adj


def calc_factor(distance_m: float, avg_hr: float, rf_adj: float, prev_rf_trend: float = None,
                days_since_last_run: int = None) -> float:
    """
    Calculate weighting factor for RF trend.
    Longer runs at higher HR get more weight.
    Low RF outliers get reduced weight to prevent single bad runs dragging down trend.
    
    v44.3: If days_since_last_run > factor_reduction_gap_days, don't reduce factor
    for low RF - the runner may have detrained and this RF is the new reality.
    """
    if rf_adj is None or not np.isfinite(rf_adj):
        return np.nan
    
    if distance_m is None or not np.isfinite(distance_m) or distance_m <= 0:
        return np.nan
    
    if avg_hr is None or not np.isfinite(avg_hr) or avg_hr <= RF_CONSTANTS['factor_hr_baseline']:
        return np.nan
    
    hr_base = RF_CONSTANTS['factor_hr_baseline']
    hr_boost_threshold = RF_CONSTANTS['factor_hr_boost_threshold']
    hr_boost_power = RF_CONSTANTS['factor_hr_boost_power']
    divisor = RF_CONSTANTS['factor_divisor']
    
    hr_boost = max(1, avg_hr / hr_boost_threshold) ** hr_boost_power
    
    base_factor = (distance_m * (avg_hr - hr_base) * hr_boost) / divisor
    
    # Outlier adjustment: reduce factor for RF significantly below previous trend
    # x% below trend -> reduce factor by min(x^2, 90)%
    # v44.3: Skip this reduction if there's been a long gap (detraining is real)
    gap_threshold = RF_CONSTANTS.get('factor_reduction_gap_days', 7)
    apply_reduction = True
    
    if days_since_last_run is not None and days_since_last_run > gap_threshold:
        apply_reduction = False  # Trust this RF, even if below trend
    
    if apply_reduction and prev_rf_trend is not None and np.isfinite(prev_rf_trend) and prev_rf_trend > 0:
        pct_below = max(0, (prev_rf_trend - rf_adj) / prev_rf_trend * 100)  # % below trend
        if pct_below > 0:
            reduction_pct = min(pct_below ** 2, RF_CONSTANTS['factor_reduction_max_pct'])  # Square it, cap at max
            base_factor = base_factor * (1 - reduction_pct / 100)
    
    return base_factor


def calc_tss(moving_time_s: float, avg_hr: float, rfl: float, terrain_adj: float, 
             distance_m: float, is_race: bool) -> float:
    """
    Calculate Training Stress Score.
    Based on time, HR intensity, and current fitness level.
    """
    if moving_time_s is None or not np.isfinite(moving_time_s) or moving_time_s <= 0:
        return np.nan
    
    hr_base = RF_CONSTANTS['tss_hr_baseline']
    divisor = RF_CONSTANTS['tss_divisor']
    
    if avg_hr is None or not np.isfinite(avg_hr):
        avg_hr = hr_base + 30  # Default assumption
    
    if distance_m is None or not np.isfinite(distance_m):
        distance_m = 0
    
    if terrain_adj is None or not np.isfinite(terrain_adj):
        terrain_adj = 1.0
    
    if rfl is not None and np.isfinite(rfl) and rfl > 0:
        # Full formula with RFL
        tss = (moving_time_s * (avg_hr - hr_base) ** 2 / (divisor * rfl) * terrain_adj + 
               distance_m / 200) / 1.5
    else:
        # Simplified formula without RFL
        if is_race:
            tss = distance_m / 100
        else:
            tss = distance_m / 200
    
    return max(0, tss)


def calc_rf_trend_for_row(df: pd.DataFrame, row_idx: int, days_back: int = None) -> float:
    """
    Calculate weighted RF trend for a specific row with quadratic time decay.
    Uses Factor as weight, RF_adj as value, with additional time-based decay.
    
    v45: Extended to 42 days with quadratic decay: weight = 1 - (days_ago/42)^2
    This keeps recent runs strong (97% at 1 week, 89% at 2 weeks, 75% at 3 weeks)
    while smoothly tapering to 0% at 42 days (matching CTL time constant).
    
    Formula: SUMPRODUCT(decay_weight * Factor * RF_adj) / SUMPRODUCT(decay_weight * Factor)
    """
    if days_back is None:
        days_back = RF_CONSTANTS['days_back']
    
    current_date = df.loc[row_idx, 'date']
    if pd.isna(current_date):
        return np.nan
    
    cutoff_date = current_date - pd.Timedelta(days=days_back)
    
    # Get rows in window (up to and including current row)
    mask = (df['date'] >= cutoff_date) & (df['date'] <= current_date) & (df.index <= row_idx)
    window_df = df.loc[mask].copy()
    
    if len(window_df) == 0:
        return np.nan
    
    # Calculate days ago for each row
    window_df['days_ago'] = (current_date - window_df['date']).dt.total_seconds() / 86400
    
    # Quadratic decay: 1 - (days_ago / days_back)^2
    window_df['time_decay'] = (1 - (window_df['days_ago'] / days_back) ** 2).clip(lower=0)
    
    factors = window_df['Factor'].fillna(0)
    rf_adjs = window_df['RF_adj'].fillna(0)
    time_decay = window_df['time_decay']
    
    # Only include rows where both Factor and RF_adj are valid
    valid = (factors > 0) & (rf_adjs > 0)
    
    # Combined weight = Factor * time_decay
    combined_weights = factors[valid] * time_decay[valid]
    
    numerator = (combined_weights * rf_adjs[valid]).sum()
    denominator = combined_weights.sum()
    
    if denominator == 0:
        return np.nan
    
    return numerator / denominator


def calc_rolling_metrics(df: pd.DataFrame) -> pd.DataFrame:
    """
    Calculate RFL metrics from existing RF_Trend.
    
    v44.5: Simplified - Power Score now affects RF_adj directly (floor) and Factor (boost)
    in the main calculation loop. This function just calculates RFL from RF_Trend.
    
    - RFL = RF_adj / peak_RF (per-activity)
    - RFL_Trend = RF_Trend / peak_RF (rolling)
    - Power_Score kept for diagnostics but no longer creates separate RFL_Combined
    
    Returns DataFrame with RFL, RFL_Trend columns added.
    """
    print("  Calculating RFL metrics...")
    
    # Initialize columns
    df['RFL'] = np.nan
    df['RFL_Trend'] = np.nan
    
    # Find all-time peak RF_Trend
    peak_rf_trend = df['RF_Trend'].max()
    print(f"  Peak RF_Trend: {peak_rf_trend:.4f}")
    
    if peak_rf_trend > 0 and np.isfinite(peak_rf_trend):
        # Calculate RFL (per-activity RF as % of peak)
        df['RFL'] = df['RF_adj'] / peak_rf_trend
        
        # Calculate RFL_Trend (rolling RF as % of peak)
        df['RFL_Trend'] = df['RF_Trend'] / peak_rf_trend
    
    # Report Power Score stats
    if 'Power_Score' in df.columns:
        ps_valid = df['Power_Score'].notna()
        if ps_valid.sum() > 0:
            print(f"  Power Score range: {df.loc[ps_valid, 'Power_Score'].min():.1f} - {df.loc[ps_valid, 'Power_Score'].max():.1f}")
    
    # Count how often Power Score floor was applied (RF_adj was boosted)
    if 'PS_Floor_Applied' in df.columns:
        floor_count = df['PS_Floor_Applied'].sum()
        print(f"  Power Score floor applied: {floor_count} runs")
    
    return df


def calc_easy_rf_metrics(df: pd.DataFrame) -> pd.DataFrame:
    """v51: Calculate Easy RF EMA, z-score, RFL_Trend_Delta, and Easy_RFL_Gap.
    
    Easy RF EMA is an exponentially weighted mean of RF_adj for easy runs only
    (avg_hr 120-150, distance >= 4km, non-race, RF_adj valid). This tracks 
    underlying aerobic fitness without contamination from races or hard efforts.
    
    New columns:
      Easy_RF_EMA       - 15-run EWM of easy-run RF_adj, forward-filled to all runs
      Easy_RF_z         - Z-score of easy RF vs trailing 30-run easy-run baseline
      RFL_Trend_Delta   - Change in RFL_Trend from previous run
      Easy_RFL_Gap      - Gap between Easy_RF_EMA (normalised) and RFL_Trend
    """
    print("\n=== v51: Calculating Easy RF metrics ===")
    
    # --- Easy run mask (v51: power-based ceiling replaces HR ceiling) ---
    # nPower < 85% CP captures genuine easy efforts even with elevated HR (post-illness, heat)
    # VI cap (nPower/avg_power < 1.10) excludes interval sessions with low average power
    # CP may not be populated yet, so derive from RFL_Trend * PEAK_CP_WATTS
    from config import PEAK_CP_WATTS as _peak_cp
    cp_series = df['RFL_Trend'] * _peak_cp
    easy_mask = (
        (df['npower_w'].notna()) & (df['avg_power_w'] > 0) & (cp_series > 0) &
        (df['npower_w'] < cp_series * EASY_RF_NP_CP_MAX) &
        ((df['npower_w'] / df['avg_power_w']) < EASY_RF_VI_MAX) &
        (df['avg_hr'] >= EASY_RF_HR_MIN) &
        (df['distance_km'] >= EASY_RF_DIST_MIN_KM) &
        (df['race_flag'] != 1) &
        (df['RF_adj'].notna())
    )
    n_easy = easy_mask.sum()
    n_total = len(df)
    print(f"  Easy runs: {n_easy} of {n_total} ({100*n_easy/n_total:.0f}%)")
    
    # --- Easy RF EMA (span=15 easy runs, then forward-fill) ---
    df['Easy_RF_EMA'] = np.nan
    if n_easy > 0:
        easy_rf = df.loc[easy_mask, 'RF_adj'].copy()
        ema = easy_rf.ewm(span=EASY_RF_EMA_SPAN, adjust=True).mean()
        df.loc[easy_mask, 'Easy_RF_EMA'] = ema
        df['Easy_RF_EMA'] = df['Easy_RF_EMA'].ffill()
        print(f"  Easy RF EMA: {ema.iloc[-1]:.4f} (latest)")
    
    # --- Easy RF z-score (vs trailing 30-run easy-run baseline) ---
    df['Easy_RF_z'] = np.nan
    if n_easy >= EASY_RF_Z_WINDOW:
        easy_rf_vals = df.loc[easy_mask, 'RF_adj'].copy()
        rolling_mean = easy_rf_vals.rolling(window=EASY_RF_Z_WINDOW, min_periods=10).mean()
        rolling_std = easy_rf_vals.rolling(window=EASY_RF_Z_WINDOW, min_periods=10).std()
        z_scores = (easy_rf_vals - rolling_mean) / rolling_std.replace(0, np.nan)
        df.loc[easy_mask, 'Easy_RF_z'] = z_scores
    
    # --- RFL_Trend_Delta (run-to-run change) ---
    df['RFL_Trend_Delta'] = np.nan
    if 'RFL_Trend' in df.columns:
        rfl_trend = df['RFL_Trend'].copy()
        df['RFL_Trend_Delta'] = rfl_trend - rfl_trend.shift(1)
        # Report distribution
        valid_delta = df['RFL_Trend_Delta'].dropna()
        if len(valid_delta) > 50:
            p5 = valid_delta.quantile(0.05)
            p95 = valid_delta.quantile(0.95)
            big_boost = (valid_delta > 0.01).sum()
            big_drag = (valid_delta < -0.01).sum()
            print(f"  RFL_Trend_Delta: P5={p5*100:.2f}%, P95={p95*100:.2f}%, "
                  f"big boosts={big_boost}, big drags={big_drag}")
    
    # --- Easy_RFL_Gap (normalised Easy RF EMA minus RFL_Trend) ---
    df['Easy_RFL_Gap'] = np.nan
    if 'RFL_Trend' in df.columns and 'Easy_RF_EMA' in df.columns:
        peak_rf_trend = df['RF_Trend'].max() if 'RF_Trend' in df.columns else np.nan
        if pd.notna(peak_rf_trend) and peak_rf_trend > 0:
            easy_rfl = df['Easy_RF_EMA'] / peak_rf_trend
            df['Easy_RFL_Gap'] = easy_rfl - df['RFL_Trend']
            latest_gap = df['Easy_RFL_Gap'].dropna()
            if len(latest_gap) > 0:
                print(f"  Easy_RFL_Gap: {latest_gap.iloc[-1]*100:.1f}% (latest)")
    
    return df


def calc_alert_columns(df: pd.DataFrame) -> pd.DataFrame:
    """v51: Calculate historical alert columns for every run in the Master.
    
    Each alert column is True when the alert condition is active, False otherwise.
    Alerts persist (stay True) on every run while the condition holds.
    
    New columns:
      Alert_1      - CTL rising, RFL falling (4-week divergence)
      Alert_1b     - Taper not working (RFL below 90-day peak near race)
      Alert_2      - Sustained deep-negative TSB (3 of last 5 runs < -15)
      Alert_3b     - Easy run outlier (z < -2.0 on this easy run)
      Alert_5      - Easy RF / RFL_Trend gap < -3%
    """
    print("\n=== v51: Calculating alert columns ===")
    
    # Initialise all alert columns as False
    for col in ['Alert_1', 'Alert_1b', 'Alert_2', 'Alert_3b', 'Alert_5']:
        df[col] = False
    
    dates = df['date'].values  # numpy datetime64
    n = len(df)
    
    # --- Alert 1: CTL rising while RFL falling (4-week window) ---
    if 'CTL' in df.columns and 'RFL_Trend' in df.columns:
        rfl = df['RFL_Trend'].values
        ctl = df['CTL'].values
        a1_count = 0
        for i in range(1, n):
            # Find run ~28 days ago
            cutoff = dates[i] - np.timedelta64(ALERT1_WINDOW_DAYS, 'D')
            # Find the closest run at or before the cutoff
            earlier = np.where((dates[:i] <= cutoff))[0]
            if len(earlier) == 0:
                continue
            j = earlier[-1]
            if np.isfinite(rfl[i]) and np.isfinite(rfl[j]) and np.isfinite(ctl[i]) and np.isfinite(ctl[j]):
                rfl_drop = rfl[j] - rfl[i]  # positive = dropped
                ctl_rise = ctl[i] - ctl[j]   # positive = rose
                if rfl_drop > ALERT1_RFL_DROP and ctl_rise > ALERT1_CTL_RISE:
                    df.iat[i, df.columns.get_loc('Alert_1')] = True
                    a1_count += 1
        print(f"  Alert 1 (CTL up/RFL down): {a1_count} runs flagged")
    
    # --- Alert 1b: Taper not working (near race, RFL below 90d peak) ---
    if 'RFL_Trend' in df.columns and 'race_flag' in df.columns:
        rfl = df['RFL_Trend'].values
        race_flags = df['race_flag'].values
        dist_km = df['distance_km'].values
        a1b_count = 0
        for i in range(1, n):
            # Check if this is a race >= 20km
            if race_flags[i] != 1 or dist_km[i] < 20:
                continue
            # Find 90-day peak RFL_Trend before this race
            cutoff_90d = dates[i] - np.timedelta64(ALERT1B_PEAK_WINDOW_DAYS, 'D')
            window = np.where((dates[:i+1] >= cutoff_90d))[0]
            if len(window) < 3:
                continue
            rfl_window = rfl[window]
            peak = np.nanmax(rfl_window)
            if np.isfinite(peak) and np.isfinite(rfl[i]):
                gap = peak - rfl[i]
                if gap > ALERT1B_RFL_GAP:
                    df.iat[i, df.columns.get_loc('Alert_1b')] = True
                    a1b_count += 1
        print(f"  Alert 1b (taper): {a1b_count} runs flagged")
    
    # --- Alert 2: Sustained deep-negative TSB (3 of last 5 < -15) ---
    if 'TSB' in df.columns:
        tsb = df['TSB'].values
        deep = tsb < ALERT2_TSB_THRESHOLD
        a2_count = 0
        for i in range(ALERT2_WINDOW - 1, n):
            window_deep = deep[max(0, i - ALERT2_WINDOW + 1):i + 1]
            if np.sum(window_deep) >= ALERT2_COUNT:
                df.iat[i, df.columns.get_loc('Alert_2')] = True
                a2_count += 1
        print(f"  Alert 2 (deep TSB): {a2_count} runs flagged")
    
    # --- Alert 3b: Easy run outlier (z < -2.0) ---
    if 'Easy_RF_z' in df.columns:
        z_vals = df['Easy_RF_z'].values
        a3b_count = 0
        for i in range(n):
            if np.isfinite(z_vals[i]) and z_vals[i] < ALERT3B_Z_THRESHOLD:
                df.iat[i, df.columns.get_loc('Alert_3b')] = True
                a3b_count += 1
        print(f"  Alert 3b (easy outlier): {a3b_count} runs flagged")
    
    # --- Alert 5: Easy RF / RFL_Trend gap < -3% ---
    if 'Easy_RFL_Gap' in df.columns:
        gap = df['Easy_RFL_Gap'].values
        a5_count = 0
        for i in range(n):
            if np.isfinite(gap[i]) and gap[i] < ALERT5_GAP_THRESHOLD:
                df.iat[i, df.columns.get_loc('Alert_5')] = True
                a5_count += 1
        print(f"  Alert 5 (easy RF gap): {a5_count} runs flagged")
    
    # Summary
    any_alert = df[['Alert_1', 'Alert_1b', 'Alert_2', 'Alert_3b', 'Alert_5']].any(axis=1)
    total_flagged = any_alert.sum()
    print(f"  Total runs with any alert: {total_flagged} ({100*total_flagged/n:.1f}%)")
    
    return df


def get_current_alerts(df: pd.DataFrame) -> list:
    """v51: Check which alerts are active on the most recent run(s).
    
    Returns a list of alert dicts for console output and dashboard.
    """
    alerts = []
    if len(df) < 20:
        return alerts
    
    latest = df.iloc[-1]
    latest_date = latest['date']
    
    if latest.get('Alert_1', False):
        # Get details
        lookback = df[df['date'] >= latest_date - pd.Timedelta(days=ALERT1_WINDOW_DAYS)]
        rfl_drop = lookback.iloc[0]['RFL_Trend'] - latest['RFL_Trend'] if len(lookback) > 0 else 0
        ctl_rise = latest['CTL'] - lookback.iloc[0]['CTL'] if len(lookback) > 0 else 0
        alerts.append({
            'alert': 'Alert 1: Training more, scoring worse',
            'level': 'concern',
            'icon': '!!',
            'message': (f"RFL dropped {rfl_drop*100:.1f}% over {ALERT1_WINDOW_DAYS}d "
                        f"while CTL rose {ctl_rise:.0f}."),
        })
    
    if latest.get('Alert_1b', False):
        alerts.append({
            'alert': 'Alert 1b: Taper not working',
            'level': 'concern',
            'icon': '!!',
            'message': f"RFL below 90-day peak at race distance.",
        })
    
    if latest.get('Alert_2', False):
        last5 = df.tail(ALERT2_WINDOW)
        deep_count = (last5['TSB'] < ALERT2_TSB_THRESHOLD).sum()
        worst = last5['TSB'].min()
        alerts.append({
            'alert': 'Alert 2: Deep fatigue',
            'level': 'watch',
            'icon': '(!)',
            'message': (f"TSB < {ALERT2_TSB_THRESHOLD} on {deep_count}/{ALERT2_WINDOW} "
                        f"recent runs (worst: {worst:.0f})."),
        })
    
    if latest.get('Alert_3b', False):
        z = latest.get('Easy_RF_z', np.nan)
        if np.isfinite(z):
            alerts.append({
                'alert': 'Alert 3b: Easy run outlier',
                'level': 'watch',
                'icon': '(!)',
                'message': f"Last easy run z={z:.1f} vs baseline.",
            })
    
    if latest.get('Alert_5', False):
        gap = latest.get('Easy_RFL_Gap', np.nan)
        rfl_t = latest.get('RFL_Trend', np.nan)
        alerts.append({
            'alert': 'Alert 5: Easy RF divergence',
            'level': 'concern',
            'icon': '!!',
            'message': (f"Easy RF {abs(gap)*100:.1f}% below RFL Trend ({rfl_t*100:.1f}%). "
                        f"Early fatigue signal."),
        })
    
    return alerts


def print_alerts(alerts: list) -> None:
    """Print alert summary to console."""
    if not alerts:
        print("\n=== v51 Health Check: All clear ===")
        return
    
    level_order = {'concern': 0, 'watch': 1, 'info': 2}
    alerts.sort(key=lambda a: level_order.get(a['level'], 9))
    
    print(f"\n=== v51 Health Check: {len(alerts)} alert(s) ===")
    for a in alerts:
        print(f"  {a['icon']} {a['alert']}")
        print(f"      {a['message']}")
    print()


def _read_athlete_csv(path: str) -> pd.DataFrame:
    """Read athlete_data.csv, handling malformed comment headers.
    
    The BFW export can produce files where comment lines and the header row
    are concatenated onto one line (e.g. '# comment#date,weight_kg,...').
    Plain comment='#' would skip that entire line, losing the header.
    
    Fix: strip everything up to and including the last '#' on any line that
    starts with '#', keeping the real header if it's been appended.
    """
    import io
    with open(path, 'r') as f:
        raw_lines = f.readlines()
    
    clean_lines = []
    for line in raw_lines:
        stripped = line.lstrip()
        if stripped.startswith('#'):
            # BFW export bug: header sometimes concatenated after comment
            # e.g. '# comment#date,weight_kg,garmin_tr,non_running_tss'
            last_hash = stripped.rfind('#')
            remainder = stripped[last_hash + 1:].strip()
            # Only treat as data if it looks like a CSV header (contains commas)
            if remainder and ',' in remainder and not remainder.startswith('#'):
                clean_lines.append(remainder + '\n')
            # Otherwise skip the comment line entirely
        else:
            clean_lines.append(line)
    
    return pd.read_csv(io.StringIO(''.join(clean_lines)))


def calc_ctl_atl(df: pd.DataFrame, athlete_data_path: str = None) -> tuple:
    """
    Calculate CTL, ATL, TSB using exponential weighted moving average.
    
    CTL = CTL_yesterday + (TSS_today - CTL_yesterday) / 42
    ATL = ATL_yesterday + (TSS_today - ATL_yesterday) / 7
    TSB = CTL - ATL
    
    Note: This operates on DAILY data. Multiple activities on same day sum their TSS.
    
    v46: Reads non-running TSS from athlete_data.csv (column: non_running_tss)
    
    Returns:
        tuple: (df with CTL/ATL/TSB columns, daily_df with full calendar day time series)
    """
    print("  Calculating CTL/ATL/TSB...")
    
    # Get date range
    min_date = df['date'].min()
    max_date = df['date'].max()
    
    if pd.isna(min_date) or pd.isna(max_date):
        print("  WARNING: No valid dates for CTL/ATL calculation")
        df['CTL'] = np.nan
        df['ATL'] = np.nan
        df['TSB'] = np.nan
        return df, pd.DataFrame()
    
    # Create daily aggregation
    # Use string dates for reliable matching
    df['_date_str'] = df['date'].dt.strftime('%Y-%m-%d')
    
    # Aggregate multiple runs on same day
    daily_agg = df.groupby('_date_str').agg({
        'TSS': 'sum',
        'distance_km': 'sum',
        'RF_Trend': 'last',  # Take the latest RF_Trend for that day
        'RFL_Trend': 'last',
    }).reset_index()
    daily_agg.columns = ['date_str', 'tss_running', 'distance_km', 'RF_Trend', 'RFL_Trend']
    
    # v46: Load non-running TSS from athlete_data.csv
    non_running_tss = {}
    if athlete_data_path and os.path.exists(athlete_data_path):
        try:
            ad_df = _read_athlete_csv(athlete_data_path)
            if 'non_running_tss' in ad_df.columns:
                ad_df['date_str'] = pd.to_datetime(ad_df['date']).dt.strftime('%Y-%m-%d')
                nr_df = ad_df[ad_df['non_running_tss'].notna()]
                non_running_tss = nr_df.groupby('date_str')['non_running_tss'].sum().to_dict()
                print(f"  Loaded {len(non_running_tss)} days of non-running TSS from athlete_data.csv")
            elif 'tss' in ad_df.columns:
                # Fallback: legacy non_running_tss.csv format
                ad_df['date_str'] = pd.to_datetime(ad_df['date']).dt.strftime('%Y-%m-%d')
                non_running_tss = ad_df.groupby('date_str')['tss'].sum().to_dict()
                print(f"  Loaded {len(non_running_tss)} days of non-running TSS (legacy format)")
        except Exception as e:
            print(f"  WARNING: Could not load non-running TSS from {athlete_data_path}: {e}")
    
    # Create full date range as strings (extend to today + 30 days for CTL/ATL projection)
    from datetime import datetime
    today = pd.Timestamp(datetime.now().date())
    projection_days = 30
    end_date = max(max_date, today) + pd.Timedelta(days=projection_days)
    
    date_range = pd.date_range(start=min_date, end=end_date, freq='D')
    daily_df = pd.DataFrame({
        'date': date_range,
        'date_str': [d.strftime('%Y-%m-%d') for d in date_range]
    })
    daily_df = daily_df.merge(daily_agg, on='date_str', how='left')
    daily_df['tss_running'] = daily_df['tss_running'].fillna(0)
    daily_df['distance_km'] = daily_df['distance_km'].fillna(0)
    
    # Forward-fill RF_Trend, RFL_Trend (carry over from last run day)
    daily_df['RF_Trend'] = daily_df['RF_Trend'].ffill()
    daily_df['RFL_Trend'] = daily_df['RFL_Trend'].ffill()
    
    # Add non-running TSS
    daily_df['tss_other'] = daily_df['date_str'].map(non_running_tss).fillna(0)
    daily_df['tss_total'] = daily_df['tss_running'] + daily_df['tss_other']
    
    # Debug: show some stats
    total_tss = daily_df['tss_total'].sum()
    days_with_tss = (daily_df['tss_total'] > 0).sum()
    print(f"  Daily TSS: {days_with_tss} days with activity, total TSS={total_tss:.0f}")
    
    # Calculate CTL/ATL
    ctl_tc = RF_CONSTANTS['ctl_time_constant']
    atl_tc = RF_CONSTANTS['atl_time_constant']
    
    ctl = 0.0
    atl = 0.0
    
    ctl_values = []
    atl_values = []
    tsb_values = []
    
    for _, row in daily_df.iterrows():
        tss = row['tss_total']
        
        ctl = ctl + (tss - ctl) / ctl_tc
        atl = atl + (tss - atl) / atl_tc
        tsb = ctl - atl
        
        ctl_values.append(ctl)
        atl_values.append(atl)
        tsb_values.append(tsb)
    
    daily_df['CTL'] = ctl_values
    daily_df['ATL'] = atl_values
    daily_df['TSB'] = tsb_values
    
    # Create lookup dicts for mapping back to activity df
    ctl_by_date = dict(zip(daily_df['date_str'], daily_df['CTL']))
    atl_by_date = dict(zip(daily_df['date_str'], daily_df['ATL']))
    tsb_by_date = dict(zip(daily_df['date_str'], daily_df['TSB']))
    
    # Map back to activity DataFrame using string date keys
    df['CTL'] = df['_date_str'].map(ctl_by_date)
    df['ATL'] = df['_date_str'].map(atl_by_date)
    df['TSB'] = df['_date_str'].map(tsb_by_date)
    
    # Clean up temp column
    df.drop(columns=['_date_str'], inplace=True)
    
    # Clean up daily_df for output
    daily_df = daily_df[['date', 'distance_km', 'tss_running', 'tss_other', 'tss_total', 
                          'CTL', 'ATL', 'TSB', 'RF_Trend', 'RFL_Trend']]
    daily_df.columns = ['Date', 'Distance_km', 'TSS_Running', 'TSS_Other', 'TSS_Total',
                        'CTL', 'ATL', 'TSB', 'RF_Trend', 'RFL_Trend']
    
    latest_ctl = df['CTL'].iloc[-1] if len(df) > 0 else 0
    latest_atl = df['ATL'].iloc[-1] if len(df) > 0 else 0
    print(f"  Latest CTL={latest_ctl:.1f}, ATL={latest_atl:.1f}")
    print(f"  Daily summary: {len(daily_df)} days from {daily_df['Date'].min().strftime('%Y-%m-%d')} to {daily_df['Date'].max().strftime('%Y-%m-%d')}")
    
    return df, daily_df


def generate_weekly_summary(df: pd.DataFrame) -> pd.DataFrame:
    """
    Generate weekly summary from activity data.
    
    Returns DataFrame with: Week_Start, Runs, Distance_km, TSS, RF_Trend, RFL_Trend
    """
    df = df.copy()
    df['week_start'] = df['date'].dt.to_period('W-MON').dt.start_time
    
    weekly = df.groupby('week_start').agg({
        'date': 'count',  # number of runs
        'distance_km': 'sum',
        'TSS': 'sum',
        'RF_Trend': 'mean',  # average RF_Trend for the week
        'RFL_Trend': 'mean',  # average RFL_Trend for the week
    }).reset_index()
    
    weekly.columns = ['Week_Start', 'Runs', 'Distance_km', 'TSS', 'RF_Trend', 'RFL_Trend']
    weekly = weekly.sort_values('Week_Start')
    
    # Round numeric columns
    weekly['Distance_km'] = weekly['Distance_km'].round(1)
    weekly['TSS'] = weekly['TSS'].round(0)
    weekly['RF_Trend'] = weekly['RF_Trend'].round(4)
    weekly['RFL_Trend'] = weekly['RFL_Trend'].round(4)
    
    print(f"  Weekly summary: {len(weekly)} weeks")
    return weekly


def generate_monthly_summary(df: pd.DataFrame) -> pd.DataFrame:
    """
    Generate monthly summary from activity data.
    
    Returns DataFrame with: Month, Runs, Distance_km, TSS, RF_Trend, RFL_Trend
    """
    df = df.copy()
    df['month'] = df['date'].dt.to_period('M').dt.start_time
    
    monthly = df.groupby('month').agg({
        'date': 'count',  # number of runs
        'distance_km': 'sum',
        'TSS': 'sum',
        'RF_Trend': 'mean',  # average RF_Trend for the month
        'RFL_Trend': 'mean',  # average RFL_Trend for the month
    }).reset_index()
    
    monthly.columns = ['Month', 'Runs', 'Distance_km', 'TSS', 'RF_Trend', 'RFL_Trend']
    monthly = monthly.sort_values('Month')
    
    # Round numeric columns
    monthly['Distance_km'] = monthly['Distance_km'].round(1)
    monthly['TSS'] = monthly['TSS'].round(0)
    monthly['RF_Trend'] = monthly['RF_Trend'].round(4)
    monthly['RFL_Trend'] = monthly['RFL_Trend'].round(4)
    
    print(f"  Monthly summary: {len(monthly)} months")
    return monthly


def generate_yearly_summary(df: pd.DataFrame) -> pd.DataFrame:
    """
    Generate yearly summary from activity data.
    
    Returns DataFrame with: Year, Runs, Distance_km, TSS, RF_Trend, RFL_Trend
    """
    df = df.copy()
    df['year'] = df['date'].dt.year
    
    yearly = df.groupby('year').agg({
        'date': 'count',  # number of runs
        'distance_km': 'sum',
        'TSS': 'sum',
        'RF_Trend': 'mean',  # average RF_Trend for the year
        'RFL_Trend': 'mean',  # average RFL_Trend for the year
    }).reset_index()
    
    yearly.columns = ['Year', 'Runs', 'Distance_km', 'TSS', 'RF_Trend', 'RFL_Trend']
    yearly = yearly.sort_values('Year')
    
    # Round numeric columns
    yearly['Distance_km'] = yearly['Distance_km'].round(0)
    yearly['TSS'] = yearly['TSS'].round(0)
    yearly['RF_Trend'] = yearly['RF_Trend'].round(4)
    yearly['RFL_Trend'] = yearly['RFL_Trend'].round(4)
    
    print(f"  Yearly summary: {len(yearly)} years")
    return yearly


def detect_changed_rows(df_input: pd.DataFrame, df_prev: pd.DataFrame, 
                        overrides: pd.DataFrame) -> tuple:
    """
    Detect which rows need reprocessing.
    
    Returns:
        - set of row indices to reprocess
        - earliest date requiring rolling recalculation
    """
    if df_prev is None or len(df_prev) == 0:
        return set(df_input.index), df_input['date'].min()
    
    # Build lookup of previous output by file (lowercase for case-insensitive)
    prev_by_file = {str(row['file']).lower(): idx for idx, row in df_prev.iterrows() if pd.notna(row.get('file'))}
    
    # Build lookup of overrides by file (with hash for change detection)
    override_hashes = {}
    if overrides is not None and len(overrides) > 0:
        for _, row in overrides.iterrows():
            file_name = str(row.get('file', '')).lower()
            if file_name:
                # Hash key override fields
                hash_input = f"{row.get('race_flag','')}{row.get('parkrun','')}{row.get('official_distance_km','')}{row.get('surface','')}{row.get('surface_adj','')}{row.get('temp_override','')}"
                override_hashes[file_name] = hash(hash_input)
    
    # Build lookup of previous override state from prev output
    # We can detect changes by comparing prev values to current override
    prev_override_state = {}
    for idx, row in df_prev.iterrows():
        file_name = str(row.get('file', '')).lower()
        if file_name:
            hash_input = f"{row.get('race_flag','')}{row.get('parkrun','')}{row.get('official_distance_km','')}{row.get('surface','')}{row.get('surface_adj','')}{row.get('temp_override','')}"
            prev_override_state[file_name] = hash(hash_input)
    
    rows_to_process = set()
    earliest_change_date = None
    
    for idx, row in df_input.iterrows():
        file_name = str(row.get('file', '')).lower()
        
        # New row?
        if file_name not in prev_by_file:
            rows_to_process.add(idx)
            row_date = row.get('date')
            if row_date is not None and (earliest_change_date is None or row_date < earliest_change_date):
                earliest_change_date = row_date
            continue
        
        # Check if override changed for this file
        current_override_hash = override_hashes.get(file_name, hash(''))
        prev_override_hash = prev_override_state.get(file_name, hash(''))
        
        if current_override_hash != prev_override_hash:
            rows_to_process.add(idx)
            row_date = row.get('date')
            if row_date is not None and (earliest_change_date is None or row_date < earliest_change_date):
                earliest_change_date = row_date
            continue
    
    print(f"  Incremental: {len(rows_to_process)} rows to process, {len(df_input) - len(rows_to_process)} cached")
    
    return rows_to_process, earliest_change_date


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Step B v50: HR correction + RF metrics + adjustments + rolling metrics")
    p.add_argument("--master", required=True, help="Input master XLSX")
    p.add_argument("--persec-cache-dir", required=True, help="Directory with per-second .npz caches")
    p.add_argument("--out", required=True, help="Output master XLSX")
    p.add_argument("--model-json", required=True, help="RE model JSON (used when cache power is missing)")
    p.add_argument("--override-file", default="activity_overrides.xlsx", help="Activity overrides Excel file")
    p.add_argument("--era-adjusters", default="stryd_era_adjusters.csv", help="Era adjusters CSV (optional, uses defaults if missing)")
    p.add_argument("--athlete-data", default="athlete_data.csv", help="Athlete data CSV (weight, Garmin TR, non-running TSS)")
    p.add_argument("--strava", default="activities.csv", help="Strava activities.csv for refreshing activity names")
    p.add_argument("--mass-kg", type=float, default=76.0)
    p.add_argument("--tz", default="Europe/Stockholm")
    p.add_argument("--match-tol-s", type=int, default=900)
    p.add_argument("--progress-every", type=int, default=50, help="Print progress every N processed rows (0 disables)")
    p.add_argument("--progress-per-run", action="store_true", help="Print one progress line per activity")
    p.add_argument("--incremental", action="store_true", help="Incremental mode: only process new/changed rows")
    p.add_argument("--prev-output", default=None, help="Previous output file for incremental comparison (auto-detected if not set)")
    p.add_argument("--runner-age", type=int, default=55, help="Runner's current age for age grade calculations")
    p.add_argument("--runner-gender", default="male", choices=["male", "female"], help="Runner's gender for age grade")
    p.add_argument("--runner-dob", default=None, help="Runner's DOB (YYYY-MM-DD) for precise age calculations")
    return p.parse_args()


def _hash_file(path: str) -> str:
    with open(path, "rb") as f:
        b = f.read()
    return hashlib.sha1(b).hexdigest()[:8]


def _ensure_1d(a, dtype=float) -> np.ndarray:
    return np.asarray(a, dtype=dtype).reshape(-1)


def _roll_med(a: np.ndarray, win: int) -> np.ndarray:
    a = np.asarray(a, float)
    try:
        s = pd.Series(a)
        return s.rolling(int(win), center=True, min_periods=1).median().to_numpy(dtype=float)
    except Exception:
        return a.astype(float, copy=False)


def _build_cache_index(cache_dir: str) -> Tuple[List[int], List[str]]:
    """Build sorted (keys, paths) for bisect cache lookup.
    v50: Uses _cache_index.json for fast startup instead of opening all .npz files."""
    try:
        from rebuild_from_fit_zip import build_cache_index_fast
        return build_cache_index_fast(cache_dir)
    except ImportError:
        pass
    # Fallback: scan .npz files directly (slow but always works)
    items: List[Tuple[int, str]] = []
    for fn in os.listdir(cache_dir):
        if not fn.lower().endswith(".npz"):
            continue
        fp = os.path.join(cache_dir, fn)
        try:
            z = np.load(fp, allow_pickle=True)
            ts = z.get("ts", None)
            if ts is None:
                continue
            ts0 = float(_ensure_1d(ts)[0])
            items.append((int(round(ts0)), fp))
        except Exception:
            continue
    items.sort(key=lambda x: x[0])
    return [k for k, _ in items], [p for _, p in items]


def _find_cache(keys: List[int], paths: List[str], epoch_s: float, tol_s: int) -> Optional[str]:
    if not keys:
        return None
    k = int(round(float(epoch_s)))
    j = bisect.bisect_left(keys, k)
    best = None
    best_dt = None
    for idx in (j - 1, j, j + 1):
        if 0 <= idx < len(keys):
            dt = abs(keys[idx] - k)
            if dt <= tol_s and (best_dt is None or dt < best_dt):
                best_dt = dt
                best = paths[idx]
    return best


def _calc_np(power_w: np.ndarray, win: int = 30) -> float:
    p = np.asarray(power_w, dtype=float).reshape(-1)
    if p.size == 0:
        return float("nan")
    p = np.where(np.isfinite(p), p, np.nan)
    if np.isnan(p).all():
        return float("nan")
    fill = float(np.nanmedian(p))
    if not np.isfinite(fill):
        fill = 0.0
    p = np.nan_to_num(p, nan=fill, posinf=fill, neginf=fill)
    if len(p) < win:
        return float(np.mean(p))
    c = np.cumsum(np.insert(p, 0, 0.0))
    rm = (c[win:] - c[:-win]) / win
    return float(np.mean(rm ** 4) ** 0.25)


def _active_mask_for_re(t: np.ndarray, v: np.ndarray, p: np.ndarray, mass_kg: float) -> np.ndarray:
    """Active-running mask for RE."""
    t = np.asarray(t, float).reshape(-1)
    v = np.asarray(v, float).reshape(-1)
    p = np.asarray(p, float).reshape(-1)
    wkg = p / max(float(mass_kg), 1e-9)
    active = np.isfinite(p) & (p > 0) & np.isfinite(wkg) & (wkg >= 1.0)
    v_cov = float(np.mean(np.isfinite(v))) if v.size else 0.0
    if v_cov >= 0.2:
        active &= np.isfinite(v) & (v >= 0.3) & (v <= 6.5)
    return active


def _calc_re_active(t: np.ndarray, v: np.ndarray, p: np.ndarray, mass_kg: float, row: pd.Series, dist_m_arr: np.ndarray = None) -> tuple[float, float, float, float]:
    """Return (re_avg, re_norm, speed_active_mps, active_time_s).
    
    v44.2: Calculate RE using consistent active periods for both distance and power.
    Uses distance array (cumulative) to get active-only distance, avoiding speed field issues.
    Falls back to overall distance/time if distance array unavailable.
    """
    active = _active_mask_for_re(t, v, p, mass_kg)
    active_n = int(active.sum())
    if active_n < 120:
        return (float('nan'), float('nan'), float('nan'), float(active_n))

    t = np.asarray(t, float).reshape(-1)
    dt = np.diff(t, prepend=t[0])
    dt = np.where(np.isfinite(dt), dt, 1.0)
    # v44.2: Widen clipping to handle sparse data (e.g., 3-4 second intervals from Garmin)
    # Old: 0.5-1.5s assumed 1-second data
    # New: 0.5-10s allows for sparse recordings while still capping crazy gaps
    dt = np.clip(dt, 0.5, 10.0)

    p = np.asarray(p, float).reshape(-1)
    
    # Active time from the mask (same period used for power averaging)
    active_time_s = float(np.nansum(dt[active]))
    
    # Try to get distance from distance array (cumulative) - more robust than speed
    if dist_m_arr is not None and len(dist_m_arr) == len(t):
        d = np.asarray(dist_m_arr, float).reshape(-1)
        # Calculate distance covered during active seconds
        # Use diff of cumulative distance, masked by active
        d_diff = np.diff(d, prepend=0)
        d_diff = np.where(np.isfinite(d_diff) & (d_diff >= 0) & (d_diff < 50), d_diff, 0)  # Cap at 50m/s sanity check
        dist_m = float(np.sum(d_diff[active]))
    else:
        # Fallback: use overall distance and active time ratio
        # This assumes distance is proportional to active time (reasonable for steady runs)
        dist_km = pd.to_numeric(row.get('distance_km', np.nan), errors='coerce')
        moving_time_s = pd.to_numeric(row.get('moving_time_s', np.nan), errors='coerce')
        
        if np.isfinite(dist_km) and np.isfinite(moving_time_s) and moving_time_s > 0:
            # Scale total distance by ratio of active time to moving time
            dist_m = float(dist_km * 1000.0) * (active_time_s / moving_time_s)
        elif np.isfinite(dist_km):
            dist_m = float(dist_km * 1000.0)
        else:
            dist_m = float('nan')
    
    if not (np.isfinite(dist_m) and np.isfinite(active_time_s) and active_time_s > 0):
        return (float('nan'), float('nan'), float('nan'), float(active_time_s))

    speed_mps = dist_m / active_time_s
    p_active = p[active]
    p_active = p_active[np.isfinite(p_active) & (p_active > 0)]
    if p_active.size < 120:
        return (float('nan'), float('nan'), float(speed_mps), float(active_time_s))

    p_mean = float(np.nanmean(p_active))
    p_np = _calc_np(p_active, win=30)
    if not (np.isfinite(p_mean) and p_mean > 0 and np.isfinite(p_np) and p_np > 0):
        return (float('nan'), float('nan'), float(speed_mps), float(active_time_s))

    re_avg = (speed_mps * float(mass_kg)) / p_mean
    re_norm = (speed_mps * float(mass_kg)) / p_np
    return (float(re_avg), float(re_norm), float(speed_mps), float(active_time_s))


def _rf_metrics(t: np.ndarray, v: np.ndarray, p: np.ndarray, hr: np.ndarray, mass_kg: float) -> Dict[str, float | str | bool]:
    out: Dict[str, float | str | bool] = {
        "RF_window_start_s": np.nan,
        "RF_window_end_s": np.nan,
        "RF_window_shifted": False,
        "RF_select_code": "0.0",
        "RF_dead_frac": np.nan,
        "RF_adjusted_mean_W_per_bpm": np.nan,
        "RF_adjusted_median_W_per_bpm": np.nan,
        "RF_drift_pct_per_min": np.nan,
        "RF_drift_r2": np.nan,
        "RF_r2": np.nan,
    }

    t = np.asarray(t, float).reshape(-1)
    v = np.asarray(v, float).reshape(-1)
    p = np.asarray(p, float).reshape(-1)
    hr = np.asarray(hr, float).reshape(-1)
    n = len(t)
    if n < 200:
        return out

    p_thr = 2.5 * float(mass_kg)
    dead = (~np.isfinite(p)) | (p < p_thr) | (~np.isfinite(v)) | (v < 0.3)
    running = (~dead) & np.isfinite(v) & (v <= 6.5) & np.isfinite(p) & (p > 0)
    if running.sum() < 120:
        return out

    tmax = float(np.nanmax(t))
    
    # RF window parameters
    RF_WINDOW_DURATION = RF_CONSTANTS['rf_window_duration_s']
    PAUSE_THRESHOLD = RF_CONSTANTS['rf_window_pause_threshold_s']
    
    # Scaled warmup skip based on run duration:
    # 0 min -> 120s, 15 min -> 300s, 30+ min -> 600s (linear interpolation)
    if tmax <= 900.0:  # <= 15 min
        skip_s = 120.0 + (tmax / 900.0) * 180.0
    elif tmax <= 1800.0:  # 15-30 min
        skip_s = 300.0 + ((tmax - 900.0) / 900.0) * 300.0
    else:  # 30+ min
        skip_s = 600.0
    
    start_s = skip_s
    end_s = min(start_s + RF_WINDOW_DURATION, tmax)
    
    # Detect pauses for "Swiss cheese" exclusion - two types:
    # 1. Time gaps (watch paused) - gaps in time array > threshold
    # 2. Standing rests (continuous recording) - periods where dead=True for > threshold
    
    pause_periods = []  # List of (pause_start, pause_end, duration) tuples
    
    # Type 1: Time gaps
    dt = np.diff(t)
    time_gap_indices = np.where(dt > PAUSE_THRESHOLD)[0]
    for idx in time_gap_indices:
        gap_start = float(t[idx])
        gap_end = float(t[idx + 1])
        gap_duration = gap_end - gap_start
        pause_periods.append((gap_start, gap_end, gap_duration))
    
    # Type 2: Standing rests (consecutive dead samples)
    if dead.sum() > 0:
        dead_int = dead.astype(int)
        dead_diff = np.diff(dead_int, prepend=0, append=0)
        dead_starts = np.where(dead_diff == 1)[0]
        dead_ends = np.where(dead_diff == -1)[0]
        
        for ds, de in zip(dead_starts, dead_ends):
            if de <= len(t) and ds < len(t):
                dead_start_t = float(t[ds])
                dead_end_t = float(t[min(de, len(t)-1)])
                dead_duration = dead_end_t - dead_start_t
                if dead_duration > PAUSE_THRESHOLD:
                    pause_periods.append((dead_start_t, dead_end_t, dead_duration))
    
    # Sort pause periods by start time
    pause_periods.sort(key=lambda x: x[0])
    
    # Build exclusion mask for pauses within our window
    # Exclude: pause duration + min(pause_duration, 60s) settling time
    exclude_mask = np.zeros(n, dtype=bool)
    
    for pause_start, pause_end, pause_duration in pause_periods:
        # Only process pauses that overlap with our window
        if pause_end < start_s or pause_start > end_s:
            continue
        
        # Settling time after pause: min(pause_duration, 60s) + HR_LAG_S
        settling_time = min(pause_duration, 60.0) + HR_LAG_S
        exclude_end = pause_end + settling_time
        
        # Mark samples to exclude
        exclude_mask |= (t >= pause_start) & (t <= exclude_end)
        out["RF_window_shifted"] = True  # Flag that we had to exclude pauses
    
    # Final window mask: within time bounds, running, and not excluded
    win = running & (t >= start_s) & (t <= end_s) & (~exclude_mask)
    
    if win.sum() < 120:
        # v47: Try shifting the RF window to start after the largest pause
        # A watch pause is like a restart — treat post-pause as a fresh warmup
        if pause_periods:
            # Find the largest pause that overlaps with the original window
            window_pauses = [(ps, pe, pd) for ps, pe, pd in pause_periods 
                             if pe > start_s and ps < end_s]
            if window_pauses:
                biggest = max(window_pauses, key=lambda x: x[2])
                settling_time = min(biggest[2], 60.0) + HR_LAG_S
                new_start = biggest[1] + settling_time  # After pause + settling
                new_end = min(new_start + RF_WINDOW_DURATION, tmax)
                win_shifted = running & (t >= new_start) & (t <= new_end) & (~exclude_mask)
                if win_shifted.sum() >= 120:
                    win = win_shifted
                    out["RF_window_shifted"] = True
    
    if win.sum() < 120:
        # Fall back: try without exclusions
        win = running & (t >= start_s) & (t <= end_s)
        if win.sum() < 120:
            # Fall back to full run minus warmup
            win = running & (t >= 120.0)
            out["RF_window_shifted"] = True
    if win.sum() < 120:
        return out

    # Calculate actual window bounds from the mask
    win_times = t[win]
    out["RF_window_start_s"] = float(np.min(win_times))
    out["RF_window_end_s"] = float(np.max(win_times))
    out["RF_dead_frac"] = float(np.mean(dead))

    hr_win = win & np.isfinite(hr) & (hr > 30)
    if hr_win.sum() < 120:
        return out

    # Smooth power and apply HR lag (HR responds to power from HR_LAG_S seconds ago)
    # Calculate actual sample rate from time array (handles old Garmin ~0.33Hz and modern 1Hz)
    dt_median = float(np.nanmedian(np.diff(t))) if n > 1 else 1.0
    dt_median = max(0.5, min(dt_median, 5.0))  # Clamp to reasonable range
    
    # Convert time-based constants to sample counts
    smooth_samples = max(3, int(round(POWER_SMOOTH_S / dt_median)))
    lag_samples = max(1, int(round(HR_LAG_S / dt_median)))
    
    kernel = np.ones(smooth_samples, dtype=float) / smooth_samples
    p_smooth = np.convolve(np.nan_to_num(p, nan=0.0), kernel, mode="same")
    
    # Shift power forward by lag_samples to align with HR response
    # (i.e., current HR corresponds to power from HR_LAG_S seconds ago)
    p_lagged = np.full_like(p_smooth, np.nan)
    if lag_samples < n:
        p_lagged[lag_samples:] = p_smooth[:-lag_samples]

    ratio = p_lagged[hr_win] / hr[hr_win]
    ratio = ratio[np.isfinite(ratio)]
    if ratio.size < 60:
        return out

    out["RF_adjusted_mean_W_per_bpm"] = float(np.mean(ratio))
    out["RF_adjusted_median_W_per_bpm"] = float(np.median(ratio))
    out["RF_select_code"] = "1.0"

    tt = t[hr_win][: ratio.size].astype(float)
    yy = ratio.astype(float)
    x = tt - float(tt.min())
    if yy.size >= 60 and np.isfinite(x).all():
        b, a = np.polyfit(x, yy, 1)
        yhat = a + b * x
        ss_res = float(np.sum((yy - yhat) ** 2))
        ss_tot = float(np.sum((yy - float(np.mean(yy))) ** 2))
        r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else np.nan
        drift_pct_per_min = (b / float(np.mean(yy))) * 60.0 * 100.0
        out["RF_drift_pct_per_min"] = float(drift_pct_per_min) if np.isfinite(drift_pct_per_min) else np.nan
        out["RF_drift_r2"] = float(r2) if np.isfinite(r2) else np.nan
        out["RF_r2"] = out["RF_drift_r2"]

    return out


def _calc_full_run_drift(t: np.ndarray, p: np.ndarray, hr: np.ndarray,
                          mass_kg: float, moving_time_s: float) -> Dict[str, float]:
    """
    Calculate P/HR drift over the full run (post-warmup, pre-cooldown).
    
    Unlike RF_drift which covers a 20-minute window, this measures drift across
    the entire run to capture cardiac drift in long efforts. Only meaningful for
    runs > 60 minutes where there's substantial data beyond the RF window.
    
    Cooldown detection: finds where 2-min rolling power drops below 65% of
    steady-state and doesn't recover for 3+ minutes.
    
    Power gating: excludes 1-min blocks where avg power < 65% of steady-state
    (catches walk breaks, water stops, traffic lights).
    
    Args:
        t: time array (seconds)
        p: power array (watts)
        hr: heart rate array (bpm)
        mass_kg: athlete mass in kg
        moving_time_s: total moving time in seconds
        
    Returns:
        Dict with full_run_drift_pct_per_min, full_run_drift_r2,
        full_run_drift_duration_min (all NaN if insufficient data)
    """
    result = {
        "full_run_drift_pct_per_min": np.nan,
        "full_run_drift_r2": np.nan,
        "full_run_drift_duration_min": np.nan,
    }
    
    # Only for runs > 60 minutes
    if moving_time_s < 3600:
        return result
    
    t = np.asarray(t, float).reshape(-1)
    p = np.asarray(p, float).reshape(-1)
    hr = np.asarray(hr, float).reshape(-1)
    
    valid = np.isfinite(p) & np.isfinite(hr) & (hr > 50) & (p > 2.5 * mass_kg)
    if valid.sum() < 300:
        return result
    
    warmup_skip_s = 600.0  # Skip first 10 min
    
    # Establish steady-state power from minutes 10-30
    ss_mask = (t >= warmup_skip_s) & (t < warmup_skip_s + 1200) & valid
    if ss_mask.sum() < 120:
        return result
    ss_power = float(np.nanmean(p[ss_mask]))
    if ss_power <= 0:
        return result
    
    power_gate_pct = 0.65
    power_threshold = ss_power * power_gate_pct
    
    # Cooldown detection: find where 2-min rolling avg drops below threshold
    # and doesn't recover for 3+ minutes
    cooldown_start = None
    t_end = float(t[-1])
    
    for check_s in range(int(warmup_skip_s) + 1200, int(t_end) - 180, 60):
        blk = (t >= check_s) & (t < check_s + 120) & valid
        if blk.sum() < 30:
            continue
        blk_power = float(np.nanmean(p[blk]))
        
        if blk_power < power_threshold:
            # Check recovery in next 3 minutes
            rec = (t >= check_s + 120) & (t < check_s + 300) & valid
            if rec.sum() > 30:
                rec_power = float(np.nanmean(p[rec]))
                if rec_power >= power_threshold:
                    continue  # Recovered — water stop, traffic light
            cooldown_start = float(check_s)
            break
    
    end_s = cooldown_start if cooldown_start else t_end
    
    # Power gating: include 1-min blocks where avg power >= threshold
    include_mask = np.zeros(len(t), dtype=bool)
    for blk_start in range(int(warmup_skip_s), int(end_s), 60):
        blk_end = min(blk_start + 60, end_s)
        blk = (t >= blk_start) & (t < blk_end) & valid
        if blk.sum() < 15:
            continue
        if float(np.nanmean(p[blk])) >= power_threshold:
            include_mask |= blk
    
    if include_mask.sum() < 300:
        return result
    
    # Linear regression of P/HR over included period
    ratio = p[include_mask] / hr[include_mask]
    tt = t[include_mask]
    x = (tt - float(tt.min())).astype(float)
    y = ratio.astype(float)
    
    if len(x) < 300 or np.std(x) == 0:
        return result
    
    b, a = np.polyfit(x, y, 1)
    yhat = a + b * x
    ss_res = float(np.sum((y - yhat) ** 2))
    ss_tot = float(np.sum((y - float(np.mean(y))) ** 2))
    r2 = 1.0 - ss_res / ss_tot if ss_tot > 0 else 0.0
    drift_pct_per_min = (b / float(np.mean(y))) * 60.0 * 100.0
    
    included_duration = (float(tt.max()) - float(tt.min())) / 60.0
    
    if np.isfinite(drift_pct_per_min) and np.isfinite(r2) and included_duration > 30:
        result["full_run_drift_pct_per_min"] = drift_pct_per_min
        result["full_run_drift_r2"] = max(0.0, r2)
        result["full_run_drift_duration_min"] = included_duration
    
    return result


def _calc_hr_cv(t: np.ndarray, p: np.ndarray, hr: np.ndarray, mass_kg: float) -> tuple[float, float]:
    """Calculate coefficient of variation of power:HR ratio.
    
    High CV (>threshold) indicates unreliable HR data (multiple dropouts, sensor issues).
    Threshold is adjusted based on power variability - interval sessions with high power CV
    get a higher threshold since HR:power variability is expected.
    
    Returns (hr_cv, power_cv) as percentages, or (NaN, NaN) if insufficient data.
    
    Note: 
    - Uses RF window: starts at 600s for runs >= 15min (skips HR ramp-up)
    - Excludes samples near pauses (gaps > 5s) where HR lag creates natural variability.
    - Returns NaN for runs < 8 minutes (HR ramp-up dominates, not useful for reliability check).
    - Returns NaN for sparse data (< 0.5 Hz) - old Garmin smart recording mode
    - Uses RAW power CV (not smoothed) to detect interval structure
    """
    t = np.asarray(t, float).reshape(-1)
    p = np.asarray(p, float).reshape(-1)
    hr = np.asarray(hr, float).reshape(-1)
    n = len(t)
    
    if n < 120:
        return (float('nan'), float('nan'))
    
    # Skip short runs - HR ramp-up dominates and inflates CV
    duration_s = float(t[-1] - t[0]) if n > 0 else 0.0
    if duration_s < 8 * 60:  # < 8 minutes
        return (float('nan'), float('nan'))
    
    # Use RF window logic: start at 600s if run >= 15min (skips HR ramp-up)
    rf_start_s = 600.0 if duration_s >= 900.0 else 0.0
    start_idx = int(np.searchsorted(t, rf_start_s))
    
    # End at 90% to exclude cooldown (for longer runs)
    end_idx = int(n * 0.9) if duration_s >= 900.0 else n
    
    # Check if we have enough samples in RF window
    # Old Garmin smart recording (~0.2 Hz) produces too few samples for reliable CV
    # But if we have 200+ samples, we can assess regardless of rate
    rf_window_samples = end_idx - start_idx
    if rf_window_samples < 200:
        sample_rate = n / duration_s if duration_s > 0 else 0.0
        if sample_rate < 0.5:  # Sparse recording AND few samples - skip
            return (float('nan'), float('nan'))
    
    # Cap power to filter GPS spikes (24 m/s GPS glitch -> ~6000W simulated power)
    # Max reasonable running power ~600W (elite sprinter short burst)
    p_capped = np.clip(p, 0, 600)
    
    # Calculate actual sample rate from time array (handles old Garmin ~0.33Hz and modern 1Hz)
    dt_median = float(np.nanmedian(np.diff(t))) if n > 1 else 1.0
    dt_median = max(0.5, min(dt_median, 5.0))  # Clamp to reasonable range
    
    # Convert time-based constants to sample counts
    smooth_samples = max(3, int(round(POWER_SMOOTH_S / dt_median)))
    lag_samples = max(1, int(round(HR_LAG_S / dt_median)))
    
    # Smooth power with lag for HR response (for ratio calculation)
    p_smooth = _roll_med(p_capped, smooth_samples)
    p_lagged = np.full_like(p_smooth, np.nan)
    if lag_samples < n:
        p_lagged[lag_samples:] = p_smooth[:-lag_samples]
    
    # Running mask
    p_thr = 2.5 * float(mass_kg)
    mask = (np.arange(n) >= start_idx) & (np.arange(n) <= end_idx)
    mask &= np.isfinite(p_lagged) & (p_lagged > p_thr)
    mask &= np.isfinite(hr) & (hr > 30)
    
    # Exclude samples near pauses (gaps significantly longer than normal sample rate)
    # HR naturally lags around stops/starts, which inflates CV
    dt = np.diff(t)
    median_dt = float(np.nanmedian(dt)) if len(dt) > 0 else 1.0
    # A "pause" is a gap > 3x the normal sample interval, minimum 10s
    pause_threshold = max(10.0, median_dt * 3.0)
    gap_idx = np.where(np.isfinite(dt) & (dt > pause_threshold))[0]
    if len(gap_idx) > 0:
        near_pause = np.zeros(n, dtype=bool)
        for g in gap_idx:
            pause_start = max(0, g - 60)
            pause_end = min(n, g + 61)
            near_pause[pause_start:pause_end] = True
        mask_with_exclusion = mask & ~near_pause
        # Only apply exclusion if we still have enough samples
        # Otherwise fall back to no exclusion (stop-start runs with many pauses)
        if mask_with_exclusion.sum() >= 60:
            mask = mask_with_exclusion
    
    if mask.sum() < 60:
        return (float('nan'), float('nan'))
    
    # Power CV: use RAW power in RF window (not smoothed/thresholded)
    # This better captures interval structure (jog recoveries stay above threshold but raw CV is higher)
    # Use capped power to avoid GPS spike influence
    p_raw_window = p_capped[start_idx:end_idx]
    p_raw_valid = p_raw_window[np.isfinite(p_raw_window) & (p_raw_window > 0)]
    if p_raw_valid.size < 60:
        return (float('nan'), float('nan'))
    power_cv = float(np.std(p_raw_valid) / np.mean(p_raw_valid) * 100.0)
    
    # HR:Power ratio CV (uses smoothed/lagged power for fair comparison)
    ratio = p_lagged[mask] / hr[mask]
    ratio = ratio[np.isfinite(ratio)]
    
    if ratio.size < 60:
        return (float('nan'), power_cv)
    
    hr_cv = float(np.std(ratio) / np.mean(ratio) * 100.0)
    return (hr_cv, power_cv)


def _calc_terrain_metrics(t: np.ndarray, g: np.ndarray, dist_km: float, 
                          elev_gain_m: float = 0.0, elev: np.ndarray = None) -> Dict[str, float]:
    """Calculate terrain undulation metrics from per-second grade and elevation data.
    
    Designed to detect "Lidingöloppet-style" courses with constant rolling undulation.
    
    v45: Changed reversal counting from grade zero-crossings to elevation-based.
    A reversal now requires a minimum elevation swing of ±2.5m before changing direction.
    This filters out GPS/barometric noise and properly distinguishes:
    - LL30-style (constant undulation): ~3 rev/km
    - Haga-style (sustained climbs/descents): ~1 rev/km
    
    Returns:
        grade_std_pct: Standard deviation of grade
        elev_gain_per_km: Elevation gain per km
        reversals_per_km: Significant direction changes per km (±2.5m threshold)
        undulation_score: Combined metric - higher = more LL30-style
    """
    t = np.asarray(t, float).reshape(-1)
    g = np.asarray(g, float).reshape(-1)
    
    out: Dict[str, float] = {
        "grade_std_pct": float('nan'),
        "elev_gain_per_km": float('nan'),
        "reversals_per_km": float('nan'),
        "undulation_score": float('nan'),
    }
    
    if len(g) < 60 or dist_km < 0.5:
        return out
    
    # Compute grade std on valid samples
    g_valid = g[np.isfinite(g)]
    if len(g_valid) < 60:
        return out
    
    grade_std = float(np.std(g_valid) * 100.0)  # Convert to %
    out["grade_std_pct"] = grade_std
    
    # Elevation gain per km
    elev_gain_per_km = float(elev_gain_m / dist_km) if dist_km > 0 else 0.0
    out["elev_gain_per_km"] = elev_gain_per_km
    
    # v45: Count significant reversals using elevation data
    # A reversal requires minimum ±2.5m elevation change before changing direction
    # This filters GPS/barometric noise and properly distinguishes course types
    MIN_SWING_M = 2.5
    
    reversals_per_km = 0.0
    if elev is not None and len(elev) >= 60:
        elev = np.asarray(elev, float).reshape(-1)
        elev_valid = elev[np.isfinite(elev)]
        
        if len(elev_valid) >= 60:
            n_reversals = 0
            direction = 0  # 0=unknown, 1=climbing, -1=descending
            last_extreme = elev_valid[0]
            
            for e in elev_valid[1:]:
                if direction == 0:
                    # Establish initial direction
                    if e - last_extreme >= MIN_SWING_M:
                        direction = 1  # climbing
                        last_extreme = e
                    elif last_extreme - e >= MIN_SWING_M:
                        direction = -1  # descending
                        last_extreme = e
                elif direction == 1:  # Currently climbing
                    if e > last_extreme:
                        last_extreme = e  # New high
                    elif last_extreme - e >= MIN_SWING_M:
                        # Significant reversal downward
                        n_reversals += 1
                        direction = -1
                        last_extreme = e
                else:  # direction == -1, currently descending
                    if e < last_extreme:
                        last_extreme = e  # New low
                    elif e - last_extreme >= MIN_SWING_M:
                        # Significant reversal upward
                        n_reversals += 1
                        direction = 1
                        last_extreme = e
            
            reversals_per_km = float(n_reversals / dist_km) if dist_km > 0 else 0.0
    else:
        # Fallback to grade-based method if no elevation data
        g_smooth = pd.Series(g).rolling(window=30, center=True, min_periods=1).mean().values
        g_sign = np.sign(g_smooth)
        g_sign_nonzero = g_sign[g_sign != 0]
        
        if len(g_sign_nonzero) > 1:
            sign_changes = np.diff(g_sign_nonzero)
            n_reversals = int(np.sum(sign_changes != 0))
            reversals_per_km = float(n_reversals / dist_km) if dist_km > 0 else 0.0
    
    out["reversals_per_km"] = reversals_per_km
    
    # Undulation score formula:
    # Combines grade variability with reversal frequency
    # 
    # v45 updated expected values with new reversal counting:
    # LL30-style: grade_std ~6%, reversals ~3/km, gain ~14m/km
    # Haga-style: grade_std ~4%, reversals ~1/km, gain ~11m/km
    # Flat road:  grade_std ~2%, reversals ~0.5/km, gain ~5m/km
    #
    # Score = grade_std × sqrt(reversals_per_km) × gain_factor
    
    if elev_gain_per_km < 8:  # Minimum gain to count as hilly
        out["undulation_score"] = 0.0
        return out
    
    # Cap grade_std at 7% for scoring (higher suggests noise or sustained steep)
    grade_std_capped = min(grade_std, 7.0)
    
    # Reversal factor: sqrt to dampen extreme values
    # v45: recalibrated for new reversal counting
    # 1/km (Haga) -> 1.0, 3/km (LL30) -> 1.7, 5/km -> 2.2
    reversal_factor = np.sqrt(max(reversals_per_km, 0.5))
    
    # Gain factor: moderate gain is good, very high gain suggests sustained climbs
    # Peak around 15-25 m/km, then decreases for very steep terrain
    if elev_gain_per_km <= 25:
        gain_factor = elev_gain_per_km / 15.0  # Increases up to 25m/km
    else:
        gain_factor = 25.0 / 15.0 * (25.0 / elev_gain_per_km)  # Decreases above 25m/km
    
    gain_factor = min(gain_factor, 2.0)  # Cap at 2x
    
    undulation_score = grade_std_capped * reversal_factor * gain_factor
    out["undulation_score"] = undulation_score
    
    return out


def _calc_hill_power_metrics(g: np.ndarray, p: np.ndarray, grade_threshold: float = 0.01) -> Dict[str, float]:
    """Calculate uphill vs downhill power metrics to distinguish hill intervals from rolling terrain.
    
    Hill intervals: Short hard bursts uphill, recovery jogs downhill
    - High uphill/downhill power ratio (>1.8)
    - High power variance
    
    Rolling terrain (LL30-style): Constant moderate effort throughout
    - Lower uphill/downhill power ratio (~1.2-1.5)
    - More consistent power
    
    Args:
        g: Per-second grade array (as fraction, not %)
        p: Per-second power array (watts)
        grade_threshold: Minimum grade to count as uphill/downhill (default 1%)
    
    Returns:
        uphill_power_mean: Average power when grade > threshold
        downhill_power_mean: Average power when grade < -threshold
        uphill_downhill_ratio: Ratio of uphill to downhill power
        uphill_pct: Percentage of time spent going uphill
        flat_power_mean: Average power when -threshold < grade < threshold
    """
    g = np.asarray(g, float).reshape(-1)
    p = np.asarray(p, float).reshape(-1)
    
    out: Dict[str, float] = {
        "uphill_power_mean": float('nan'),
        "downhill_power_mean": float('nan'),
        "uphill_downhill_ratio": float('nan'),
        "uphill_pct": float('nan'),
        "flat_power_mean": float('nan'),
    }
    
    n = min(len(g), len(p))
    if n < 60:
        return out
    
    g, p = g[:n], p[:n]
    
    # Valid samples where both grade and power are finite and power > 0
    valid = np.isfinite(g) & np.isfinite(p) & (p > 0)
    if valid.sum() < 60:
        return out
    
    g_valid = g[valid]
    p_valid = p[valid]
    
    # Classify by grade
    uphill = g_valid > grade_threshold
    downhill = g_valid < -grade_threshold
    flat = ~uphill & ~downhill
    
    # Calculate means for each category
    if uphill.sum() >= 10:
        out["uphill_power_mean"] = float(np.mean(p_valid[uphill]))
        out["uphill_pct"] = float(uphill.sum() / len(g_valid) * 100)
    
    if downhill.sum() >= 10:
        out["downhill_power_mean"] = float(np.mean(p_valid[downhill]))
    
    if flat.sum() >= 10:
        out["flat_power_mean"] = float(np.mean(p_valid[flat]))
    
    # Calculate ratio
    if np.isfinite(out["uphill_power_mean"]) and np.isfinite(out["downhill_power_mean"]) and out["downhill_power_mean"] > 0:
        out["uphill_downhill_ratio"] = out["uphill_power_mean"] / out["downhill_power_mean"]
    
    return out


def correct_hr_stepb(
    t: np.ndarray,
    power_w: np.ndarray,
    hr_bpm: np.ndarray,
    speed_mps: np.ndarray,
    mass_kg: float,
    hard_floor_s: float = None,
    hr_floor: float = None,
) -> tuple[np.ndarray, bool, float | None, str | None]:
    """Step B HR correction using a rolling-median (15s) power:HR ratio back-projection method."""
    if hard_floor_s is None:
        hard_floor_s = RF_CONSTANTS['hr_correction_hard_floor_s']
    if hr_floor is None:
        hr_floor = RF_CONSTANTS['hr_correction_floor_bpm']
    t = np.asarray(t, float).reshape(-1)
    p = np.asarray(power_w, float).reshape(-1)
    h = np.asarray(hr_bpm, float).reshape(-1)
    v = np.asarray(speed_mps, float).reshape(-1)

    n = len(t)
    if n < 600:
        return h, False, None, None

    moving = np.isfinite(v) & (v >= 0.3)
    ok = moving & np.isfinite(p) & (p > 0) & np.isfinite(h) & (h > 30)
    if ok.mean() < 0.60:
        return h, False, None, None

    dt = np.nanmedian(np.diff(t))
    if not np.isfinite(dt) or dt <= 0:
        dt = 1.0

    dt_all = np.diff(t)
    gap_idx = np.where(np.isfinite(dt_all) & (dt_all > 2.0))[0] + 1
    bad_near_gap = np.zeros(n, dtype=bool)
    if gap_idx.size:
        for g in gap_idx.astype(int, copy=False):
            a = max(0, g - int(round(30.0 / dt)))
            b = min(n, g + int(round(120.0 / dt)))
            bad_near_gap[a:b] = True

    w = int(round(POWER_SMOOTH_S / dt))  # Use POWER_SMOOTH_S for smoothing window
    w = max(5, min(w, 61))
    def roll_med(x: np.ndarray) -> np.ndarray:
        out = np.full_like(x, np.nan, dtype=float)
        half = w // 2
        for i in range(n):
            a = max(0, i - half)
            b = min(n, i + half + 1)
            xx = x[a:b]
            xx = xx[np.isfinite(xx)]
            if xx.size:
                out[i] = float(np.median(xx))
        return out

    p_smooth = roll_med(np.where(ok, p, np.nan))
    h15 = roll_med(np.where(ok, h, np.nan))
    
    # Apply HR lag: shift power forward so current HR is compared to power from HR_LAG_S ago
    lag_samples = int(round(HR_LAG_S / dt))
    p_lagged = np.full_like(p_smooth, np.nan)
    if lag_samples < n:
        p_lagged[lag_samples:] = p_smooth[:-lag_samples]

    ratio = np.where(np.isfinite(p_lagged) & np.isfinite(h15) & (h15 > 30), p_lagged / h15, np.nan)

    def ratio_med(t0: float, t1: float) -> float:
        m = (t >= t0) & (t <= t1) & np.isfinite(ratio)
        if m.sum() < 10:
            return np.nan
        return float(np.median(ratio[m]))

    drop_i = None
    drop_mag = 0.0
    max_i = int(min(n-120, (hard_floor_s + 25*60.0) / dt))
    for i in range(int(min(n-1, hard_floor_s/dt)), max_i):
        if 'bad_near_gap' in locals() and bad_near_gap[i]:
            continue
        if not (np.isfinite(h15[i]) and np.isfinite(h15[i+30])):
            continue
        dh = h15[i+30] - h15[i]
        if dh > -12:
            continue
        if not (np.isfinite(p_smooth[i]) and np.isfinite(p_smooth[i+30])):
            continue
        if abs(p_smooth[i+30] - p_smooth[i]) > 0.08 * p_smooth[i]:
            continue
        j = i + 90
        if j >= n:
            break
        if np.nanmedian(h15[i+30:j]) > (h15[i] + dh/2):
            continue
        if abs(dh) > abs(drop_mag):
            drop_mag = dh
            drop_i = i

    early_t0, early_t1 = 60.0, min(300.0, t[-1])
    base_t0, base_t1 = min(360.0, t[-1]*0.6), min(720.0, t[-1]-30.0)
    r_base = ratio_med(base_t0, base_t1)
    r_early = ratio_med(early_t0, early_t1)

    applied = False
    corr_type: str | None = None
    drop_s: float | None = None
    h_corr = h.copy().astype(float)

    def cap_before(idx: int, ratio_series: np.ndarray, fit_from: int, fit_to: int):
        nonlocal h_corr
        x = t[fit_from:fit_to]
        y = ratio_series[fit_from:fit_to]
        m = np.isfinite(x) & np.isfinite(y)
        if m.sum() >= 20:
            A = np.vstack([np.ones(m.sum()), x[m]]).T
            a, b = np.linalg.lstsq(A, y[m], rcond=None)[0]
            rt = a + b * t[:idx]
        else:
            rt = np.full(idx, np.nanmedian(ratio_series[fit_from:fit_to]))
        rt = np.where(np.isfinite(rt) & (rt > 0), rt, np.nan)
        exp = np.where(np.isfinite(rt) & np.isfinite(p_lagged[:idx]) & (p_lagged[:idx] > 0), p_lagged[:idx] / rt, np.nan)
        mcap = np.isfinite(exp) & np.isfinite(h_corr[:idx]) & (h_corr[:idx] > exp + 2.0)
        h_corr[:idx][mcap] = exp[mcap] + 2.0

    def lift_after(idx: int, r_pre: float):
        nonlocal h_corr
        if not (np.isfinite(r_pre) and r_pre > 0):
            return
        exp = np.where(np.isfinite(p_lagged[idx:]) & (p_lagged[idx:] > 0), p_lagged[idx:] / r_pre, np.nan)
        mlift = np.isfinite(exp) & np.isfinite(h_corr[idx:]) & (h_corr[idx:] < exp - 2.0)
        h_corr[idx:][mlift] = exp[mlift] - 2.0

    if drop_i is not None:
        drop_s = float(t[drop_i])
        r_pre = ratio_med(t[drop_i]-240.0, t[drop_i]-30.0)
        r_post = ratio_med(t[drop_i]+30.0, t[drop_i]+360.0)

        dip_win = int(round(20.0 / dt))
        a_dip = max(0, drop_i - dip_win)
        b_dip = min(n, drop_i + dip_win + 1)
        p_local_min = float(np.nanmin(p_smooth[a_dip:b_dip])) if np.isfinite(p_smooth[a_dip:b_dip]).any() else float("nan")
        dip_w = 1.8 * float(mass_kg)
        has_transient_dip = np.isfinite(p_local_min) and (p_local_min < dip_w)

        # v40.1: Improved late_dropout vs early_artifact discrimination
        # Key insight: after a genuine dropout, HR is unrealistically LOW (ratio > 2.5 W/bpm)
        # After an early spike settles, HR is at a NORMAL level (ratio 1.5-2.5 W/bpm)
        #
        # Old logic: r_post > 1.10 * r_pre -> late_dropout
        # Problem: both early spike recovery AND late dropout show this pattern!
        #
        # New logic:
        # - If r_post > 2.5: post-drop HR is suspiciously LOW -> genuine dropout, lift
        # - If r_post < 2.5 and r_pre < 1.5: pre-drop HR was HIGH (artifact), post is normal -> early_artifact, cap
        # - If ambiguous: don't correct, log as "ambiguous_hr_drop"
        
        is_late_dropout = False
        is_early_artifact = False
        
        if np.isfinite(r_pre) and np.isfinite(r_post) and (r_post > 1.10 * r_pre) and (not has_transient_dip):
            if r_post > 2.5:
                # Post-drop HR is unrealistically low for the power - genuine dropout
                is_late_dropout = True
            elif r_pre < 1.5:
                # Pre-drop ratio was low (HR was too high) and post-drop is reasonable
                # This is an early spike that settled, not a dropout
                is_early_artifact = True
            elif r_post > 2.2 and r_pre > 1.6:
                # Post-drop ratio is quite high AND pre-drop was reasonable
                # Likely a genuine dropout
                is_late_dropout = True
            else:
                # Ambiguous - ratio change could be either cause
                # Default to early_artifact if drop is early, otherwise don't correct
                if drop_s < 8*60:
                    is_early_artifact = True
                # else: leave uncorrected (ambiguous)
        
        if is_late_dropout:
            lift_after(drop_i, r_pre)
            applied = True
            corr_type = "late_dropout"
        elif is_early_artifact or (drop_s < 12*60 and not is_late_dropout):
            # Treat as early artifact - cap the high values before the drop
            if drop_s < 12*60:
                fit_from = min(n-1, drop_i + int(round(30.0/dt)))
                fit_to = min(n, drop_i + int(round(360.0/dt)))
                cap_before(drop_i, ratio, fit_from, fit_to)
                applied = True
                corr_type = "early_artifact"

    if (not applied) and np.isfinite(r_base) and np.isfinite(r_early) and (r_base > 1.05 * r_early):
        idx = int(round(min(hard_floor_s, 480.0) / dt))
        idx = max(1, min(idx, n))
        exp = np.where(np.isfinite(p_lagged[:idx]) & (p_lagged[:idx] > 0), p_lagged[:idx] / r_base, np.nan)
        mcap = np.isfinite(exp) & np.isfinite(h_corr[:idx]) & (h_corr[:idx] > exp + 2.0)
        h_corr[:idx][mcap] = exp[mcap] + 2.0
        applied = True
        corr_type = "early_artifact"
        drop_s = float(t[idx-1]) if idx > 0 else None

    h_corr = np.clip(h_corr, 30.0, 230.0)

    return h_corr, applied, drop_s, corr_type


# ============================================================================
# v40: OVERRIDE FILE SUPPORT
# ============================================================================

def _resolve_date_overrides(df: pd.DataFrame, master_df: pd.DataFrame) -> pd.DataFrame:
    """
    Resolve date-based overrides to filename-based overrides using the master.
    
    Supported 'file' column formats:
        - Filename:     "2026-02-04_10-30-00.FIT"  (exact match, original behaviour)
        - Date only:    "2026-02-04"                (all runs that day)
        - Date + seq:   "2026-02-04 #2"             (2nd run that day, chronological)
    
    Date-based entries are expanded into one row per matched activity, then
    merged back with the filename-based entries.
    """
    import re
    
    # Pattern: YYYY-MM-DD optionally followed by #N
    date_pattern = re.compile(r'^(\d{4}-\d{2}-\d{2})\s*(?:#(\d+))?$')
    
    date_rows = []
    file_rows = []
    
    for idx, row in df.iterrows():
        file_val = str(row['file']).strip()
        m = date_pattern.match(file_val)
        if m:
            date_rows.append((idx, m.group(1), int(m.group(2)) if m.group(2) else None, row))
        else:
            file_rows.append(idx)
    
    if not date_rows:
        return df  # Nothing to resolve
    
    # Build date-to-files lookup from master (sorted by start time within each day)
    # Note: #N counts RUNS only — non-running activities are not in the master,
    # so a gym session before your run doesn't affect the numbering.
    master_dates = {}
    if master_df is not None and len(master_df) > 0:
        for _, mrow in master_df.iterrows():
            mfile = str(mrow.get('file', '')).strip()
            mdate = mrow.get('date')
            if not mfile or pd.isna(mdate):
                continue
            date_str = pd.Timestamp(mdate).strftime('%Y-%m-%d')
            if date_str not in master_dates:
                master_dates[date_str] = []
            master_dates[date_str].append((pd.Timestamp(mdate), mfile))
        
        # Sort each day's runs chronologically
        for d in master_dates:
            master_dates[d].sort(key=lambda x: x[0])
    
    # Resolve date-based rows into filename-based rows
    resolved = []
    for idx, date_str, seq_num, row in date_rows:
        runs_that_day = master_dates.get(date_str, [])
        
        if not runs_that_day:
            print(f"  Warning: Date override '{row['file']}' matched no activities - keeping for future use")
            # Keep as-is (won't match anything, but won't be lost)
            resolved.append(row)
            continue
        
        if seq_num is not None:
            # Specific run: #1 = first, #2 = second, etc.
            if seq_num < 1 or seq_num > len(runs_that_day):
                print(f"  Warning: Date override '{row['file']}' requests #{seq_num} but only {len(runs_that_day)} run(s) on {date_str}")
                continue
            _, target_file = runs_that_day[seq_num - 1]
            new_row = row.copy()
            new_row['file'] = target_file
            resolved.append(new_row)
            print(f"  Date override: {date_str} #{seq_num} -> {target_file}")
        else:
            # All runs that day
            for i, (_, target_file) in enumerate(runs_that_day):
                new_row = row.copy()
                new_row['file'] = target_file
                resolved.append(new_row)
            n = len(runs_that_day)
            print(f"  Date override: {date_str} -> {n} run{'s' if n > 1 else ''}")
    
    # Rebuild DataFrame: filename rows + resolved date rows
    result_rows = [df.loc[i] for i in file_rows] + resolved
    if not result_rows:
        return df.iloc[:0]  # Empty with same columns
    
    return pd.DataFrame(result_rows).reset_index(drop=True)


def load_override_file(override_path: str, master_df: pd.DataFrame = None) -> pd.DataFrame:
    """Load activity overrides from Excel file.
    
    Supports three formats in the 'file' column:
        - Filename:     "2026-02-04_10-30-00.FIT"  (exact match)
        - Date only:    "2026-02-04"                (all runs that day)
        - Date + seq:   "2026-02-04 #2"             (Nth run that day, chronological)
    
    Returns DataFrame indexed by 'file' column with columns:
        - temp_override (float, NaN if not set)
        - surface_adj (float, defaults to 1.0)
        - race_flag (int, 0 or 1)
        - official_distance_km (float, NaN if not set)
        - notes (str)
    """
    if not Path(override_path).exists():
        print(f"Note: Override file not found: {override_path} - using defaults")
        return pd.DataFrame(columns=['file', 'temp_override', 'surface_adj', 
                                      'race_flag', 'official_distance_km', 'notes']).set_index('file')
    
    try:
        df = pd.read_excel(override_path, engine='openpyxl', dtype={'file': str})
    except Exception as e:
        print(f"Warning: Could not read override file {override_path}: {e}")
        return pd.DataFrame(columns=['file', 'temp_override', 'surface_adj', 
                                      'race_flag', 'official_distance_km', 'notes']).set_index('file')
    
    # Ensure required columns exist
    if 'file' not in df.columns:
        print(f"Warning: Override file missing 'file' column")
        return pd.DataFrame(columns=['file', 'temp_override', 'surface_adj', 
                                      'race_flag', 'official_distance_km', 'notes']).set_index('file')
    
    # Set defaults for missing columns
    if 'temp_override' not in df.columns:
        df['temp_override'] = np.nan
    if 'surface_adj' not in df.columns:
        df['surface_adj'] = np.nan
    if 'race_flag' not in df.columns:
        df['race_flag'] = 0
    if 'official_distance_km' not in df.columns:
        df['official_distance_km'] = np.nan
    if 'notes' not in df.columns:
        df['notes'] = ''
    
    # Convert types
    df['temp_override'] = pd.to_numeric(df['temp_override'], errors='coerce')
    df['surface_adj'] = pd.to_numeric(df['surface_adj'], errors='coerce')
    df['race_flag'] = pd.to_numeric(df['race_flag'], errors='coerce').fillna(0).astype(int)
    df['official_distance_km'] = pd.to_numeric(df['official_distance_km'], errors='coerce')
    
    # Resolve date-based overrides to filenames using master data
    df = _resolve_date_overrides(df, master_df)
    
    print(f"Loaded {len(df)} overrides from {override_path}")
    race_count = (df['race_flag'] == 1).sum()
    dist_count = df['official_distance_km'].notna().sum()
    sadj_count = df['surface_adj'].notna().sum()
    print(f"  - {race_count} races, {dist_count} with official distance, {sadj_count} with surface_adj")
    
    # Normalize file names to lowercase for case-insensitive matching
    df['file'] = df['file'].astype(str).str.lower()
    
    # Check for duplicates after normalization
    duplicates = df[df['file'].duplicated(keep=False)]
    if len(duplicates) > 0:
        dup_files = duplicates['file'].unique()
        print(f"  Warning: {len(dup_files)} duplicate filenames (after case normalization), using first entry")
        df = df.drop_duplicates(subset='file', keep='first')
    
    return df.set_index('file')


# ============================================================================
# v46: STRAVA ELEVATION MATCHING
# ============================================================================

def match_strava_elevation(dfm: pd.DataFrame, strava_path: str, tz_local: str) -> int:
    """
    Match Strava elevation data to master rows by timestamp.
    
    v46: Strava uses corrected elevation (terrain model) which is more reliable
    than FIT barometric altitude, especially for older devices. This data is used
    as a sanity check for terrain adjustment gating.
    
    Populates strava_elev_gain_m and strava_elev_loss_m columns.
    Returns number of rows matched.
    """
    if not Path(strava_path).exists():
        print(f"Note: Strava file not found: {strava_path} - skipping elevation match")
        return 0
    
    try:
        act = pd.read_csv(strava_path)
    except Exception as e:
        print(f"Warning: Could not read Strava file {strava_path}: {e}")
        return 0
    
    # Filter to Run activities
    if 'Activity Type' in act.columns:
        act = act[act['Activity Type'] == 'Run'].copy()
    
    if len(act) == 0:
        print(f"Note: No Run activities found in {strava_path}")
        return 0
    
    # Parse Strava dates (UTC)
    act['strava_dt'] = pd.to_datetime(act['Activity Date'], format='mixed')
    
    # Parse elevation
    for col_name, target in [('Elevation Gain', 'ElevGain_m'), 
                              ('Total Elevation Gain', 'ElevGain_m'),
                              ('Elevation Loss', 'ElevLoss_m')]:
        if col_name in act.columns and target not in act.columns:
            act[target] = pd.to_numeric(act[col_name], errors='coerce')
    
    if 'ElevGain_m' not in act.columns:
        print(f"Warning: No elevation gain column found in Strava file")
        return 0
    
    # Master dates are local time (Stockholm), Strava dates are UTC.
    # Offset is 1h (CET) or 2h (CEST). Use 3h tolerance to handle both safely,
    # plus distance as tiebreaker when multiple Strava runs fall in the window.
    master_dt = pd.to_datetime(dfm['date'])
    
    # Sort Strava for binary search
    act_sorted = act.sort_values('strava_dt').reset_index(drop=True)
    act_times = act_sorted['strava_dt'].values.astype('datetime64[s]').astype(np.int64)
    
    # Also prepare Strava distances for tiebreaking
    act_sorted['_dist_km'] = pd.to_numeric(act_sorted.get('Distance', 0), errors='coerce')
    if act_sorted['_dist_km'].median(skipna=True) > 1000:
        act_sorted['_dist_km'] = act_sorted['_dist_km'] / 1000.0
    
    matched = 0
    tolerance_s = 10800  # 3 hours (covers CET/CEST offset + some slack)
    
    for i, row in dfm.iterrows():
        t = master_dt.loc[i] if i in master_dt.index else pd.NaT
        if pd.isna(t):
            continue
        
        t_int = np.datetime64(t, 's').astype(np.int64)
        idx = np.searchsorted(act_times, t_int)
        
        # Collect all candidates within tolerance
        candidates = []
        for j in range(max(0, idx - 3), min(len(act_sorted), idx + 4)):
            diff = abs(t_int - act_times[j])
            if diff < tolerance_s:
                candidates.append((diff, j))
        
        if not candidates:
            continue
        
        # If multiple candidates, prefer closest distance match
        if len(candidates) > 1 and 'distance_km' in dfm.columns:
            master_dist = pd.to_numeric(row.get('distance_km', np.nan), errors='coerce')
            if pd.notna(master_dist) and master_dist > 0:
                best_j = min(candidates, key=lambda c: abs(act_sorted.iloc[c[1]]['_dist_km'] - master_dist))[1]
            else:
                best_j = min(candidates, key=lambda c: c[0])[1]
        else:
            best_j = candidates[0][1]
        
        best_row = act_sorted.iloc[best_j]
        eg = best_row.get('ElevGain_m', np.nan)
        el = best_row.get('ElevLoss_m', np.nan)
        if pd.notna(eg):
            dfm.at[i, 'strava_elev_gain_m'] = float(eg)
            matched += 1
        if pd.notna(el):
            dfm.at[i, 'strava_elev_loss_m'] = float(el)
    
    print(f"  Matched Strava elevation for {matched}/{len(dfm)} runs")
    return matched


# ============================================================================
# v45: STRAVA ACTIVITY NAME REFRESH
# ============================================================================

def refresh_activity_names(dfm: pd.DataFrame, strava_path: str, tz_local: str) -> int:
    """Refresh activity names from Strava activities.csv.
    
    Matches activities by strava_activity_id (if present) or by timestamp.
    Only updates names where Strava has a non-empty name.
    
    Returns number of names updated.
    """
    if not Path(strava_path).exists():
        print(f"Note: Strava file not found: {strava_path} - skipping name refresh")
        return 0
    
    try:
        act = pd.read_csv(strava_path)
    except Exception as e:
        print(f"Warning: Could not read Strava file {strava_path}: {e}")
        return 0
    
    # Filter to Run activities only
    if 'Activity Type' in act.columns:
        act = act[act['Activity Type'] == 'Run'].copy()
    
    if len(act) == 0:
        print(f"Note: No Run activities found in {strava_path}")
        return 0
    
    # Build lookup by Activity ID
    id_to_name = {}
    if 'Activity ID' in act.columns and 'Activity Name' in act.columns:
        for _, row in act.iterrows():
            aid = row.get('Activity ID')
            name = row.get('Activity Name', '')
            if pd.notna(aid) and pd.notna(name) and str(name).strip():
                id_to_name[int(aid)] = str(name).strip()
    
    updated = 0
    
    # Update names where we have a strava_activity_id match
    if 'strava_activity_id' in dfm.columns and 'activity_name' in dfm.columns:
        for i, row in dfm.iterrows():
            sid = row.get('strava_activity_id')
            if pd.notna(sid):
                try:
                    sid_int = int(sid)
                    if sid_int in id_to_name:
                        new_name = id_to_name[sid_int]
                        old_name = str(row.get('activity_name', ''))
                        if new_name != old_name:
                            dfm.at[i, 'activity_name'] = new_name
                            updated += 1
                except (ValueError, TypeError):
                    pass
    
    if updated > 0:
        print(f"  Refreshed {updated} activity names from Strava")
    
    # Apply pending_activities.csv as override (user-provided names take priority)
    pending_csv = Path(strava_path).parent / "pending_activities.csv"
    if pending_csv.exists():
        try:
            pend = pd.read_csv(pending_csv, encoding='utf-8')
        except UnicodeDecodeError:
            pend = pd.read_csv(pending_csv, encoding='cp1252')
        
        if 'activity_name' in pend.columns and 'date' in pend.columns:
            pend_by_date = {}
            for _, pr in pend.iterrows():
                ds = str(pr.get('date', '')).strip()[:10]
                name = pr.get('activity_name', '')
                if ds and pd.notna(name) and str(name).strip():
                    pend_by_date[ds] = str(name).strip()
            
            pend_applied = 0
            if 'date' in dfm.columns:
                for i, row in dfm.iterrows():
                    d = pd.Timestamp(row.get('date'))
                    if not pd.isna(d):
                        ds = d.strftime('%Y-%m-%d')
                        if ds in pend_by_date:
                            old_name = str(row.get('activity_name', ''))
                            new_name = pend_by_date[ds]
                            if new_name != old_name:
                                dfm.at[i, 'activity_name'] = new_name
                                pend_applied += 1
            if pend_applied > 0:
                print(f"  Applied {pend_applied} pending activity name override(s)")
    
    return updated


def main() -> int:
    args = parse_args()

    dfm = pd.read_excel(args.master, engine="openpyxl")
    model = REModel.from_json(open(args.model_json, "r", encoding="utf-8").read())
    mid = f"simRE_s4_{_hash_file(args.model_json)}"

    # v45: Refresh activity names from Strava
    if args.strava:
        refresh_activity_names(dfm, args.strava, args.tz)
    
    # v46: Match Strava elevation data (for terrain adj sanity check)
    if args.strava:
        match_strava_elevation(dfm, args.strava, args.tz)

    # v40: Load override file (pass master for date-based override resolution)
    overrides = load_override_file(args.override_file, master_df=dfm)
    override_matches = 0
    distance_corrections = 0

    # Ensure output cols exist (v39 columns + v40 new columns + v43 new columns)
    needed_cols = [
        "hr_corrected",
        "avg_hr",
        "max_hr",
        "nPower_HR",
        "RF_window_start_s",
        "RF_window_end_s",
        "RF_window_shifted",
        "RF_select_code",
        "RF_adjusted_mean_W_per_bpm",
        "RF_adjusted_median_W_per_bpm",
        "RF_dead_frac",
        "RF_drift_pct_per_min",
        "RF_drift_r2",
        "RF_r2",
        "RE_avg",
        "RE_normalised",
        "power_source",
        "sim_model_id",
        # v40 new columns
        "race_flag",
        "official_distance_km",
        "surface_adj",
        "distance_corrected",
        "gps_distance_error_m",
        # v43 new columns
        "parkrun",
        "surface",
        "Temp_Adj",
        "Terrain_Adj",
        "Era_Adj",
        "Total_Adj",
        "Intensity_Adj",  # v44: intensity-based RF adjustment
        "Duration_Adj",   # v44: duration-based RF adjustment for long efforts
        "RF_adj",
        "Factor",
        "RF_Trend",
        "RFL",
        "RFL_Trend",
        "Power_Score",       # v44.5: Riegel-normalised power output (sets RF floor, boosts Factor)
        "PS_Floor_Applied",  # v44.5: True if Power Score floor was applied to RF_adj
        "TSS",
        "CTL",
        "ATL",
        "TSB",
    ]
    for c in needed_cols:
        if c not in dfm.columns:
            if c in ("hr_corrected", "RF_window_shifted", "distance_corrected"):
                dfm[c] = False
            elif c in ("race_flag", "parkrun"):
                dfm[c] = 0
            elif c in ("surface_adj", "Temp_Adj", "Terrain_Adj", "Elevation_Adj", "Era_Adj", "Total_Adj", "Intensity_Adj", "Duration_Adj"):
                dfm[c] = 1.0
            elif c in ("surface",):
                dfm[c] = ""
            else:
                dfm[c] = np.nan
    
    # v43: Era adjusters calculated after main loop (when RE_avg is fresh)
    era_csv_path = os.path.join(os.path.dirname(args.out), "stryd_era_adjusters.csv")

    # --- v40: Apply override file data BEFORE processing ---
    # This must happen early so distance corrections are in place for RE calculations
    for i, row in dfm.iterrows():
        fit_file = str(row.get('file', '')).lower()  # Lowercase for case-insensitive match
        if not fit_file:
            continue
            
        if fit_file in overrides.index:
            override = overrides.loc[fit_file]
            # Handle duplicate entries - take first row if multiple matches
            if isinstance(override, pd.DataFrame):
                print(f"  Warning: duplicate override entries for {fit_file}, using first")
                override = override.iloc[0]
            override_matches += 1
            
            # Race flag
            dfm.at[i, 'race_flag'] = int(override.get('race_flag', 0))
            
            # v43: Parkrun flag
            dfm.at[i, 'parkrun'] = int(override.get('parkrun', 0))
            
            # v43: Surface
            surface_val = str(override.get('surface', '')).strip().upper()
            if surface_val in VALID_SURFACES:
                dfm.at[i, 'surface'] = surface_val
            
            # Official distance (store for output regardless of era)
            if pd.notna(override.get('official_distance_km')):
                dfm.at[i, 'official_distance_km'] = float(override['official_distance_km'])
            
            # Surface adjustment
            if pd.notna(override.get('surface_adj')):
                dfm.at[i, 'surface_adj'] = float(override['surface_adj'])
            else:
                dfm.at[i, 'surface_adj'] = 1.0
            
            # Temperature override (apply to weather column)
            if pd.notna(override.get('temp_override')):
                dfm.at[i, 'avg_temp_c'] = float(override['temp_override'])
                # v41: When temp is overridden (indoor run), default humidity to 50%
                # This prevents outdoor humidity affecting indoor run adjustments
                dfm.at[i, 'avg_humidity_pct'] = 50.0
            
            # Distance correction for pre-Stryd activities (from override file)
            # Check calibration_era_id (if present) or power_source to determine if pre-Stryd
            era_id = str(row.get('calibration_era_id', '')).lower()
            power_src = str(row.get('power_source', '')).lower()
            is_pre_stryd = ('pre_stryd' in era_id) or ('sim' in power_src) or (power_src == '')
            
            if pd.notna(override.get('official_distance_km')):
                official_dist = float(override['official_distance_km'])
                gps_dist = pd.to_numeric(row.get('gps_distance_km', np.nan), errors='coerce')
                fit_dist = pd.to_numeric(row.get('distance_km', np.nan), errors='coerce')
                # Use GPS distance for pre-Stryd, FIT distance for Stryd era
                ref_dist = gps_dist if is_pre_stryd else fit_dist
                
                if pd.notna(ref_dist) and ref_dist > 0:
                    # v40.1: Sanity check - reject unreasonable corrections (>10% error)
                    error_pct = abs(ref_dist - official_dist) / official_dist
                    
                    if error_pct > 0.10:
                        print(f"  WARNING: Skipping unreasonable distance correction for {fit_file}: "
                              f"{ref_dist:.3f}km vs {official_dist:.3f}km ({error_pct*100:.1f}% error)")
                        # Still store official_distance for reference, but don't flag as corrected
                        continue
                    
                    # Apply distance correction
                    dfm.at[i, 'distance_km'] = official_dist
                    
                    # Recalculate pace: use elapsed_time for races (chip time),
                    # moving_time for non-races
                    is_race = dfm.at[i, 'race_flag'] == 1
                    if is_race:
                        time_for_pace = pd.to_numeric(row.get('elapsed_time_s', np.nan), errors='coerce')
                    else:
                        time_for_pace = pd.to_numeric(row.get('moving_time_s', np.nan), errors='coerce')
                    if pd.notna(time_for_pace) and time_for_pace > 0:
                        dfm.at[i, 'avg_pace_min_per_km'] = time_for_pace / 60.0 / official_dist
                    
                    # Flag correction
                    dfm.at[i, 'distance_corrected'] = True
                    dfm.at[i, 'gps_distance_error_m'] = (ref_dist - official_dist) * 1000.0
                    distance_corrections += 1
                    
                    era_label = "pre-Stryd" if is_pre_stryd else "Stryd-era"
                    print(f"  Distance corrected ({era_label}): {fit_file}: {ref_dist:.3f}km -> {official_dist:.3f}km "
                          f"({dfm.at[i, 'gps_distance_error_m']:.0f}m error)")
    
    # --- v40.1: Also flag Strava distance corrections for races ---
    # These affect RE calculation even though power is measured
    # Compare Strava vs FIT distance - only for races where user may have corrected in Strava
    # v47: Gate on race_flag — Strava/FIT differences of 1-2% are normal GPS noise,
    # not user corrections. Without this gate, ~180 non-race runs get false official_distance_km.
    strava_corrections = 0
    if 'strava_distance_km' in dfm.columns:
        for i, row in dfm.iterrows():
            # v47: Only apply Strava distance corrections for races
            if row.get('race_flag', 0) != 1:
                continue
            
            # Skip if already distance-corrected (from override file)
            if row.get('distance_corrected', False):
                continue
            
            # Skip if override file already set official_distance_km
            override_dist = pd.to_numeric(row.get('official_distance_km', np.nan), errors='coerce')
            if pd.notna(override_dist):
                continue
            
            strava_dist = pd.to_numeric(row.get('strava_distance_km', np.nan), errors='coerce')
            gps_dist = pd.to_numeric(row.get('gps_distance_km', np.nan), errors='coerce')
            fit_dist = pd.to_numeric(row.get('distance_km', np.nan), errors='coerce')
            
            # Check if Strava distance differs significantly from FIT distance
            # This indicates user correction in Strava (not just GPS vs Stryd difference)
            if pd.notna(strava_dist) and pd.notna(fit_dist) and fit_dist > 0 and pd.notna(gps_dist):
                strava_vs_fit_diff = abs(strava_dist - fit_dist) / fit_dist
                
                if strava_vs_fit_diff > 0.01:  # >1% difference suggests user correction
                    dfm.at[i, 'distance_corrected'] = True
                    dfm.at[i, 'official_distance_km'] = strava_dist
                    dfm.at[i, 'gps_distance_error_m'] = (gps_dist - strava_dist) * 1000.0
                    strava_corrections += 1
    
    if strava_corrections > 0:
        print(f"  Strava distance corrections flagged: {strava_corrections}")

    print(f"\nOverride file: {override_matches} activities matched, {distance_corrections} distance corrections applied")

    # --- v39: dtype-safe columns ---
    float_cols = [
        "avg_hr", "max_hr", "avg_power_w", "npower_w", "nPower_HR",
        "RE_avg", "RE_normalised",
        "RF_window_start_s", "RF_window_end_s",
        "RF_adjusted_mean_W_per_bpm", "RF_adjusted_median_W_per_bpm",
        "RF_dead_frac", "RF_drift_pct_per_min", "RF_drift_r2", "RF_r2",
        "full_run_drift_pct_per_min", "full_run_drift_r2", "full_run_drift_duration_min",
        "official_distance_km", "surface_adj", "gps_distance_error_m",
    ]
    for c in float_cols:
        if c in dfm.columns:
            dfm[c] = pd.to_numeric(dfm[c], errors="coerce").astype("float64")

    bool_cols = ["hr_corrected", "RF_window_shifted", "distance_corrected"]
    for c in bool_cols:
        if c in dfm.columns:
            dfm[c] = dfm[c].fillna(False).astype(bool)

    int_cols = ["race_flag"]
    for c in int_cols:
        if c in dfm.columns:
            dfm[c] = pd.to_numeric(dfm[c], errors="coerce").fillna(0).astype(int)

    obj_cols = ["RF_select_code", "power_source", "sim_model_id"]
    for c in obj_cols:
        if c in dfm.columns:
            dfm[c] = dfm[c].astype("object")

    keys, paths = _build_cache_index(args.persec_cache_dir)

    updated = 0
    missing_cache = 0
    processed = 0
    last_progress_reported = -1
    total = int(len(dfm))
    progress_every = int(getattr(args, "progress_every", 250) or 0)
    per_run = bool(getattr(args, "progress_per_run", False))

    def _progress(i: int, label: str) -> None:
        nonlocal last_progress_reported
        if per_run:
            print(f"[{processed}/{total}] {label} | updated={updated} | missing_cache={missing_cache}")
            return
        if progress_every and (processed % progress_every == 0 or processed == total):
            if processed != last_progress_reported:
                last_progress_reported = processed
                print(f"Progress {processed}/{total} | updated={updated} | missing_cache={missing_cache}")

    for i, row in dfm.iterrows():
        dt_local = row.get("date", None)
        if dt_local is None or (isinstance(dt_local, float) and not np.isfinite(dt_local)):
            continue
        tloc = pd.Timestamp(dt_local)
        if getattr(tloc, "tzinfo", None) is None:
            tloc = tloc.tz_localize(args.tz, ambiguous="NaT", nonexistent="shift_forward")
        tutc = tloc.tz_convert("UTC")
        epoch = float(tutc.timestamp())

        processed += 1
        label = f"{tloc.strftime('%Y-%m-%d')} {str(row.get('activity_name',''))}".strip()
        cache_fp = _find_cache(keys, paths, epoch, tol_s=int(args.match_tol_s))
        if cache_fp is None:
            missing_cache += 1
            _progress(i, label + " (no cache)")
            continue

        z = np.load(cache_fp, allow_pickle=True)
        t = _ensure_1d(z.get("t", np.array([])))
        v = _ensure_1d(z.get("speed_mps", np.array([])))
        d = _ensure_1d(z.get("dist_m", np.array([])))  # v44.2: Load distance array for RE calc
        g = _ensure_1d(z.get("grade", np.zeros_like(v)))
        elev = _ensure_1d(z.get("elev_m_clean", np.array([])))  # v45: Load elevation for reversal calc
        heading = _ensure_1d(z.get("heading_deg", np.full_like(v, np.nan)))
        hr = _ensure_1d(z.get("hr_bpm", np.full_like(v, np.nan)))
        p_meas = _ensure_1d(z.get("power_w", np.full_like(v, np.nan)))

        n = int(min(len(t), len(v), len(g), len(heading), len(hr), len(p_meas)))
        if n <= 0:
            continue
        t, v, g, heading, hr, p_meas = t[:n], v[:n], g[:n], heading[:n], hr[:n], p_meas[:n]

        # --- v40.1: Scale per-second speed for distance-corrected runs ---
        # If we have an official/Strava distance correction, scale per-second speeds so that
        # RE is calculated correctly. For pre-Stryd runs, this also corrects simulated power and RF.
        # For post-Stryd runs, RF uses measured power (unaffected) but RE uses speed (affected).
        meas_ok = np.isfinite(p_meas) & (p_meas > 0)
        use_meas = float(np.mean(meas_ok)) >= 0.5
        speed_scaled = False
        
        if row.get('distance_corrected', False):
            official_dist = pd.to_numeric(row.get('official_distance_km', np.nan), errors='coerce')
            gps_dist = pd.to_numeric(row.get('gps_distance_km', np.nan), errors='coerce')
            
            if pd.notna(official_dist) and pd.notna(gps_dist) and gps_dist > 0:
                speed_scale = official_dist / gps_dist
                
                if abs(speed_scale - 1.0) > 0.005:  # Only scale if >0.5% difference
                    v = v * speed_scale
                    speed_scaled = True
                    power_note = "power+RE+RF" if not use_meas else "RE only (power measured)"
                    print(f"  Speed scaled by {speed_scale:.4f} for {row.get('file', '?')} "
                          f"({official_dist:.3f}km vs {gps_dist:.3f}km GPS) - affects {power_note}")

        if use_meas:
            p = p_meas
            dfm.at[i, "power_source"] = str(row.get("power_source", "")) or "stryd"
        else:
            temp_c = pd.to_numeric(row.get("avg_temp_c", np.nan), errors="coerce")
            rh_pct = pd.to_numeric(row.get("avg_humidity_pct", np.nan), errors="coerce")
            wind_ms = pd.to_numeric(row.get("avg_wind_ms", 0.0), errors="coerce")
            wind_dir = pd.to_numeric(row.get("avg_wind_dir_deg", np.nan), errors="coerce")

            if not np.isfinite(temp_c):
                temp_c = 10.0
            if not np.isfinite(rh_pct):
                rh_pct = 70.0
            rh = float(rh_pct) / 100.0
            rho = air_density_kg_m3(float(temp_c), 1013.25, float(rh))
            rho_arr = np.full_like(v, rho, dtype=float)

            wa = np.zeros_like(v, dtype=float)

            v_s = _roll_med(v, 15)
            g_s = _roll_med(g, 15)
            g_s = np.clip(g_s, -0.12, 0.12)
            p = _ensure_1d(simulate_power_series(model, v_s, g_s, wa, rho_arr, mass_kg=float(args.mass_kg), smooth_win=15))
            if len(p) != len(v):
                p = p[:len(v)]
            dfm.at[i, "power_source"] = "sim_v1"
            dfm.at[i, "sim_model_id"] = mid

        # --- v38.7: RE recompute using ACTIVE running only ---
        try:
            re_avg, re_norm, _re_spd, _re_t = _calc_re_active(t, v, p, float(args.mass_kg), row, dist_m_arr=d)
        except Exception:
            re_avg, re_norm = (float("nan"), float("nan"))
        if np.isfinite(re_avg):
            dfm.at[i, "RE_avg"] = float(re_avg)
        else:
            dfm.at[i, "RE_avg"] = np.nan  # v44.2: Clear old value if calc fails
        if np.isfinite(re_norm):
            dfm.at[i, "RE_normalised"] = float(re_norm)
        else:
            dfm.at[i, "RE_normalised"] = np.nan  # v44.2: Clear old value if calc fails

        # --- v44: Terrain undulation metrics ---
        dist_km = pd.to_numeric(row.get('distance_km', np.nan), errors='coerce')
        if not np.isfinite(dist_km):
            dist_km = 0.0
        elev_gain_m = pd.to_numeric(row.get('elev_gain_m', np.nan), errors='coerce')
        if not np.isfinite(elev_gain_m):
            elev_gain_m = 0.0
        terrain = _calc_terrain_metrics(t, g, dist_km, elev_gain_m, elev=elev)
        for k, val in terrain.items():
            dfm.at[i, k] = val
        
        # --- v45: Hill power metrics (distinguish intervals from rolling terrain) ---
        hill_metrics = _calc_hill_power_metrics(g, p)
        for k, val in hill_metrics.items():
            dfm.at[i, k] = val

        # v43: Adjustment calculations moved to after main loop (second pass)
        # This ensures era adjusters are calculated from fresh RE_avg values

        # HR correction (Step B)
        prev_hr_corrected = bool(row.get("hr_corrected", False))
        prev_corr_type = str(row.get("hr_corr_type", "") or "")
        prev_corr_drop = row.get("hr_corr_drop_s", np.nan)
        if prev_hr_corrected and (prev_corr_type == "early_artifact") and pd.notna(prev_corr_drop) and float(prev_corr_drop) > 12*60:
            prev_hr_corrected = False
        prev_avg_hr = row.get("avg_hr", np.nan)
        prev_max_hr = row.get("max_hr", np.nan)
        prev_npower_hr = row.get("nPower_HR", np.nan)
        prev_rf_mean = row.get("RF_adjusted_mean_W_per_bpm", np.nan)
        prev_rf_median = row.get("RF_adjusted_median_W_per_bpm", np.nan)
        prev_rf_dead = row.get("RF_dead_frac", np.nan)
        prev_rf_drift = row.get("RF_drift_pct_per_min", np.nan)
        prev_rf_r2 = row.get("RF_drift_r2", np.nan)
        prev_frd_drift = row.get("full_run_drift_pct_per_min", np.nan)
        prev_frd_r2 = row.get("full_run_drift_r2", np.nan)
        prev_frd_dur = row.get("full_run_drift_duration_min", np.nan)

        hr_corr, applied, drop_s, corr_type = correct_hr_stepb(t, p, hr, v, mass_kg=float(args.mass_kg))

        preserve_prior_hr = (prev_hr_corrected and (not applied))

        if applied:
            dfm.at[i, "hr_corrected"] = True
            hr_used = hr_corr
        else:
            dfm.at[i, "hr_corrected"] = bool(prev_hr_corrected)
            hr_used = hr

        running_mask = np.isfinite(v) & (v >= 0.3) & np.isfinite(p) & (p > 0)

        if preserve_prior_hr:
            if pd.notna(prev_avg_hr): dfm.at[i, "avg_hr"] = float(prev_avg_hr)
            if pd.notna(prev_max_hr): dfm.at[i, "max_hr"] = float(prev_max_hr)
            if pd.notna(prev_npower_hr): dfm.at[i, "nPower_HR"] = float(prev_npower_hr)
            if pd.notna(prev_rf_mean): dfm.at[i, "RF_adjusted_mean_W_per_bpm"] = float(prev_rf_mean)
            if pd.notna(prev_rf_median): dfm.at[i, "RF_adjusted_median_W_per_bpm"] = float(prev_rf_median)
            if pd.notna(prev_rf_dead): dfm.at[i, "RF_dead_frac"] = float(prev_rf_dead)
            if pd.notna(prev_rf_drift): dfm.at[i, "RF_drift_pct_per_min"] = float(prev_rf_drift)
            if pd.notna(prev_rf_r2): dfm.at[i, "RF_drift_r2"] = float(prev_rf_r2)
            if pd.notna(prev_frd_drift): dfm.at[i, "full_run_drift_pct_per_min"] = float(prev_frd_drift)
            if pd.notna(prev_frd_r2): dfm.at[i, "full_run_drift_r2"] = float(prev_frd_r2)
            if pd.notna(prev_frd_dur): dfm.at[i, "full_run_drift_duration_min"] = float(prev_frd_dur)
        else:
            hrun = hr_used[running_mask]
            hrun = hrun[np.isfinite(hrun) & (hrun > 30)]
            if hrun.size:
                dfm.at[i, "avg_hr"] = float(np.nanmean(hrun))
                dfm.at[i, "max_hr"] = float(np.nanmax(hrun))

            p_run = p[running_mask]
            npw = _calc_np(p_run, win=30)
            if np.isfinite(npw) and hrun.size:
                dfm.at[i, "npower_w"] = float(npw)
                dfm.at[i, "avg_power_w"] = float(np.nanmean(p_run))
                dfm.at[i, "nPower_HR"] = float(npw / float(np.nanmean(hrun)))

            rf = _rf_metrics(t, v, p, hr_used, mass_kg=float(args.mass_kg))
            for k, val in rf.items():
                if k in dfm.columns:
                    dfm.at[i, k] = val
                else:
                    dfm[k] = (np.nan if k.endswith("_s") else "")
                    dfm.at[i, k] = val
            
            # v45: Calculate terrain metrics for RF window only (not whole run)
            # This ensures terrain_adj reflects conditions where RF was measured
            rf_start_s = rf.get('RF_window_start_s', np.nan)
            rf_end_s = rf.get('RF_window_end_s', np.nan)
            if np.isfinite(rf_start_s) and np.isfinite(rf_end_s) and rf_end_s > rf_start_s:
                # Slice arrays to RF window
                rf_mask = (t >= rf_start_s) & (t <= rf_end_s)
                if rf_mask.sum() >= 60:
                    t_rf = t[rf_mask]
                    g_rf = g[rf_mask]
                    p_rf = p[rf_mask]
                    v_rf = v[rf_mask]
                    elev_rf = elev[rf_mask] if len(elev) == len(t) else np.array([])
                    
                    # Calculate RF window distance (approximate from time and overall pace)
                    rf_duration_s = rf_end_s - rf_start_s
                    overall_pace = dist_km / (pd.to_numeric(row.get('moving_time_s', 1), errors='coerce') / 1000)
                    rf_dist_km = rf_duration_s / 1000 * overall_pace if overall_pace > 0 else rf_duration_s / 1000 * 3.0
                    
                    # Calculate RF window elev gain (approximate from grade)
                    g_rf_valid = g_rf[np.isfinite(g_rf)]
                    if len(g_rf_valid) > 0:
                        # Sum positive grade changes (uphill)
                        rf_elev_gain_m = np.sum(np.maximum(0, g_rf_valid)) * (rf_dist_km * 1000 / len(g_rf_valid))
                    else:
                        rf_elev_gain_m = 0.0
                    
                    rf_terrain = _calc_terrain_metrics(t_rf, g_rf, rf_dist_km, rf_elev_gain_m, elev=elev_rf)
                    dfm.at[i, 'rf_window_undulation_score'] = rf_terrain.get('undulation_score', np.nan)
                    dfm.at[i, 'rf_window_grade_std_pct'] = rf_terrain.get('grade_std_pct', np.nan)
                    
                    # Also calculate hill power metrics for RF window
                    rf_hill = _calc_hill_power_metrics(g_rf, p_rf)
                    dfm.at[i, 'rf_window_uphill_downhill_ratio'] = rf_hill.get('uphill_downhill_ratio', np.nan)
                    
                    # v45: Calculate RE for RF window
                    # RE = speed / (power/mass) = speed * mass / power
                    v_rf_valid = v_rf[np.isfinite(v_rf) & (v_rf > 0.5)]
                    p_rf_valid = p_rf[np.isfinite(p_rf) & (p_rf > 50)]
                    if len(v_rf_valid) >= 60 and len(p_rf_valid) >= 60:
                        rf_speed_mps = float(np.mean(v_rf_valid))
                        rf_power_w = float(np.mean(p_rf_valid))
                        rf_re = rf_speed_mps * float(args.mass_kg) / rf_power_w
                        dfm.at[i, 'rf_window_RE'] = rf_re

        # v47: Calculate full-run P/HR drift for long runs (>60 min)
        # Must be outside preserve_prior_hr block — new metric, no prior values to copy
        moving_time_s_frd = pd.to_numeric(row.get('moving_time_s', 0), errors='coerce')
        frd = _calc_full_run_drift(t, p, hr_used,
                                    mass_kg=float(args.mass_kg),
                                    moving_time_s=float(moving_time_s_frd) if pd.notna(moving_time_s_frd) else 0.0)
        for k, val in frd.items():
            if k not in dfm.columns:
                dfm[k] = np.nan
            dfm.at[i, k] = val

        # v40.1: Check HR reliability via power:HR ratio CV
        # Threshold is adjusted based on power variability - intervals get more leeway
        hr_cv, power_cv = _calc_hr_cv(t, p, hr_used, mass_kg=float(args.mass_kg))
        dfm.at[i, "hr_cv_pct"] = float(hr_cv) if np.isfinite(hr_cv) else np.nan
        
        # Adjusted threshold: base 12% + extra allowance for high power variability (intervals)
        # Uses RAW power CV (baseline 8% for steady running; intervals can be 15-50%+)
        # Capped at 35% - no need for higher thresholds
        adjusted_threshold = HR_CV_THRESHOLD + max(0.0, power_cv - 8.0) * 1.2 if np.isfinite(power_cv) else HR_CV_THRESHOLD
        adjusted_threshold = min(adjusted_threshold, 35.0)
        
        if not np.isfinite(hr_cv):
            # Can't assess HR reliability (not enough samples, too messy) - blank RF
            # Exceptions that should NOT be marked unreliable:
            # - Runs < 8 min (NaN is intentional, HR ramp-up issue)
            # - Sparse data with few RF window samples (old Garmin smart recording)
            run_duration_s = float(t[-1] - t[0]) if len(t) > 1 else 0.0
            rf_start_s = 600.0 if run_duration_s >= 900.0 else 0.0
            # v42: Match _calc_hr_cv logic - apply 90% truncation for runs >= 15min
            # Previously this counted all samples from rf_start_s, but _calc_hr_cv
            # truncates at 90% to exclude cooldown, causing false "unreliable" flags
            n_samples = len(t)
            start_idx = int(np.searchsorted(t, rf_start_s))
            end_idx = int(n_samples * 0.9) if run_duration_s >= 900.0 else n_samples
            rf_window_samples = end_idx - start_idx
            sample_rate = n_samples / run_duration_s if run_duration_s > 0 else 0.0
            # Only mark unreliable if: run >= 8min AND (enough samples OR not sparse)
            is_sparse_with_few_samples = (rf_window_samples < 200) and (sample_rate < 0.5)
            if run_duration_s >= 8 * 60 and not is_sparse_with_few_samples:
                corr_type = "unreliable"
                applied = True
                dfm.at[i, "RF_adjusted_mean_W_per_bpm"] = np.nan
                dfm.at[i, "RF_adjusted_median_W_per_bpm"] = np.nan
                dfm.at[i, "nPower_HR"] = np.nan
        elif hr_cv > adjusted_threshold:
            # HR data is unreliable - blank out RF values
            corr_type = "unreliable"
            applied = True  # Mark as "corrected" so hr_corr_type gets set
            dfm.at[i, "RF_adjusted_mean_W_per_bpm"] = np.nan
            dfm.at[i, "RF_adjusted_median_W_per_bpm"] = np.nan
            dfm.at[i, "nPower_HR"] = np.nan

        dfm.at[i, "hr_corr_drop_s"] = float(drop_s) if applied and drop_s is not None else np.nan
        dfm.at[i, "hr_corr_type"] = str(corr_type) if applied and corr_type is not None else ""

        # --- v43: RF_adj, Factor, TSS calculations (after RF_raw is computed) ---
        rf_raw = dfm.at[i, 'RF_adjusted_median_W_per_bpm']  # Use median to match BFW
        total_adj = dfm.at[i, 'Total_Adj']
        
        # For RF_adj we need the previous RF_Trend - but that requires sequential processing
        # For now, calculate without the jump cap (will be applied in rolling phase)
        rf_adj = calc_rf_adj(rf_raw, total_adj, None)  # No jump cap in first pass
        dfm.at[i, 'RF_adj'] = float(rf_adj) if np.isfinite(rf_adj) else np.nan
        
        # Factor (weighting for RF trend)
        distance_m = pd.to_numeric(row.get('distance_km', np.nan), errors='coerce') * 1000
        avg_hr = dfm.at[i, 'avg_hr']
        factor = calc_factor(distance_m, avg_hr, rf_adj)
        dfm.at[i, 'Factor'] = float(factor) if np.isfinite(factor) else np.nan
        
        # TSS (will be refined after RFL is calculated)
        moving_time_s = pd.to_numeric(row.get('moving_time_s', np.nan), errors='coerce')
        terrain_adj = dfm.at[i, 'Terrain_Adj']
        is_race = bool(dfm.at[i, 'race_flag'])
        # RFL not yet available, use None for initial TSS
        tss = calc_tss(moving_time_s, avg_hr, None, terrain_adj, distance_m, is_race)
        dfm.at[i, 'TSS'] = float(tss) if np.isfinite(tss) else np.nan

        updated += 1
        _progress(i, label + " (done)")

    # rounding
    for c in ("avg_hr", "max_hr"):
        if c in dfm.columns:
            dfm[c] = pd.to_numeric(dfm[c], errors="coerce").round(1)
    for c in ("avg_power_w", "npower_w"):
        if c in dfm.columns:
            dfm[c] = pd.to_numeric(dfm[c], errors="coerce").round(0)
    for c in ("nPower_HR", "RE_avg", "RE_normalised", "RF_adjusted_mean_W_per_bpm", "RF_adjusted_median_W_per_bpm", "RF_dead_frac", "RF_drift_pct_per_min", "RF_drift_r2", "RF_r2", "full_run_drift_pct_per_min", "full_run_drift_r2", "full_run_drift_duration_min"):
        if c in dfm.columns:
            dfm[c] = pd.to_numeric(dfm[c], errors="coerce").round(3)
    for c in ("hr_cv_pct",):
        if c in dfm.columns:
            dfm[c] = pd.to_numeric(dfm[c], errors="coerce").round(1)
    for c in ("surface_adj",):
        if c in dfm.columns:
            dfm[c] = pd.to_numeric(dfm[c], errors="coerce").round(3)
    for c in ("gps_distance_error_m",):
        if c in dfm.columns:
            dfm[c] = pd.to_numeric(dfm[c], errors="coerce").round(0)

    # --- v43: Rolling RF calculations ---
    # First, calculate era adjusters from fresh RE_avg values
    print("\n=== v43: Calculating era adjusters from fresh RE_avg ===")
    era_adjusters = calculate_era_adjusters_from_data(dfm)
    export_era_adjusters_csv(era_adjusters, era_csv_path)
    
    # Recalculate adjustment columns with correct era adjusters
    print("  Recalculating adjustment factors...")
    
    # Sort by date for rolling calculations
    dfm = dfm.sort_values('date').reset_index(drop=True)
    
    # v50: Vectorized Temp_Trend calculation (replaces iterrows loop)
    # Uses cumulative sums for O(n) performance instead of O(n²) window lookups
    print("  Calculating weighted Temp_Trend (vectorized)...")
    days_back = RF_CONSTANTS['days_back']
    
    dates_s = pd.to_datetime(dfm['date']).values.astype('datetime64[s]').astype('int64')  # epoch seconds (pandas 3.0 safe)
    moving_times = pd.to_numeric(dfm['moving_time_s'], errors='coerce').fillna(0).values
    temps_raw = pd.to_numeric(dfm['avg_temp_c'], errors='coerce').fillna(0).values
    wt_product = moving_times * temps_raw
    
    temp_trend_arr = np.full(len(dfm), np.nan)
    cutoff_seconds = days_back * 86400
    
    # Use searchsorted for efficient window boundary finding
    for i in range(len(dfm)):
        if not np.isfinite(dates_s[i]):
            continue
        cutoff_epoch = dates_s[i] - cutoff_seconds
        j_start = np.searchsorted(dates_s[:i+1], cutoff_epoch, side='left')
        w_sum = moving_times[j_start:i+1].sum()
        if w_sum > 0:
            temp_trend_arr[i] = wt_product[j_start:i+1].sum() / w_sum
    
    dfm['Temp_Trend'] = temp_trend_arr
    
    # v50: Merged adjustment factor calculation (was separate iterrows loop)
    print("  Calculating adjustment factors...")
    for i, row in dfm.iterrows():
        temp_c = pd.to_numeric(row.get('avg_temp_c', np.nan), errors='coerce')
        humidity_pct = pd.to_numeric(row.get('avg_humidity_pct', np.nan), errors='coerce')
        temp_trend = pd.to_numeric(row.get('Temp_Trend', np.nan), errors='coerce')
        era_id = str(row.get('calibration_era_id', '')).lower().strip()
        undulation_score = pd.to_numeric(row.get('undulation_score', np.nan), errors='coerce')
        elev_gain_m = pd.to_numeric(row.get('elev_gain_m', np.nan), errors='coerce')
        distance_km = pd.to_numeric(row.get('distance_km', np.nan), errors='coerce')
        elev_gain_per_km = elev_gain_m / distance_km if distance_km > 0 else 0
        
        # v44.2: Get RE_avg and era median for terrain gating
        re_avg = pd.to_numeric(row.get('RE_avg', np.nan), errors='coerce')
        era_data = era_adjusters.get(era_id, {})
        era_median_re = era_data.get('re_median', None)
        
        # Calculate full Temp_Adj (stored for diagnostics and Power Score)
        temp_adj_full = calc_temp_adj(temp_c, humidity_pct, temp_trend)
        
        # v45: Scale Temp_Adj for RF based on RF window duration
        # RF is measured early in the run, so heat hasn't fully accumulated yet
        # Use same formula as Power Score: heat_mult = duration / reference_mins, but fixed at RF window
        # Window midpoint = warmup_skip (10min) + half the window duration
        rf_window_mid_mins = (600.0 + RF_CONSTANTS['rf_window_duration_s'] / 2.0) / 60.0
        heat_ref_mins = RF_CONSTANTS['heat_reference_mins']
        rf_heat_multiplier = rf_window_mid_mins / heat_ref_mins
        temp_adj_for_rf = 1.0 + (temp_adj_full - 1.0) * rf_heat_multiplier
        
        # v45: Use RF window terrain metrics (not whole run) for terrain_adj
        # Also gate by uphill/downhill ratio to exclude hill intervals
        rf_window_undulation = row.get('rf_window_undulation_score', np.nan)
        rf_window_ratio = row.get('rf_window_uphill_downhill_ratio', np.nan)
        
        # v45: Sanity check - RF window undulation can't be much higher than whole run
        # If whole run is flat (und < 5) but RF window claims to be hilly, it's likely GPS noise
        # v46.1: Tightened cap for low-undulation runs to prevent terrain leaking onto
        # genuinely flat runs (parkruns, track, treadmill) where rf_window catches a
        # local hill. Old cap (und + 5) allowed und=0 -> cap=5 -> passed threshold(4).
        # New cap: max(und * 2, und + 2) -> und=0 caps at 2, und=2 caps at 4, etc.
        if pd.notna(rf_window_undulation) and pd.notna(undulation_score):
            if undulation_score < 5:
                # Flat run - don't trust high RF window undulation
                rf_window_undulation = min(rf_window_undulation, max(undulation_score * 2, undulation_score + 2))
            else:
                # Hilly run - RF window can be somewhat higher but not extreme
                rf_window_undulation = min(rf_window_undulation, undulation_score * 2)
        
        # Use RF window undulation if available, otherwise fall back to whole run.
        # RF window is what the RF calculation actually experienced.
        undulation_for_adj = rf_window_undulation if pd.notna(rf_window_undulation) else undulation_score
        
        # Gate: if uphill/downhill ratio > threshold, it's likely hill intervals - no terrain boost
        # Hill intervals have hard sprints up and easy jogs down (ratio 1.6-2.3)
        # Rolling terrain like LL30 has consistent effort (ratio ~1.0-1.1)
        ratio_threshold = RF_CONSTANTS['terrain_hill_interval_ratio']
        
        # v46: Strava elevation sanity check
        # FIT barometric elevation can be inflated by noise (especially older devices).
        # If Strava elevation is available and much lower, skip terrain adjustment.
        strava_elev_gate_pass = True
        strava_elev_m = pd.to_numeric(row.get('strava_elev_gain_m', np.nan), errors='coerce')
        if pd.notna(strava_elev_m) and distance_km > 0:
            strava_elev_per_km = strava_elev_m / distance_km
            if strava_elev_per_km < RF_CONSTANTS['terrain_strava_elev_min']:
                strava_elev_gate_pass = False
        
        if pd.notna(rf_window_ratio) and rf_window_ratio > ratio_threshold:
            terrain_adj = 1.0  # No terrain adjustment for hill intervals
        elif not strava_elev_gate_pass:
            terrain_adj = 1.0  # v46: No terrain adj when Strava elev too low (bad barometric data)
        else:
            terrain_adj = calc_terrain_adj(undulation_for_adj)
        
        elevation_adj = calc_elevation_adj(elev_gain_per_km)
        # v49: Use regression-based power_adjuster_to_S4 as Era_Adj (speed-controlled)
        power_adj_to_s4 = pd.to_numeric(row.get('power_adjuster_to_S4', np.nan), errors='coerce')
        era_adj = calc_era_adj(era_id, era_adjusters, power_adj_to_s4=power_adj_to_s4)
        # v51.6: Sim power is already S4-calibrated — don't apply Stryd-era correction
        # for runs outside the primary sim eras (those have their own calibration path)
        if str(row.get('power_source', '')) == 'sim_v1' and era_id not in ('pre_stryd', 'v1_late'):
            era_adj = 1.0
        # Own_Adj is just surface_adj
        surface_adj = pd.to_numeric(row.get('surface_adj', 1.0), errors='coerce')
        own_adj = surface_adj if surface_adj and np.isfinite(surface_adj) else 1.0
        
        # Use scaled temp_adj for RF's total_adj
        total_adj = calc_total_adj(temp_adj_for_rf, terrain_adj, era_adj, own_adj, elevation_adj=elevation_adj)
        
        # Store the FULL Temp_Adj for diagnostics and Power Score calculation
        dfm.at[i, 'Temp_Adj'] = float(temp_adj_full)
        dfm.at[i, 'Terrain_Adj'] = float(terrain_adj)
        dfm.at[i, 'Elevation_Adj'] = float(elevation_adj)
        dfm.at[i, 'Era_Adj'] = float(era_adj)
        dfm.at[i, 'Total_Adj'] = float(total_adj)
    
    # Calculate RF_adj with jump cap - needs to be done sequentially with RF_Trend
    # because each RF_adj depends on the previous RF_Trend
    print("  Calculating RF_adj with jump cap and RF_Trend...")
    dfm['RF_adj'] = np.nan
    dfm['RF_Trend'] = np.nan
    dfm['Factor'] = np.nan
    dfm['Power_Score'] = np.nan  # v44.5: Power Score for each run
    dfm['PS_Floor_Applied'] = False  # v44.5: Track when Power Score floor was used
    
    days_back = RF_CONSTANTS['days_back']
    max_jump_pct = RF_CONSTANTS['rf_max_jump_pct']
    
    prev_run_date = None  # v44.3: Track previous run date for gap calculation
    peak_rf_trend = 0.0   # Running max of RF_Trend (for intensity_adj RFL conversion)
    
    # v44.5: Power Score parameters for RF_adj floor and Factor boost
    ps_riegel_k = POWER_SCORE_RIEGEL_K
    ps_reference_dist_km = POWER_SCORE_REFERENCE_DIST_KM
    ps_rf_divisor = POWER_SCORE_RF_DIVISOR  # RF_adj floor = Power_Score / this
    ps_factor_threshold = RF_CONSTANTS['power_score_threshold']  # Only boost Factor above this
    ps_factor_boost = POWER_SCORE_FACTOR_BOOST  # Factor multiplier for high Power Score runs
    
    for i, row in dfm.iterrows():
        current_date = row['date']
        
        # v44.3: Calculate days since last run
        days_since_last_run = None
        if prev_run_date is not None and pd.notna(current_date):
            days_since_last_run = (current_date - prev_run_date).days
        
        # Get RF_raw and total_adj
        rf_raw = pd.to_numeric(row.get('RF_adjusted_median_W_per_bpm', np.nan), errors='coerce')
        total_adj = row.get('Total_Adj', 1.0)
        
        # v44.5: Calculate Power Score for this row
        # v45: Use distance (not time) for Riegel factor, plus duration-scaled heat adjustment
        avg_power = pd.to_numeric(row.get('avg_power_w', np.nan), errors='coerce')
        avg_air_power = pd.to_numeric(row.get('avg_air_power_w', np.nan), errors='coerce')
        era_adj = row.get('Era_Adj', 1.0)
        temp_adj = row.get('Temp_Adj', 1.0)
        moving_time_s = pd.to_numeric(row.get('moving_time_s', np.nan), errors='coerce')
        distance_km = pd.to_numeric(row.get('distance_km', np.nan), errors='coerce')
        
        # Use official_distance_km if available (for races with corrected distance)
        official_dist = pd.to_numeric(row.get('official_distance_km', np.nan), errors='coerce')
        if pd.notna(official_dist) and official_dist > 0:
            distance_km = official_dist
        
        power_score = np.nan
        if pd.notna(avg_power) and pd.notna(distance_km) and distance_km > 0 and pd.notna(moving_time_s) and moving_time_s > 0:
            # v45: Apply diminishing returns to air power above threshold
            # Air power can be inflated at high paces or in headwinds
            # Above threshold: cut excess by factor (e.g., 9.3% -> 4% + 2.65% = 6.65%)
            power_for_ps = avg_power
            if pd.notna(avg_air_power) and avg_air_power > 0 and avg_power > 0:
                air_pct = avg_air_power / avg_power
                if air_pct > POWER_SCORE_AIR_THRESHOLD:
                    excess_air_pct = air_pct - POWER_SCORE_AIR_THRESHOLD
                    capped_air_pct = POWER_SCORE_AIR_THRESHOLD + excess_air_pct * POWER_SCORE_AIR_EXCESS_FACTOR
                    # Reduce power by the difference
                    air_reduction = (air_pct - capped_air_pct) * avg_power
                    power_for_ps = avg_power - air_reduction
            
            adj_power = power_for_ps * era_adj
            # v45: Use distance for Riegel factor (not time)
            # This ensures same-distance races are comparable regardless of finish time
            distance_factor = (distance_km / ps_reference_dist_km) ** ps_riegel_k
            
            # v45: Duration-scaled heat adjustment for Power Score
            # Heat impact scales with duration: full effect at heat_reference_mins, capped at heat_max_multiplier
            # Short races (5K, 10K) get less heat adjustment than long races (HM, marathon)
            heat_ref_s = RF_CONSTANTS['heat_reference_mins'] * 60.0
            ps_heat_multiplier = min(RF_CONSTANTS['heat_max_multiplier'], moving_time_s / heat_ref_s)
            ps_heat_adj = 1.0 + (temp_adj - 1.0) * ps_heat_multiplier
            
            power_score = adj_power * distance_factor * ps_heat_adj
            dfm.at[i, 'Power_Score'] = power_score
        
        if pd.notna(rf_raw) and rf_raw > 0 and np.isfinite(total_adj):
            # Get previous RF_Trend (from row i-1) for intensity calculation
            prev_rf_trend = dfm.at[i-1, 'RF_Trend'] if i > 0 else np.nan
            
            # v44: Calculate intensity adjustment based on power vs current CP
            # CP = peak_cp * RFL_Trend, where RFL_Trend ~ RF_Trend / peak_RF
            # For intensity calc, we use previous RF_Trend as proxy for current fitness
            if pd.notna(prev_rf_trend) and prev_rf_trend > 0 and peak_rf_trend > 0:
                # Convert RF_Trend to RFL_Trend using peak RF calculated from data
                rfl_for_intensity = prev_rf_trend / peak_rf_trend
            else:
                # First few runs or missing trend - use default RFL
                rfl_for_intensity = RF_CONSTANTS['intensity_default_rfl']
            
            # v45: Get RF window undulation for terrain-adjusted intensity threshold
            # Use the same sanity-checked value we used for terrain_adj
            rf_und_for_intensity = undulation_for_adj if pd.notna(undulation_for_adj) else 0.0
            intensity_adj = calc_intensity_adj(avg_power, rfl_for_intensity, rf_und_for_intensity)
            
            # v47: Calculate duration adjustment using full-run P/HR drift
            rf_drift = pd.to_numeric(row.get('RF_drift_pct_per_min', np.nan), errors='coerce')
            frd = pd.to_numeric(row.get('full_run_drift_pct_per_min', np.nan), errors='coerce')
            frd_r2 = pd.to_numeric(row.get('full_run_drift_r2', np.nan), errors='coerce')
            duration_adj = calc_duration_adj(moving_time_s, rf_drift,
                                             full_run_drift=frd, full_run_drift_r2=frd_r2)
            
            # Apply intensity and duration adjustments to total_adj
            rf_adj_uncapped = rf_raw * total_adj * intensity_adj * duration_adj
            
            # Apply jump cap: max 110% of previous RF_Trend (or 2.0 if no previous)
            if pd.notna(prev_rf_trend) and prev_rf_trend > 0:
                max_rf = prev_rf_trend * (1 + max_jump_pct / 100)
                rf_adj = min(rf_adj_uncapped, max_rf)
            else:
                rf_adj = min(rf_adj_uncapped, 2.0)
            
            # v44.5: Apply Power Score floor - RF_adj can't be lower than Power_Score / 180
            # This ensures great race performances are reflected even if RF was suppressed
            ps_floor_applied = False
            if pd.notna(power_score) and power_score > 0:
                rf_floor = power_score / ps_rf_divisor
                if rf_adj < rf_floor:
                    rf_adj = rf_floor
                    ps_floor_applied = True
            
            dfm.at[i, 'RF_adj'] = float(rf_adj)
            dfm.at[i, 'PS_Floor_Applied'] = ps_floor_applied
            
            # Store adjustments for diagnostics
            dfm.at[i, 'Intensity_Adj'] = float(intensity_adj)
            dfm.at[i, 'Duration_Adj'] = float(duration_adj)
            
            # v47: Update Total_Adj to include ALL adjusters
            dfm.at[i, 'Total_Adj'] = float(total_adj * intensity_adj * duration_adj)
            
            # Calculate Factor for this row (with outlier adjustment based on prev_rf_trend)
            # v44.3: Pass days_since_last_run to skip reduction if there's been a long gap
            distance_m = pd.to_numeric(row.get('distance_km', 0), errors='coerce') * 1000
            avg_hr = pd.to_numeric(row.get('avg_hr', np.nan), errors='coerce')
            factor = calc_factor(distance_m, avg_hr, rf_adj, prev_rf_trend, days_since_last_run)
            
            # v44.5: Boost Factor for high Power Score runs
            # Good races get more weight in RF_Trend
            if pd.notna(power_score) and power_score > ps_factor_threshold:
                factor = factor * (1 + ps_factor_boost)
            
            if np.isfinite(factor):
                dfm.at[i, 'Factor'] = float(factor)
        
        # v44.3: Update previous run date
        if pd.notna(current_date):
            prev_run_date = current_date
        
        # Calculate RF_Trend up to this point
        # v45: days_back-day window with quadratic decay: weight = 1 - (days_ago/days_back)^2
        # v45: Days-off penalty: periods > gap_threshold days without activity add penalty to effective days
        current_date = row['date']
        gap_threshold = RF_CONSTANTS['days_off_gap_threshold']
        penalty_rate = RF_CONSTANTS['days_off_penalty_rate']
        if pd.notna(current_date):
            cutoff_date = current_date - pd.Timedelta(days=days_back)
            mask = (dfm['date'] >= cutoff_date) & (dfm['date'] <= current_date) & (dfm.index <= i)
            window_df = dfm.loc[mask].copy()
            
            if len(window_df) > 0:
                # Calculate days ago for each row
                days_ago = (current_date - window_df['date']).dt.total_seconds() / 86400
                
                # v45: Find inactive periods > gap_threshold days in window and calculate penalties
                # For each activity, add penalty for any breaks that occurred AFTER that activity
                window_dates = window_df['date'].sort_values()
                
                # Calculate gaps between consecutive activities
                gaps = []
                prev_date = None
                for activity_date in window_dates:
                    if prev_date is not None:
                        gap_days = (activity_date - prev_date).total_seconds() / 86400
                        if gap_days > gap_threshold:
                            gaps.append((prev_date, activity_date, gap_days))
                    prev_date = activity_date
                
                # Also check gap from last activity to current_date
                if len(window_dates) > 0:
                    last_activity = window_dates.iloc[-1]
                    gap_to_now = (current_date - last_activity).total_seconds() / 86400
                    if gap_to_now > gap_threshold:
                        gaps.append((last_activity, current_date, gap_to_now))
                
                # For each activity, calculate total penalty from breaks that occurred after it
                effective_days_ago = days_ago.copy()
                for idx in window_df.index:
                    activity_date = window_df.loc[idx, 'date']
                    penalty = 0
                    for gap_start, gap_end, gap_days in gaps:
                        # If this gap occurred after the activity, apply penalty
                        if gap_start >= activity_date:
                            penalty += (gap_days - gap_threshold) * penalty_rate
                    effective_days_ago.loc[idx] = days_ago.loc[idx] + penalty
                
                # Quadratic decay using effective days (capped at days_back for the formula)
                time_decay = (1 - (effective_days_ago.clip(upper=days_back) / days_back) ** 2).clip(lower=0)
                
                factors = window_df['Factor'].fillna(0)
                rf_adjs = window_df['RF_adj'].fillna(0)
                
                valid = (factors > 0) & (rf_adjs > 0)
                if valid.any():
                    # Combined weight = Factor * time_decay
                    combined_weights = factors[valid] * time_decay[valid]
                    numerator = (combined_weights * rf_adjs[valid]).sum()
                    denominator = combined_weights.sum()
                    if denominator > 0:
                        rf_trend_val = numerator / denominator
                        dfm.at[i, 'RF_Trend'] = rf_trend_val
                        if np.isfinite(rf_trend_val) and rf_trend_val > peak_rf_trend:
                            peak_rf_trend = rf_trend_val
    
    print("\n=== v43: Calculating rolling metrics ===")
    dfm = calc_rolling_metrics(dfm)
    
    # Recalculate TSS with RFL now available
    print("  Recalculating TSS with RFL...")
    for i, row in dfm.iterrows():
        rfl = row.get('RFL')
        if pd.notna(rfl) and rfl > 0:
            moving_time_s = pd.to_numeric(row.get('moving_time_s', np.nan), errors='coerce')
            avg_hr = row.get('avg_hr')
            distance_m = pd.to_numeric(row.get('distance_km', np.nan), errors='coerce') * 1000
            terrain_adj = row.get('Terrain_Adj', 1.0)
            is_race = bool(row.get('race_flag', 0))
            tss = calc_tss(moving_time_s, avg_hr, rfl, terrain_adj, distance_m, is_race)
            if np.isfinite(tss):
                dfm.at[i, 'TSS'] = float(tss)
    
    # Calculate CTL/ATL/TSB
    athlete_data_path = args.athlete_data if hasattr(args, 'athlete_data') else "athlete_data.csv"
    dfm, daily_df = calc_ctl_atl(dfm, athlete_data_path)
    
    # v51: Join weight_kg from athlete_data.csv
    if os.path.exists(athlete_data_path):
        try:
            ad_df = _read_athlete_csv(athlete_data_path)
            if 'weight_kg' in ad_df.columns:
                ad_df['date'] = pd.to_datetime(ad_df['date'], dayfirst=True, format='mixed')
                wt = ad_df[['date', 'weight_kg']].dropna(subset=['weight_kg']).copy()
                wt['date_only'] = wt['date'].dt.normalize()
                wt = wt.drop_duplicates(subset='date_only', keep='last')
                # v51.6: Drop existing weight_kg to avoid merge column collision (_x/_y)
                if 'weight_kg' in dfm.columns:
                    dfm.drop(columns=['weight_kg'], inplace=True)
                dfm['date_only'] = dfm['date'].dt.normalize()
                dfm = dfm.merge(wt[['date_only', 'weight_kg']], on='date_only', how='left')
                dfm.drop(columns=['date_only'], inplace=True)
                # DEBUG: show athlete_data values for 1st-of-month dates
                _dec1 = wt[wt['date_only'] == pd.Timestamp('2025-12-01')]
                _jan12 = wt[wt['date_only'] == pd.Timestamp('2025-01-12')]
                print(f"  DEBUG athlete_data Dec 1: {_dec1['weight_kg'].tolist()}")
                print(f"  DEBUG athlete_data Jan 12: {_jan12['weight_kg'].tolist()}")
                _m = dfm[dfm['date'].between('2025-11-28', '2025-12-05')]
                for _, _r in _m.iterrows():
                    print(f"  DEBUG Master {str(_r['date'])[:10]} weight={_r['weight_kg']}")
                # Forward-fill weight for runs on days without weight data
                dfm['weight_kg'] = dfm['weight_kg'].ffill()
                n_wt = dfm['weight_kg'].notna().sum()
                latest_wt = dfm['weight_kg'].dropna().iloc[-1] if n_wt > 0 else np.nan
                print(f"  Joined weight_kg: {n_wt} runs, latest={latest_wt:.1f}kg")
        except Exception as e:
            print(f"  WARNING: Could not join weight_kg: {e}")
    
    # v51: Calculate Easy RF metrics (needs RFL_Trend and RF_adj)
    dfm = calc_easy_rf_metrics(dfm)
    
    # v51: Calculate historical alert columns (needs CTL, TSB, RFL_Trend, Easy_RF_EMA, Easy_RFL_Gap)
    dfm = calc_alert_columns(dfm)
    
    # v51: Report current alert state
    alerts = get_current_alerts(dfm)
    print_alerts(alerts)
    
    # Generate summary sheets
    print("\n=== Generating Summary Sheets ===")
    weekly_df = generate_weekly_summary(dfm)
    monthly_df = generate_monthly_summary(dfm)
    yearly_df = generate_yearly_summary(dfm)
    
    # ==========================================================================
    # v43: Calculate CP, Race Predictions, and Age Grades
    # ==========================================================================
    if AGE_GRADE_AVAILABLE:
        print("\n=== v43: Calculating CP, Race Predictions, Age Grades ===")
        
        # v44.5: Use RFL_Trend (Power Score now affects RF_adj directly)
        latest_rfl = dfm['RFL_Trend'].iloc[-1] if 'RFL_Trend' in dfm.columns else np.nan
        
        # CP = RFL_Trend * peak_CP (370W)
        if pd.notna(latest_rfl) and latest_rfl > 0:
            cp_current = latest_rfl * PEAK_CP_WATTS
            print(f"  Current CP: {cp_current:.0f}W (RFL_Trend={latest_rfl*100:.1f}%)")
            
            # Get 90th percentile RE from recent s5 era runs
            recent_re = dfm[dfm['calibration_era_id'] == 's5']['RE_avg'].dropna()
            if len(recent_re) < 10:
                # Fall back to all recent runs
                recent_re = dfm.tail(50)['RE_avg'].dropna()
            re_p90 = recent_re.quantile(0.90) if len(recent_re) > 0 else 0.95
            print(f"  RE p90: {re_p90:.4f}")
            
            # Race predictions (need mass_kg from args)
            mass_kg = args.mass_kg
            pred_5k = calc_race_prediction(latest_rfl, '5k', re_p90, PEAK_CP_WATTS, mass_kg)
            pred_10k = calc_race_prediction(latest_rfl, '10k', re_p90, PEAK_CP_WATTS, mass_kg)
            pred_hm = calc_race_prediction(latest_rfl, 'hm', re_p90, PEAK_CP_WATTS, mass_kg)
            pred_mara = calc_race_prediction(latest_rfl, 'marathon', re_p90, PEAK_CP_WATTS, mass_kg)
            
            print(f"  Race predictions: 5K={format_time(pred_5k)}, 10K={format_time(pred_10k)}, "
                  f"HM={format_time(pred_hm)}, Mar={format_time(pred_mara)}")
            
            # v51: Populate predictions on ALL runs (not just last row)
            # CP and predictions are simple functions of RFL_Trend with fixed constants
            rfl_valid = dfm['RFL_Trend'].notna() & (dfm['RFL_Trend'] > 0)
            dfm.loc[rfl_valid, 'CP'] = (dfm.loc[rfl_valid, 'RFL_Trend'] * PEAK_CP_WATTS).round(0)
            
            for dist_key, col_name in [('5k', 'pred_5k_s'), ('10k', 'pred_10k_s'), 
                                        ('hm', 'pred_hm_s'), ('marathon', 'pred_marathon_s')]:
                dfm.loc[rfl_valid, col_name] = dfm.loc[rfl_valid, 'RFL_Trend'].apply(
                    lambda rfl: round(calc_race_prediction(rfl, dist_key, re_p90, PEAK_CP_WATTS, mass_kg), 0)
                )
            
            n_pred = rfl_valid.sum()
            print(f"  Populated predictions on {n_pred} runs (all with RFL_Trend)")
            
            # Predicted 5K age grade (calculate age from DOB if available)
            runner_gender = getattr(args, 'runner_gender', 'male')
            runner_dob = getattr(args, 'runner_dob', None)
            runner_age_default = getattr(args, 'runner_age', 55)
            
            if runner_dob:
                try:
                    dob = pd.Timestamp(runner_dob)
                    today = pd.Timestamp.now()
                    runner_age = int((today - dob).days / 365.25)
                except:
                    runner_age = runner_age_default
            else:
                runner_age = runner_age_default
            
            pred_5k_ag = calc_age_grade(pred_5k, 5.0, runner_age, runner_gender, 'road')
            if pred_5k_ag:
                dfm.at[dfm.index[-1], 'pred_5k_age_grade'] = round(pred_5k_ag, 1)
                print(f"  Predicted 5K Age Grade: {pred_5k_ag:.1f}% (age {runner_age})")
        
        # Calculate age grades for races and parkruns
        print("  Calculating age grades for races/parkruns...")
        runner_dob = getattr(args, 'runner_dob', None)
        runner_age_default = getattr(args, 'runner_age', 55)
        
        # Parse DOB if provided
        if runner_dob:
            try:
                dob = pd.Timestamp(runner_dob)
            except:
                dob = None
                print(f"  Warning: Could not parse DOB '{runner_dob}', using default age {runner_age_default}")
        else:
            dob = None
        
        age_grade_count = 0
        for i, row in dfm.iterrows():
            # Skip if no time data
            moving_s = row.get('moving_time_s')
            if pd.isna(moving_s) or moving_s <= 0:
                continue
            
            # Determine if this is a race or parkrun
            is_race = bool(row.get('race_flag', 0))
            is_parkrun = bool(row.get('parkrun', 0)) or bool(row.get('hf_parkrun', 0))
            
            if not (is_race or is_parkrun):
                continue
            
            # Get distance (prefer official distance for races)
            official_dist = row.get('official_distance_km')
            dist_km = official_dist if pd.notna(official_dist) and official_dist > 0 else row.get('distance_km', 0)
            
            if not (dist_km and dist_km > 0):
                continue
            
            # Determine surface
            surface = str(row.get('surface', '')).upper().strip()
            if surface in ('TRACK', 'INDOOR_TRACK'):
                surface_type = surface.lower()
            else:
                surface_type = 'road'  # Default to road for parkruns and road races
            
            # Calculate age at race date
            race_date = row.get('date')
            if dob and pd.notna(race_date):
                race_ts = pd.Timestamp(race_date)
                age_at_race = int((race_ts - dob).days / 365.25)
            elif pd.notna(race_date):
                # Approximate from default age and year difference
                race_year = pd.Timestamp(race_date).year
                age_at_race = runner_age_default - (2026 - race_year)
            else:
                age_at_race = runner_age_default
            
            ag = calc_age_grade(moving_s, dist_km, age_at_race, runner_gender, surface_type)
            if ag:
                dfm.at[i, 'age_grade_pct'] = round(ag, 2)
                age_grade_count += 1
        
        print(f"  Calculated {age_grade_count} age grades")
    else:
        print("  Age grade calculations skipped (module not available)")
    
    # Round v43 columns
    for c in ("Temp_Adj", "Terrain_Adj", "Elevation_Adj", "Era_Adj", "Total_Adj", "Intensity_Adj", "Duration_Adj"):
        if c in dfm.columns:
            dfm[c] = pd.to_numeric(dfm[c], errors="coerce").round(4)
    for c in ("RF_adj", "RF_Trend", "RFL", "RFL_Trend"):
        if c in dfm.columns:
            dfm[c] = pd.to_numeric(dfm[c], errors="coerce").round(4)
    for c in ("Factor", "TSS", "CTL", "ATL", "TSB"):
        if c in dfm.columns:
            dfm[c] = pd.to_numeric(dfm[c], errors="coerce").round(1)
    for c in ("CP", "pred_5k_s", "pred_10k_s", "pred_hm_s", "pred_marathon_s"):
        if c in dfm.columns:
            dfm[c] = pd.to_numeric(dfm[c], errors="coerce").round(0)
    for c in ("pred_5k_age_grade", "age_grade_pct"):
        if c in dfm.columns:
            dfm[c] = pd.to_numeric(dfm[c], errors="coerce").round(2)
    # v51: Round Easy RF metrics
    for c in ("Easy_RF_EMA",):
        if c in dfm.columns:
            dfm[c] = pd.to_numeric(dfm[c], errors="coerce").round(4)
    for c in ("Easy_RF_z", "Easy_RFL_Gap"):
        if c in dfm.columns:
            dfm[c] = pd.to_numeric(dfm[c], errors="coerce").round(3)
    for c in ("RFL_Trend_Delta",):
        if c in dfm.columns:
            dfm[c] = pd.to_numeric(dfm[c], errors="coerce").round(4)
    for c in ("weight_kg",):
        if c in dfm.columns:
            dfm[c] = pd.to_numeric(dfm[c], errors="coerce").round(1)

    # --- v38.1: remove legacy/ambiguous HR correction flag column ---
    if "hr_corr_applied" in dfm.columns:
        dfm.drop(columns=["hr_corr_applied"], inplace=True)

    # --- v41/v43: Canonical column ordering ---
    # Ensures consistent column positions regardless of when columns were added
    # v43: New columns added at END to avoid breaking BFW column references
    CANONICAL_ORDER = [
        # Core identification
        "date", "activity_name", "shoe", "notes",
        # Distance metrics
        "distance_km", "official_distance_km", "distance_corrected", "gps_distance_error_m",
        "gps_distance_km", "gps_coverage", "gps_distance_ratio", "gps_max_seg_m", 
        "gps_p99_speed_mps", "gps_outlier_frac",
        # Time metrics
        "elapsed_time_s", "moving_time_s", "avg_pace_min_per_km",
        # Power metrics
        "avg_power_w", "power_mean_w", "power_median_w", "npower_w", 
        "avg_air_power_wkg", "avg_air_power_w",
        # HR metrics
        "avg_hr", "max_hr", "hr_corrected", "hr_corr_type", "hr_corr_drop_s", "hr_cv_pct",
        # Efficiency metrics
        "RE_avg", "RE_normalised", "nPower_HR",
        # Calibration
        "calibration_era_id", "power_adjuster_to_S4",
        # Elevation
        "elev_gain_m", "elev_loss_m", "grade_mean_pct",
        # Terrain metrics (v44)
        "grade_std_pct", "elev_gain_per_km", "reversals_per_km", "undulation_score",
        # Hill power metrics (v45) - distinguish intervals from rolling terrain
        "uphill_power_mean", "downhill_power_mean", "uphill_downhill_ratio", "uphill_pct", "flat_power_mean",
        # RF window terrain metrics (v45) - terrain in RF measurement window
        "rf_window_undulation_score", "rf_window_grade_std_pct", "rf_window_uphill_downhill_ratio", "rf_window_RE",
        # RF metrics (raw)
        "RF_window_start_s", "RF_window_end_s", "RF_window_shifted", "RF_select_code",
        "RF_adjusted_mean_W_per_bpm", "RF_adjusted_median_W_per_bpm",
        "RF_dead_frac", "RF_drift_pct_per_min", "RF_drift_r2", "RF_r2",
        "full_run_drift_pct_per_min", "full_run_drift_r2", "full_run_drift_duration_min",
        # Weather
        "avg_temp_c", "avg_humidity_pct", "avg_wind_ms", "avg_wind_dir_deg",
        # Source tracking
        "time_source", "elev_source", "alt_quality", "power_source", 
        "sim_model_id", "sim_ok", "sim_coverage", "sim_re_med",
        # Flags (v42)
        "race_flag", "surface_adj", "hf_parkrun",
        # Strava data
        "strava_match_type", "strava_activity_id", "strava_distance_km",
        "strava_moving_time_s", "strava_elapsed_time_s", 
        "strava_elev_gain_m", "strava_elev_loss_m",
        # Device info
        "stryd_manufacturer", "stryd_product", "stryd_serial_number", "stryd_ant_device_number",
        # File
        "file",
        # v43 NEW columns (at end to preserve BFW column references)
        "Temp_Adj", "Temp_Trend", "Terrain_Adj", "Elevation_Adj", "Era_Adj", "Total_Adj", "Intensity_Adj", "Duration_Adj",
        "RF_adj", "Factor", "RF_Trend", "RFL", "RFL_Trend",
        # v44.3: Power Score and Combined RFL
        "Power_Score", "Power_Score_Decayed", "RFL_from_Power", "RFL_Combined", "RF_Combined",
        "TSS", "CTL", "ATL", "TSB",
        "weight_kg",
        # v51: Easy RF metrics and alerts
        "Easy_RF_EMA", "Easy_RF_z", "RFL_Trend_Delta", "Easy_RFL_Gap",
        "Alert_1", "Alert_1b", "Alert_2", "Alert_3b", "Alert_5",
        "parkrun", "surface",
        # v43 Age grade and race predictions
        "CP", "pred_5k_s", "pred_10k_s", "pred_hm_s", "pred_marathon_s",
        "pred_5k_age_grade", "age_grade_pct",
    ]
    
    # Reorder: canonical columns first (in order), then any extras at the end
    final_cols = [c for c in CANONICAL_ORDER if c in dfm.columns]
    extra_cols = [c for c in dfm.columns if c not in CANONICAL_ORDER]
    if extra_cols:
        print(f"  Note: {len(extra_cols)} extra columns not in canonical order: {extra_cols}")
    dfm = dfm[final_cols + extra_cols]

    # --- v41: Write Excel with formatting ---
    # v44: Include summary sheets (Daily, Weekly, Monthly, Yearly)
    _write_formatted_excel(dfm, args.out, 
                           daily_df=daily_df, 
                           weekly_df=weekly_df, 
                           monthly_df=monthly_df, 
                           yearly_df=yearly_df)

    print(f"\nWrote: {args.out}")
    print(f"Updated rows: {updated}")
    print(f"Missing cache: {missing_cache}")
    print(f"Model id: {mid}")
    print(f"Override file matches: {override_matches}")
    print(f"Distance corrections applied: {distance_corrections}")
    
    # v43: Print rolling metrics summary
    if 'RF_Trend' in dfm.columns:
        latest_rf_trend = dfm['RF_Trend'].iloc[-1] if len(dfm) > 0 else np.nan
        latest_rfl = dfm['RFL_Trend'].iloc[-1] if len(dfm) > 0 else np.nan
        print(f"Latest RF_Trend: {latest_rf_trend:.4f}" if np.isfinite(latest_rf_trend) else "Latest RF_Trend: N/A")
        print(f"Latest RFL_Trend: {latest_rfl*100:.1f}%" if np.isfinite(latest_rfl) else "Latest RFL_Trend: N/A")
    if 'CTL' in dfm.columns:
        latest_ctl = dfm['CTL'].iloc[-1] if len(dfm) > 0 else np.nan
        latest_atl = dfm['ATL'].iloc[-1] if len(dfm) > 0 else np.nan
        latest_tsb = dfm['TSB'].iloc[-1] if len(dfm) > 0 else np.nan
        print(f"Latest CTL={latest_ctl:.1f}, ATL={latest_atl:.1f}, TSB={latest_tsb:.1f}" if np.isfinite(latest_ctl) else "CTL/ATL: N/A")
    # v51: Easy RF summary
    if 'Easy_RF_EMA' in dfm.columns:
        latest_easy = dfm['Easy_RF_EMA'].iloc[-1] if len(dfm) > 0 else np.nan
        latest_gap = dfm['Easy_RFL_Gap'].iloc[-1] if 'Easy_RFL_Gap' in dfm.columns and len(dfm) > 0 else np.nan
        if np.isfinite(latest_easy):
            parts = [f"Easy_RF_EMA={latest_easy:.4f}"]
            if np.isfinite(latest_gap):
                parts.append(f"Gap={latest_gap*100:.1f}%")
            print(f"Latest {', '.join(parts)}")
    
    return 0


def _write_formatted_excel(df: pd.DataFrame, output_path: str,
                           daily_df: pd.DataFrame = None,
                           weekly_df: pd.DataFrame = None,
                           monthly_df: pd.DataFrame = None,
                           yearly_df: pd.DataFrame = None) -> None:
    """Write DataFrame to Excel with formatting and summary sheets."""
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # --- Master sheet ---
    ws = wb.active
    ws.title = "Master"
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Column widths: A-D = 25, rest = 12
    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        if col_idx <= 4:
            ws.column_dimensions[col_letter].width = 25
        else:
            ws.column_dimensions[col_letter].width = 12
    
    # Freeze top row
    ws.freeze_panes = "A2"
    
    # Zoom 70%
    ws.sheet_view.zoomScale = 70
    
    # --- Helper function to write summary sheets ---
    def write_summary_sheet(wb, sheet_name, summary_df, date_col_width=12):
        if summary_df is None or len(summary_df) == 0:
            return
        
        ws = wb.create_sheet(title=sheet_name)
        
        for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Column widths
        for col_idx in range(1, len(summary_df.columns) + 1):
            col_letter = get_column_letter(col_idx)
            if col_idx == 1:  # Date column
                ws.column_dimensions[col_letter].width = date_col_width
            else:
                ws.column_dimensions[col_letter].width = 12
        
        # Freeze top row
        ws.freeze_panes = "A2"
        
        # Zoom 85%
        ws.sheet_view.zoomScale = 85
    
    # --- Write summary sheets ---
    write_summary_sheet(wb, "Daily", daily_df, date_col_width=12)
    write_summary_sheet(wb, "Weekly", weekly_df, date_col_width=12)
    write_summary_sheet(wb, "Monthly", monthly_df, date_col_width=12)
    write_summary_sheet(wb, "Yearly", yearly_df, date_col_width=8)
    
    wb.save(output_path)
    print(f"  Wrote sheets: Master, Daily, Weekly, Monthly, Yearly")


if __name__ == "__main__":
    raise SystemExit(main())
