#!/usr/bin/env python3
"""
add_run.py - Add one or more FIT files and run pipeline UPDATE.

Usage:
    python add_run.py file.FIT "Name" "Shoe" "FLAGS"
    python add_run.py file1.FIT "Name 1" "Shoe 1" file2.FIT "Name 2" "Shoe 2" ...
    python add_run.py file1.FIT file2.FIT file3.FIT           (no names, generic)
    python add_run.py                                          (interactive mode)
    
With flags:
    python add_run.py warmup.FIT "Warmup" "Nike" race.FIT "Marathon" "Vaporfly" "RACE,42.195"
    python add_run.py parkrun.FIT "Parkrun" "Pegasus" "PARKRUN,5.0"
    
The script detects FIT files by extension and treats following non-FIT arguments
as name, shoe, and flags until the next FIT file.

Flag values (comma-separated):
    RACE         - Mark as race (race_flag=1)
    PARKRUN      - Mark as parkrun (implies RACE)
    TRACK        - Track surface
    INDOOR_TRACK - Indoor track surface
    TRAIL        - Trail surface
    SNOW         - Snow conditions
    HEAVY_SNOW   - Heavy snow conditions
    <number>     - Official distance in km (e.g., 5.0, 42.195)

Examples:
    python add_run.py run.FIT "Easy 10k" "Nike Pegasus"
    python add_run.py run.FIT "Marathon" "Vaporfly" "RACE,42.195"
    python add_run.py warmup.FIT "Warmup" race.FIT "London Marathon" "Vaporfly" "RACE,42.195"
    python add_run.py *.FIT
"""

import os
import sys
import shutil
import zipfile
import subprocess
from pathlib import Path
from datetime import datetime

# ============================================================================
# CONFIGURATION - Edit these paths to match your setup
# ============================================================================
TOTAL_HISTORY_FOLDER = r"C:\Users\paul.collyer\Dropbox\Running and Cycling\DataPipeline\TotalHistory"
TOTAL_HISTORY_ZIP = r"C:\Users\paul.collyer\Dropbox\Running and Cycling\DataPipeline\TotalHistory.zip"
PENDING_ACTIVITIES_CSV = r"C:\Users\paul.collyer\Dropbox\Running and Cycling\DataPipeline\pending_activities.csv"
OVERRIDE_FILE = r"C:\Users\paul.collyer\Dropbox\Running and Cycling\DataPipeline\activity_overrides.xlsx"
PIPELINE_BAT = r"C:\Users\paul.collyer\Dropbox\Running and Cycling\DataPipeline\Run_Full_Pipeline.bat"
STAGING_FOLDER = r"C:\Users\paul.collyer\Dropbox\Running and Cycling\DataPipeline\NewRuns"

# Valid surface and flag values
VALID_SURFACES = {'TRACK', 'INDOOR_TRACK', 'TRAIL', 'SNOW', 'HEAVY_SNOW'}
FLAG_KEYWORDS = {'RACE', 'PARKRUN'} | VALID_SURFACES


def ensure_pending_csv_exists():
    """Create pending_activities.csv if it doesn't exist."""
    if not os.path.exists(PENDING_ACTIVITIES_CSV):
        with open(PENDING_ACTIVITIES_CSV, 'w', encoding='utf-8') as f:
            f.write("file,activity_name,shoe\n")
        print(f"Created: {PENDING_ACTIVITIES_CSV}")


def add_to_pending_activities(fit_filename: str, activity_name: str, shoe: str = ""):
    """Add entry to pending_activities.csv."""
    if not activity_name:
        return
        
    ensure_pending_csv_exists()
    
    try:
        with open(PENDING_ACTIVITIES_CSV, 'r', encoding='utf-8') as f:
            lines = f.readlines()
    except UnicodeDecodeError:
        with open(PENDING_ACTIVITIES_CSV, 'r', encoding='cp1252') as f:
            lines = f.readlines()
    
    # Remove existing entry for this file
    new_lines = [lines[0]]
    for line in lines[1:]:
        if not line.startswith(f"{fit_filename},"):
            new_lines.append(line)
    
    # Escape commas
    name_out = f'"{activity_name}"' if ',' in activity_name else activity_name
    shoe_out = f'"{shoe}"' if shoe and ',' in shoe else shoe
    
    new_lines.append(f"{fit_filename},{name_out},{shoe_out}\n")
    
    with open(PENDING_ACTIVITIES_CSV, 'w', encoding='utf-8') as f:
        f.writelines(new_lines)


def is_flags_string(arg: str) -> bool:
    """Check if argument looks like a flags string."""
    if not arg:
        return False
    
    parts = [p.strip().upper() for p in arg.split(',')]
    for part in parts:
        if part in FLAG_KEYWORDS:
            return True
        try:
            val = float(part)
            if 0.1 < val < 500:
                return True
        except ValueError:
            pass
    return False


def parse_flags(flags_str: str) -> dict:
    """Parse comma-separated flags string.
    
    Examples:
        RACE
        PARKRUN
        RACE,5.0
        TRAIL,0.95  (surface with surface_adj)
        RACE,TRAIL,0.95,5.0  (race, trail surface, 0.95 adj, 5km official)
    """
    result = {'race_flag': 0, 'parkrun': 0, 'surface': '', 'surface_adj': None, 'official_distance_km': None}
    
    if not flags_str:
        return result
    
    for flag in [f.strip().upper() for f in flags_str.split(',')]:
        if flag == 'RACE':
            result['race_flag'] = 1
        elif flag == 'PARKRUN':
            result['race_flag'] = 1
            result['parkrun'] = 1
            if result['official_distance_km'] is None:
                result['official_distance_km'] = 5.0
        elif flag in VALID_SURFACES:
            result['surface'] = flag
        elif flag.startswith('ADJ=') or flag.startswith('SADJ='):
            # surface_adj=0.95 or adj=0.95
            try:
                result['surface_adj'] = float(flag.split('=')[1])
            except (ValueError, IndexError):
                print(f"WARNING: Invalid surface_adj '{flag}' - ignoring")
        else:
            try:
                val = float(flag)
                # Distinguish between surface_adj (0.8-1.2) and distance (>1.5)
                if 0.5 < val < 1.5:
                    result['surface_adj'] = val
                elif val > 1.5:
                    result['official_distance_km'] = val
            except ValueError:
                print(f"WARNING: Unknown flag '{flag}' - ignoring")
    
    return result


def add_to_overrides(fit_filename: str, flags: dict):
    """Add entry to activity_overrides.xlsx."""
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import numbers
    
    # Skip if no meaningful flags
    if (flags['race_flag'] == 0 and flags['surface'] == '' and 
        flags['official_distance_km'] is None and flags.get('surface_adj') is None):
        return
    
    # Ensure filename has .FIT/.fit extension (guard against downloads without extension)
    if not fit_filename.lower().endswith('.fit'):
        fit_filename = fit_filename + '.FIT'
    
    if not os.path.exists(OVERRIDE_FILE):
        df = pd.DataFrame(columns=['file', 'race_flag', 'parkrun', 'official_distance_km', 
                                   'surface', 'surface_adj', 'temp_override', 'notes'])
    else:
        df = pd.read_excel(OVERRIDE_FILE, dtype={'file': str})
    
    df = df[df['file'] != fit_filename]
    
    new_row = pd.DataFrame([{
        'file': fit_filename,
        'race_flag': flags['race_flag'],
        'parkrun': flags['parkrun'],
        'official_distance_km': flags['official_distance_km'],
        'surface': flags['surface'],
        'surface_adj': flags.get('surface_adj'),
        'temp_override': None,
        'notes': f"Added via add_run.py on {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    }])
    
    if len(df) == 0:
        df = new_row
    else:
        df = pd.concat([df, new_row], ignore_index=True)
    df.to_excel(OVERRIDE_FILE, index=False)
    
    # Force 'file' column to Text format so Excel won't auto-interpret date-like filenames
    wb = load_workbook(OVERRIDE_FILE)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        cell.number_format = '@'  # Text format
    wb.save(OVERRIDE_FILE)
    
    parts = []
    if flags['race_flag']: parts.append('RACE')
    if flags['parkrun']: parts.append('PARKRUN')
    if flags['surface']: parts.append(f"surface={flags['surface']}")
    if flags.get('surface_adj'): parts.append(f"surface_adj={flags['surface_adj']}")
    if flags['official_distance_km']: parts.append(f"distance={flags['official_distance_km']}km")
    print(f"  -> overrides: {fit_filename} [{', '.join(parts)}]")


def find_fit_file(fit_path: str) -> str:
    """Find FIT file - checks staging folder if not found directly."""
    if os.path.exists(fit_path):
        return fit_path
    
    staging_path = os.path.join(STAGING_FOLDER, os.path.basename(fit_path))
    if os.path.exists(staging_path):
        return staging_path
    
    if STAGING_FOLDER and os.path.isdir(STAGING_FOLDER):
        for f in os.listdir(STAGING_FOLDER):
            if f.lower() == os.path.basename(fit_path).lower():
                return os.path.join(STAGING_FOLDER, f)
    
    return fit_path


def copy_fit_to_history(fit_path: str) -> str:
    """Copy FIT file to TotalHistory folder. Returns the filename."""
    fit_path = Path(fit_path)
    if not fit_path.exists():
        raise FileNotFoundError(f"FIT file not found: {fit_path}")
    
    dest = Path(TOTAL_HISTORY_FOLDER) / fit_path.name
    shutil.copy2(fit_path, dest)
    return fit_path.name


def recreate_zip():
    """Delete and recreate TotalHistory.zip from all FIT files including subfolders."""
    if os.path.exists(TOTAL_HISTORY_ZIP):
        os.remove(TOTAL_HISTORY_ZIP)
    
    folder = Path(TOTAL_HISTORY_FOLDER)
    # Use rglob to recursively find FIT files in all subfolders (era folders)
    fit_files_raw = list(folder.rglob("*.fit")) + list(folder.rglob("*.FIT"))
    # Deduplicate (Windows is case-insensitive, both globs return same files)
    seen_paths = set()
    fit_files = []
    for f in fit_files_raw:
        key = str(f).lower()
        if key not in seen_paths:
            seen_paths.add(key)
            fit_files.append(f)
    
    # Track filenames to avoid duplicates (same filename in different subfolders)
    seen_names = set()
    duplicates = 0
    
    with zipfile.ZipFile(TOTAL_HISTORY_ZIP, 'w', zipfile.ZIP_DEFLATED) as zf:
        for fit_file in fit_files:
            # Store with just filename (no subfolder path) to maintain flat structure in zip
            name = fit_file.name
            if name in seen_names:
                duplicates += 1
                # Skip duplicate - first one wins
                continue
            seen_names.add(name)
            zf.write(fit_file, name)
    
    with zipfile.ZipFile(TOTAL_HISTORY_ZIP, 'r') as zf:
        print(f"Created zip: {len(zf.namelist())} files")
    
    if duplicates > 0:
        print(f"  Warning: Skipped {duplicates} duplicate filenames (same file in multiple folders)")


def run_pipeline_update():
    """Run the pipeline UPDATE."""
    print("\n" + "="*60)
    print("Running pipeline UPDATE...")
    print("="*60 + "\n")
    
    return subprocess.run([PIPELINE_BAT, "UPDATE"], shell=True, 
                         cwd=os.path.dirname(PIPELINE_BAT)).returncode


def is_fit_file(arg: str) -> bool:
    """Check if argument is a FIT file."""
    return arg.lower().endswith('.fit') and os.path.exists(find_fit_file(arg))


def parse_args(args: list) -> list:
    """Parse arguments into list of (fit_path, name, shoe, flags) tuples."""
    runs = []
    i = 0
    
    while i < len(args):
        if is_fit_file(args[i]):
            fit_path = find_fit_file(args[i])
            name, shoe, flags = "", "", ""
            
            # Collect args until next FIT file
            j = i + 1
            non_fit_args = []
            while j < len(args) and not is_fit_file(args[j]):
                non_fit_args.append(args[j])
                j += 1
            
            # Separate flags from name/shoe
            remaining = []
            for arg in non_fit_args:
                if is_flags_string(arg):
                    flags = arg
                else:
                    remaining.append(arg)
            
            if len(remaining) >= 1: name = remaining[0]
            if len(remaining) >= 2: shoe = remaining[1]
            
            runs.append((fit_path, name, shoe, flags))
            i = j
        else:
            i += 1
    
    return runs


def interactive_mode() -> list:
    """Get runs interactively."""
    runs = []
    print("\nEnter runs (empty path to finish):\n")
    
    while True:
        fit_path = input("FIT file (or Enter to finish): ").strip().strip('"')
        if not fit_path:
            break
        
        fit_path = find_fit_file(fit_path)
        if not os.path.exists(fit_path):
            print(f"  Not found: {fit_path}")
            continue
        
        name = input("  Name (Enter to skip): ").strip()
        shoe = input("  Shoe (Enter to skip): ").strip() if name else ""
        flags = input("  Flags - RACE,PARKRUN,TRACK,<dist> (Enter to skip): ").strip() if name else ""
        
        runs.append((fit_path, name, shoe, flags))
        print()
    
    return runs


def staging_folder_mode() -> list:
    """Check staging folder for FIT files."""
    if not STAGING_FOLDER or not os.path.isdir(STAGING_FOLDER):
        return []
    
    fit_files = list(Path(STAGING_FOLDER).glob("*.fit")) + list(Path(STAGING_FOLDER).glob("*.FIT"))
    if not fit_files:
        return []
    
    print(f"\nFound {len(fit_files)} FIT file(s) in staging folder:")
    for f in fit_files:
        print(f"  {f.name}")
    
    choice = input("\nAdd? (Y=generic, i=interactive, n=cancel): ").strip().lower()
    
    if choice == 'n':
        return []
    elif choice == 'i':
        runs = []
        for fit_file in fit_files:
            print(f"\n{fit_file.name}:")
            name = input("  Name (Enter for generic): ").strip()
            shoe = input("  Shoe (Enter to skip): ").strip() if name else ""
            flags = input("  Flags (Enter to skip): ").strip() if name else ""
            runs.append((str(fit_file), name, shoe, flags))
        return runs
    else:
        return [(str(f), "", "", "") for f in fit_files]


def main():
    print("="*60)
    print("Add Run")
    print("="*60)
    
    # Parse args or interactive
    if len(sys.argv) > 1:
        runs = parse_args(sys.argv[1:])
    else:
        runs = staging_folder_mode()
        if not runs:
            runs = interactive_mode()
    
    if not runs:
        print("\nNo runs to add.")
        return 0
    
    # Show summary
    print(f"\n{len(runs)} run(s) to add:")
    for fit_path, name, shoe, flags in runs:
        parts = [Path(fit_path).name]
        if name: parts.append(f'"{name}"')
        if shoe: parts.append(f'[{shoe}]')
        if flags: parts.append(f'<{flags}>')
        print(f"  {' '.join(parts)}")
    
    if input("\nProceed? (Y/n): ").strip().lower() == 'n':
        return 0
    
    print()
    
    try:
        for fit_path, name, shoe, flags in runs:
            fit_filename = copy_fit_to_history(fit_path)
            print(f"Copied: {fit_filename}")
            
            if name:
                add_to_pending_activities(fit_filename, name, shoe)
                print(f"  -> pending: {name}")
            
            if flags:
                add_to_overrides(fit_filename, parse_flags(flags))
        
        print()
        recreate_zip()
        return run_pipeline_update()
        
    except Exception as e:
        print(f"\nERROR: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
