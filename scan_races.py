#!/usr/bin/env python3
"""
scan_races.py — Detect races from Master xlsx and generate activity_overrides.xlsx.

Uses a multi-signal approach:
  1. Activity name patterns (race keywords boost, training keywords suppress)
  2. Parkrun detection (Saturday 9am, ~5K, name patterns)
  3. Pace-based detection calibrated per distance band
  4. HR confirmation (real races have high HR)

Usage:
    python scan_races.py athletes/PaulTest/output/Master_FULL.xlsx athletes/PaulTest/activity_overrides.xlsx
    python scan_races.py master.xlsx overrides.xlsx 192
"""
import sys
import re
import numpy as np
import pandas as pd
from pathlib import Path

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Activity name patterns ────────────────────────────────────────────────

# Strong race indicators — if present AND pace/HR are consistent, flag as race
RACE_POS_STRONG = re.compile(
    r'\b(?:race|PB|pb|SB|sb|lopp|tävling|varvet|loppet|marathon|halvmarathon'
    r'|half(?!\s+(?:a|an|of|marathon\s+distance|distance))'
    r'|competition|relay|stafett|place|winner|V\d+\s+winner'
    r'|chase\s+the\s+sun|winter\s+series|RunThrough|halloween\s+\d+k'
    r'|midnatts|höstrusket|hässelby|djurgårds|hackney|big\s+half'
    r'|premiärhalven|göteborgsvarvet|reading|windsor|oxford|royal\s+parks'
    r'|copenhagen|stockholm\s+\d+k|victoria\s+park\s+\d+k'
    r'|Stockholms\s+Bästa|LFOTM|LFoTM|assembly\s+league|sri\s+chimnoy'
    r'|veterans\s+AC|sommarspelen|broloppet|spåret'
    r'|BMAF|english\s+athletics)\b',
    re.IGNORECASE
)

# Weak race indicators — suggestive but not sufficient alone
RACE_POS_WEAK = re.compile(
    r'\b(?:event|official|chip\s+time|bib|finisher|medal|trophy)\b',
    re.IGNORECASE
)

# Training suppressors — if present, almost certainly NOT a race
TRAINING_SUPPRESS = re.compile(
    r'\b(?:interval|tempo|threshold|fartlek|session|warmup|warm\s*up|cooldown|cool\s*down'
    r'|WU|CD|recovery|easy|steady|jog|progressive|pyramid|repeats'
    r'|strides|bursts|drill|test|fitness\s*test'
    r'|chat\s+with|relaxed|gentle|comfortable|cruise'
    r'|before\s+PT|after\s+PT|morning\s+.*jog|evening\s+.*jog'
    r'|preamble|pre\-?\s*race|shakeout|mini\s+tempo'
    r'|long\s+run|training\s+run|pre\s+marathon|simulator'
    r'|park\s+to\s+park|yasso'
    r'|marathon\s+effort|at\s+.*effort|sub\s+\d+\s+marathon\s+effort)\b'
    r'|@\s*~?ME\b|@\s*~?HME\b|@\s*~?\d+k\s*E\b'
    r'|\d+\s*[xX×]\s*\d+',
    re.IGNORECASE
)

# Parkrun-specific suppressors
PARKRUN_SUPPRESS = re.compile(
    r'\b(?:buggy|pram|pushchair|jog\s+to|jog\s+from|walk|walked'
    r'|easy\s+post|post\s+parkrun|after\s+parkrun|before\s+parkrun'
    r'|jog\s+back|back\s+from|preamble)\b',
    re.IGNORECASE
)

# Non-run activities that might slip through
NON_RUN_SUPPRESS = re.compile(
    r'\b(?:duathlon|triathlon|cycle|swim|bike|hike|walk(?:ing)?|yoga)\b',
    re.IGNORECASE
)


def _name_score(name, is_parkrun_candidate=False):
    """Return a name-based score: positive = race-like, negative = training-like."""
    if not name:
        return 0.0
    
    score = 0.0
    
    if NON_RUN_SUPPRESS.search(name):
        return -5.0
    
    has_race = bool(RACE_POS_STRONG.search(name))
    has_training = bool(TRAINING_SUPPRESS.search(name))
    
    if has_training:
        score -= 1.5
    
    if has_race:
        score += 1.5
    
    # When both race and training keywords present, lean toward race
    # (e.g. "Hackney Half...steady 16k then hammered it" = race with pacing description)
    if has_race and has_training:
        score += 0.5
    
    if RACE_POS_WEAK.search(name):
        score += 0.5
    
    if is_parkrun_candidate and PARKRUN_SUPPRESS.search(name):
        score -= 2.0
    
    # "parkrun" in name (not suppressed) is a strong indicator — but only for ~5K runs
    # (e.g. "penance for not getting out of bed for parkrun" at 10K is NOT a parkrun)
    if is_parkrun_candidate and 'parkrun' in name.lower() and not PARKRUN_SUPPRESS.search(name):
        score += 1.0
    
    return score


def detect_races_from_master(df, max_hr=None):
    """Auto-detect likely race efforts from master DataFrame."""
    df = df.copy()
    df['date'] = pd.to_datetime(df['date'])
    
    df['_pace'] = pd.to_numeric(df.get('avg_pace_min_per_km'), errors='coerce')
    df['_dist'] = pd.to_numeric(df.get('distance_km'), errors='coerce')
    df['_hr'] = pd.to_numeric(df.get('avg_hr'), errors='coerce')
    df['_file'] = df.get('file', df.get('filename', ''))
    df['_name'] = df.get('activity_name', '')
    
    valid = df[df['_pace'].notna() & df['_dist'].notna() & (df['_pace'] > 2.5)].copy()
    
    if len(valid) < 20:
        print("  Not enough valid runs for pace calibration")
        return []
    
    # ── Pace calibration ──────────────────────────────────────────────
    pace_p05 = valid['_pace'].quantile(0.05)
    pace_p10 = valid['_pace'].quantile(0.10)
    pace_p15 = valid['_pace'].quantile(0.15)
    pace_p20 = valid['_pace'].quantile(0.20)
    pace_p25 = valid['_pace'].quantile(0.25)
    pace_median = valid['_pace'].median()
    
    if max_hr is None:
        max_hr = df['_hr'].quantile(0.98) if df['_hr'].notna().sum() > 50 else 192
    
    print(f"  Pace calibration: p5={pace_p05:.2f}, p10={pace_p10:.2f}, p15={pace_p15:.2f}, p25={pace_p25:.2f}, median={pace_median:.2f}")
    print(f"  Max HR estimate: {max_hr:.0f} bpm")
    
    # ── Distance-specific thresholds ──────────────────────────────────
    # Pace thresholds: what percentile of overall pace to use as cutoff
    # HR thresholds: min HR fraction for race confirmation, and "strong" HR fraction
    race_distances = [
        # (dmin, dmax, official_km, label, pace_threshold, hr_min_frac, hr_boost_frac)
        (2.8, 3.3,  3.0,    "3K",       pace_p15,                0.91, 0.94),
        (4.5, 5.5,  5.0,    "5K",       pace_p15,                0.90, 0.93),
        (9.5, 10.8, 10.0,   "10K",      pace_p15,                0.88, 0.92),
        (14.5, 16.5, 15.0,  "15K",      pace_p20,                0.85, 0.90),
        (20.0, 22.0, 21.097, "HM",      pace_p25,                0.82, 0.88),
        (25.0, 27.5, 26.2,  "26K",      pace_p25,                0.80, 0.85),
        (29.5, 31.5, 30.0,  "30K",      pace_median * 0.92,      0.78, 0.83),
        (41.0, 44.0, 42.195, "Marathon", pace_median * 0.95,      0.75, 0.80),
    ]
    
    print(f"\n  Distance-specific pace thresholds:")
    for _, _, _, label, pace_thr, hr_min, hr_boost in race_distances:
        print(f"    {label:>8}: pace < {pace_thr:.2f}/km, HR > {max_hr * hr_min:.0f} (min) / {max_hr * hr_boost:.0f} (boost)")
    
    races = []
    
    for _, row in valid.iterrows():
        dist = row['_dist']
        pace = row['_pace']
        hr = row['_hr']
        dt = row['date']
        fname = str(row['_file'])
        name = str(row['_name']) if pd.notna(row['_name']) else ''
        
        # ── Parkrun detection ─────────────────────────────────────────
        is_saturday = dt.weekday() == 5
        start_hour = dt.hour + dt.minute / 60
        is_parkrun_time = (9.0 <= start_hour <= 9.17) or (9.5 <= start_hour <= 9.67)
        is_5k_dist = 4.5 <= dist <= 5.5
        
        name_sc = _name_score(name, is_parkrun_candidate=(is_saturday and is_5k_dist))
        
        # Skip non-run activities
        if name_sc <= -4.0:
            continue
        
        is_race = False
        race_distance = None
        race_type = None
        confidence = "low"
        
        # Parkrun: Saturday + time + 5K distance
        if is_saturday and is_parkrun_time and is_5k_dist and name_sc >= -1.0:
            if pace < pace_p10 * 1.05:
                is_race = True
                race_distance = 5.0
                race_type = "parkrun"
                confidence = "high"
            elif pace < pace_p25:
                is_race = True
                race_distance = 5.0
                race_type = "parkrun"
                confidence = "medium" if name_sc >= 0 else "low"
        
        # Name-based parkrun (any day, "parkrun" in name, not suppressed)
        if not is_race and 'parkrun' in name.lower() and is_5k_dist:
            if PARKRUN_SUPPRESS.search(name):
                pass
            elif pace < pace_p10 * 1.05:
                is_race = True
                race_distance = 5.0
                race_type = "parkrun"
                confidence = "high"
            elif pace < pace_p25 and (pd.isna(hr) or hr > max_hr * 0.82):
                is_race = True
                race_distance = 5.0
                race_type = "parkrun"
                confidence = "medium"
        
        # ── Standard race distances ──────────────────────────────────
        if not is_race:
            for dmin, dmax, official, label, pace_thr, hr_min_frac, hr_boost_frac in race_distances:
                if dmin <= dist <= dmax:
                    pace_ok = pace < pace_thr
                    hr_ok = pd.notna(hr) and hr > max_hr * hr_min_frac
                    hr_strong = pd.notna(hr) and hr > max_hr * hr_boost_frac
                    name_positive = name_sc > 0.5
                    name_negative = name_sc < -0.5
                    
                    if name_negative and not name_positive:
                        break  # Training keywords, no race keywords → skip
                    
                    if name_positive and (pace_ok or hr_strong):
                        is_race = True
                        race_distance = official
                        race_type = label
                        confidence = "high" if (pace_ok and hr_strong) else "medium"
                    elif pace_ok and hr_ok:
                        is_race = True
                        race_distance = official
                        race_type = label
                        confidence = "high" if hr_strong else "medium"
                    elif pace_ok and name_sc >= 0 and pd.isna(hr):
                        is_race = True
                        race_distance = official
                        race_type = label
                        confidence = "medium"
                    
                    break
        
        if is_race:
            elapsed = pd.to_numeric(row.get('elapsed_time_s', row.get('moving_time_s', 0)), errors='coerce')
            duration_min = elapsed / 60 if pd.notna(elapsed) and elapsed > 0 else 0
            
            races.append({
                "file": fname,
                "date": dt.strftime('%Y-%m-%d'),
                "distance_km": race_distance,
                "actual_distance_km": round(dist, 3),
                "duration_min": round(duration_min, 1),
                "pace_min_km": round(pace, 2),
                "avg_hr": round(hr) if pd.notna(hr) else None,
                "race_type": race_type,
                "confidence": confidence,
                "activity_name": name[:60],
                "name_score": round(name_sc, 1),
            })
    
    return races


def generate_override_xlsx(path, races=None):
    """Generate activity_overrides.xlsx with detected races pre-populated."""
    if not HAS_OPENPYXL:
        print(f"  openpyxl not installed!")
        return
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "overrides"
    
    headers = [
        "file", "race_flag", "parkrun", "official_distance_km", "surface", "surface_adj",
        "temp_override", "power_override_w", "notes"
    ]
    ws.append(headers)
    
    from openpyxl.styles import Font, PatternFill
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    if races:
        for r in races:
            is_parkrun = r.get("race_type") == "parkrun"
            ws.append([
                r["file"],
                1,
                1 if is_parkrun else 0,
                r["distance_km"],
                None, None, None, None,
                f"Auto: {r['race_type']} ({r['confidence']}, ns={r.get('name_score', '?')}) — "
                f"{r['duration_min']:.1f}min, {r['pace_min_km']:.2f}/km, HR {r.get('avg_hr', '?')}  "
                f"{r.get('activity_name', '')}"
            ])
    
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 60)
    
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main():
    if len(sys.argv) < 3:
        print("Usage: python scan_races.py <master.xlsx> <output_overrides.xlsx> [max_hr]")
        print("  master.xlsx: Master_FULL.xlsx or Master_FULL_post.xlsx")
        print("  output_overrides.xlsx: where to write the overrides file")
        print("  max_hr: optional, athlete's max heart rate (auto-estimated if omitted)")
        sys.exit(1)
    
    master_path = sys.argv[1]
    output_path = sys.argv[2]
    max_hr = float(sys.argv[3]) if len(sys.argv) > 3 else None
    
    print(f"Loading master: {master_path}")
    df = pd.read_excel(master_path, sheet_name=0)
    print(f"  {len(df)} runs loaded")
    
    races = detect_races_from_master(df, max_hr=max_hr)
    
    high = [r for r in races if r["confidence"] == "high"]
    med = [r for r in races if r["confidence"] == "medium"]
    low = [r for r in races if r["confidence"] == "low"]
    
    print(f"\nDetected {len(races)} potential races:")
    print(f"  High confidence: {len(high)}")
    print(f"  Medium confidence: {len(med)}")
    print(f"  Low confidence: {len(low)}")
    
    from collections import Counter
    type_counts = Counter(r["race_type"] for r in races)
    print(f"\nBy type: {dict(type_counts)}")
    
    print(f"\n{'Conf':>4}  {'Date':>10}  {'Type':>8}  {'Dist':>5}  {'Pace':>5}  {'HR':>3}  {'NS':>3}  Name")
    print(f"{'----':>4}  {'----------':>10}  {'--------':>8}  {'-----':>5}  {'-----':>5}  {'---':>3}  {'---':>3}  ----")
    for r in sorted(races, key=lambda x: x['date']):
        conf_icon = {"high": "✓", "medium": "~", "low": "?"}[r["confidence"]]
        hr_str = str(r['avg_hr']) if r['avg_hr'] else '?'
        ns = r.get('name_score', 0)
        ns_str = f"{ns:+.0f}" if ns != 0 else " 0"
        print(f"   {conf_icon}  {r['date']}  {r['race_type']:>8}  "
              f"{r['actual_distance_km']:5.1f}  {r['pace_min_km']:5.2f}  {hr_str:>3}  {ns_str:>3}  "
              f"{r.get('activity_name', '')[:50]}")
    
    races_to_write = [r for r in races if r["confidence"] in ("high", "medium")]
    print(f"\nWriting {len(races_to_write)} races (high+medium) to {output_path}")
    generate_override_xlsx(output_path, races_to_write)
    print("Done!")


if __name__ == "__main__":
    main()
