#!/usr/bin/env python3
"""
onboard_athlete.py — Generate all files needed for a new athlete.

Usage:
    python onboard_athlete.py                          # Interactive mode
    python onboard_athlete.py --config new_athlete.json # From JSON file
    python onboard_athlete.py --scan-fits path/to/fits  # Scan FIT files for HR/race hints

Generates:
    athletes/<FolderName>/athlete.yml          — athlete configuration
    athletes/<FolderName>/activity_overrides.xlsx — empty override file (with races if scanned)
    athletes/<FolderName>/athlete_data.csv     — empty athlete data file
    .github/workflows/<slug>_pipeline.yml      — GitHub Actions workflow

Also prints:
    - Dropbox folder setup commands
    - GitHub secrets to configure
    - First-run instructions
"""

import argparse
import json
import os
import re
import sys
import textwrap
from datetime import datetime, date
from pathlib import Path

# ─── Optional imports for FIT scanning ────────────────────────────────────
try:
    from fitparse import FitFile
    HAS_FITPARSE = True
except ImportError:
    HAS_FITPARSE = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import numpy as np
    HAS_NUMPY = True
except ImportError:
    HAS_NUMPY = False


# ═════════════════════════════════════════════════════════════════════════════
# FIT SCANNING — extract HR stats and auto-detect races
# ═════════════════════════════════════════════════════════════════════════════

def scan_fit_file(fit_path: str) -> dict | None:
    """Extract summary stats from a single FIT file. Returns None if not a run.
    Auto-detects gzip compression (Strava exports use .fit.gz or .fit that are gzipped)."""
    import gzip
    import tempfile
    
    actual_path = fit_path
    tmp_path = None
    
    try:
        # Check if file is gzip-compressed (magic bytes 1f 8b)
        with open(fit_path, 'rb') as f:
            magic = f.read(2)
        
        if magic == b'\x1f\x8b':
            # Decompress to temp file
            tmp_fd, tmp_path = tempfile.mkstemp(suffix='.fit')
            os.close(tmp_fd)
            with gzip.open(fit_path, 'rb') as gz_in:
                with open(tmp_path, 'wb') as f_out:
                    import shutil
                    shutil.copyfileobj(gz_in, f_out)
            actual_path = tmp_path
        
        fit = FitFile(actual_path)
        # Get session data
        sessions = list(fit.get_messages("session"))
        if not sessions:
            return None
        sess = sessions[0]
        
        sport = sess.get_value("sport")
        if sport is not None and str(sport).lower() not in ("running", "trail_running"):
            return None
        
        def get(field):
            try:
                return sess.get_value(field)
            except:
                return None
        
        avg_hr = get("avg_heart_rate")
        max_hr = get("max_heart_rate")
        distance = get("total_distance")  # metres
        elapsed = get("total_elapsed_time")  # seconds
        timestamp = get("start_time")  # datetime
        avg_speed = get("enhanced_avg_speed") or get("avg_speed")  # m/s
        avg_power = get("avg_power")
        total_ascent = get("total_ascent")
        
        if distance is None or elapsed is None or elapsed < 60:
            return None
        
        distance_km = distance / 1000.0
        duration_min = elapsed / 60.0
        pace_min_km = duration_min / distance_km if distance_km > 0.1 else None
        
        filename = os.path.basename(fit_path)
        
        return {
            "file": filename,
            "date": timestamp.strftime("%Y-%m-%d") if timestamp else None,
            "timestamp": timestamp,
            "distance_km": round(distance_km, 3),
            "duration_min": round(duration_min, 1),
            "pace_min_km": round(pace_min_km, 2) if pace_min_km else None,
            "avg_hr": int(avg_hr) if avg_hr else None,
            "max_hr": int(max_hr) if max_hr else None,
            "avg_power": int(avg_power) if avg_power else None,
            "total_ascent_m": round(total_ascent, 0) if total_ascent else None,
            "fit_path": fit_path,
        }
    except Exception as e:
        return None
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except:
                pass


def scan_fit_folder(folder: str, verbose: bool = True) -> list[dict]:
    """Scan all FIT files in a folder (or zip). Returns list of run summaries."""
    import zipfile
    import tempfile
    import gzip
    import shutil
    
    runs = []
    fit_files = []
    
    # Handle zip file
    if folder.endswith(".zip") and os.path.isfile(folder):
        tmpdir = tempfile.mkdtemp(prefix="fit_scan_")
        if verbose:
            print(f"Extracting {folder}...")
        with zipfile.ZipFile(folder, "r") as zf:
            members = [m for m in zf.namelist() if m.upper().endswith(".FIT") or m.upper().endswith(".FIT.GZ")]
            if verbose:
                print(f"  Found {len(members)} FIT files in zip")
            for m in members:
                zf.extract(m, tmpdir)
                extracted = os.path.join(tmpdir, m)
                if extracted.upper().endswith(".FIT.GZ"):
                    # Decompress .fit.gz → .fit
                    fit_path = extracted[:-3]  # strip .gz
                    with gzip.open(extracted, 'rb') as gz_in:
                        with open(fit_path, 'wb') as f_out:
                            shutil.copyfileobj(gz_in, f_out)
                    os.remove(extracted)
                    fit_files.append(fit_path)
                else:
                    fit_files.append(extracted)
    else:
        # Walk directory
        for root, dirs, files in os.walk(folder):
            for f in files:
                fpath = os.path.join(root, f)
                if f.upper().endswith(".FIT"):
                    fit_files.append(fpath)
                elif f.upper().endswith(".FIT.GZ"):
                    # Decompress in place
                    fit_path = fpath[:-3]
                    try:
                        with gzip.open(fpath, 'rb') as gz_in:
                            with open(fit_path, 'wb') as f_out:
                                shutil.copyfileobj(gz_in, f_out)
                        fit_files.append(fit_path)
                    except Exception:
                        pass
        if verbose:
            print(f"Found {len(fit_files)} FIT files")
    
    total = len(fit_files)
    for i, fp in enumerate(sorted(fit_files)):
        if verbose and (i + 1) % 100 == 0:
            print(f"  Scanned {i+1}/{total}...")
        result = scan_fit_file(fp)
        if result:
            runs.append(result)
    
    if verbose:
        print(f"  → {len(runs)} running activities found")
    
    runs.sort(key=lambda r: r.get("timestamp") or datetime.min)
    return runs


def estimate_hr_zones(runs: list[dict]) -> dict:
    """Estimate LTHR and max HR from run data."""
    max_hrs = [r["max_hr"] for r in runs if r.get("max_hr") and r["max_hr"] > 100]
    avg_hrs = [r["avg_hr"] for r in runs if r.get("avg_hr") and r["avg_hr"] > 80]
    
    if not max_hrs:
        return {"max_hr": None, "lthr": None, "confidence": "none", "note": "No HR data found"}
    
    # Max HR: take 95th percentile of max HR values (avoids single-spike outliers)
    if HAS_NUMPY:
        max_hr = int(np.percentile(max_hrs, 95))
    else:
        sorted_hrs = sorted(max_hrs)
        idx = int(len(sorted_hrs) * 0.95)
        max_hr = sorted_hrs[min(idx, len(sorted_hrs) - 1)]
    
    # LTHR estimation: look at runs that are likely tempo/threshold efforts
    # These are runs 20-60 min, moderate-to-hard effort (avg HR > 75% of max)
    threshold_hrs = []
    for r in runs:
        if (r.get("avg_hr") and r.get("duration_min") 
            and 20 <= r["duration_min"] <= 60
            and r["avg_hr"] > max_hr * 0.75):
            threshold_hrs.append(r["avg_hr"])
    
    if threshold_hrs:
        if HAS_NUMPY:
            lthr = int(np.percentile(threshold_hrs, 85))
        else:
            sorted_th = sorted(threshold_hrs)
            idx = int(len(sorted_th) * 0.85)
            lthr = sorted_th[min(idx, len(sorted_th) - 1)]
        confidence = "moderate" if len(threshold_hrs) > 10 else "low"
    else:
        # Fallback: 89% of max HR (Karvonen approximation)
        lthr = int(max_hr * 0.89)
        confidence = "estimated"
    
    n_with_hr = len(avg_hrs)
    n_total = len(runs)
    
    return {
        "max_hr": max_hr,
        "lthr": lthr,
        "confidence": confidence,
        "n_with_hr": n_with_hr,
        "n_total": n_total,
        "note": f"Based on {n_with_hr}/{n_total} runs with HR data. "
                f"Confidence: {confidence}. Review and adjust if needed."
    }


def detect_races(runs: list[dict]) -> list[dict]:
    """Auto-detect likely race efforts from run data."""
    if not runs:
        return []
    
    # Build pace distribution for calibration
    paces = [r["pace_min_km"] for r in runs if r.get("pace_min_km") and r["pace_min_km"] > 2.5]
    if not paces or len(paces) < 20:
        return []
    
    if HAS_NUMPY:
        pace_p10 = np.percentile(paces, 10)  # fast end
        pace_median = np.median(paces)
    else:
        sorted_paces = sorted(paces)
        pace_p10 = sorted_paces[int(len(sorted_paces) * 0.10)]
        pace_median = sorted_paces[len(sorted_paces) // 2]
    
    races = []
    
    for r in runs:
        if not r.get("pace_min_km") or not r.get("avg_hr") or not r.get("distance_km"):
            continue
        
        dist = r["distance_km"]
        pace = r["pace_min_km"]
        hr = r["avg_hr"]
        dur = r.get("duration_min", 0)
        
        # Heuristics for race detection:
        is_race = False
        race_distance = None
        race_type = None
        confidence = "low"
        
        # 1. Parkrun: ~5K distance, fast pace
        if 4.8 <= dist <= 5.5 and pace < pace_p10 * 1.08:
            is_race = True
            race_distance = 5.0
            race_type = "parkrun"
            confidence = "high" if pace < pace_p10 * 1.03 else "medium"
        
        # 2. Standard race distances at race pace
        race_distances = [
            (2.8, 3.2, 3.0, "3K"),
            (4.8, 5.5, 5.0, "5K"),
            (9.5, 10.8, 10.0, "10K"),
            (14.5, 16.5, 15.0, "15K"),
            (20.5, 22.0, 21.097, "HM"),
            (25.0, 27.5, 26.2, "26K"),  # rare
            (29.5, 31.5, 30.0, "30K"),
            (41.0, 43.5, 42.195, "Marathon"),
        ]
        
        for dmin, dmax, official, label in race_distances:
            if dmin <= dist <= dmax:
                # Must be faster than typical training pace for this duration
                # And relatively even pace (races are more consistent)
                if pace < pace_median * 0.92:
                    is_race = True
                    race_distance = official
                    race_type = label
                    confidence = "medium"
                    break
        
        # 3. HR-based confirmation: avg HR > 88% of max is likely racing
        max_hrs_list = [r2["max_hr"] for r2 in runs if r2.get("max_hr") and r2["max_hr"] > 100]
        if max_hrs_list:
            est_max = sorted(max_hrs_list)[-1]
            if hr > est_max * 0.88 and is_race:
                confidence = "high"
        
        if is_race:
            races.append({
                "file": r["file"],
                "date": r["date"],
                "distance_km": race_distance,
                "actual_distance_km": round(dist, 3),
                "duration_min": r["duration_min"],
                "pace_min_km": r["pace_min_km"],
                "avg_hr": r["avg_hr"],
                "race_type": race_type,
                "confidence": confidence,
            })
    
    return races


def detect_power_meter(runs: list[dict]) -> dict:
    """Check if athlete has a power meter."""
    with_power = [r for r in runs if r.get("avg_power") and r["avg_power"] > 50]
    n_total = len(runs)
    n_power = len(with_power)
    
    if n_power == 0:
        return {"has_power": False, "mode": "gap", "note": "No power data found → GAP mode"}
    
    pct = n_power / n_total * 100 if n_total > 0 else 0
    
    # Check if it's likely Stryd (consistent power) vs Garmin (erratic)
    powers = [r["avg_power"] for r in with_power]
    if HAS_NUMPY and len(powers) > 10:
        cv = np.std(powers) / np.mean(powers)
        if cv > 0.4:  # Very high variability = likely Garmin wrist estimate
            return {
                "has_power": True, "mode": "gap",
                "n_power": n_power, "pct": round(pct, 1),
                "note": f"Power data in {n_power}/{n_total} runs but high variability "
                        f"(CV={cv:.2f}) suggests Garmin wrist estimate → GAP mode recommended"
            }
    
    if pct > 50:
        return {
            "has_power": True, "mode": "stryd",
            "n_power": n_power, "pct": round(pct, 1),
            "note": f"Power data in {n_power}/{n_total} runs ({pct:.0f}%) → Stryd mode"
        }
    else:
        return {
            "has_power": True, "mode": "gap",
            "n_power": n_power, "pct": round(pct, 1),
            "note": f"Power data in only {n_power}/{n_total} runs ({pct:.0f}%) → GAP mode recommended"
        }


# ═════════════════════════════════════════════════════════════════════════════
# FILE GENERATION
# ═════════════════════════════════════════════════════════════════════════════

def make_folder_name(name: str) -> str:
    """Convert 'Ian Lilley' → 'IanLilley'."""
    return re.sub(r"[^a-zA-Z0-9]", "", name.title().replace(" ", ""))


def make_slug(name: str) -> str:
    """Convert 'Ian Lilley' → 'ian' (for workflow files, pages paths, cache keys)."""
    parts = name.strip().lower().split()
    return parts[0] if parts else "athlete"


def generate_athlete_yml(cfg: dict) -> str:
    """Generate athlete.yml content."""
    
    # Planned races section
    if cfg.get("planned_races"):
        races_yaml = "\n".join([
            f'  - name: "{r["name"]}"\n'
            f'    date: "{r["date"]}"\n'
            f'    distance_km: {r["distance_km"]}\n'
            f'    priority: {r.get("priority", "B")}\n'
            f'    surface: {r.get("surface", "road")}'
            for r in cfg["planned_races"]
        ])
    else:
        races_yaml = "[]"
    
    mode = cfg.get("power_mode", "gap")
    peak_cp = cfg.get("peak_cp_watts", 300)
    
    # Intervals section
    intervals_section = ""
    if cfg.get("data_source") == "intervals":
        intervals_section = f"""  intervals:
    athlete_id: ""        # Set via environment variable INTERVALS_ATHLETE_ID
    api_key: ""           # Set via environment variable {{SECRET_PREFIX}}_INTERVALS_API_KEY"""
    else:
        intervals_section = """  intervals:
    athlete_id: ""
    api_key: \"\""""
    
    source = cfg.get("data_source", "fit_folder")
    
    yml = f"""# =============================================================================
# ATHLETE CONFIGURATION — {cfg['name']} ({mode.upper()} Mode)
# =============================================================================
# Generated by onboard_athlete.py on {datetime.now().strftime('%Y-%m-%d %H:%M')}

athlete:
  name: "{cfg['name']}"
  mass_kg: {cfg['mass_kg']}
  date_of_birth: "{cfg['dob']}"
  gender: "{cfg['gender']}"
  timezone: "{cfg.get('timezone', 'Europe/London')}"

  # HR thresholds
  lthr: {cfg['lthr']}
  max_hr: {cfg['max_hr']}

planned_races:
{races_yaml}

power:
  mode: "{mode}"

  stryd:
    peak_cp_watts: {peak_cp}
    eras: {{}}
    mass_corrections: []
    re_reference_era: "s4"

  gap:
    re_constant: {cfg.get('re_constant', 0.92)}

data:
  source: "{source}"
{intervals_section}
  fit_folder:
    path: "./data/fits"
  strava_export:
    zip_path: ""

pipeline:
  rf_window_duration_s: 2400
  rf_trend_window_days: 42
  rf_trend_min_periods: 5
  ctl_time_constant: 42
  atl_time_constant: 7
  temp_baseline_c: {cfg.get('temp_baseline_c', 10.0)}
  easy_rf:
    hr_min: {max(120, int(cfg['lthr'] * 0.77))}
    hr_max: 0
"""
    return yml


def generate_workflow_yml(cfg: dict) -> str:
    """Generate GitHub Actions workflow YAML."""
    
    name = cfg["name"]
    folder = cfg["folder_name"]
    slug = cfg["slug"]
    mode = cfg.get("power_mode", "gap")
    mass = cfg["mass_kg"]
    dob = cfg["dob"]
    gender = cfg["gender"]
    tz = cfg.get("timezone", "Europe/London")
    source = cfg.get("data_source", "fit_folder")
    
    secret_prefix = slug.upper()
    
    # ── Header ────────────────────────────────────────────────
    header = f"""# .github/workflows/{slug}_pipeline.yml
# Pipeline for {name} — {mode.upper()} mode{"" if source == "fit_folder" else ", intervals.icu sync"}
#
# Triggers:
#   - {"Scheduled check 8x/day (skips quickly if no new activities)" if source == "intervals" else "Manual (workflow_dispatch) with mode selection"}{"" if source == "fit_folder" else chr(10) + "#   - Manual (workflow_dispatch) with mode selection"}
#
# Secrets required:
#   DROPBOX_TOKEN / DROPBOX_REFRESH_TOKEN / DROPBOX_APP_KEY / DROPBOX_APP_SECRET"""
    
    if source == "intervals":
        header += f"""
#   {secret_prefix}_INTERVALS_API_KEY    — {name}'s intervals.icu API key
#   {secret_prefix}_INTERVALS_ATHLETE_ID — {name}'s intervals.icu athlete ID"""
    
    # ── Triggers ──────────────────────────────────────────────
    if source == "intervals":
        triggers = f"""
on:
  schedule:
    - cron: '0 9,11,13,15,17,19,21,23 * * *'
  workflow_dispatch:
    inputs:
      mode:
        description: 'Pipeline mode'
        type: choice
        options:
          - UPDATE
          - FULL
          - FROM_STEPB
          - DASHBOARD
        default: UPDATE
      sync:
        description: 'Sync from intervals.icu'
        type: boolean
        default: true"""
    else:
        triggers = """
on:
  workflow_dispatch:
    inputs:
      mode:
        description: 'Pipeline mode'
        type: choice
        options:
          - FULL
          - FROM_STEPB
          - DASHBOARD
        default: FROM_STEPB"""
    
    # ── Env block ─────────────────────────────────────────────
    env_block = f"""
    env:
      DROPBOX_TOKEN: ${{{{ secrets.DROPBOX_TOKEN }}}}
      DROPBOX_REFRESH_TOKEN: ${{{{ secrets.DROPBOX_REFRESH_TOKEN }}}}
      DROPBOX_APP_KEY: ${{{{ secrets.DROPBOX_APP_KEY }}}}
      DROPBOX_APP_SECRET: ${{{{ secrets.DROPBOX_APP_SECRET }}}}"""
    
    if source == "intervals":
        env_block += f"""
      INTERVALS_API_KEY: ${{{{ secrets.{secret_prefix}_INTERVALS_API_KEY }}}}
      INTERVALS_ATHLETE_ID: ${{{{ secrets.{secret_prefix}_INTERVALS_ATHLETE_ID }}}}"""
    
    env_block += f"""
      ATHLETE_DIR: athletes/{folder}
      ATHLETE_CONFIG_PATH: athletes/{folder}/athlete.yml
      DB_BASE: /Running and Cycling/DataPipeline/athletes/{folder}
      TZ_LOCAL: {tz}"""
    
    if source == "intervals":
        env_block += f"""
      PIPELINE_MODE: ${{{{ github.event.inputs.mode || 'UPDATE' }}}}"""
    
    # ── Common steps ──────────────────────────────────────────
    common_steps = f"""
    steps:
      - name: Checkout repository
        uses: actions/checkout@v4

      - name: Setup Python
        uses: actions/setup-python@v5
        with:
          python-version: '3.12'
          cache: 'pip'

      - name: Install dependencies
        run: |
          pip install --upgrade pip
          pip install -r requirements.txt

      - name: Create directories
        run: |
          mkdir -p ${{{{ env.ATHLETE_DIR }}}}/data/FIT_downloads
          mkdir -p ${{{{ env.ATHLETE_DIR }}}}/persec_cache
          mkdir -p ${{{{ env.ATHLETE_DIR }}}}/output/dashboard"""
    
    # ── Intervals.icu sync block ──────────────────────────────
    if source == "intervals":
        sync_block = f"""
      # ═══════════════════════════════════════════════════════════
      # LIGHTWEIGHT CHECK — runs every time (~30s)
      # ═══════════════════════════════════════════════════════════

      - name: Restore sync state
        uses: actions/cache/restore@v4
        with:
          path: |
            ${{{{ env.ATHLETE_DIR }}}}/fit_sync_state.json
            ${{{{ env.ATHLETE_DIR }}}}/pending_activities.csv
          key: {slug}-sync-state-${{{{ github.run_number }}}}
          restore-keys: |
            {slug}-sync-state-

      - name: Download fits.zip for dedup
        if: ${{{{ github.event.inputs.sync != 'false' && env.PIPELINE_MODE != 'DASHBOARD' && env.PIPELINE_MODE != 'FROM_STEPB' }}}}
        run: |
          python ci/dropbox_sync.py download \\
            --dropbox-base "${{{{ env.DB_BASE }}}}" \\
            --local-prefix "${{{{ env.ATHLETE_DIR }}}}" \\
            --items data/fits.zip
        continue-on-error: true

      - name: Check for new activities
        id: fetch
        if: ${{{{ github.event.inputs.sync != 'false' && env.PIPELINE_MODE != 'DASHBOARD' && env.PIPELINE_MODE != 'FROM_STEPB' }}}}
        run: |
          python fetch_fit_files.py \\
            --fit-dir ${{{{ env.ATHLETE_DIR }}}}/data/FIT_downloads \\
            --zip ${{{{ env.ATHLETE_DIR }}}}/data/fits.zip \\
            --state-file ${{{{ env.ATHLETE_DIR }}}}/fit_sync_state.json \\
            --pending-csv ${{{{ env.ATHLETE_DIR }}}}/pending_activities.csv \\
            --refresh-names 14

          if [ -f _new_runs_added.tmp ]; then
            echo "new_runs=true" >> $GITHUB_OUTPUT
            COUNT=$(cat _new_runs_added.tmp)
            echo "  → $COUNT new run(s) detected"
          else
            echo "new_runs=false" >> $GITHUB_OUTPUT
            echo "  → No new runs"
          fi
        continue-on-error: true

      - name: Decide whether to continue
        id: gate
        run: |
          if [ "${{{{ github.event_name }}}}" = "workflow_dispatch" ]; then
            echo "continue=true" >> $GITHUB_OUTPUT
          elif [ "${{{{ steps.fetch.outputs.new_runs }}}}" = "true" ]; then
            echo "continue=true" >> $GITHUB_OUTPUT
          else
            echo "continue=false" >> $GITHUB_OUTPUT
          fi

      - name: Save sync state
        uses: actions/cache/save@v4
        if: always()
        with:
          path: |
            ${{{{ env.ATHLETE_DIR }}}}/fit_sync_state.json
            ${{{{ env.ATHLETE_DIR }}}}/pending_activities.csv
          key: {slug}-sync-state-${{{{ github.run_number }}}}"""
        gate_condition = "${{ steps.gate.outputs.continue == 'true' }}"
    else:
        sync_block = ""
        gate_condition = None  # Not used for FIT-only
    
    # ── Pipeline steps ────────────────────────────────────────
    gc = gate_condition
    
    if source == "intervals":
        pipeline_block = f"""
      # ═══════════════════════════════════════════════════════════
      # FULL PIPELINE — only if new data or manual dispatch
      # ═══════════════════════════════════════════════════════════

      - name: Restore persec cache
        if: {gc}
        uses: actions/cache/restore@v4
        with:
          path: ${{{{ env.ATHLETE_DIR }}}}/persec_cache
          key: {slug}-persec-${{{{ github.run_number }}}}
          restore-keys: |
            {slug}-persec-

      - name: Download data from Dropbox
        if: {gc}
        run: |
          python ci/dropbox_sync.py download \\
            --dropbox-base "${{{{ env.DB_BASE }}}}" \\
            --local-prefix "${{{{ env.ATHLETE_DIR }}}}" \\
            --items data/activities.csv \\
                    activity_overrides.xlsx \\
                    athlete_data.csv \\
                    output/Master_FULL.xlsx \\
                    output/Master_FULL_post.xlsx \\
                    output/_weather_cache_openmeteo/openmeteo_cache.sqlite \\
            --cache-dir ${{{{ env.ATHLETE_DIR }}}}/persec_cache
        continue-on-error: true

      - name: Sync athlete data from intervals.icu
        if: ${{{{ {gc.strip('${{ }}')} && github.event.inputs.sync != 'false' && env.PIPELINE_MODE != 'DASHBOARD' }}}}
        run: |
          python sync_athlete_data.py \\
            --athlete-data ${{{{ env.ATHLETE_DIR }}}}/athlete_data.csv \\
            --nr-tss-oldest 2020-01-01
        continue-on-error: true

      - name: Rebuild from FIT files (FULL)
        if: ${{{{ {gc.strip('${{ }}')} && env.PIPELINE_MODE == 'FULL' }}}}
        run: |
          python rebuild_from_fit_zip.py \\
            --fit-zip ${{{{ env.ATHLETE_DIR }}}}/data/fits.zip \\
            --template master_template.xlsx \\
            --out ${{{{ env.ATHLETE_DIR }}}}/output/Master_FULL.xlsx \\
            --persec-cache-dir ${{{{ env.ATHLETE_DIR }}}}/persec_cache \\
            --strava ${{{{ env.ATHLETE_DIR }}}}/data/activities.csv \\
            --pending-activities ${{{{ env.ATHLETE_DIR }}}}/pending_activities.csv \\
            --override-file ${{{{ env.ATHLETE_DIR }}}}/activity_overrides.xlsx \\
            --tz ${{{{ env.TZ_LOCAL }}}} \\
            --weight {mass}

      - name: Rebuild from FIT files (UPDATE — append only)
        if: ${{{{ {gc.strip('${{ }}')} && env.PIPELINE_MODE == 'UPDATE' }}}}
        run: |
          python rebuild_from_fit_zip.py \\
            --fit-zip ${{{{ env.ATHLETE_DIR }}}}/data/fits.zip \\
            --template master_template.xlsx \\
            --out ${{{{ env.ATHLETE_DIR }}}}/output/Master_FULL.xlsx \\
            --append-master-in ${{{{ env.ATHLETE_DIR }}}}/output/Master_FULL.xlsx \\
            --persec-cache-dir ${{{{ env.ATHLETE_DIR }}}}/persec_cache \\
            --strava ${{{{ env.ATHLETE_DIR }}}}/data/activities.csv \\
            --pending-activities ${{{{ env.ATHLETE_DIR }}}}/pending_activities.csv \\
            --override-file ${{{{ env.ATHLETE_DIR }}}}/activity_overrides.xlsx \\
            --tz ${{{{ env.TZ_LOCAL }}}} \\
            --weight {mass}

      - name: Add GAP power
        if: ${{{{ {gc.strip('${{ }}')} && (env.PIPELINE_MODE == 'FULL' || env.PIPELINE_MODE == 'UPDATE') }}}}
        run: |
          python add_gap_power.py \\
            --master ${{{{ env.ATHLETE_DIR }}}}/output/Master_FULL.xlsx \\
            --cache-dir ${{{{ env.ATHLETE_DIR }}}}/persec_cache \\
            --out ${{{{ env.ATHLETE_DIR }}}}/output/Master_FULL.xlsx \\
            --mass-kg {mass} \\
            --re-constant 0.92

      - name: StepB post-processing
        if: ${{{{ {gc.strip('${{ }}')} && env.PIPELINE_MODE != 'DASHBOARD' }}}}
        run: |
          python StepB_PostProcess.py \\
            --master ${{{{ env.ATHLETE_DIR }}}}/output/Master_FULL.xlsx \\
            --persec-cache-dir ${{{{ env.ATHLETE_DIR }}}}/persec_cache \\
            --out ${{{{ env.ATHLETE_DIR }}}}/output/Master_FULL_post.xlsx \\
            --model-json re_model_generic.json \\
            --override-file ${{{{ env.ATHLETE_DIR }}}}/activity_overrides.xlsx \\
            --athlete-data ${{{{ env.ATHLETE_DIR }}}}/athlete_data.csv \\
            --mass-kg {mass} \\
            --tz ${{{{ env.TZ_LOCAL }}}} \\
            --runner-dob "{dob}" \\
            --runner-gender {gender} \\
            --strava ${{{{ env.ATHLETE_DIR }}}}/data/activities.csv \\
            --progress-every 50

      - name: Generate dashboard
        if: {gc}
        env:
          MASTER_FILE: ${{{{ env.ATHLETE_DIR }}}}/output/Master_FULL_post.xlsx
          OUTPUT_FILE: ${{{{ env.ATHLETE_DIR }}}}/output/dashboard/index.html
          ATHLETE_DATA_FILE: ${{{{ env.ATHLETE_DIR }}}}/athlete_data.csv
        run: python generate_dashboard.py
        continue-on-error: true

      - name: Save persec cache
        uses: actions/cache/save@v4
        if: {gc}
        with:
          path: ${{{{ env.ATHLETE_DIR }}}}/persec_cache
          key: {slug}-persec-${{{{ github.run_number }}}}

      - name: Upload fits.zip to Dropbox (only if new FITs added)
        if: ${{{{ {gc.strip('${{ }}')} && steps.fetch.outputs.new_runs == 'true' }}}}
        run: |
          python ci/dropbox_sync.py upload \\
            --dropbox-base "${{{{ env.DB_BASE }}}}" \\
            --local-prefix "${{{{ env.ATHLETE_DIR }}}}" \\
            --items data/fits.zip
        continue-on-error: true

      - name: Upload results to Dropbox
        if: {gc}
        run: |
          python ci/dropbox_sync.py upload \\
            --dropbox-base "${{{{ env.DB_BASE }}}}" \\
            --local-prefix "${{{{ env.ATHLETE_DIR }}}}" \\
            --items activity_overrides.xlsx \\
                    athlete_data.csv \\
                    fit_sync_state.json \\
                    pending_activities.csv \\
                    output/Master_FULL.xlsx \\
                    output/Master_FULL_post.xlsx \\
                    output/dashboard/index.html \\
                    output/_weather_cache_openmeteo/openmeteo_cache.sqlite \\
            --cache-dir ${{{{ env.ATHLETE_DIR }}}}/persec_cache"""
    else:
        # FIT-folder mode: simpler, no gating
        pipeline_block = f"""
      # ─── Restore persec cache ───────────────────────────────
      - name: Restore persec cache
        uses: actions/cache/restore@v4
        with:
          path: ${{{{ env.ATHLETE_DIR }}}}/persec_cache
          key: {slug}-persec-${{{{ github.run_number }}}}
          restore-keys: |
            {slug}-persec-

      # ─── Download data from Dropbox ─────────────────────────
      - name: Download data from Dropbox
        run: |
          python ci/dropbox_sync.py download \\
            --dropbox-base "${{{{ env.DB_BASE }}}}" \\
            --local-prefix "${{{{ env.ATHLETE_DIR }}}}" \\
            --items data/fits.zip \\
                    data/activities.csv \\
                    activity_overrides.xlsx \\
                    athlete_data.csv \\
                    output/Master_FULL.xlsx \\
                    output/Master_FULL_post.xlsx \\
                    output/_weather_cache_openmeteo/openmeteo_cache.sqlite \\
            --cache-dir ${{{{ env.ATHLETE_DIR }}}}/persec_cache
        continue-on-error: true

      # ─── Step 1: Rebuild from FIT files ─────────────────────
      - name: Rebuild from FIT files
        if: ${{{{ github.event.inputs.mode == 'FULL' }}}}
        run: |
          python rebuild_from_fit_zip.py \\
            --fit-zip ${{{{ env.ATHLETE_DIR }}}}/data/fits.zip \\
            --template master_template.xlsx \\
            --out ${{{{ env.ATHLETE_DIR }}}}/output/Master_FULL.xlsx \\
            --persec-cache-dir ${{{{ env.ATHLETE_DIR }}}}/persec_cache \\
            --strava ${{{{ env.ATHLETE_DIR }}}}/data/activities.csv \\
            --override-file ${{{{ env.ATHLETE_DIR }}}}/activity_overrides.xlsx \\
            --tz ${{{{ env.TZ_LOCAL }}}} \\
            --weight {mass}

      # ─── Step 2: Add GAP simulated power ────────────────────
      - name: Add GAP power
        if: ${{{{ github.event.inputs.mode == 'FULL' }}}}
        run: |
          python add_gap_power.py \\
            --master ${{{{ env.ATHLETE_DIR }}}}/output/Master_FULL.xlsx \\
            --cache-dir ${{{{ env.ATHLETE_DIR }}}}/persec_cache \\
            --out ${{{{ env.ATHLETE_DIR }}}}/output/Master_FULL.xlsx \\
            --mass-kg {mass} \\
            --re-constant 0.92

      # ─── Step 3: StepB post-processing ──────────────────────
      - name: StepB post-processing
        if: ${{{{ github.event.inputs.mode != 'DASHBOARD' }}}}
        run: |
          python StepB_PostProcess.py \\
            --master ${{{{ env.ATHLETE_DIR }}}}/output/Master_FULL.xlsx \\
            --persec-cache-dir ${{{{ env.ATHLETE_DIR }}}}/persec_cache \\
            --out ${{{{ env.ATHLETE_DIR }}}}/output/Master_FULL_post.xlsx \\
            --model-json re_model_generic.json \\
            --override-file ${{{{ env.ATHLETE_DIR }}}}/activity_overrides.xlsx \\
            --athlete-data ${{{{ env.ATHLETE_DIR }}}}/athlete_data.csv \\
            --mass-kg {mass} \\
            --tz ${{{{ env.TZ_LOCAL }}}} \\
            --runner-dob "{dob}" \\
            --runner-gender {gender} \\
            --strava ${{{{ env.ATHLETE_DIR }}}}/data/activities.csv \\
            --progress-every 50

      # ─── Step 4: Generate dashboard ─────────────────────────
      - name: Generate dashboard
        env:
          MASTER_FILE: ${{{{ env.ATHLETE_DIR }}}}/output/Master_FULL_post.xlsx
          OUTPUT_FILE: ${{{{ env.ATHLETE_DIR }}}}/output/dashboard/index.html
          ATHLETE_DATA_FILE: ${{{{ env.ATHLETE_DIR }}}}/athlete_data.csv
        run: python generate_dashboard.py
        continue-on-error: true

      # ─── Save persec cache ─────────────────────────────────
      - name: Save persec cache
        uses: actions/cache/save@v4
        if: always()
        with:
          path: ${{{{ env.ATHLETE_DIR }}}}/persec_cache
          key: {slug}-persec-${{{{ github.run_number }}}}

      # ─── Upload results to Dropbox ──────────────────────────
      - name: Upload results to Dropbox
        if: always()
        run: |
          python ci/dropbox_sync.py upload \\
            --dropbox-base "${{{{ env.DB_BASE }}}}" \\
            --local-prefix "${{{{ env.ATHLETE_DIR }}}}" \\
            --items activity_overrides.xlsx \\
                    athlete_data.csv \\
                    output/Master_FULL.xlsx \\
                    output/Master_FULL_post.xlsx \\
                    output/dashboard/index.html \\
                    output/_weather_cache_openmeteo/openmeteo_cache.sqlite \\
            --cache-dir ${{{{ env.ATHLETE_DIR }}}}/persec_cache"""
    
    # ── Pages deployment + summary ────────────────────────────
    pages_condition = gc if source == "intervals" else "success()"
    
    tail_block = f"""

      # ─── Deploy dashboard to GitHub Pages ──────────────────
      - name: Prepare dashboard for Pages
        {"if: " + gc if source == "intervals" else ""}
        run: |
          mkdir -p docs/{slug}
          cp -f ${{{{ env.ATHLETE_DIR }}}}/output/dashboard/index.html docs/{slug}/index.html 2>/dev/null || true
        continue-on-error: true

      - name: Deploy to GitHub Pages
        uses: peaceiris/actions-gh-pages@v3
        if: {pages_condition}
        with:
          github_token: ${{{{ secrets.GITHUB_TOKEN }}}}
          publish_dir: ./docs
          publish_branch: gh-pages
          keep_files: true
        continue-on-error: true

      # ─── Summary ────────────────────────────────────────────
      - name: Summary
        if: always()
        run: |
          echo "## {name} Pipeline Run Summary" >> $GITHUB_STEP_SUMMARY
          echo "" >> $GITHUB_STEP_SUMMARY
          echo "- **Trigger:** ${{{{ github.event_name }}}}" >> $GITHUB_STEP_SUMMARY
          echo "- **Mode:** ${{{{ github.event.inputs.mode || 'UPDATE' }}}}" >> $GITHUB_STEP_SUMMARY
          echo "- **Status:** ${{{{ job.status }}}}" >> $GITHUB_STEP_SUMMARY
          echo "- **Time:** $(date -u +%Y-%m-%dT%H:%M:%SZ)" >> $GITHUB_STEP_SUMMARY
          
          MASTER="${{{{ env.ATHLETE_DIR }}}}/output/Master_FULL_post.xlsx"
          if [ -f "$MASTER" ]; then
            SIZE=$(stat -c%s "$MASTER" 2>/dev/null || echo "?")
            echo "- **Master size:** $SIZE bytes" >> $GITHUB_STEP_SUMMARY
          fi
          
          CACHE="${{{{ env.ATHLETE_DIR }}}}/persec_cache"
          if [ -d "$CACHE" ]; then
            COUNT=$(find "$CACHE" -name "*.npz" | wc -l)
            echo "- **NPZ cache:** $COUNT files" >> $GITHUB_STEP_SUMMARY
          fi"""
    
    # ── Assemble ──────────────────────────────────────────────
    workflow = f"""{header}

name: {name} Pipeline
{triggers}

concurrency:
  group: {slug}-pipeline
  cancel-in-progress: false

jobs:
  pipeline:
    runs-on: ubuntu-latest
    timeout-minutes: 240
{env_block}
{common_steps}
{sync_block}
{pipeline_block}
{tail_block}
"""
    return workflow


def generate_override_xlsx(path: str, races: list[dict] = None):
    """Generate activity_overrides.xlsx with detected races pre-populated."""
    if not HAS_OPENPYXL:
        print(f"  ⚠ openpyxl not installed — creating empty override placeholder")
        # Write a minimal file marker
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        Path(path).touch()
        return
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "overrides"
    
    headers = [
        "file", "race", "parkrun", "distance_km", "surface", "surface_adj",
        "temp_c", "power_override_w", "notes"
    ]
    ws.append(headers)
    
    # Style header
    from openpyxl.styles import Font, PatternFill
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Add detected races
    if races:
        for r in races:
            is_parkrun = r.get("race_type") == "parkrun"
            ws.append([
                r["file"],
                True,
                is_parkrun,
                r["distance_km"],
                None,  # surface
                None,  # surface_adj
                None,  # temp_c
                None,  # power_override_w
                f"Auto-detected {r['race_type']} ({r['confidence']} confidence) — "
                f"{r['duration_min']:.1f}min, {r['pace_min_km']:.2f}/km, HR {r.get('avg_hr', '?')}"
            ])
    
    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)
    
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def generate_athlete_data_csv(path: str):
    """Generate empty athlete_data.csv with correct headers."""
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        f.write("date,weight_kg,resting_hr,hrv,sleep_hours,notes\n")


# ═════════════════════════════════════════════════════════════════════════════
# INTERACTIVE MODE
# ═════════════════════════════════════════════════════════════════════════════

def interactive_config() -> dict:
    """Collect athlete info interactively."""
    print("\n╔══════════════════════════════════════════╗")
    print("║   Athlete Onboarding — Running Pipeline  ║")
    print("╚══════════════════════════════════════════╝\n")
    
    cfg = {}
    
    # Basic info
    cfg["name"] = input("  Athlete name: ").strip()
    if not cfg["name"]:
        print("  ✗ Name is required")
        sys.exit(1)
    
    cfg["dob"] = input("  Date of birth (YYYY-MM-DD): ").strip()
    cfg["gender"] = input("  Gender (male/female): ").strip().lower()
    
    mass_str = input("  Weight in kg: ").strip()
    cfg["mass_kg"] = float(mass_str) if mass_str else 70.0
    
    # HR — can be auto-estimated if FIT scan is used
    print("\n  HR thresholds (leave blank to auto-estimate from FIT files):")
    lthr_str = input("    LTHR (lactate threshold HR): ").strip()
    max_hr_str = input("    Max HR: ").strip()
    cfg["lthr"] = int(lthr_str) if lthr_str else None
    cfg["max_hr"] = int(max_hr_str) if max_hr_str else None
    
    # Timezone
    tz = input("\n  Timezone [Europe/London]: ").strip()
    cfg["timezone"] = tz if tz else "Europe/London"
    
    # Data source
    print("\n  Data source:")
    print("    1. FIT files in a folder/zip (simplest)")
    print("    2. intervals.icu (automated sync)")
    src = input("  Choice [1]: ").strip()
    cfg["data_source"] = "intervals" if src == "2" else "fit_folder"
    
    # FIT scan
    fit_path = input("\n  Path to FIT files (folder or .zip) for scanning [skip]: ").strip()
    cfg["fit_scan_path"] = fit_path if fit_path else None
    
    return cfg


# ═════════════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Onboard a new athlete")
    parser.add_argument("--config", help="JSON config file (skip interactive)")
    parser.add_argument("--scan-fits", help="Path to FIT files to scan (folder or .zip)")
    parser.add_argument("--output-dir", default=".", help="Base directory for output")
    parser.add_argument("--dry-run", action="store_true", help="Print what would be created")
    args = parser.parse_args()
    
    # ── Gather config ─────────────────────────────────────────
    if args.config:
        with open(args.config) as f:
            cfg = json.load(f)
    else:
        cfg = interactive_config()
    
    # Override scan path from CLI
    if args.scan_fits:
        cfg["fit_scan_path"] = args.scan_fits
    
    # Derived names
    cfg["folder_name"] = make_folder_name(cfg["name"])
    cfg["slug"] = make_slug(cfg["name"])
    
    base = Path(args.output_dir)
    athlete_dir = base / "athletes" / cfg["folder_name"]
    workflow_dir = base / ".github" / "workflows"
    
    # ── FIT scanning ──────────────────────────────────────────
    scan_results = None
    detected_races = []
    
    if cfg.get("fit_scan_path"):
        if not HAS_FITPARSE:
            print("\n⚠ fitparse not installed. Install with: pip install fitparse")
            print("  Skipping FIT scan — you'll need to set HR zones manually.\n")
        else:
            print(f"\n📁 Scanning FIT files...")
            runs = scan_fit_folder(cfg["fit_scan_path"])
            
            if runs:
                # Date range
                dates = [r["date"] for r in runs if r.get("date")]
                print(f"\n  📊 Summary:")
                print(f"     Runs: {len(runs)}")
                print(f"     Period: {min(dates)} → {max(dates)}")
                
                # HR estimation
                hr_info = estimate_hr_zones(runs)
                print(f"\n  ❤️  HR estimation:")
                print(f"     Max HR: {hr_info['max_hr']}")
                print(f"     LTHR: {hr_info['lthr']} ({hr_info['confidence']} confidence)")
                print(f"     {hr_info['note']}")
                
                if cfg.get("lthr") is None and hr_info["lthr"]:
                    cfg["lthr"] = hr_info["lthr"]
                if cfg.get("max_hr") is None and hr_info["max_hr"]:
                    cfg["max_hr"] = hr_info["max_hr"]
                
                # Power meter detection
                power_info = detect_power_meter(runs)
                print(f"\n  ⚡ Power: {power_info['note']}")
                cfg["power_mode"] = power_info["mode"]
                
                # Race detection
                detected_races = detect_races(runs)
                if detected_races:
                    high_conf = [r for r in detected_races if r["confidence"] == "high"]
                    med_conf = [r for r in detected_races if r["confidence"] == "medium"]
                    low_conf = [r for r in detected_races if r["confidence"] == "low"]
                    print(f"\n  🏁 Detected {len(detected_races)} potential races:")
                    print(f"     High confidence: {len(high_conf)}")
                    print(f"     Medium confidence: {len(med_conf)}")
                    print(f"     Low confidence: {len(low_conf)}")
                    
                    # Show top 10
                    for r in detected_races[:10]:
                        conf_icon = {"high": "✓", "medium": "~", "low": "?"}[r["confidence"]]
                        print(f"     {conf_icon} {r['date']}  {r['race_type']:>8}  "
                              f"{r['duration_min']:6.1f}min  {r['pace_min_km']:.2f}/km  "
                              f"HR {r.get('avg_hr', '?')}")
                    if len(detected_races) > 10:
                        print(f"     ... and {len(detected_races) - 10} more")
                
                scan_results = {
                    "n_runs": len(runs),
                    "date_range": (min(dates), max(dates)),
                    "hr_info": hr_info,
                    "power_info": power_info,
                    "n_races": len(detected_races),
                }
    
    # ── Defaults for missing values ───────────────────────────
    if cfg.get("lthr") is None:
        cfg["lthr"] = 160  # generic default
        print("  ⚠ Using default LTHR=160. Update in athlete.yml once known.")
    if cfg.get("max_hr") is None:
        cfg["max_hr"] = 185  # generic default
        print("  ⚠ Using default max_hr=185. Update in athlete.yml once known.")
    if cfg.get("power_mode") is None:
        cfg["power_mode"] = "gap"
    
    # ── Generate files ────────────────────────────────────────
    print(f"\n{'═' * 60}")
    print(f"  Generating files for {cfg['name']}")
    print(f"{'═' * 60}\n")
    
    files_to_create = {
        str(athlete_dir / "athlete.yml"): generate_athlete_yml(cfg),
        str(workflow_dir / f"{cfg['slug']}_pipeline.yml"): generate_workflow_yml(cfg),
    }
    
    if args.dry_run:
        for path, content in files_to_create.items():
            print(f"  Would create: {path}")
            print(f"    ({len(content)} chars)")
        print(f"  Would create: {athlete_dir}/activity_overrides.xlsx")
        print(f"    ({len(detected_races)} races pre-populated)")
        print(f"  Would create: {athlete_dir}/athlete_data.csv")
        return
    
    # Create directories
    athlete_dir.mkdir(parents=True, exist_ok=True)
    (athlete_dir / "data").mkdir(exist_ok=True)
    (athlete_dir / "persec_cache").mkdir(exist_ok=True)
    (athlete_dir / "output" / "dashboard").mkdir(parents=True, exist_ok=True)
    workflow_dir.mkdir(parents=True, exist_ok=True)
    
    # Write text files
    for path, content in files_to_create.items():
        with open(path, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"  ✓ {path}")
    
    # Write override xlsx
    override_path = str(athlete_dir / "activity_overrides.xlsx")
    generate_override_xlsx(override_path, detected_races)
    print(f"  ✓ {override_path}")
    if detected_races:
        print(f"    → {len(detected_races)} races pre-populated (review & edit before running)")
    
    # Write athlete_data.csv
    data_path = str(athlete_dir / "athlete_data.csv")
    generate_athlete_data_csv(data_path)
    print(f"  ✓ {data_path}")
    
    # ── Post-creation instructions ────────────────────────────
    slug = cfg["slug"]
    folder = cfg["folder_name"]
    secret_prefix = slug.upper()
    
    print(f"\n{'─' * 60}")
    print(f"  SETUP CHECKLIST")
    print(f"{'─' * 60}\n")
    
    print("  1. REVIEW athlete.yml")
    print(f"     → {athlete_dir}/athlete.yml")
    print(f"     → Check LTHR ({cfg['lthr']}), max HR ({cfg['max_hr']}), weight ({cfg['mass_kg']}kg)")
    
    if detected_races:
        print(f"\n  2. REVIEW activity_overrides.xlsx")
        print(f"     → {override_path}")
        print(f"     → {len(detected_races)} auto-detected races — verify and correct")
        print(f"     → Remove false positives, add any missed races")
    
    print(f"\n  3. DROPBOX — create folder structure:")
    print(f"     /Running and Cycling/DataPipeline/athletes/{folder}/")
    print(f"       ├── data/fits.zip          ← athlete's FIT files")
    print(f"       ├── data/activities.csv    ← empty or from Strava export")
    print(f"       ├── activity_overrides.xlsx")
    print(f"       ├── athlete_data.csv")
    print(f"       └── athlete.yml")
    
    if cfg.get("data_source") == "intervals":
        print(f"\n  4. GITHUB SECRETS — add to repository:")
        print(f"     {secret_prefix}_INTERVALS_API_KEY    = <intervals.icu API key>")
        print(f"     {secret_prefix}_INTERVALS_ATHLETE_ID = <intervals.icu athlete ID>")
    
    print(f"\n  5. FIT FILES — get athlete's data into fits.zip:")
    print(f"     Option A: Garmin Connect → export all activities → zip the FITs")
    print(f"     Option B: Strava → Download your data → use the activities/ folder")
    print(f"     Option C: intervals.icu → bulk export (if synced)")
    
    print(f"\n  6. FIRST RUN — trigger FULL pipeline:")
    print(f"     GitHub Actions → {cfg['name']} Pipeline → Run workflow → mode: FULL")
    
    print(f"\n  7. DASHBOARD — will be at:")
    print(f"     https://dobssi.github.io/Fitness-Data/{slug}/")
    
    # Save config for reference
    config_path = str(athlete_dir / "onboard_config.json")
    with open(config_path, "w", encoding="utf-8") as f:
        save_cfg = {k: v for k, v in cfg.items() if k != "fit_scan_path"}
        if scan_results:
            save_cfg["scan_results"] = {
                "n_runs": scan_results["n_runs"],
                "date_range": scan_results["date_range"],
                "n_races": scan_results["n_races"],
            }
        json.dump(save_cfg, f, indent=2, default=str)
    print(f"\n  💾 Config saved: {config_path}")
    
    print(f"\n{'═' * 60}")
    print(f"  Done! Review the files, set up Dropbox, and trigger the pipeline.")
    print(f"{'═' * 60}\n")


if __name__ == "__main__":
    main()
